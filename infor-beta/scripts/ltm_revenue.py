"""Build a standalone LTM revenue breakdown workbook for the earnings update.

The earnings-update overview slide (library slide 7) reserves the lower-left
quadrant for an LTM revenue pie. The chart itself is deferred, but the
companion workbook captures the underlying breakdown so the analyst can build
or refresh the chart. Segmentation is by service / product line when the
filing discloses it, otherwise by geography (or any other available basis).

Arithmetic lives in cell formulas (% of total, the total row) so the workbook
stays analyst-auditable, matching the cap-table convention.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# INFOR mid-blue header fill / Palatino body, mirroring the deck brand.
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(name="Palatino Linotype", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Palatino Linotype", size=14, bold=True, color="1F3864")
_META_FONT = Font(name="Palatino Linotype", size=10, italic=True, color="595959")
_BODY_FONT = Font(name="Palatino Linotype", size=11)
_TOTAL_FONT = Font(name="Palatino Linotype", size=11, bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@dataclass(frozen=True)
class RevenueSegment:
    """One row of the LTM revenue breakdown."""

    name: str
    ltm_revenue: float


def _safe_name(value: str) -> str:
    safe = re.sub(r"[/\\:*?\"<>|]+", "-", value).strip()
    safe = re.sub(r"\s+", " ", safe)
    return safe or "Company"


def build_ltm_revenue_workbook(
    *,
    company_name: str,
    period_label: str,
    currency: str,
    segmentation_basis: str,
    segments: list[RevenueSegment] | list[tuple[str, float]],
    output_dir: Path | str,
    file_stem: str | None = None,
) -> Path:
    """Write an LTM revenue breakdown .xlsx and return its path.

    `segments` may be RevenueSegment objects or (name, value) tuples.
    `segmentation_basis` is a human label such as "Service line" or "Geography".
    """
    rows = [
        seg if isinstance(seg, RevenueSegment) else RevenueSegment(seg[0], float(seg[1]))
        for seg in segments
    ]
    if not rows:
        raise ValueError("LTM revenue workbook requires at least one segment")

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = file_stem or f"{_safe_name(company_name)} - LTM Revenue Breakdown"
    output_path = out_dir / f"{stem}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "LTM Revenue"
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"{company_name} — LTM Revenue Breakdown"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Period: {period_label}"
    ws["A3"] = f"Segmentation: {segmentation_basis}"
    ws["A4"] = f"Figures in {currency}"
    for cell in ("A2", "A3", "A4"):
        ws[cell].font = _META_FONT

    header_row = 6
    headers = ["Segment", f"LTM Revenue ({currency})", "% of Total"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    first_data = header_row + 1
    for i, seg in enumerate(rows):
        r = first_data + i
        ws.cell(row=r, column=1, value=seg.name).font = _BODY_FONT
        v = ws.cell(row=r, column=2, value=seg.ltm_revenue)
        v.font = _BODY_FONT
        v.number_format = "#,##0.0"
        ws.cell(row=r, column=1).border = _BORDER
        v.border = _BORDER

    last_data = first_data + len(rows) - 1
    total_row = last_data + 1
    total_cell_ref = f"B{total_row}"

    # % of total formulas reference the total cell so the workbook recomputes.
    for i in range(len(rows)):
        r = first_data + i
        p = ws.cell(row=r, column=3, value=f"=B{r}/{total_cell_ref}")
        p.font = _BODY_FONT
        p.number_format = "0.0%"
        p.border = _BORDER
        p.alignment = Alignment(horizontal="center")

    ws.cell(row=total_row, column=1, value="Total").font = _TOTAL_FONT
    tv = ws.cell(row=total_row, column=2, value=f"=SUM(B{first_data}:B{last_data})")
    tv.font = _TOTAL_FONT
    tv.number_format = "#,##0.0"
    tp = ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data}:C{last_data})")
    tp.font = _TOTAL_FONT
    tp.number_format = "0.0%"
    tp.alignment = Alignment(horizontal="center")
    for col in (1, 2, 3):
        ws.cell(row=total_row, column=col).border = _BORDER

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14

    wb.save(output_path)
    return output_path
