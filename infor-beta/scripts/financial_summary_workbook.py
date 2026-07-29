"""Write the deal workbook's `financial-summary` data tab for the pitch deck.

The tab holds the data behind the deck's Financial Summary
slide(s) (slide-library entry 8): the **metrics the `financial-summary` skill selected**
— four per deck slide, so 4 by default or 8 when the deck spec asks for two Financial
Summary slides — each with its **last 5 fiscal years** plus an **LTM** column. The
layout is deliberately *chart-ready* so a later step can drop native Excel charts on
top with no reshaping.

Chart-ready layout contract (rely on this from the future chart step)::

    A1 : "<company> — Financial Summary"                         (title)
    A2 : "<currency note>"  e.g. "Figures in US$MM unless noted"  (meta)
    A3 : "<period note>"    e.g. "FY = fiscal year; LTM = trailing twelve months as of Q3 2026"
    row 4: (blank spacer)
    row 5 (HEADER — a single contiguous period axis):
        A5 = "Metric"   B5..F5 = the 5 fiscal-year labels, oldest -> newest
        G5 = "LTM"      (only when the LTM column is shown — see suppression below)
        <last col> = "Units"
    rows 6..5+N (DATA — one metric per row; N = 4 per deck slide, so rows 6-9
        for the default four metrics, rows 6-13 for eight):
        A = metric label (identical to the deck tile label)
        B..F = five NUMERIC fiscal-year values, chronological
        LTM cell:
            * flow metric  -> =INDEX('ltm-metrics'!$B:$B, MATCH("(=) <result_label>", 'ltm-metrics'!$A:$A, 0))
              (a label-keyed link into the `ltm-metrics` tab of the same deal
              workbook — internal, and live as soon as that tab is written)
            * non-flow metric (balance / ratio that has no LTM bridge) -> the latest
              reported value, written as a literal number (point-in-time fallback)
        <last col> = the metric's units string (e.g. "US$MM", "%")

LTM suppression: when the most recent filing is a fiscal year-end 10-K with no later
interim 10-Q stub, LTM == the latest fiscal year. The caller passes ``show_ltm=False``;
the LTM column is dropped and, for flow metrics, the **most-recent fiscal-year cell**
carries the `ltm-metrics` link instead (req 5).

The data block (B5 down through the last metric row, across every period column) carries
**no merged cells** and **numeric** value cells, so the chart step can select the header
row + one metric row and chart it directly. Arithmetic / LTM derivation lives on the
`ltm-metrics` tab and is *linked* here — Excel does the math, not the LLM.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from comment_citations import cite_cell
from deal_workbook import TAB_FINANCIAL_SUMMARY, TAB_LTM_METRICS, TabSpec, write_tab
from ltm_metrics import ltm_total_formula
from provenance import FigureSource, ProvenanceError, ProvenanceLedger

# No template_layout anchors here: this workbook is authored from scratch (no
# shipped template to shift), and the chart step re-derives its geometry by
# label (financial_charts.period_axis_columns / metric_data_rows on the row-5
# 'Units' header and column-A metric labels), never by stored address.

# INFOR mid-blue header fill / Palatino body, mirroring `ltm_metrics.py` so this
# tab matches the rest of the deal workbook.
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(name="Palatino Linotype", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Palatino Linotype", size=14, bold=True, color="1F3864")
_META_FONT = Font(name="Palatino Linotype", size=10, italic=True, color="595959")
_BODY_FONT = Font(name="Palatino Linotype", size=11)
_LABEL_FONT = Font(name="Palatino Linotype", size=11, bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

#: The `ltm-metrics` tab this tab's LTM links point at. Both tabs live in the
#: same deal workbook since Phase D, so the link is internal from the start.
_DEFAULT_LTM_SHEET = TAB_LTM_METRICS

_METRIC_COUNT = 4  # tiles per Financial Summary slide; the default single-slide deck shows four
# Currency value format for the metric cells (FY values, the LTM literal fallback,
# and the LTM link cell): "$#,##0.0" positive, "($#,##0.0)" negative, "--" zero.
_VALUE_FORMAT = '$#,##0.0_);($#,##0.0);"--"'


@dataclass(frozen=True)
class MetricSeries:
    """One metric row of the Financial Summary tab.

    - ``label``: the metric NAME (e.g. ``"Revenue"``). Doubles as the deck tile
      label and column A. No amounts / currency tokens.
    - ``units``: the unit string for this metric's values (e.g. ``"US$MM"``,
      ``"%"``). Constant down the row, shown in the Units column.
    - ``fiscal_values``: the five fiscal-year values, **chronological (oldest ->
      newest)**; length must match the workbook's ``fiscal_labels``. Each value
      is either a number **or** an Excel formula string beginning with ``"="``.
      A **combined** metric (one that sums two or more reported figures — e.g.
      "Ending Combined Loan & Advance Bal." = loans + advances) is passed as a
      formula of its components (``"=9000+800"``), **never pre-summed**, so the
      arithmetic lives in the cell and stays analyst-auditable.
    - ``result_label``: for a **flow** metric, the exact `ltm-metrics` bridge
      result label (e.g. ``"LTM Revenue"``) — the LTM cell links to its
      ``(=) <result_label>`` row. ``None`` for a non-flow metric.
    - ``ltm_value``: for a **non-flow** metric (``result_label is None``), the
      latest reported value used as the point-in-time "LTM" figure — a number,
      or a ``"="`` formula when that latest value is itself combined. Ignored
      when ``result_label`` is set.
    - ``sources``: per-value :class:`provenance.FigureSource` records, aligned
      with ``fiscal_values`` (same length; ``None`` entries skip that cell). Each
      names the filing, the statement and — where known — the page the figure came
      from, e.g. ``FigureSource(filing="FY2023 10-K", statement="Consolidated
      Statements of Operations", page=42)``. The record goes in the run's
      provenance ledger and the ``Source: …`` cell comment is *rendered from* it,
      so the artefact and the machine-readable record cannot disagree. A bare
      string is rejected — see `provenance.py`.
    - ``ltm_source``: source record for the LTM cell — used when the cell carries
      a literal ``ltm_value`` (non-flow metric). A flow metric's LTM cell is a
      link into the `ltm-metrics` tab, whose bridge components carry their own
      records, so it normally needs no source here.
    """

    label: str
    units: str
    fiscal_values: list[float | str] = field(default_factory=list)
    result_label: str | None = None
    ltm_value: float | str | None = None
    sources: "list[FigureSource | None] | None" = None
    ltm_source: FigureSource | None = None


def _coerce_value(value: "float | int | str") -> float | str:
    """Coerce one cell value: keep an Excel formula string, else cast to float.

    A string value must begin with ``"="`` (a formula such as ``"=9000+800"``);
    a bare text value would make the chart-ready data block non-numeric and
    break the downstream charts, so it is rejected.
    """
    if isinstance(value, str):
        if not value.startswith("="):
            raise ValueError(
                f"string metric value {value!r} must be an Excel formula "
                f'starting with "=" (e.g. "=9000+800"); bare text is not allowed'
            )
        return value
    return float(value)


def _coerce_sources(label: str, sources) -> "list[FigureSource | None] | None":
    """Validate a metric's per-value source records.

    A citation string used to BE the record; since Phase G it is rendered from
    one, so a string here would silently produce a record with the whole citation
    stuffed into ``filing`` and no statement or page — provenance that looks fine
    and cannot be followed. Reject it, naming the fix.
    """
    if sources is None:
        return None
    out: list[FigureSource | None] = []
    for source in sources:
        if source is None or isinstance(source, FigureSource):
            out.append(source)
            continue
        raise ProvenanceError(
            f"metric {label!r} has a source of type {type(source).__name__} "
            f"({source!r}); pass FigureSource(filing=…, statement=…, page=…) — a "
            f"citation string is no longer a source record."
        )
    return out


def _normalize_metric(metric: "MetricSeries | dict") -> MetricSeries:
    if isinstance(metric, MetricSeries):
        # Re-validate formula strings even on a pre-built instance.
        return MetricSeries(
            label=metric.label,
            units=metric.units,
            fiscal_values=[_coerce_value(v) for v in metric.fiscal_values],
            result_label=metric.result_label,
            ltm_value=(None if metric.ltm_value is None else _coerce_value(metric.ltm_value)),
            sources=_coerce_sources(metric.label, metric.sources),
            ltm_source=_coerce_sources(metric.label, [metric.ltm_source])[0]
            if metric.ltm_source is not None
            else None,
        )
    return MetricSeries(
        label=metric["label"],
        units=metric.get("units", ""),
        fiscal_values=[_coerce_value(v) for v in metric.get("fiscal_values", [])],
        result_label=metric.get("result_label"),
        ltm_value=(None if metric.get("ltm_value") is None else _coerce_value(metric["ltm_value"])),
        sources=_coerce_sources(metric["label"], metric.get("sources")),
        ltm_source=_coerce_sources(metric["label"], [metric.get("ltm_source")])[0]
        if metric.get("ltm_source") is not None
        else None,
    )


def _ref(cell) -> str:
    """`"financial-summary!B6"` — a provenance record's location for a written cell."""
    return f"{TAB_FINANCIAL_SUMMARY}!{cell.coordinate}"


def _ltm_link(ltm_sheet: str, result_label: str) -> str:
    """A label-keyed lookup of a bridge total on the deal workbook's LTM tab.

    An ordinary internal reference: both tabs are in the same file, so this
    resolves as soon as `ltm-metrics` is written (`ltm-metrics` is scheduled
    before `financial-summary`'s consumers). Keyed on the bridge's
    ``(=) <result_label>`` row so it survives the bridge's variable row position.

    The formula itself is built by `ltm_metrics.ltm_total_formula` — the tab that
    owns those rows owns how they are addressed, and the cap table's LTM
    valuation cells link through the same builder.
    """
    return ltm_total_formula(result_label, sheet=ltm_sheet)


def build_financial_summary_workbook(
    *,
    company_name: str,
    currency_note: str,
    period_note: str,
    fiscal_labels: list[str],
    metrics: "list[MetricSeries | dict]",
    metric_count: int = _METRIC_COUNT,
    show_ltm: bool = True,
    ltm_sheet_name: str = _DEFAULT_LTM_SHEET,
    deal_workbook: Path | str,
    provenance: ProvenanceLedger | None = None,
) -> Path:
    """Write the chart-ready `financial-summary` tab into the deal workbook.

    ``provenance`` is filled **in place** with one record per figure written —
    every fiscal value, plus each LTM cell — and each cell's ``Source: …`` comment
    is rendered from its record. Pass the stage's ledger and write it afterwards
    (`ledger.write(io.stage_dir)`); pass nothing and the records are still built
    (the comments come from them either way), just not kept.

    Returns the deal workbook's path. Since Phase D each flow metric's
    `=INDEX('ltm-metrics'!…)` LTM link resolves as soon as the `ltm-metrics` tab
    exists in the same file — it is no longer `#N/A` until an aggregation step.

    ``fiscal_labels`` are the five fiscal-year column headers in chronological
    order (oldest -> newest). ``metrics`` must hold exactly ``metric_count``
    ``MetricSeries`` (or dicts) — a positive multiple of 4 because every deck
    Financial Summary slide shows four tiles (default 4 = the single-slide deck;
    8 = the two-slide deck) — each supplying one value per fiscal label. When
    ``show_ltm`` is True an LTM column is appended; flow metrics (those with a
    ``result_label``) link to the ``ltm-metrics`` tab and non-flow metrics show
    their ``ltm_value``. When ``show_ltm`` is False the LTM column is dropped and
    a flow metric's most-recent fiscal-year cell carries the LTM-tab link
    instead (req 5).

    Raises ValueError on the wrong metric count, a fiscal-value count that does
    not match ``fiscal_labels``, or a flow metric missing its ``result_label``
    while a non-flow metric has neither a link nor an ``ltm_value``.
    """
    if metric_count <= 0 or metric_count % 4 != 0:
        raise ValueError(
            f"metric_count must be a positive multiple of 4 (four tiles per "
            f"Financial Summary slide); got {metric_count}"
        )
    rows = [_normalize_metric(m) for m in metrics]
    if len(rows) != metric_count:
        raise ValueError(
            f"Financial Summary holds exactly {metric_count} metrics; got {len(rows)}"
        )
    if not fiscal_labels:
        raise ValueError("at least one fiscal-year label is required")
    n_fy = len(fiscal_labels)
    for m in rows:
        if len(m.fiscal_values) != n_fy:
            raise ValueError(
                f"metric {m.label!r} has {len(m.fiscal_values)} fiscal values; "
                f"expected {n_fy} to match the fiscal-year labels"
            )
        if m.sources is not None and len(m.sources) != n_fy:
            raise ValueError(
                f"metric {m.label!r} has {len(m.sources)} source citations; "
                f"expected {n_fy} to match the fiscal values (use None for a "
                f"value without a citation)"
            )
        if not m.label.strip():
            raise ValueError("metric label cannot be blank")
        if m.result_label is None and m.ltm_value is None and show_ltm:
            raise ValueError(
                f"non-flow metric {m.label!r} needs an ltm_value (it has no "
                f"result_label to link to the ltm-metrics tab)"
            )

    ledger = provenance if provenance is not None else ProvenanceLedger(stage="financial-summary")

    def _write(_wb, ws) -> None:
        _fill_financial_summary_tab(
            ws,
            company_name=company_name,
            currency_note=currency_note,
            period_note=period_note,
            fiscal_labels=fiscal_labels,
            rows=rows,
            n_fy=n_fy,
            show_ltm=show_ltm,
            ltm_sheet_name=ltm_sheet_name,
            ledger=ledger,
        )

    write_tab(deal_workbook, TAB_FINANCIAL_SUMMARY, TabSpec(create=True, write=_write))
    return Path(deal_workbook)


def _fill_financial_summary_tab(
    ws: Worksheet,
    *,
    company_name: str,
    currency_note: str,
    period_note: str,
    fiscal_labels: list[str],
    rows: list[MetricSeries],
    n_fy: int,
    show_ltm: bool,
    ltm_sheet_name: str,
    ledger: ProvenanceLedger,
) -> None:
    """Write the chart-ready tab. Layout unchanged from the standalone workbook."""
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"{company_name} — Financial Summary"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = currency_note
    ws["A3"] = period_note
    for cell in ("A2", "A3"):
        ws[cell].font = _META_FONT

    # --- Header row 5: Metric | <FY labels> | [LTM] | Units ------------------
    header_row = 5
    first_data = header_row + 1

    period_labels = list(fiscal_labels) + (["LTM"] if show_ltm else [])
    headers = ["Metric", *period_labels, "Units"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    units_col = 1 + len(period_labels) + 1  # Metric + periods + Units
    # The most-recent fiscal-year column (last FY, before any LTM column).
    last_fy_col = 1 + n_fy
    ltm_col = last_fy_col + 1 if show_ltm else None

    # --- Data rows: one metric per row ---------------------------------------
    for i, m in enumerate(rows):
        r = first_data + i
        label_cell = ws.cell(row=r, column=1, value=m.label)
        label_cell.font = _LABEL_FONT
        label_cell.border = _BORDER

        for j, value in enumerate(m.fiscal_values):
            c = 2 + j
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = _BODY_FONT
            cell.number_format = _VALUE_FORMAT
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
            source = m.sources[j] if m.sources is not None else None
            if source is None:
                continue
            # The record first, the comment rendered from it — never the reverse.
            cite_cell(
                cell,
                ledger.record(
                    f"{m.label} {fiscal_labels[j]}",
                    sources=source,
                    value=value,
                    units=m.units,
                    location=_ref(cell),
                ),
            )

        if show_ltm:
            cell = ws.cell(row=r, column=ltm_col)
            if m.result_label is not None:
                cell.value = _ltm_link(ltm_sheet_name, m.result_label)
            else:
                cell.value = m.ltm_value
            cell.font = _LABEL_FONT
            cell.number_format = _VALUE_FORMAT
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if m.result_label is not None:
                # A flow metric's LTM cell is a link: its provenance is the
                # `ltm-metrics` bridge it points at, whose components carry the
                # filing records. Record the chain, not a source it does not have.
                ledger.record(
                    f"{m.label} LTM",
                    value=cell.value,
                    units=m.units,
                    location=_ref(cell),
                    derivation=(
                        f"link to the {ltm_sheet_name} tab's "
                        f"'(=) {m.result_label}' bridge total"
                    ),
                )
            elif m.ltm_source is not None:
                cite_cell(
                    cell,
                    ledger.record(
                        f"{m.label} LTM",
                        sources=m.ltm_source,
                        value=m.ltm_value,
                        units=m.units,
                        location=_ref(cell),
                    ),
                )
        elif m.result_label is not None:
            # Suppression: LTM == latest FY, so the most-recent FY cell carries
            # the link to the LTM tab instead of the literal value (req 5).
            # The FY cell's source comment (if any) stays — LTM == that FY
            # figure by definition here, so the citation still describes the
            # number shown.
            cell = ws.cell(row=r, column=last_fy_col)
            cell.value = _ltm_link(ltm_sheet_name, m.result_label)
            cell.number_format = _VALUE_FORMAT
            ledger.record(
                f"{m.label} {fiscal_labels[-1]} (LTM == latest FY)",
                value=cell.value,
                units=m.units,
                location=_ref(cell),
                derivation=(
                    f"link to the {ltm_sheet_name} tab's '(=) {m.result_label}' bridge "
                    f"total; the LTM column is suppressed because LTM equals the latest "
                    f"fiscal year"
                ),
            )

        units_cell = ws.cell(row=r, column=units_col, value=m.units)
        units_cell.font = _META_FONT
        units_cell.border = _BORDER
        units_cell.alignment = Alignment(horizontal="center")

    # Column widths: metric label wide, period/units columns even.
    ws.column_dimensions["A"].width = 34
    for col in range(2, units_col):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(units_col)].width = 10
