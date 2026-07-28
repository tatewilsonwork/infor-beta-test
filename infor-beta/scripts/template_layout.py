"""The shipped templates' layout map: named ranges, sentinels, slide markers.

Nothing here is addressed blind. Every load-bearing cell in the four workbook
templates is reachable by a **defined name** the template itself carries, and
every library slide the assemblers touch is located by a **marker shape**
rather than an index. The point of both is the same: an analyst re-saving a
template or inserting a library slide must not require a code change.

**Named ranges (Phase C).** Writers resolve an address through
``resolve_name_cell`` / ``resolve_name_range`` instead of hardcoding it. A
defined name is metadata that Excel moves with its cell when rows or columns
are inserted, so ``infor_fx_rate`` follows the FX cell from ``F7`` to ``F9``
and the writer never notices. The names are all ``infor_``-prefixed and
**worksheet-scoped**: the cap table ships 34 Capital IQ defined names
(``CIQWBGuid``, ``IQ_LTM``, …) and the comps template 1,246 legacy artefacts
(``_______AOL2``, ``__123Graph_A``, …), so a bare ``fx_rate`` would be neither
obviously ours nor safe from collision — and a workbook-scoped name travelling
with a sheet through the aggregator's COM merge is how phantom
external-workbook aliases get created.

**Sentinels (Phase A/B, kept as a cross-check for this release).** Each
protected address is also paired with the label text the shipped template
carries in a stable anchor cell beside it (discovered by opening the templates,
never invented). ``verify_anchors`` checks the sentinel *and*, when the anchor
declares a ``name``, asserts the defined name resolves to the very same cell —
so the Phase C migration cannot silently have mis-mapped an address. A
disagreement raises :class:`TemplateLayoutError` naming both. The sentinel
tables are scheduled for deletion in a follow-up release; the cross-check is
what earns that deletion.

Style model: the Bloomberg-export ``C13 = "Holder Name"`` header check in
``ownership_workbook.read_bloomberg_export`` — cheap, targeted cell reads (no
full-sheet scans), with a message that tells the analyst what to fix.

**Slide markers.** The pitch flow's self-discovery of Financial Summary slides
by their ``Rectangle 17`` marker shape is generalised here into
``find_slides_by_marker`` / ``find_slide_by_marker``, and every hardcoded slide
index now goes through it — the earnings assembler's kept-library set (which
needed three manual migrations), its filled-slide positions, the pitch layout
arithmetic, and ``financial_charts``' overview slide. Library markers identify
a *blank* library slide; the ``MARKER_BUILT_*`` markers identify a slide in an
*assembled* deck, where the titles have been filled and only the deferred
placeholders survive.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


class TemplateLayoutError(RuntimeError):
    """A shipped template's layout no longer matches the hardcoded addresses.

    Raised by the verify helpers when a sentinel anchor cell / marker shape
    does not carry its expected label — i.e. the template was re-saved with
    rows/columns inserted, or the slide library was re-ordered. The message
    names the template, the protected address, the anchor, what was expected,
    and what was actually found.
    """


# ─── Cell anchors ─────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class CellAnchor:
    """One protected address, its defined name, and the sentinel that anchors it.

    - ``target``: the load-bearing cell/range/rows being protected (what a
      writer reads or writes).
    - ``label_addr``: the nearby cell that carries the sentinel label in the
      shipped template.
    - ``expected``: the sentinel text (verbatim from the shipped template).
    - ``contains``: substring match instead of exact — for anchors whose cell
      holds a formula string (e.g. the cap table's ``B16`` share-price label
      is a ``CONCATENATE`` formula containing ``"Share Price"``).
    - ``name``: the ``infor_``-prefixed defined name the shipped template
      carries for ``target``. Writers resolve through it; ``verify_anchors``
      cross-checks that it still points at ``target``. ``None`` for anchors
      that only *witness* a region another anchor names (the cap table's
      ``D33 = "LTM"`` column header, say), so a region is named exactly once.
    """

    target: str
    label_addr: str
    expected: str
    contains: bool = False
    name: str | None = None


def _anchor_text(value) -> str | None:
    """Normalise a cell value for sentinel comparison.

    Plain strings (labels and formula strings) pass through; openpyxl
    ``ArrayFormula`` objects contribute their formula text; everything else
    (numbers, None) is returned as-is / None.
    """
    if value is None or isinstance(value, str):
        return value
    text = getattr(value, "text", None)  # openpyxl ArrayFormula
    if isinstance(text, str):
        return text
    return str(value)


def _anchor_matches(anchor: CellAnchor, found: str | None) -> bool:
    if found is None:
        return False
    if anchor.contains:
        return anchor.expected in found
    return found.strip() == anchor.expected


# ─── Defined-name resolution ─────────────────────────────────────────────────


def normalize_ref(ref: str) -> str:
    """A defined name's destination as a plain A1 reference.

    Strips the absolute-reference ``$`` markers Excel stores (``$B$15:$F$40`` ->
    ``B15:F40``) and upper-cases, so a resolved name compares directly against
    the ``CellAnchor.target`` addresses the writers used before Phase C.
    """
    return ref.replace("$", "").strip().upper()


def _lookup_defined_name(ws, name: str):
    """The ``DefinedName`` for ``name``, worksheet scope first, or ``None``.

    Our names are authored worksheet-scoped; the workbook-scope fallback is
    there so a template an analyst re-created with a workbook-scoped name of
    the same spelling still resolves rather than failing closed for a reason
    that is not a layout shift.
    """
    local = getattr(ws, "defined_names", None)
    if local is not None and name in local:
        return local[name]
    workbook = getattr(ws, "parent", None)
    if workbook is not None and name in workbook.defined_names:
        return workbook.defined_names[name]
    return None


def defined_name_ref(ws, name: str) -> str | None:
    """The A1 range ``name`` resolves to on ``ws``, or ``None`` if it does not.

    ``None`` covers all three "this sheet has no such usable name" cases: the
    name is absent, it resolves to several areas, or it points at a different
    sheet. Callers that need a hard failure use ``resolve_name_range``.
    """
    defined = _lookup_defined_name(ws, name)
    if defined is None:
        return None
    try:
        destinations = list(defined.destinations)
    except Exception:  # a constant / formula name has no destinations
        return None
    if len(destinations) != 1:
        return None
    sheet, ref = destinations[0]
    if sheet != ws.title:
        return None
    return normalize_ref(ref)


def resolve_name_range(ws, name: str, *, template: str | None = None) -> str:
    """The A1 range ``name`` resolves to on ``ws``; raise if it does not.

    This is what the writers call instead of hardcoding an address. The error
    names the template and the missing name, because the remedy is to re-run
    ``tools/add_template_named_ranges.py`` against the re-saved template.
    """
    ref = defined_name_ref(ws, name)
    if ref is None:
        raise TemplateLayoutError(
            f"{template or 'workbook'}, sheet {ws.title!r}: defined name {name!r} is "
            f"missing (or does not resolve to a single range on this sheet). The "
            f"shipped templates carry it; a re-saved template that dropped it must be "
            f"re-stamped with tools/add_template_named_ranges.py before it can be used."
        )
    return ref


def resolve_name_cell(ws, name: str, *, template: str | None = None) -> str:
    """The single cell ``name`` resolves to on ``ws`` (e.g. ``"F7"``); raise otherwise."""
    ref = resolve_name_range(ws, name, template=template)
    if ":" in ref:
        raise TemplateLayoutError(
            f"{template or 'workbook'}, sheet {ws.title!r}: defined name {name!r} "
            f"resolves to the range {ref!r}, but a single cell was expected."
        )
    return ref


def verify_cell_anchor(ws, anchor: CellAnchor, *, template: str) -> None:
    """Verify one sentinel anchor on an open openpyxl worksheet; raise on mismatch."""
    verify_anchors(ws, (anchor,), template=template)


def verify_anchors(ws, anchors, *, template: str, require_names: bool = False) -> None:
    """Verify sentinel anchors — and their defined names — on an open worksheet.

    Two independent checks per anchor, both reported together so a shifted
    template names every displaced address at once rather than the first:

    1. **Sentinel** — the anchor cell still carries its expected label.
    2. **Name ↔ sentinel cross-check** — when the anchor declares a ``name``
       *and* the workbook carries it, the name must resolve to exactly
       ``anchor.target``. A disagreement means the Phase C migration mapped a
       name to the wrong cell, or the template moved one without the other;
       either way the writers must not proceed.

    ``require_names`` additionally makes a *missing* name an error. Writers
    resolving through names pass it (they are looking at a freshly copied
    shipped template, which always carries them); the built-artefact callers —
    the assemblers' picture-range checks, the aggregator's relink pre-flight —
    leave it off, so a workbook produced before Phase C still passes on its
    sentinels alone.
    """
    problems: list[str] = []
    for anchor in anchors:
        found = _anchor_text(ws[anchor.label_addr].value)
        if not _anchor_matches(anchor, found):
            mode = "containing" if anchor.contains else "="
            problems.append(
                f"{anchor.target} expects anchor {anchor.label_addr} {mode} "
                f"{anchor.expected!r} but found {found!r}"
            )
        if anchor.name is None:
            continue
        resolved = defined_name_ref(ws, anchor.name)
        if resolved is None:
            if require_names:
                problems.append(
                    f"{anchor.target} expects defined name {anchor.name!r}, which is "
                    f"missing from this workbook"
                )
        elif resolved != normalize_ref(anchor.target):
            problems.append(
                f"defined name {anchor.name!r} resolves to {resolved} but its sentinel "
                f"{anchor.label_addr}={anchor.expected!r} pins {anchor.target}"
            )
    if problems:
        raise TemplateLayoutError(
            f"{template}, sheet {ws.title!r}: layout verification failed — "
            + "; ".join(problems)
            + ". The template layout has shifted; refusing to read/write these "
            "addresses."
        )


def resolve_workbook_range(
    workbook_path: Path | str, *, sheet: str, name: str, fallback: str
) -> str:
    """The A1 range ``name`` resolves to on a saved workbook, else ``fallback``.

    For the callers holding a path to a **built artefact** rather than a
    template: the deck assemblers' Excel picture ranges, and the aggregator's
    relink cells. A workbook a builder produced this release carries the names
    its template did, so this resolves; ``fallback`` keeps a workbook from an
    earlier run — an analyst re-running the deck stage over an existing deal
    directory — working on the pre-Phase-C address rather than failing for a
    reason that is not a layout shift.

    Safe to fall back because the caller has already run
    ``verify_workbook_anchors``, which sentinel-checks the same region and
    fails hard if a *present* name disagrees with it.
    """
    from openpyxl import load_workbook

    wb = load_workbook(Path(workbook_path), read_only=False, data_only=False)
    try:
        if sheet not in wb.sheetnames:
            return fallback
        return defined_name_ref(wb[sheet], name) or fallback
    finally:
        wb.close()


def verify_workbook_anchors(
    workbook_path: Path | str,
    *,
    sheet: str,
    anchors,
    template: str | None = None,
    require_names: bool = False,
) -> None:
    """Open a workbook and verify sentinel anchors on one sheet.

    Convenience wrapper for callers holding only a path (the assemblers'
    picture-range inserts, the aggregator's relink pre-flight). ``template``
    defaults to the workbook's filename. Raises :class:`TemplateLayoutError`
    when the sheet is missing or an anchor mismatches; file-level errors
    (missing/corrupt file) propagate as their native exceptions.
    """
    from openpyxl import load_workbook

    path = Path(workbook_path)
    name = template or path.name
    wb = load_workbook(path, data_only=False)
    try:
        if sheet not in wb.sheetnames:
            raise TemplateLayoutError(
                f"{name}: layout verification failed — expected sheet {sheet!r} "
                f"(have {wb.sheetnames})"
            )
        verify_anchors(wb[sheet], anchors, template=name, require_names=require_names)
    finally:
        wb.close()


# ─── Cap table (INFOR Cap Table Template.xlsx, sheet 'Cap with Links') ───────

CAP_TABLE_TEMPLATE = "INFOR Cap Table Template.xlsx"
CAP_TABLE_SHEET = "Cap with Links"

# Defined names carried by the shipped cap-table template. Writers resolve
# through these; `CAP_TABLE_*_ANCHOR.name` cross-checks each one against its
# sentinel. Addresses in the comments are the shipped positions, for orientation
# only — nothing reads them.
NAME_CAP_TICKER = "infor_cap_ticker"                        # F3
NAME_CAP_OUTPUT_CCY = "infor_cap_output_ccy"                # F5
NAME_FX_RATE = "infor_fx_rate"                              # F7
NAME_SHARE_PRICE = "infor_share_price"                      # F16
NAME_BASIC_SHARES = "infor_basic_shares"                    # F17
NAME_LTM_REVENUE_VALUATION = "infor_ltm_revenue_valuation"  # D47
NAME_LTM_EBITDA_VALUATION = "infor_ltm_ebitda_valuation"    # D48
NAME_CAP_PICTURE_RANGE = "infor_cap_picture_range"          # B15:F40
NAME_CAP_SHARE_INPUTS = "infor_cap_share_inputs"            # F168:F185

# Header input cells the captable skill writes (Steps 3/3b) and the aggregator
# reads/references at relink time.
CAP_TABLE_TICKER_ANCHOR = CellAnchor("F3", "B3", "Ticker:", name=NAME_CAP_TICKER)
CAP_TABLE_OUTPUT_CCY_ANCHOR = CellAnchor(
    "F5", "B5", "Output Currency:", name=NAME_CAP_OUTPUT_CCY
)
CAP_TABLE_FX_ANCHOR = CellAnchor("F7", "B7", "FX Rate:", name=NAME_FX_RATE)
# B16 is a CONCATENATE formula ('Share Price (<date>)'), so substring-match it.
CAP_TABLE_SHARE_PRICE_ANCHOR = CellAnchor(
    "F16", "B16", "Share Price", contains=True, name=NAME_SHARE_PRICE
)
CAP_TABLE_BASIC_SHARES_ANCHOR = CellAnchor(
    "F17", "B17", "Basic Shares Outstanding", name=NAME_BASIC_SHARES
)

CAP_TABLE_HEADER_ANCHORS = (
    CAP_TABLE_TICKER_ANCHOR,
    CAP_TABLE_OUTPUT_CCY_ANCHOR,
    CAP_TABLE_FX_ANCHOR,
    CAP_TABLE_SHARE_PRICE_ANCHOR,
    CAP_TABLE_BASIC_SHARES_ANCHOR,
)

# LTM valuation inputs (captable Step 6b; relinked to the ltm-metrics tab by
# the aggregator). D33 is the Financial Metrics block's LTM column header — a
# witness for the pair, so it carries no name of its own.
CAP_TABLE_LTM_ANCHORS = (
    CellAnchor("D47", "B47", "Revenue", name=NAME_LTM_REVENUE_VALUATION),
    CellAnchor("D48", "B48", "Adj. EBITDA", name=NAME_LTM_EBITDA_VALUATION),
    CellAnchor("D47:D48", "D33", "LTM"),
)

# Slide picture range (both deck assemblers paste 'Cap with Links'!B15:F40).
# B15 is the range's own top-left label; B40 pins the bottom row.
CAP_TABLE_PICTURE_RANGE = "B15:F40"
CAP_TABLE_PICTURE_ANCHORS = (
    CellAnchor(
        CAP_TABLE_PICTURE_RANGE, "B15", "Company Ticker:", name=NAME_CAP_PICTURE_RANGE
    ),
    CellAnchor(CAP_TABLE_PICTURE_RANGE, "B40", "EV / Adj. EBITDA"),
)

# Section VII basic-share input rows (read by
# ownership_workbook.read_basic_shares_from_cap_table as F168:F185).
CAP_TABLE_SECTION_VII_ROWS = (168, 185)
CAP_TABLE_SECTION_VII_ANCHORS = (
    CellAnchor(
        "F168:F185", "B166", "VII. BASIC SHARES OUTSTANDING", name=NAME_CAP_SHARE_INPUTS
    ),
    CellAnchor("F168:F185", "B167", "Description"),
    CellAnchor("F168:F185", "F167", "Amount"),
    CellAnchor("F168:F185", "B186", "Total Basic Shares Outstanding"),
)


def verify_cap_table_before_write(ws) -> None:
    """Verify every cap-table region the captable skill writes.

    Called by the captable skill (SKILL.md Step 3) on the freshly copied
    template before any cell write: the header inputs (F3/F5/F7/F16), the LTM
    valuation cells (D47/D48), and the Section VII block the section writes
    land around. Requires the defined names — this is a shipped template, so a
    missing name is itself the layout problem. Raises
    :class:`TemplateLayoutError` if the template layout has shifted.
    """
    verify_anchors(
        ws,
        CAP_TABLE_HEADER_ANCHORS + CAP_TABLE_LTM_ANCHORS + CAP_TABLE_SECTION_VII_ANCHORS,
        template=CAP_TABLE_TEMPLATE,
        require_names=True,
    )


# ─── Ownership (INFOR Ownership Template.xlsx) ────────────────────────────────

OWNERSHIP_TEMPLATE = "INFOR Ownership Template.xlsx"
OWNERSHIP_SHEET = "Ownership"
OWNERSHIP_BBG_SHEET = "Bloomberg Output"

NAME_OWN_INSIDER_BLOCK = "infor_own_insider_block"                    # B39:J65
NAME_OWN_TOTAL_SHARES = "infor_own_total_shares"                      # F35
NAME_OWN_BBG_LINK_BLOCK = "infor_own_bbg_link_block"                  # H68:J185
NAME_OWN_BBG_HOLDER_BLOCK = "infor_own_bbg_holder_block"              # C14:AC131
NAME_OWN_INSIDERS_PICTURE = "infor_own_insiders_picture_range"        # B4:G17
NAME_OWN_INSTITUTIONS_PICTURE = "infor_own_institutions_picture_range"  # B19:G35

# Select-Insiders data block (rows 39–65, cols B/F/G/H/J): anchored by its
# header row 38 and bounded below by the row-67 Bloomberg link-block header.
OWNERSHIP_INSIDER_BLOCK_ANCHORS = (
    CellAnchor("B39:J65", "B38", "SEDI Name", name=NAME_OWN_INSIDER_BLOCK),
    CellAnchor("B39:J65", "F38", "Basic"),
    CellAnchor("B39:J65", "G38", "Date"),
    CellAnchor("B39:J65", "H38", "Include"),
    CellAnchor("B39:J65", "J38", "Adj. Name"),
    CellAnchor("B39:J65", "B67", "From Bloomberg"),
)

# % denominator (written by the ownership skill; relinked to the cap table's
# basic shares by the aggregator).
OWNERSHIP_TOTAL_SHARES_ANCHORS = (
    CellAnchor(
        "F35", "B35", "Total Basic Shares Outstanding", name=NAME_OWN_TOTAL_SHARES
    ),
)

# Bloomberg link rows 68–185 (cols H/J written; B/F/G neutralised on unused
# rows): anchored by the row-67 header and the first/last pre-wired link
# formulas, which pin the 118-row extent against 'Bloomberg Output'!C14:C131.
OWNERSHIP_BBG_LINK_ANCHORS = (
    CellAnchor("H68:J185", "B67", "From Bloomberg", name=NAME_OWN_BBG_LINK_BLOCK),
    CellAnchor("H68:J185", "B68", "'Bloomberg Output'!C14", contains=True),
    CellAnchor("H68:J185", "B185", "'Bloomberg Output'!C131", contains=True),
)

# The template's own 'Bloomberg Output' tab mirrors the BBG add-in Summary
# View; the export-side header check in read_bloomberg_export validates the
# attachment, this validates the template copy the holder rows land on.
OWNERSHIP_BBG_TEMPLATE_ANCHORS = (
    CellAnchor("C14:AC131", "C13", "Holder Name", name=NAME_OWN_BBG_HOLDER_BLOCK),
    CellAnchor("C14:AC131", "L13", "Position"),
    CellAnchor("C14:AC131", "N13", "Filing Date"),
    CellAnchor("C14:AC131", "R13", "Insider Status"),
)

# Slide picture ranges (pitch deck ownership slide).
OWNERSHIP_INSIDERS_PICTURE_RANGE = "B4:G17"
OWNERSHIP_INSIDERS_PICTURE_ANCHORS = (
    CellAnchor(
        OWNERSHIP_INSIDERS_PICTURE_RANGE, "B4", "Select Insiders",
        name=NAME_OWN_INSIDERS_PICTURE,
    ),
    CellAnchor(OWNERSHIP_INSIDERS_PICTURE_RANGE, "B17", "Subtotal"),
)
OWNERSHIP_INSTITUTIONS_PICTURE_RANGE = "B19:G35"
OWNERSHIP_INSTITUTIONS_PICTURE_ANCHORS = (
    CellAnchor(
        OWNERSHIP_INSTITUTIONS_PICTURE_RANGE, "B19", "Select Institutions",
        name=NAME_OWN_INSTITUTIONS_PICTURE,
    ),
    CellAnchor(OWNERSHIP_INSTITUTIONS_PICTURE_RANGE, "B35", "Total Basic Shares Outstanding"),
)


# ─── Comps (INFOR Comps Template.xlsx, sheet 'Comps') ─────────────────────────

COMPS_TEMPLATE = "INFOR Comps Template.xlsx"
COMPS_SHEET = "Comps"

NAME_COMPS_OUTPUT_CCY = "infor_comps_output_ccy"  # F3
# One label + one block name per vertical. The builder derives the ticker
# (column B) and description (column AA) rows from the block's own extent, so
# a template that grows a vertical from six rows to eight needs no code change.
NAME_COMPS_GROUP_LABELS = (
    "infor_comps_group1_label",  # D9
    "infor_comps_group2_label",  # D19
    "infor_comps_group3_label",  # D29
)
NAME_COMPS_GROUP_BLOCKS = (
    "infor_comps_group1_block",  # B10:AA15
    "infor_comps_group2_block",  # B20:AA25
    "infor_comps_group3_block",  # B30:AA35
)

# The three vertical blocks (label D9/D19/D29; tickers B10:B15/B20:B25/B30:B35;
# descriptions AA10…): anchored by the row-7 column headers above the first
# block and the 'Group Average' row that closes each 6-row block. The ticker
# column B itself has no adjacent header label in the template — its position
# is pinned transitively by the D/AA headers and the block bounds.
#
# Each label cell is its own sentinel: the shipped template carries the
# '[Group #N]' placeholder there, and the builder verifies before it overwrites
# it, so the check always runs against a pristine copy.
COMPS_BLOCK_ANCHORS = (
    CellAnchor("D9", "D9", "[Group #1]", name=NAME_COMPS_GROUP_LABELS[0]),
    CellAnchor("D19", "D19", "[Group #2]", name=NAME_COMPS_GROUP_LABELS[1]),
    CellAnchor("D29", "D29", "[Group #3]", name=NAME_COMPS_GROUP_LABELS[2]),
    CellAnchor("B10:AA15", "D7", "Company", name=NAME_COMPS_GROUP_BLOCKS[0]),
    CellAnchor("B10:AA15", "AA7", "Description"),
    CellAnchor("B10:AA15", "D17", "Group Average"),
    CellAnchor("B20:AA25", "D27", "Group Average", name=NAME_COMPS_GROUP_BLOCKS[1]),
    CellAnchor("B30:AA35", "D37", "Group Average", name=NAME_COMPS_GROUP_BLOCKS[2]),
)

# Output-currency cell the aggregator relinks to the cap table's F5.
COMPS_OUTPUT_CCY_ANCHORS = (
    CellAnchor("F3", "E3", "Currency:", name=NAME_COMPS_OUTPUT_CCY),
)


# ─── Precedents (INFOR Precedents Template.xlsx, sheet 'Precedents') ──────────

PRECEDENTS_TEMPLATE = "INFOR Precedents Template.xlsx"
PRECEDENTS_SHEET = "Precedents"

# The plan calls this one `precedents_input_ccy`; the template labels the cell
# "Output:" and the aggregator relinks it to the cap table's *output* currency,
# so the name follows the artefact rather than the plan's shorthand.
NAME_PREC_OUTPUT_CCY = "infor_prec_output_ccy"  # C2
NAME_PREC_GROUP_LABELS = (
    "infor_prec_group1_label",  # E7
    "infor_prec_group2_label",  # E16
)
NAME_PREC_GROUP_BLOCKS = (
    "infor_prec_group1_block",  # B8:AI13
    "infor_prec_group2_block",  # B17:AI22
)

# Output-currency cell (written by the builder; relinked by the aggregator).
PRECEDENTS_OUTPUT_CCY_ANCHORS = (
    CellAnchor("C2", "B2", "Output:", name=NAME_PREC_OUTPUT_CCY),
)

# The two transaction blocks (rows 8–13 / 17–22) and the input columns the
# builder writes, anchored by the row-4/5 header labels and the 'Group
# Average' row that closes each block. Column AI (HQ code) has its own row-5
# header; column H is left empty by design (no anchor needed). Like comps, each
# group label cell is its own sentinel — the builder verifies the pristine
# '[Group #N]' placeholder before overwriting it.
PRECEDENTS_BLOCK_ANCHORS = (
    CellAnchor("E7", "E7", "[Group #1]", name=NAME_PREC_GROUP_LABELS[0]),
    CellAnchor("E16", "E16", "[Group #2]", name=NAME_PREC_GROUP_LABELS[1]),
    CellAnchor("B8:B22", "B5", "Currency"),
    CellAnchor("E8:E22", "E5", "Ann. Date"),
    CellAnchor("F8:F22", "F5", "Target"),
    CellAnchor("G8:G22", "G5", "Acquiror"),
    CellAnchor("I8:I22", "I4", "TEV"),
    CellAnchor("I8:I22", "I5", "Source FX"),
    CellAnchor("K8:L22", "K4", "Revenue - Source FX"),
    CellAnchor("M8:N22", "M4", "Net Income - Source FX"),
    CellAnchor("O8:P22", "O4", "Adj. EBITDA - Source FX"),
    CellAnchor("Q8:R22", "Q4", "Book Value - Source FX"),
    CellAnchor("S8:V22", "S4", "EV / Revenue"),
    CellAnchor("S8:V22", "U4", "EV / EBITDA"),
    CellAnchor("W8:X22", "W4", "P/ E"),
    CellAnchor("Y8:Y22", "Y4", "P / B"),
    CellAnchor("Z8:Z22", "Z4", "P / TBV"),
    CellAnchor("AB8:AG22", "AB4", "TEV"),
    CellAnchor("AB8:AG22", "AG4", "P / TBV"),
    CellAnchor("AI8:AI22", "AI5", "HQ"),
    CellAnchor("B8:AI13", "E14", "Group Average", name=NAME_PREC_GROUP_BLOCKS[0]),
    CellAnchor("B17:AI22", "E23", "Group Average", name=NAME_PREC_GROUP_BLOCKS[1]),
)


# ─── The defined-name registry ────────────────────────────────────────────────
# Derived from the anchors above rather than written out a second time: the
# prep tool that stamps the templates and the test that verifies them both read
# this, so a name can never be declared in one place and stamped in another.


def named_targets(*anchor_groups) -> dict[str, str]:
    """``{defined name: target address}`` for every named anchor in the groups.

    Raises :class:`TemplateLayoutError` when one name is declared against two
    different targets — a contradiction the writers could not resolve.
    """
    targets: dict[str, str] = {}
    for group in anchor_groups:
        for anchor in group:
            if anchor.name is None:
                continue
            existing = targets.get(anchor.name)
            if existing is not None and existing != anchor.target:
                raise TemplateLayoutError(
                    f"defined name {anchor.name!r} is declared against both "
                    f"{existing} and {anchor.target}"
                )
            targets[anchor.name] = anchor.target
    return targets


# template filename -> sheet -> {name: A1 target}. The order here is the order
# the prep tool stamps them in.
TEMPLATE_NAMED_RANGES: dict[str, dict[str, dict[str, str]]] = {
    CAP_TABLE_TEMPLATE: {
        CAP_TABLE_SHEET: named_targets(
            CAP_TABLE_HEADER_ANCHORS,
            CAP_TABLE_LTM_ANCHORS,
            CAP_TABLE_PICTURE_ANCHORS,
            CAP_TABLE_SECTION_VII_ANCHORS,
        )
    },
    OWNERSHIP_TEMPLATE: {
        OWNERSHIP_SHEET: named_targets(
            OWNERSHIP_INSIDER_BLOCK_ANCHORS,
            OWNERSHIP_TOTAL_SHARES_ANCHORS,
            OWNERSHIP_BBG_LINK_ANCHORS,
            OWNERSHIP_INSIDERS_PICTURE_ANCHORS,
            OWNERSHIP_INSTITUTIONS_PICTURE_ANCHORS,
        ),
        OWNERSHIP_BBG_SHEET: named_targets(OWNERSHIP_BBG_TEMPLATE_ANCHORS),
    },
    COMPS_TEMPLATE: {
        COMPS_SHEET: named_targets(COMPS_BLOCK_ANCHORS, COMPS_OUTPUT_CCY_ANCHORS)
    },
    PRECEDENTS_TEMPLATE: {
        PRECEDENTS_SHEET: named_targets(
            PRECEDENTS_BLOCK_ANCHORS, PRECEDENTS_OUTPUT_CCY_ANCHORS
        )
    },
}


# ─── Slide library (INFOR Slide Library.pptx) ─────────────────────────────────

SLIDE_LIBRARY_TEMPLATE = "INFOR Slide Library.pptx"
SLIDE_LIBRARY_SLIDE_COUNT = 17


@dataclass(frozen=True)
class SlideMarker:
    """A slide's identifying shape + expected text fragment.

    Generalises the pitch flow's FS-slide self-discovery (the ``Rectangle 17``
    metric-chart placeholder in ``financial_charts``): a slide is *located* and
    verified by a named shape whose text contains a fragment unique to that
    slide concept.

    Two families, because a marker has to survive whatever the caller is
    looking at. ``MARKER_*`` identify slides in the **blank library**, where
    every title is still its placeholder. ``MARKER_BUILT_*`` identify slides in
    an **assembled deck**, where the titles have been filled and only the
    deferred placeholders remain — so they key off those instead.
    """

    shape_name: str
    contains: str
    description: str


MARKER_COVER = SlideMarker("Title 1", "[CLIENT NAME]", "cover")
MARKER_OVERVIEW = SlideMarker("Title 6", "Introduction to [Client Name]", "public-company overview")
MARKER_EARNINGS_SUMMARY = SlideMarker("Title 1", "Earnings Summary", "earnings summary")
# Same marker shape financial_charts uses to self-discover FS slides.
MARKER_FINANCIAL_SUMMARY = SlideMarker(
    "Rectangle 17", "[Placeholder for Metric #1 Chart]", "financial summary"
)
MARKER_OWNERSHIP = SlideMarker("Title 6", "Ownership", "insider ownership")
MARKER_RISKS = SlideMarker("Title 2", "Considerations and Mitigants", "acquirer considerations/mitigants")
MARKER_COMPS = SlideMarker("Title 6", "Comparable Companies Analysis", "comparable companies")
MARKER_PRECEDENTS = SlideMarker("Title 6", "Precedent Transactions Analysis", "precedent transactions")
MARKER_KIH = SlideMarker("Title 1", "Key Investment Highlights", "key investment highlights")
MARKER_MARKET_ENTRY = SlideMarker("Title 1", "Market Entry Targets", "market entry targets")
MARKER_DISCLAIMER = SlideMarker("Title 1", "Disclaimer", "disclaimer")
MARKER_CONTACT = SlideMarker("Title 6", "Contact", "contact")

# Raw (blank) library slide index -> marker, for every slide the assemblers
# clone, delete, keep, or fill. Discovered from the shipped 17-slide library.
#
# Nothing addresses a library slide by these indices any more — the assemblers
# locate every slide with `find_slide_by_marker`. The map survives as the
# library's inventory: `verify_library_slide` uses it, `deck_contract` reads it
# to name a matched baseline slide, and `test_template_layout` asserts each
# marker still lands on exactly the index recorded here, which is what turns a
# re-ordered library into a caught regression rather than a silent one.
LIBRARY_SLIDE_MARKERS: dict[int, SlideMarker] = {
    0: MARKER_COVER,
    6: MARKER_OVERVIEW,
    7: MARKER_EARNINGS_SUMMARY,
    8: MARKER_FINANCIAL_SUMMARY,
    9: MARKER_OWNERSHIP,
    10: MARKER_RISKS,
    11: MARKER_COMPS,
    12: MARKER_PRECEDENTS,
    13: MARKER_KIH,
    14: MARKER_MARKET_ENTRY,
    15: MARKER_DISCLAIMER,
    16: MARKER_CONTACT,
}

# Markers for an ASSEMBLED deck. The library markers above key off titles that
# assembly overwrites ("Introduction to [Client Name]" becomes "Introduction to
# Acme Corp"), so a post-fill lookup needs a shape the fill leaves alone. The
# deferred chart placeholders are exactly that, and they are also what the
# post-assembly stages are looking for: `financial_charts` finds the overview
# slide by the pie placeholder it is about to replace, the same way it already
# found the Financial Summary slides by their Metric #1 placeholder.
MARKER_BUILT_OVERVIEW = SlideMarker(
    "Rectangle 4", "[Pie Chart Placeholder]", "public-company overview"
)
MARKER_BUILT_FINANCIAL_SUMMARY = MARKER_FINANCIAL_SUMMARY


def verify_slide_marker(
    slide, marker: SlideMarker, *, template: str, slide_index: int | None = None
) -> None:
    """Verify a slide carries its identifying marker shape; raise on mismatch.

    ``slide_index`` (zero-based, when known) is included in the error message.
    The marker shape is looked up among the slide's top-level shapes (every
    library marker is top-level).
    """
    where = f"slide index {slide_index}" if slide_index is not None else "slide"
    shape = next((s for s in slide.shapes if s.name == marker.shape_name), None)
    if shape is None:
        names = sorted({s.name for s in slide.shapes})
        raise TemplateLayoutError(
            f"{template}: layout verification failed — {where} was expected to be "
            f"the {marker.description} slide (shape {marker.shape_name!r} containing "
            f"{marker.contains!r}) but has no shape named {marker.shape_name!r} "
            f"(shapes: {names}). The slide library has been re-ordered or edited; "
            "refusing to clone/delete/fill by index blind."
        )
    text = shape.text_frame.text if getattr(shape, "has_text_frame", False) else ""
    if marker.contains not in text:
        raise TemplateLayoutError(
            f"{template}: layout verification failed — {where} was expected to be "
            f"the {marker.description} slide, but shape {marker.shape_name!r} reads "
            f"{text!r} instead of containing {marker.contains!r}. The slide library "
            "has been re-ordered or edited; refusing to clone/delete/fill by index blind."
        )


def slide_matches_marker(slide, marker: SlideMarker) -> bool:
    """True when ``slide`` carries ``marker``'s shape with its text fragment."""
    shape = next((s for s in slide.shapes if s.name == marker.shape_name), None)
    if shape is None or not getattr(shape, "has_text_frame", False):
        return False
    return marker.contains in shape.text_frame.text


def find_slides_by_marker(prs, marker: SlideMarker) -> list[int]:
    """Zero-based indices of every slide carrying ``marker``, in deck order.

    The general form of ``financial_charts._find_financial_summary_slides``.
    Returns ``[]`` rather than raising — callers that require a hit use
    ``find_slide_by_marker``, and the sections that legitimately repeat
    (Financial Summary, market entry) want the whole list.
    """
    return [i for i, slide in enumerate(prs.slides) if slide_matches_marker(slide, marker)]


def find_slide_by_marker(
    prs, marker: SlideMarker, *, template: str = SLIDE_LIBRARY_TEMPLATE
) -> int:
    """The index of the one slide carrying ``marker``; raise otherwise.

    This is what replaced the hardcoded slide indices. Insisting on **exactly
    one** match is the substance of the check: zero means the slide was removed
    or renamed, and more than one means the marker no longer identifies a
    single slide, which would make the caller's choice arbitrary. Both are
    layout errors, and both used to be silent.
    """
    hits = find_slides_by_marker(prs, marker)
    if len(hits) == 1:
        return hits[0]
    if not hits:
        raise TemplateLayoutError(
            f"{template}: no slide carries the {marker.description} marker "
            f"(shape {marker.shape_name!r} containing {marker.contains!r}) in a "
            f"{len(prs.slides)}-slide deck. The slide was removed, renamed, or its "
            f"marker shape was edited."
        )
    raise TemplateLayoutError(
        f"{template}: the {marker.description} marker (shape {marker.shape_name!r} "
        f"containing {marker.contains!r}) matches {len(hits)} slides {hits}, but it "
        f"must identify exactly one. The marker is no longer unique."
    )


def find_optional_slide_by_marker(
    prs, marker: SlideMarker, *, template: str = SLIDE_LIBRARY_TEMPLATE
) -> int | None:
    """``find_slide_by_marker``, but ``None`` when the slide is absent.

    For slides the deck spec can legitimately drop — today only Key Investment
    Highlights. Two or more matches still raise: an ambiguous marker is a
    layout error whether or not the slide is optional.
    """
    hits = find_slides_by_marker(prs, marker)
    if not hits:
        return None
    return find_slide_by_marker(prs, marker, template=template)


def verify_library_slide(prs, index: int, *, template: str = SLIDE_LIBRARY_TEMPLATE) -> None:
    """Verify the slide at a raw library index is the concept the map expects.

    ``prs`` is an open python-pptx ``Presentation`` of the blank library.
    Raises :class:`TemplateLayoutError` on an out-of-range index, an index this
    map does not cover, or a marker mismatch.
    """
    marker = LIBRARY_SLIDE_MARKERS.get(index)
    if marker is None:
        raise TemplateLayoutError(
            f"{template}: no layout marker is defined for library slide index {index} "
            f"(known indices: {sorted(LIBRARY_SLIDE_MARKERS)})"
        )
    if index >= len(prs.slides):
        raise TemplateLayoutError(
            f"{template}: layout verification failed — expected the "
            f"{marker.description} slide at index {index}, but the library has only "
            f"{len(prs.slides)} slides (expected {SLIDE_LIBRARY_SLIDE_COUNT})."
        )
    verify_slide_marker(prs.slides[index], marker, template=template, slide_index=index)
