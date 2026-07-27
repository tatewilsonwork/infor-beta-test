"""Aggregate every Excel workbook produced during a deliverable into one file.

The conductor runs the workbook-producing stages (captable, ltm-metrics,
comps, ...) as siblings, each emitting a standalone `.xlsx` under the deal's
`artefacts/`. This helper is the final consolidation stage: it merges those
workbooks into a single combined workbook named `<deliverable>-<deal name>.xlsx`
(e.g. `earningsupdate-Project Atlas.xlsx`, `pitch-Project Atlas.xlsx`), with
each source contributing its sheets under a tab named after the producing
skill. The individual source workbooks are deleted only once the merge is
VERIFIED — it ran on the platform's intended backend, the cross-tab relink
succeeded, and the combined file carries no external-workbook references —
otherwise they are preserved so the merge can be retried (see
`combine_workbooks` / `CombineResult`).

Two merge backends, mirroring `excel_to_powerpoint.py`:

  - **Excel COM** (Windows + Excel) — when a `captable` source is present it is
    opened as the **base** workbook (and saved as the combined file), so the cap
    table's theme, CapIQ links and intra-workbook references survive intact; the
    other skills' sheets are copied in after it. (A blank workbook was used as
    the base before v0.5.9, which shifted theme colours and, on the openpyxl
    fallback, dropped CapIQ links — making the combined file hard to format and
    link.) CapIQ's very-hidden `__snloffice` helper sheet is dropped so it never
    surfaces as a tab. Each source's content sheets are copied **as a group in a
    single operation** so a source's intra-workbook cross-sheet references stay
    internal (the ownership `Ownership` sheet's hundreds of `='Bloomberg
    Output'!…` lookups would otherwise become external links to the soon-deleted
    source and resolve to `#REF`); the copy destination is given positionally,
    because the named `After=` argument is silently dropped by some Excel builds
    (they then copy into a brand-new workbook — a no-op append that, pre-v0.5.10,
    forced the openpyxl fallback and lost the theme).

  - **openpyxl** (Cowork / Linux / macOS, or Windows without Excel) — a
    best-effort cell-and-style copy. External data connections (CapIQ) and
    charts do NOT survive this path; use it only when COM is unavailable. The
    openpyxl copy writes formula strings with NO cached values, so after the
    merge `_recalc_with_libreoffice` re-saves the combined file through headless
    LibreOffice (recalc-on-load) to cache evaluated values while preserving the
    formulas — otherwise downstream stages (`financial-charts`) would read the
    cross-tab links as `None`. That recalc is best-effort: when LibreOffice is
    absent the workbook simply keeps its un-evaluated formulas (the COM path needs
    no such step — Excel recalcs natively on save).

Theme: the combined workbook is stamped with the INFOR brand theme
(`templates/INFORFG.thmx`) on both backends — `ApplyTheme` under COM,
`loaded_theme` under openpyxl — so it carries INFOR colours/fonts even when the
merge base is a blank workbook (no cap table) or the openpyxl fallback runs.

Cross-tab links: once every workbook is one file, a relink pass rewrites the
skills' standalone scalar handoffs into live cross-tab formulas, so the
analyst's combined workbook stays internally linked — the cap table's LTM
Revenue / Adj. EBITDA cells (`D47`/`D48`) point at the `ltm-metrics` bridge
totals, the ownership % denominator (`F35`) points at the cap table's basic
shares, and the comps (`F3`) / precedents (`C2`) output-currency cells point at
the cap table's output currency (`F5`) so the whole workbook shows one currency.
See `_relink_cross_tab_openpyxl` / `_relink_cross_tab_com`.

Hyperlinks + comments: the openpyxl merge copies each cell's hyperlink and
comment alongside its value and style (`_copy_sheet`), so the precedents source
links on `AB`–`AG` and the cap table's commented CapIQ refresh formulas
(`F7`/`F16`) — including the "Source: <url> — retrieved <date>" citation line
the captable skill appends below each formula (v0.5.31, `comment_citations`) —
survive the off-Windows merge; the COM backend preserves both natively via
Excel's copy.

Tab naming: a single-sheet source becomes one tab named after the skill
(`captable`, `ltm-metrics`); a multi-sheet source contributes one tab per
sheet named `<skill>-<sheet>`. Excel's 31-char / forbidden-character / unique
constraints are enforced in both backends.
"""

from __future__ import annotations

import re
import shutil
import sys
import tempfile
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping

from naming import safe_filename
from template_layout import (
    CAP_TABLE_BASIC_SHARES_ANCHOR,
    CAP_TABLE_FX_ANCHOR,
    CAP_TABLE_LTM_ANCHORS,
    CAP_TABLE_OUTPUT_CCY_ANCHOR,
    CAP_TABLE_SHEET,
    COMPS_OUTPUT_CCY_ANCHORS,
    COMPS_SHEET,
    OWNERSHIP_SHEET,
    OWNERSHIP_TOTAL_SHARES_ANCHORS,
    PRECEDENTS_OUTPUT_CCY_ANCHORS,
    PRECEDENTS_SHEET,
    verify_anchors,
)

# xlOpenXMLWorkbook — the .xlsx SaveAs file format for Excel COM.
_XL_OPEN_XML_WORKBOOK = 51
_EXCEL_SHEET_NAME_MAX = 31
_FORBIDDEN_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]+")

# INFOR brand theme, applied to every combined workbook so it keeps INFOR
# colours/fonts regardless of merge base or backend. Ships beside the templates.
_THEME_FILENAME = "INFORFG.thmx"
# A .thmx is a zip; the theme XML Excel/openpyxl want lives at this member path.
_THMX_THEME_MEMBER = "theme/theme/theme1.xml"

# CapIQ's Excel add-in stows formula metadata in a very-hidden helper sheet
# named "__snloffice". Copied verbatim it surfaces as a garbled (CJK-looking)
# tab in the combined workbook, so it's excluded from aggregation. The add-in
# regenerates it on refresh, so dropping it from the merged file is harmless.
_CAPIQ_HELPER_SHEET = re.compile(r"^__snl", re.IGNORECASE)


def _is_capiq_helper_sheet(name: str) -> bool:
    return bool(_CAPIQ_HELPER_SHEET.match(name.strip()))


def _default_theme_path() -> Path:
    """Path to the shipped INFOR theme (`templates/INFORFG.thmx`)."""
    return Path(__file__).resolve().parent.parent / "templates" / _THEME_FILENAME


def _resolve_theme_path(theme_path: Path | str | None) -> Path | None:
    """Return an existing theme path (caller override or the shipped default), or
    None when no theme file is available — theming is then skipped."""
    candidate = Path(theme_path) if theme_path is not None else _default_theme_path()
    return candidate if candidate.exists() else None


def _extract_theme_xml(theme_path: Path) -> bytes | None:
    """Return the `theme1.xml` bytes from a `.thmx` package for openpyxl theme
    injection, or None if the file is missing / not a valid theme zip."""
    try:
        with zipfile.ZipFile(theme_path) as z:
            return z.read(_THMX_THEME_MEMBER)
    except (OSError, KeyError, zipfile.BadZipFile):
        return None


def combined_filename(deliverable_type: str, deal_name: str) -> str:
    """Return `<deliverable>-<deal name>.xlsx`.

    The deliverable prefix drops hyphens so `earnings-update` reads as
    `earningsupdate` (matching the analyst-facing naming), while `pitch`
    stays `pitch`.
    """
    prefix = deliverable_type.replace("-", "").strip() or "deliverable"
    return f"{prefix}-{safe_filename(deal_name, default='Deal')}.xlsx"


def _excel_safe_sheet_name(name: str) -> str:
    """Sanitize a tab name to Excel's rules: no []:*?/\\, <=31 chars, non-empty."""
    safe = _FORBIDDEN_SHEET_CHARS.sub("-", name).strip().strip("'")
    safe = safe[:_EXCEL_SHEET_NAME_MAX].strip()
    return safe or "Sheet"


def _unique_sheet_name(name: str, used: set[str]) -> str:
    """Return a sheet name unique within `used` (case-insensitive), <=31 chars."""
    candidate = _excel_safe_sheet_name(name)
    if candidate.casefold() not in used:
        used.add(candidate.casefold())
        return candidate
    for i in range(2, 1000):
        suffix = f" ({i})"
        trimmed = candidate[: _EXCEL_SHEET_NAME_MAX - len(suffix)].strip()
        attempt = f"{trimmed}{suffix}"
        if attempt.casefold() not in used:
            used.add(attempt.casefold())
            return attempt
    raise ValueError(f"could not derive a unique sheet name from {name!r}")


def _tab_name(skill: str, original_sheet: str, sheet_count: int) -> str:
    """Single-sheet source -> the skill name; multi-sheet -> the original sheet
    names, unprefixed.

    A single-sheet workbook's sheet is usually generic (`Sheet1`, `Cap with
    Links`), so the skill name is the more useful tab label. A multi-sheet source
    carries self-describing sheet names (`Ownership`, `Bloomberg Output`) that the
    analyst expects to see verbatim — and prefixing them would force a rename that
    breaks the source's intra-workbook cross-sheet references (the ownership
    `Ownership` sheet's `='Bloomberg Output'!…` lookups -> `#REF`). So multi-sheet
    sources keep their sheet names; only the `_unique_sheet_name` collision guard
    can alter them."""
    return skill if sheet_count == 1 else original_sheet


def _resolve_sources(
    sources: Mapping[str, str | Path | None],
) -> tuple[list[tuple[str, Path]], list[str]]:
    """Drop None / missing-file entries; return (kept, skipped-skill-names)."""
    kept: list[tuple[str, Path]] = []
    skipped: list[str] = []
    for skill, raw in sources.items():
        if raw is None:
            skipped.append(skill)
            continue
        path = Path(str(raw)).expanduser()
        if not path.exists():
            skipped.append(skill)
            continue
        kept.append((skill, path.resolve()))
    return kept, skipped


@dataclass(frozen=True)
class CombineResult:
    """Outcome of `combine_workbooks` — the combined path plus the verification
    facts the source-deletion gate was decided on, so the calling stage can put
    them in its outputs instead of the analyst discovering a degraded merge from
    a broken workbook later.

    Implements `__fspath__`, so the result can be passed anywhere a path is
    accepted (`Path(result)`, `load_workbook(result)`), but callers should
    prefer the explicit `output_path`.
    """

    output_path: Path
    backend: str                    # "com" | "openpyxl" — the backend that ran
    degraded: bool                  # win32 only: COM failed, openpyxl fallback ran
    relink_ok: bool                 # the cross-tab relink pass completed
    external_refs: tuple[str, ...]  # external-workbook references still present
    sources_deleted: bool           # the individual source workbooks were removed
    kept_sources: tuple[Path, ...]  # source workbooks still on disk
    warnings: tuple[str, ...]       # analyst-facing warnings (also on stderr)

    def __fspath__(self) -> str:
        return str(self.output_path)


def combine_workbooks(
    *,
    sources: Mapping[str, str | Path | None],
    output_dir: Path | str,
    deliverable_type: str,
    deal_name: str,
    delete_sources: bool = True,
    theme_path: Path | str | None = None,
) -> CombineResult:
    """Merge the source workbooks into one combined `.xlsx`.

    Returns a `CombineResult` carrying the combined workbook's path plus the
    merge-verification facts (backend, degradation, relink success, external
    references, whether the sources were deleted).

    `sources` maps a producing-skill name (used as the tab name) to that
    skill's workbook path. None values and non-existent paths are skipped, so
    optional upstream workbooks (e.g. a comps workbook that wasn't produced)
    can be wired in unconditionally. Raises ValueError if nothing is left to
    combine.

    When `delete_sources` is True (the default), the individual source files
    are removed only after the merge is VERIFIED — all of:

      - the merge ran on the platform's intended backend (openpyxl IS the
        intended backend off-Windows; on Windows, a COM failure that fell back
        to the lossy openpyxl merge is a degraded result — the full-fidelity
        sources are the only way to retry with CapIQ links/charts intact);
      - the cross-tab relink pass succeeded;
      - the combined workbook carries no external-workbook references (an
        `xl/externalLinks` part or a `[n]`-indexed sheet formula would point at
        the very files being deleted — a permanent #REF!/#N/A).

    Otherwise the sources are preserved and a warning is printed to stderr
    (and returned on the result) so the analyst can retry the merge.
    `delete_sources=False` always keeps the sources, as before.

    `theme_path` overrides the brand theme stamped on the combined workbook;
    it defaults to the shipped `templates/INFORFG.thmx`. Theming is skipped
    silently if no theme file is found.
    """
    kept, _skipped = _resolve_sources(sources)
    if not kept:
        raise ValueError("no workbooks to combine — every source was None or missing")

    # Layout pre-flight for the relink pass, on the shared layer BOTH backends
    # pass through (the sources are still openpyxl-readable .xlsx files here).
    # A re-saved template with shifted rows would otherwise make the relink
    # write its cross-tab formulas into the wrong cells — a silent wrong number.
    _verify_relink_layout(kept)

    out_dir = Path(output_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / combined_filename(deliverable_type, deal_name)

    theme = _resolve_theme_path(theme_path)
    used_openpyxl = False
    if sys.platform == "win32":
        try:
            relink_ok = _combine_via_com(kept, output_path, theme)
        except RuntimeError as exc:
            print(
                f"[workbook-aggregator] Excel COM merge failed ({exc}); falling "
                "back to the openpyxl merge (CapIQ connections and charts will "
                "not survive).",
                file=sys.stderr,
            )
            relink_ok = _combine_via_openpyxl(kept, output_path, theme)
            used_openpyxl = True
    else:
        relink_ok = _combine_via_openpyxl(kept, output_path, theme)
        used_openpyxl = True

    # The openpyxl merge writes formula strings with NO cached values, so the
    # cross-tab links (the financial-summary `=INDEX('ltm-metrics'!…)` LTM lookups,
    # the cap table's relinked LTM cells, CapIQ-degraded cells) sit un-evaluated
    # until the analyst opens the file — and a downstream stage like
    # `financial-charts` would read `None`. Recalc the merged file on the
    # non-Windows path with LibreOffice so it carries evaluated values (the COM
    # path already recalcs natively via Excel on save). Formulas are preserved.
    if used_openpyxl:
        _recalc_with_libreoffice(output_path)

    # ── Source-deletion gate ────────────────────────────────────────────────
    # Deleting the sources is the one irreversible act in this stage, so it is
    # gated on VERIFIED success: intended backend, relink succeeded, and no
    # external-workbook references left in the combined file. Anything short of
    # that preserves the sources for a retry — a degraded/partially-linked
    # combined workbook plus deleted sources is unrecoverable.
    degraded = used_openpyxl and sys.platform == "win32"
    external_refs = tuple(_find_external_workbook_refs(output_path))

    warnings: list[str] = []
    if degraded:
        warnings.append(
            "the merge DEGRADED to the openpyxl fallback (Excel COM failed on "
            "this Windows machine): live CapIQ connections, charts, and full "
            "formatting did not survive"
        )
    if not relink_ok:
        warnings.append(
            "the cross-tab relink FAILED: the combined workbook's cap-table LTM "
            "/ ownership-denominator / output-currency / financial-summary links "
            "may be broken or still bound to the standalone source workbooks"
        )
    if external_refs:
        shown = "; ".join(external_refs[:5])
        more = f" (+{len(external_refs) - 5} more)" if len(external_refs) > 5 else ""
        warnings.append(
            "the combined workbook still carries EXTERNAL workbook references — "
            f"deleting the sources would break them permanently: {shown}{more}"
        )
    for message in warnings:
        print(f"[workbook-aggregator] WARNING: {message}", file=sys.stderr)

    deletable = [p for _skill, p in kept if p.resolve() != output_path.resolve()]
    sources_deleted = False
    kept_sources: tuple[Path, ...] = tuple(deletable)
    if delete_sources:
        if warnings:
            print(
                "[workbook-aggregator] the individual source workbooks were "
                "PRESERVED (not deleted) so the merge can be retried: "
                + ", ".join(str(p) for p in deletable),
                file=sys.stderr,
            )
        else:
            for path in deletable:
                path.unlink(missing_ok=True)
            sources_deleted = True
            kept_sources = ()

    return CombineResult(
        output_path=output_path,
        backend="openpyxl" if used_openpyxl else "com",
        degraded=degraded,
        relink_ok=relink_ok,
        external_refs=external_refs,
        sources_deleted=sources_deleted,
        kept_sources=kept_sources,
        warnings=tuple(warnings),
    )


def _recalc_with_libreoffice(workbook_path: Path) -> bool:
    """Recalc an openpyxl-merged workbook in place with headless LibreOffice.

    Loads the workbook in LibreOffice with recalc-on-load and re-saves it, caching
    the evaluated values **while preserving the formula strings** (LibreOffice's
    `.xlsx` export keeps formulas, so analyst auditability is intact). Downstream
    stages then read evaluated numbers instead of `None`. Reuses the recalc-on-load
    `_soffice_convert` helper that `financial_charts` / `excel_to_powerpoint` use.

    Best-effort and non-fatal: returns ``True`` if the recalc ran and replaced the
    file, ``False`` if LibreOffice (`soffice`/`libreoffice`) is unavailable or the
    conversion produced nothing — in which case the merged workbook keeps its
    un-evaluated formulas (the analyst's Excel recalcs them on open) rather than the
    stage aborting. Never raises: ANY failure inside (the conversion, the file
    replace, or the post-recalc `_strip_lo_union_operators` fix — which can raise
    `zipfile.BadZipFile`/`OSError`/`UnicodeDecodeError` on a bad LibreOffice
    export) degrades to that same path. If the merged file was already replaced
    by the LibreOffice re-save when the failure hit, the pre-recalc original is
    restored — the re-save may carry the `~` union operators the strip step could
    not fix, and Excel would repair-strip those formulas on open.
    """
    # Shared locator, not a bare shutil.which: the Windows MSI never puts soffice
    # on PATH, which would silently skip the recalc on a dev box that has it.
    from excel_to_powerpoint import find_soffice

    soffice = find_soffice()
    if soffice is None:
        print(
            "[workbook-aggregator] LibreOffice (soffice/libreoffice) not found on "
            "PATH or in the standard install locations; the combined workbook keeps "
            "its un-evaluated cross-tab formulas (Excel will recalc them on open).",
            file=sys.stderr,
        )
        return False

    original_bytes: bytes | None = None
    replaced = False
    try:
        from excel_to_powerpoint import _soffice_convert  # recalc-on-load helper

        original_bytes = workbook_path.read_bytes()
        with tempfile.TemporaryDirectory() as tmp_dir:
            _soffice_convert(
                soffice, workbook_path, "xlsx:Calc MS Excel 2007 XML", Path(tmp_dir)
            )
            recalced = Path(tmp_dir) / f"{workbook_path.stem}.xlsx"
            if not recalced.exists():
                print(
                    "[workbook-aggregator] LibreOffice produced no recalculated "
                    "workbook; leaving formulas un-evaluated.",
                    file=sys.stderr,
                )
                return False
            shutil.copyfile(recalced, workbook_path)
            replaced = True
        _strip_lo_union_operators(workbook_path)
        return True
    except Exception as exc:
        if replaced and original_bytes is not None:
            try:
                workbook_path.write_bytes(original_bytes)
            except OSError:
                pass  # restore is best-effort; the note below still fires
        print(
            f"[workbook-aggregator] LibreOffice recalc failed ({exc}); leaving the "
            "combined workbook's formulas un-evaluated.",
            file=sys.stderr,
        )
        return False


# Formula elements in a worksheet part. Formula text never contains a raw '<',
# so a non-'<' body match is exact.
_SHEET_FORMULA_RE = re.compile(r"(<f\b[^>]*>)([^<]*)(</f>)")


def _excel_union_commas(formula: str) -> str:
    """Rewrite LibreOffice's ``~`` range-union operator to Excel's ``,``.

    Skips string literals (between double quotes, where ``~`` is Excel's
    wildcard-escape character and must survive). The formula text comes straight
    from sheet XML, so a quote may appear raw (``"``) or entity-escaped
    (``&quot;``) depending on the writer.
    """
    parts = re.split(r'("|&quot;)', formula)
    in_string = False
    out: list[str] = []
    for part in parts:
        if part in ('"', "&quot;"):
            in_string = not in_string
            out.append(part)
        else:
            out.append(part if in_string else part.replace("~", ","))
    return "".join(out)


def _strip_lo_union_operators(workbook_path: Path) -> int:
    """Fix LibreOffice's ``~`` union operator in every sheet formula, in place.

    LibreOffice's ``.xlsx`` export writes a multi-area union inside a
    parenthesized reference argument with its own ``~`` operator — e.g. the
    comps / precedents quartile rows come back as
    ``PERCENTILE.INC((L10:L15~L20:L25~L30:L35),0.25)`` — which Excel cannot
    parse: it opens the workbook with a "Removed Records: Formula from
    /xl/worksheets/sheetN.xml" repair that strips every such formula. Rewrites
    ``~`` to ``,`` at the XML level (string literals excluded) so the recalc's
    cached values are preserved byte-for-byte. Returns the number of sheet
    parts rewritten.
    """
    with zipfile.ZipFile(workbook_path) as zin:
        if not any(
            b"~" in zin.read(item.filename)
            for item in zin.infolist()
            if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", item.filename)
        ):
            return 0
        entries = []
        fixed = 0
        for item in zin.infolist():
            data = zin.read(item.filename)
            if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", item.filename) and b"~" in data:
                xml = data.decode("utf-8")
                new_xml = _SHEET_FORMULA_RE.sub(
                    lambda m: m.group(1) + _excel_union_commas(m.group(2)) + m.group(3), xml
                )
                if new_xml != xml:
                    data = new_xml.encode("utf-8")
                    fixed += 1
            entries.append((item, data))
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", dir=workbook_path.parent, delete=False
    ) as tmp:
        tmp_path = Path(tmp.name)
    try:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item, data in entries:
                zout.writestr(item, data)
        shutil.move(tmp_path, workbook_path)
    finally:
        tmp_path.unlink(missing_ok=True)
    if fixed:
        print(
            f"[workbook-aggregator] rewrote LibreOffice '~' range unions to ',' in "
            f"{fixed} sheet part(s) so Excel opens the combined workbook without repair.",
            file=sys.stderr,
        )
    return fixed


# ─── External-workbook reference verification ───────────────────────────────
# An external-workbook reference in a SAVED .xlsx sheet formula is stored in
# the indexed form `[n]` (e.g. `'[1]ltm-metrics'!$B$3`), where n names an
# `xl/externalLinks/externalLinkN.xml` part. A digit-only bracket never appears
# in a purely internal A1-style stored formula (structured table references
# carry column names, not bare digits), so the pattern is an exact signal.
_EXTERNAL_WORKBOOK_REF = re.compile(r"\[\d+\]")


def _formula_has_external_ref(formula: str) -> bool:
    """True when a sheet formula carries a `[n]`-indexed external-workbook ref.

    Bracket text inside string literals is ignored (same raw-`"` / `&quot;`
    string-walking as `_excel_union_commas` — the formula text comes straight
    from sheet XML, so quotes may appear in either form).
    """
    parts = re.split(r'("|&quot;)', formula)
    in_string = False
    for part in parts:
        if part in ('"', "&quot;"):
            in_string = not in_string
        elif not in_string and _EXTERNAL_WORKBOOK_REF.search(part):
            return True
    return False


def _find_external_workbook_refs(workbook_path: Path) -> list[str]:
    """Scan the combined workbook for external-workbook references.

    Returns one finding per hit: an `xl/externalLinks/…` part (a workbook-level
    link record) or a sheet formula still carrying an indexed `[n]` external
    reference (a live cell binding — the v0.5.16 `'[1]ltm-metrics'` symptom a
    failed relink leaves behind). An empty list means the combined workbook is
    self-contained and the standalone sources are safe to delete. Fail-closed:
    a scan error is itself returned as a finding, so an unreadable combined
    workbook blocks source deletion instead of green-lighting it.
    """
    findings: list[str] = []
    try:
        with zipfile.ZipFile(workbook_path) as z:
            names = z.namelist()
            findings.extend(
                n for n in names if n.startswith("xl/externalLinks/") and n.endswith(".xml")
            )
            for name in names:
                if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
                    continue
                xml = z.read(name).decode("utf-8", "replace")
                for match in _SHEET_FORMULA_RE.finditer(xml):
                    if _formula_has_external_ref(match.group(2)):
                        findings.append(f"{name}: ={match.group(2)[:70]}")
    except Exception as exc:
        findings.append(f"external-reference scan failed ({exc})")
    return findings


# ─── Cross-tab relink ────────────────────────────────────────────────────────
# Once every workbook is merged into one file, these cells are rewritten from
# the skills' standalone scalar handoffs to live cross-tab formulas, so the
# analyst's combined workbook stays internally linked: edit the LTM bridge and
# the cap table follows; the ownership % tracks the cap table's share count.
_LTM_REVENUE_LABEL = "(=) LTM Revenue"
_LTM_EBITDA_LABELS = ("(=) LTM Adj. EBITDA", "(=) LTM EBITDA")
# The financial-summary tab's LTM cells are authored as `=INDEX('ltm-metrics'!…)`
# label lookups (sheet token below). The COM copy rewrites that cross-source
# reference as an EXTERNAL link (`'[1]ltm-metrics'!` / a full source path), which
# resolves to #N/A; the relink re-points it at the sibling `ltm-metrics` tab now
# in the combined workbook. (The openpyxl path copies the formula string verbatim,
# so it is already internal — the relink is a harmless no-op there.)
_FS_SKILL = "financial-summary"
_LTM_SKILL = "ltm-metrics"
_FS_LTM_REF_SHEET = "ltm-metrics"  # the sheet token the financial-summary links use
_CAP_LTM_REVENUE_CELL = "D47"   # cap table LTM Revenue (millions, F5 currency)
_CAP_LTM_EBITDA_CELL = "D48"    # cap table LTM Adj. EBITDA
_CAP_BASIC_SHARES_CELL = "F17"  # cap table basic shares outstanding (millions)
_OWN_DENOM_CELL = "F35"         # ownership % denominator (full units)
_OWN_SHARES_SCALE = 1_000_000   # cap table is in millions; ownership is full units

# Output-currency cells that should mirror the cap table's output currency (F5)
# rather than each skill's standalone literal, so the combined workbook shows one
# consistent currency (and updating F5 flows through). Maps producing-skill -> the
# cell on that skill's tab. They are restyled to match the cap table's F5
# (Palatino 9, blue) since the bare comps template cell is Calibri 11.
_CAP_OUTPUT_CCY_CELL = "F5"
_OUTPUT_CCY_LINKS = {"comps": "F3", "precedents": "C2"}
_OUTPUT_CCY_LINK_FONT = ("Palatino Linotype", 9.0, "0000FF")  # name, size, aRGB-less hex

# Shared relink-failure trace. Best-effort by design — a relink failure must not
# lose the successfully merged workbook — but never silent: an unrelinked
# financial-summary tab is the v0.5.16 blank-LTM-bar symptom. The caller
# (`combine_workbooks`) additionally keeps the standalone sources when the
# relink reports failure, so the links stay repairable.
_RELINK_FAILURE_NOTE = (
    "[workbook-aggregator] cross-tab relink failed ({exc}); the combined "
    "workbook is saved, but its cap-table LTM / ownership-denominator / "
    "currency / financial-summary links may be broken or still bound to the "
    "standalone source workbooks."
)


def _relink_source_sheet(wb, preferred: str):
    """The sheet a source contributes as its relink target: the preferred
    template sheet name when present, else the first non-CapIQ-helper sheet
    (which is what the merge names after the skill)."""
    if preferred in wb.sheetnames:
        return wb[preferred]
    for name in wb.sheetnames:
        if not _is_capiq_helper_sheet(name):
            return wb[name]
    return wb.active


def _verify_relink_layout(kept: list[tuple[str, Path]]) -> None:
    """Verify the relinked cells' sentinel anchors on the source workbooks.

    Runs BEFORE either merge backend, so the Excel-COM and openpyxl paths share
    one verification (the sources are plain .xlsx files here; nothing has been
    merged or deleted yet, so raising loses no work). Only the relinks that
    will actually fire are checked: every relink is keyed off the cap table, and
    each partner tab adds its own cells — captable+ltm-metrics ⇒ D47/D48 (and
    the F7 FX rate the written formulas multiply by), captable+ownership ⇒
    F17 → F35, captable+comps ⇒ F5 → F3, captable+precedents ⇒ F5 → C2.
    Raises TemplateLayoutError when a template's layout has shifted.
    """
    from openpyxl import load_workbook

    skills = {skill for skill, _ in kept}
    if "captable" not in skills:
        return  # every relink is keyed off the cap table

    cap_anchors = []
    if "ltm-metrics" in skills:
        cap_anchors += [*CAP_TABLE_LTM_ANCHORS, CAP_TABLE_FX_ANCHOR]
    if "ownership" in skills:
        cap_anchors.append(CAP_TABLE_BASIC_SHARES_ANCHOR)
    if "comps" in skills or "precedents" in skills:
        cap_anchors.append(CAP_TABLE_OUTPUT_CCY_ANCHOR)

    partner_checks = {
        "ownership": (OWNERSHIP_SHEET, OWNERSHIP_TOTAL_SHARES_ANCHORS),
        "comps": (COMPS_SHEET, COMPS_OUTPUT_CCY_ANCHORS),
        "precedents": (PRECEDENTS_SHEET, PRECEDENTS_OUTPUT_CCY_ANCHORS),
    }
    for skill, path in kept:
        if skill == "captable" and cap_anchors:
            wb = load_workbook(path, data_only=False)
            try:
                verify_anchors(
                    _relink_source_sheet(wb, CAP_TABLE_SHEET), cap_anchors, template=path.name
                )
            finally:
                wb.close()
        elif skill in partner_checks:
            preferred, anchors = partner_checks[skill]
            wb = load_workbook(path, data_only=False)
            try:
                verify_anchors(_relink_source_sheet(wb, preferred), anchors, template=path.name)
            finally:
                wb.close()


def _quote_sheet(name: str) -> str:
    """Single-quote a sheet name for use in a formula (handles spaces/hyphens)."""
    return "'" + name.replace("'", "''") + "'"


def _internalize_external_sheet_ref(formula: str, ext_sheet: str, internal_tab: str) -> str:
    """Re-point a formula's external-workbook ref to ``ext_sheet`` at ``internal_tab``.

    Matches a quoted external sheet reference — `'…[<wb>]<ext_sheet>'` (the `[<wb>]`
    is the external workbook index or path COM inserts) — and rewrites it to the
    internal `'<internal_tab>'`. No-op when the formula has no such external ref.
    """
    pattern = r"'[^']*\[[^\]]*\]" + re.escape(ext_sheet) + r"'"
    return re.sub(pattern, _quote_sheet(internal_tab), formula)


def _relink_financial_summary_openpyxl(combined, skill_to_tab: dict[str, str]) -> None:
    """Re-point the financial-summary tab's LTM links at the sibling ltm-metrics tab."""
    fs = skill_to_tab.get(_FS_SKILL)
    ltm = skill_to_tab.get(_LTM_SKILL)
    names = set(combined.sheetnames)
    if fs not in names or ltm not in names:
        return
    ws = combined[fs]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and _FS_LTM_REF_SHEET in v:
                new = _internalize_external_sheet_ref(v, _FS_LTM_REF_SHEET, ltm)
                if new != v:
                    cell.value = new


def _relink_financial_summary_com(combined, skill_to_tab: dict[str, str]) -> None:
    """COM counterpart of `_relink_financial_summary_openpyxl`.

    Copying the financial-summary sheet binds its LTM-link formulas to an EXTERNAL
    workbook relationship (the soon-deleted source). Excel's `.Formula` getter
    collapses that to an internal-looking `'ltm-metrics'!` string, but the cell
    still resolves to #N/A because the binding is external — so a string-compare
    rewrite is a no-op and the bug survives. Re-*assigning* the (internalized)
    formula forces Excel to re-bind it to the sibling `ltm-metrics` tab, after
    which it resolves on recalc. So re-set every LTM-link cell unconditionally.
    """
    fs = skill_to_tab.get(_FS_SKILL)
    ltm = skill_to_tab.get(_LTM_SKILL)
    names = {combined.Sheets(i + 1).Name for i in range(combined.Sheets.Count)}
    if fs not in names or ltm not in names:
        return
    ws = combined.Worksheets(fs)
    try:
        used = ws.UsedRange
        last_row = used.Row + used.Rows.Count - 1
        last_col = used.Column + used.Columns.Count - 1
    except Exception:
        last_row, last_col = 12, 10
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.Cells(r, c)
            f = cell.Formula
            if isinstance(f, str) and f.startswith("=") and _FS_LTM_REF_SHEET in f and "MATCH(" in f:
                cell.Formula = _internalize_external_sheet_ref(f, _FS_LTM_REF_SHEET, ltm)


def _find_label_row_openpyxl(ws, prefixes) -> int | None:
    """Row (1-based) of the first col-A cell whose text starts with any prefix."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        value = row[0].value
        if isinstance(value, str) and any(value.strip().startswith(p) for p in prefixes):
            return row[0].row
    return None


def _relink_cross_tab_openpyxl(combined, skill_to_tab: dict[str, str]) -> bool:
    """Wire the cap table's LTM cells + ownership denominator to sibling tabs.

    No-op unless the relevant tabs exist. The LTM bridge total rows are dynamic
    (they depend on the segment/component counts), so they are located by their
    `(=) LTM Revenue` / `(=) LTM Adj. EBITDA` labels rather than a fixed cell.
    CapIQ links are already lost on the openpyxl path, so this only restores the
    cross-tab references; the COM path preserves the live CapIQ formulas too.

    Best-effort like its COM counterpart — never raises; a relink failure must
    not lose the successfully merged workbook. Returns True when the pass
    completed, False on failure (with a stderr trace) so the caller can gate
    source deletion on it.
    """
    try:
        cap = skill_to_tab.get("captable")
        ltm = skill_to_tab.get("ltm-metrics")
        own = skill_to_tab.get("ownership")
        names = set(combined.sheetnames)
        if cap in names and ltm in names:
            ltm_ws = combined[ltm]
            rev = _find_label_row_openpyxl(ltm_ws, (_LTM_REVENUE_LABEL,))
            ebitda = _find_label_row_openpyxl(ltm_ws, _LTM_EBITDA_LABELS)
            q = _quote_sheet(ltm)
            if rev is not None:
                combined[cap][_CAP_LTM_REVENUE_CELL] = f"={q}!B{rev}*F7"
            if ebitda is not None:
                combined[cap][_CAP_LTM_EBITDA_CELL] = f"={q}!B{ebitda}*F7"
        if cap in names and own in names:
            combined[own][_OWN_DENOM_CELL] = (
                f"={_quote_sheet(cap)}!{_CAP_BASIC_SHARES_CELL}*{_OWN_SHARES_SCALE}"
            )
        if cap in names:
            from openpyxl.styles import Font

            name, size, color = _OUTPUT_CCY_LINK_FONT
            for skill, cell_ref in _OUTPUT_CCY_LINKS.items():
                tab = skill_to_tab.get(skill)
                if tab in names:
                    cell = combined[tab][cell_ref]
                    cell.value = f"={_quote_sheet(cap)}!{_CAP_OUTPUT_CCY_CELL}"
                    cell.font = Font(name=name, size=size, color=color)

        _relink_financial_summary_openpyxl(combined, skill_to_tab)
        return True
    except Exception as exc:
        print(_RELINK_FAILURE_NOTE.format(exc=exc), file=sys.stderr)
        return False


def _find_label_row_com(ws, prefixes) -> int | None:
    """COM counterpart of `_find_label_row_openpyxl`."""
    try:
        used = ws.UsedRange
        last = used.Row + used.Rows.Count - 1
    except Exception:
        last = 200
    for r in range(1, last + 1):
        value = ws.Cells(r, 1).Value
        if isinstance(value, str) and any(value.strip().startswith(p) for p in prefixes):
            return r
    return None


def _relink_cross_tab_com(combined, skill_to_tab: dict[str, str]) -> bool:
    """COM counterpart of `_relink_cross_tab_openpyxl`; best-effort (never raises
    — a relink failure must not lose the successfully merged workbook). Returns
    True when the pass completed, False on failure (with a stderr trace) so the
    caller can gate source deletion on it."""
    try:
        cap = skill_to_tab.get("captable")
        ltm = skill_to_tab.get("ltm-metrics")
        own = skill_to_tab.get("ownership")
        names = {combined.Sheets(i + 1).Name for i in range(combined.Sheets.Count)}
        if cap in names and ltm in names:
            ltm_ws = combined.Worksheets(ltm)
            rev = _find_label_row_com(ltm_ws, (_LTM_REVENUE_LABEL,))
            ebitda = _find_label_row_com(ltm_ws, _LTM_EBITDA_LABELS)
            cap_ws = combined.Worksheets(cap)
            q = _quote_sheet(ltm)
            if rev is not None:
                cap_ws.Range(_CAP_LTM_REVENUE_CELL).Formula = f"={q}!B{rev}*F7"
            if ebitda is not None:
                cap_ws.Range(_CAP_LTM_EBITDA_CELL).Formula = f"={q}!B{ebitda}*F7"
        if cap in names and own in names:
            combined.Worksheets(own).Range(_OWN_DENOM_CELL).Formula = (
                f"={_quote_sheet(cap)}!{_CAP_BASIC_SHARES_CELL}*{_OWN_SHARES_SCALE}"
            )
        if cap in names:
            name, size, hex_color = _OUTPUT_CCY_LINK_FONT
            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
            bgr = r + (g << 8) + (b << 16)  # Excel Font.Color is a BGR-packed long
            for skill, cell_ref in _OUTPUT_CCY_LINKS.items():
                tab = skill_to_tab.get(skill)
                if tab in names:
                    rng = combined.Worksheets(tab).Range(cell_ref)
                    rng.Formula = f"={_quote_sheet(cap)}!{_CAP_OUTPUT_CCY_CELL}"
                    rng.Font.Name = name
                    rng.Font.Size = size
                    rng.Font.Color = bgr
        _relink_financial_summary_com(combined, skill_to_tab)
        return True
    except Exception as exc:
        print(_RELINK_FAILURE_NOTE.format(exc=exc), file=sys.stderr)
        return False


def _open_workbook(excel, path: Path, *, read_only: bool):
    """Open a workbook robustly under COM.

    Works around two pywin32 quirks: `Workbooks.Open` sometimes returns None even
    though the workbook opened (grab the latest one — but only when the open count
    actually rose, so a genuine failure raises instead of silently handing back
    the wrong workbook), and a transient file-lock/COM race can fail the first
    open (one short retry absorbs it). Raises RuntimeError on real failure so the
    caller falls back to the openpyxl backend."""
    last_exc: Exception | None = None
    for attempt in range(2):
        before = excel.Workbooks.Count
        try:
            wb = excel.Workbooks.Open(str(path), ReadOnly=read_only, UpdateLinks=0)
        except Exception as exc:  # COM error — retry once, then surface
            last_exc = exc
            wb = None
        if wb is None and excel.Workbooks.Count > before:
            wb = excel.Workbooks(excel.Workbooks.Count)
        if wb is not None:
            return wb
        time.sleep(0.4 * (attempt + 1))
    raise RuntimeError(f"Excel failed to open workbook: {path} ({last_exc})")


def _rename_and_clean_base(combined, skill: str, used_names: set, skill_to_tab: dict) -> None:
    """Rename the base (cap table) workbook's content sheets to the skill tab name
    and drop any CapIQ helper sheet, so the base reads like any merged source."""
    original = [combined.Worksheets(i + 1).Name for i in range(combined.Worksheets.Count)]
    content_count = sum(1 for n in original if not _is_capiq_helper_sheet(n))
    drop: list = []
    for idx in range(1, combined.Sheets.Count + 1):
        sheet = combined.Sheets(idx)
        if _is_capiq_helper_sheet(sheet.Name):
            drop.append(sheet)
            continue
        new_name = _unique_sheet_name(_tab_name(skill, sheet.Name, content_count), used_names)
        sheet.Name = new_name
        skill_to_tab.setdefault(skill, new_name)
    for sheet in drop:
        sheet.Visible = -1  # xlSheetVisible — can't delete a very-hidden sheet
        sheet.Delete()


def _combine_via_com(
    sources: list[tuple[str, Path]], output_path: Path, theme_path: Path | None = None
) -> bool:
    """Merge with Excel COM, preserving formulas, links, charts, and formatting.

    When a `captable` source is present it is opened as the BASE workbook (and
    saved as the combined file), so the cap table's theme, CapIQ links and
    intra-workbook references survive intact; the other skills' sheets are copied
    in after it. Without a cap table, a blank workbook is the base (legacy
    behaviour). After the merge, a best-effort cross-tab relink wires the cap
    table's LTM cells to the ltm-metrics tab and the ownership denominator to the
    cap table, and (when `theme_path` is given) the INFOR brand theme is stamped
    on so a blank-base merge doesn't ship the default Office theme.

    Returns True when the cross-tab relink pass succeeded, False when it failed
    (the merged file is still saved) — the caller gates source deletion on it.
    """
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError(
            "pywin32 is required for COM-based workbook aggregation "
            "(Windows + Microsoft Excel only)"
        ) from exc

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        used_names: set[str] = set()
        skill_to_tab: dict[str, str] = {}

        # The cap table (when present) is the base workbook so its formatting and
        # CapIQ links survive; the remaining sources are copied into it.
        base = next(((s, p) for s, p in sources if s == "captable"), None)
        if base is not None:
            rest = [entry for entry in sources if entry != base]
            combined = _open_workbook(excel, base[1], read_only=False)
            seed_sheets: list[str] = []
            _rename_and_clean_base(combined, base[0], used_names, skill_to_tab)
        else:
            rest = list(sources)
            combined = excel.Workbooks.Add()
            # Sheets present in the blank workbook; deleted once real sheets land.
            seed_sheets = [combined.Sheets(i + 1).Name for i in range(combined.Sheets.Count)]

        # Copy each source's content sheets into `combined`. All of a source's
        # sheets are copied in ONE operation so any intra-workbook cross-sheet
        # references survive as internal references (copying sheet-by-sheet turns
        # them into external links to the soon-deleted source -> #REF; this is the
        # ownership `Ownership` -> `Bloomberg Output` case). The destination MUST
        # be the active workbook or Excel copies into a NEW workbook instead of
        # appending here, so re-activate `combined` first; and the destination is
        # passed POSITIONALLY (Before=None, After=<last sheet>) because the named
        # `After=` form is silently dropped by some Excel builds, which then copy
        # into a stray new workbook (a no-op append). CapIQ `__snl*` helper sheets
        # are skipped at the source so they never surface as a tab.
        for skill, path in rest:
            src = _open_workbook(excel, path, read_only=True)
            try:
                content = [
                    src.Worksheets(i + 1).Name
                    for i in range(src.Worksheets.Count)
                    if not _is_capiq_helper_sheet(src.Worksheets(i + 1).Name)
                ]
                if not content:
                    continue
                before = combined.Sheets.Count
                combined.Activate()
                selector = content[0] if len(content) == 1 else content
                src.Worksheets(selector).Copy(None, combined.Sheets(combined.Sheets.Count))
                added = combined.Sheets.Count - before
                if added < len(content):
                    # Excel copied to a stray workbook instead of appending; bail
                    # to the openpyxl fallback rather than ship a partial file.
                    raise RuntimeError(
                        f"Excel did not append {len(content)} sheet(s) from {skill!r}"
                    )
                # Name each freshly-appended sheet from its OWN identity (not by
                # position), so source/destination sheet ordering can't misassign.
                for i in range(added):
                    sheet = combined.Sheets(before + i + 1)
                    new_name = _unique_sheet_name(
                        _tab_name(skill, sheet.Name, len(content)), used_names
                    )
                    if sheet.Name != new_name:
                        sheet.Name = new_name
                    skill_to_tab.setdefault(skill, sheet.Name)
            finally:
                src.Close(SaveChanges=False)

        # Drop the blank workbook's seed sheets now that real content exists.
        for name in seed_sheets:
            if combined.Sheets.Count > 1:
                combined.Sheets(name).Delete()

        # Wire the combined workbook's cross-tab links (best-effort).
        relink_ok = _relink_cross_tab_com(combined, skill_to_tab)

        # Stamp the INFOR brand theme so the combined workbook keeps INFOR
        # colours/fonts even when the base is a blank workbook (no cap table).
        # Best-effort: a theme failure must never lose the merged workbook.
        if theme_path is not None:
            try:
                combined.ApplyTheme(str(theme_path))
            except Exception:
                pass

        if output_path.exists():
            output_path.unlink()
        combined.SaveAs(str(output_path), FileFormat=_XL_OPEN_XML_WORKBOOK)
        combined.Close(SaveChanges=False)
        return relink_ok
    except RuntimeError:
        raise  # already normalized (e.g. _open_workbook, the partial-append bail)
    except Exception as exc:
        # COM failures surface as pywintypes.com_error (no Excel install, a dead
        # instance, a failed Copy/SaveAs), not RuntimeError — normalize so the
        # caller's `except RuntimeError` fallback to the openpyxl merge actually
        # engages, matching the documented "or Windows without Excel" behaviour.
        raise RuntimeError(f"Excel COM workbook aggregation failed: {exc}") from exc
    finally:
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def _combine_via_openpyxl(
    sources: list[tuple[str, Path]], output_path: Path, theme_path: Path | None = None
) -> bool:
    """Best-effort merge with openpyxl. CapIQ links and charts do NOT survive.

    Returns True when the cross-tab relink pass succeeded, False when it failed
    (the merged file is still saved) — the caller gates source deletion on it.
    """
    from openpyxl import Workbook, load_workbook

    combined = Workbook()
    combined.remove(combined.active)  # drop the default empty sheet
    used_names: set[str] = set()
    skill_to_tab: dict[str, str] = {}

    for skill, path in sources:
        src = load_workbook(path, data_only=False)
        content_sheets = [n for n in src.sheetnames if not _is_capiq_helper_sheet(n)]
        for original in content_sheets:
            src_ws = src[original]
            title = _unique_sheet_name(
                _tab_name(skill, original, len(content_sheets)), used_names
            )
            dst_ws = combined.create_sheet(title=title)
            _copy_sheet(src_ws, dst_ws)
            skill_to_tab.setdefault(skill, title)

    if not combined.sheetnames:
        combined.create_sheet(title="Sheet")
    relink_ok = _relink_cross_tab_openpyxl(combined, skill_to_tab)

    # Stamp the INFOR brand theme (a fresh openpyxl Workbook carries the default
    # Office theme, so copied cells' theme-colour refs would otherwise resolve
    # against Office colours). Best-effort: skip silently if the theme is missing.
    if theme_path is not None:
        theme_xml = _extract_theme_xml(Path(theme_path))
        if theme_xml is not None:
            combined.loaded_theme = theme_xml

    combined.save(output_path)
    return relink_ok


def _copy_sheet(src_ws, dst_ws) -> None:
    """Copy cell values/formulas, styles, dimensions, merges between sheets."""
    from copy import copy

    for row in src_ws.iter_rows():
        for cell in row:
            dst = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst.font = copy(cell.font)
                dst.fill = copy(cell.fill)
                dst.border = copy(cell.border)
                dst.alignment = copy(cell.alignment)
                dst.number_format = cell.number_format
                dst.protection = copy(cell.protection)
            # Carry the cell's hyperlink across — openpyxl does not copy it with
            # the style, so without this the precedents AB–AG source links (and
            # any other linked cell) are silently dropped on the openpyxl merge.
            if cell.hyperlink is not None:
                dst.hyperlink = copy(cell.hyperlink)
            # Carry the cell's comment across too — the cap-table template stores
            # the analyst's CapIQ refresh formulas as comments on F7/F16 and
            # ownership annotates its F35 denominator; openpyxl does not copy
            # comments with the style either (an openpyxl Comment binds to one
            # cell, so it must be copied, not shared).
            if cell.comment is not None:
                dst.comment = copy(cell.comment)

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width
        dst_ws.column_dimensions[key].hidden = dim.hidden
    for key, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key].height = dim.height
        dst_ws.row_dimensions[key].hidden = dim.hidden

    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
