"""Populate the INFOR precedents workbook from researched M&A transactions.

Companion-workbook builder for the pitch deck's precedent-transactions slide,
mirroring ``comps_workbook.py`` / ``ownership_workbook.py``. Researching the
deals, choosing the metric family, and sourcing each figure is the
``precedents`` skill's LLM/web work; this module is the deterministic write.

The template's ``Precedents`` sheet holds two peer-group blocks of six
transaction rows each (group #1 rows 8–13, group #2 rows 17–22). Per row the
skill writes the deal identity (input currency, announce date, target,
acquiror, TEV, 3-letter HQ code), the source-FX $ metric inputs for the chosen
family, source hyperlinks, and — when a multiple is disclosed in the deal PR —
the multiple itself written **directly over** the template's ratio formula.
Everything else (the CapIQ FX array formula in column C, the ``=+I*C`` TEV
conversion in J, the per-row ratio formulas in S–Z, and the group / global
statistic rows) is template-owned and left untouched.

Two metric families share the same columns; the agent fills exactly one,
chosen by the target's industry:
  * operating companies  -> Revenue (K/L) + Adj. EBITDA (O/P)  -> EV/Revenue,
    EV/EBITDA (S/T, U/V)
  * financial institutions -> Net Income (M/N) + Book Value (Q) + Tangible
    Book Value (R) -> P/E, P/B, P/TBV (W/X, Y, Z)

Sourcing preference per figure: a multiple disclosed in the deal PR (write it
straight into the S–Z cell) beats a disclosed $ metric (write the source-FX
value and let the ratio formula compute). When a clean LTM/NTM figure isn't
disclosed, use the most recent reported figure as the LTM/NTM proxy — do NOT
reconstruct LTM from multiple filings. **Every included deal must yield at least
one multiple** — a disclosed multiple, or a $ metric the ratio formula turns
into one; the builder rejects a deal that carries only a TEV (it would just add
an empty row). Don't include precedents you can't value.

The shipped template was pre-stripped of ~58.7k legacy CapIQ defined names and
174 vestigial external-workbook links (6 MB -> ~10 KB) so the openpyxl output
stays Excel-openable — the same treatment the ownership template needed. A
regression test guards that the live FX / ratio / statistic formulas survive
the build.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl.styles import Font
from openpyxl.utils import range_boundaries

from deal_workbook import TAB_PRECEDENTS, TabSpec, write_tab
from template_layout import (
    NAME_PREC_GROUP_BLOCKS,
    NAME_PREC_GROUP_LABELS,
    NAME_PREC_OUTPUT_CCY,
    PRECEDENTS_SHEET,
    PRECEDENTS_TEMPLATE,
    PRECEDENTS_WRITE_NAMES,
    resolve_name_cell,
    resolve_name_range,
    verify_names,
)

_SHEET = PRECEDENTS_SHEET
_MAX_GROUPS = 2

# The shipped template's target/acquiror cells (F/G) are Calibri 11, inconsistent
# with the Palatino 9 the rest of the table uses; set it explicitly so the names
# match the template's intended body font instead of inheriting the stray default.
_NAME_FONT = Font(name="Palatino Linotype", size=9)


def _group_slots(ws) -> list[tuple[str, range]]:
    """``(label cell, data rows)`` per peer group, resolved from the defined names.

    Shipped: ``E7`` + rows 8-13 and ``E16`` + rows 17-22. Read from
    ``infor_prec_groupN_label`` / ``infor_prec_groupN_block`` so a re-saved
    template that moved or resized a block needs no code change.
    """
    slots: list[tuple[str, range]] = []
    for label_name, block_name in zip(NAME_PREC_GROUP_LABELS, NAME_PREC_GROUP_BLOCKS):
        label = resolve_name_cell(ws, label_name, template=PRECEDENTS_TEMPLATE)
        block = resolve_name_range(ws, block_name, template=PRECEDENTS_TEMPLATE)
        _, first_row, _, last_row = range_boundaries(block)
        slots.append((label, range(first_row, last_row + 1)))
    return slots



# Deal-identity columns (always written for a populated row).
_COL_CURRENCY = "B"      # input currency, ISO-3 — drives the column-C FX formula
_COL_DATE = "E"          # announce date — FX is taken as of the workday before it
_COL_TARGET = "F"
_COL_ACQUIROR = "G"
_COL_TEV = "I"           # TEV, source-FX $MM (J = "=+I*C" converts to output ccy)
_COL_HQ = "AI"           # target HQ, 3-letter country code

# Source-FX $ metric inputs -> the per-row ratio formula in S–Z reads these.
# Fill only the chosen family's fields; the rest stay blank.
_VALUE_FIELDS = (
    ("revenue_ltm", "K"),
    ("revenue_ntm", "L"),
    ("net_income_ltm", "M"),
    ("net_income_ntm", "N"),
    ("ebitda_ltm", "O"),
    ("ebitda_ntm", "P"),
    ("book_value", "Q"),
    ("tangible_book_value", "R"),
)

# Disclosed multiples -> written as a literal straight over the template's ratio
# formula in that cell (the group / global statistic rows average it in).
_MULTIPLE_FIELDS = (
    ("ev_revenue_ltm", "S"),
    ("ev_revenue_ntm", "T"),
    ("ev_ebitda_ltm", "U"),
    ("ev_ebitda_ntm", "V"),
    ("pe_ltm", "W"),
    ("pe_ntm", "X"),
    ("pb", "Y"),
    ("ptbv", "Z"),
)

# One source-link column per metric concept (AB–AG). The cell keeps its "Link"
# display text and gains a hyperlink to the source; unused link cells in a
# populated row are cleared so nothing dangles.
_LINK_FIELDS = (
    ("tev_link", "AB"),
    ("revenue_link", "AC"),
    ("ebitda_link", "AD"),
    ("net_income_link", "AE"),
    ("book_value_link", "AF"),
    ("tangible_book_value_link", "AG"),
)
_LINK_DISPLAY = "Link"

_VALUE_ATTRS = tuple(a for a, _ in _VALUE_FIELDS)
_MULTIPLE_ATTRS = tuple(a for a, _ in _MULTIPLE_FIELDS)
_LINK_ATTRS = tuple(a for a, _ in _LINK_FIELDS)
_URL_RE = re.compile(r"^https?://", re.IGNORECASE)


@dataclass
class PrecedentTransaction:
    """One precedent M&A transaction — a single data row.

    Required identity: ``input_currency`` (ISO-3, drives FX), ``announce_date``
    (a ``datetime.date``), ``target``, ``acquiror``, ``tev`` (source-FX $MM),
    ``hq_country`` (3-letter ISO code -> column AI).

    Optional metric inputs (source-FX $MM) — fill only the chosen family:
    ``revenue_ltm`` / ``revenue_ntm``, ``net_income_ltm`` / ``net_income_ntm``,
    ``ebitda_ltm`` / ``ebitda_ntm``, ``book_value``, ``tangible_book_value``.

    Optional disclosed multiples (dimensionless ×) — written straight over the
    ratio formula: ``ev_revenue_ltm`` / ``ev_revenue_ntm``, ``ev_ebitda_ltm`` /
    ``ev_ebitda_ntm``, ``pe_ltm`` / ``pe_ntm``, ``pb``, ``ptbv``.

    Optional source links (http(s) URLs) -> AB–AG: ``tev_link``,
    ``revenue_link``, ``ebitda_link``, ``net_income_link``, ``book_value_link``,
    ``tangible_book_value_link``.
    """

    input_currency: str
    announce_date: date
    target: str
    acquiror: str
    tev: float
    hq_country: str

    revenue_ltm: float | None = None
    revenue_ntm: float | None = None
    net_income_ltm: float | None = None
    net_income_ntm: float | None = None
    ebitda_ltm: float | None = None
    ebitda_ntm: float | None = None
    book_value: float | None = None
    tangible_book_value: float | None = None

    ev_revenue_ltm: float | None = None
    ev_revenue_ntm: float | None = None
    ev_ebitda_ltm: float | None = None
    ev_ebitda_ntm: float | None = None
    pe_ltm: float | None = None
    pe_ntm: float | None = None
    pb: float | None = None
    ptbv: float | None = None

    tev_link: str | None = None
    revenue_link: str | None = None
    ebitda_link: str | None = None
    net_income_link: str | None = None
    book_value_link: str | None = None
    tangible_book_value_link: str | None = None


@dataclass
class PrecedentGroup:
    """One peer group: a label plus up to six precedent transactions.

    - ``name``: the group label -> ``E7`` / ``E16``.
    - ``transactions``: up to six ``PrecedentTransaction`` entries -> the block.
    """

    name: str
    transactions: list[PrecedentTransaction] = field(default_factory=list)


def _coerce_date(value: "date | str") -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return date.fromisoformat(value.strip())
    raise ValueError(f"announce_date must be a date or ISO string, got {value!r}")


def _normalize_transaction(tx: "PrecedentTransaction | dict") -> PrecedentTransaction:
    if isinstance(tx, PrecedentTransaction):
        return PrecedentTransaction(**{**tx.__dict__, "announce_date": _coerce_date(tx.announce_date)})
    fields = {f for f in PrecedentTransaction.__dataclass_fields__}
    unknown = set(tx) - fields
    if unknown:
        raise ValueError(f"unknown transaction field(s): {sorted(unknown)}")
    data = dict(tx)
    data["announce_date"] = _coerce_date(data["announce_date"])
    return PrecedentTransaction(**data)


def _normalize_group(group: "PrecedentGroup | dict") -> PrecedentGroup:
    if isinstance(group, PrecedentGroup):
        name, transactions = group.name, group.transactions
    else:
        name, transactions = group["name"], group.get("transactions", [])
    return PrecedentGroup(name=name, transactions=[_normalize_transaction(t) for t in transactions])


def _check_code(value: str, label: str, where: str) -> str:
    code = str(value).strip().upper()
    if not (len(code) == 3 and code.isalpha()):
        raise ValueError(f"{label} must be a 3-letter code, got {value!r} ({where})")
    return code


def _validate_transaction(tx: PrecedentTransaction, where: str) -> None:
    if not str(tx.target).strip():
        raise ValueError(f"empty target ({where})")
    if not str(tx.acquiror).strip():
        raise ValueError(f"empty acquiror ({where})")
    _check_code(tx.input_currency, "input_currency", where)
    _check_code(tx.hq_country, "hq_country", where)
    if not isinstance(tx.tev, (int, float)) or isinstance(tx.tev, bool) or tx.tev <= 0:
        raise ValueError(f"tev must be a positive number, got {tx.tev!r} ({where})")
    for attr in _VALUE_ATTRS + _MULTIPLE_ATTRS:
        v = getattr(tx, attr)
        if v is not None and (not isinstance(v, (int, float)) or isinstance(v, bool)):
            raise ValueError(f"{attr} must be a number or None, got {v!r} ({where})")
    # Every included deal must yield at least one multiple — either a disclosed
    # multiple (S–Z) or a disclosed $ metric (K–R) the ratio formula turns into
    # one. A deal with only a TEV adds an empty row to the table, so reject it:
    # only precedents you can actually value belong here.
    if not any(getattr(tx, attr) is not None for attr in _VALUE_ATTRS + _MULTIPLE_ATTRS):
        raise ValueError(
            f"deal must carry at least one disclosed multiple or $ metric so a "
            f"multiple can be shown — drop deals you can't value ({where})"
        )
    for attr in _LINK_ATTRS:
        url = getattr(tx, attr)
        if url is not None and not _URL_RE.match(str(url).strip()):
            raise ValueError(f"{attr} must be an http(s) URL, got {url!r} ({where})")


def build_precedents_workbook(
    *,
    groups: "list[PrecedentGroup | dict]",
    deal_workbook: Path | str,
    output_currency: str = "USD",
) -> Path:
    """Fill the deal workbook's `precedents` tab; return the workbook path.

    Since Phase D there is no standalone precedents file and no template to copy:
    the `precedents` tab is already in the deal workbook, carried in from
    `INFOR Deal Workbook Template.xlsx` at deal-init with its formulas and its
    `infor_prec_*` defined names intact.

    ``groups`` lists up to two peer groups, each with up to six precedent
    transactions. For each transaction the builder writes the deal identity
    (B/E/F/G/I + the 3-letter HQ code in AI), the source-FX $ metric inputs that
    are supplied, any disclosed multiples (written straight over the ratio
    formula in S–Z), and a hyperlink on each supplied source-link cell (AB–AG,
    keeping the "Link" display text); unused link cells in a populated row are
    cleared. ``output_currency`` (ISO-3, default ``"USD"``) is written to C2,
    which the column-C FX array formula keys off. The FX / TEV / ratio formulas
    and the statistic rows are left untouched; unfilled groups keep the
    template's ``[Group #N]`` placeholder. Raises ValueError on too many groups
    / transactions, a missing or malformed required field, a non-numeric metric,
    or a non-http(s) link.
    """
    groups = [_normalize_group(g) for g in groups]
    if not groups:
        raise ValueError("no groups supplied — expected 1–2 peer groups")
    if len(groups) > _MAX_GROUPS:
        raise ValueError(f"precedents template holds {_MAX_GROUPS} group blocks; got {len(groups)}")
    output_currency = _check_code(output_currency, "output_currency", "output currency")
    for group in groups:
        for i, tx in enumerate(group.transactions):
            _validate_transaction(tx, f"group {group.name!r} row {i + 1}")

    def _write(_wb, ws) -> None:
        _fill_precedents_tab(ws, groups, output_currency)

    write_tab(
        deal_workbook,
        TAB_PRECEDENTS,
        TabSpec(write=_write, verify_names=tuple(NAME_PREC_GROUP_BLOCKS)),
    )
    return Path(deal_workbook)


def _fill_precedents_tab(ws, groups: "list[PrecedentGroup]", output_currency: str) -> None:
    """Write the validated groups into the tab. Layout unchanged."""
    # Verify the output-currency cell and every group's label + block name
    # resolves before writing anything.
    verify_names(ws, PRECEDENTS_WRITE_NAMES, template=PRECEDENTS_TEMPLATE)

    ws[resolve_name_cell(ws, NAME_PREC_OUTPUT_CCY, template=PRECEDENTS_TEMPLATE)] = output_currency

    for group, (label_cell, rows) in zip(groups, _group_slots(ws)):
        if len(group.transactions) > len(rows):
            raise ValueError(
                f"group {group.name!r} has {len(group.transactions)} transactions; "
                f"the template holds {len(rows)} rows per group"
            )
        ws[label_cell] = group.name
        for row, tx in zip(rows, group.transactions):
            ws[f"{_COL_CURRENCY}{row}"] = _check_code(tx.input_currency, "input_currency", "")
            ws[f"{_COL_DATE}{row}"] = tx.announce_date
            target_cell = ws[f"{_COL_TARGET}{row}"]
            target_cell.value = str(tx.target).strip()
            target_cell.font = _NAME_FONT
            acquiror_cell = ws[f"{_COL_ACQUIROR}{row}"]
            acquiror_cell.value = str(tx.acquiror).strip()
            acquiror_cell.font = _NAME_FONT
            ws[f"{_COL_TEV}{row}"] = tx.tev
            ws[f"{_COL_HQ}{row}"] = _check_code(tx.hq_country, "hq_country", "")

            for attr, col in _VALUE_FIELDS + _MULTIPLE_FIELDS:
                value = getattr(tx, attr)
                if value is not None:
                    ws[f"{col}{row}"] = value

            for attr, col in _LINK_FIELDS:
                cell = ws[f"{col}{row}"]
                url = getattr(tx, attr)
                if url:
                    cell.value = _LINK_DISPLAY
                    cell.hyperlink = str(url).strip()
                else:
                    cell.value = None       # clear the template's "Link" placeholder
                    cell.hyperlink = None   # and any stray hyperlink, so nothing dangles
