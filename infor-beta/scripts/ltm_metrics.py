"""Write the deal workbook's `ltm-metrics` tab.

The tab stacks three blocks, separated by a blank spacer row:

1. **LTM Revenue overview** — revenue split by segment, with `% of total` and a
   total row. Feeds the overview slide's deferred LTM revenue pie.
2. **LTM Revenue bridge** — how the LTM revenue total is derived:
   `LTM = FY + current-year YTD − prior-year YTD`.
3. **LTM Adj. EBITDA bridge** — the same FY + YTD − prior-YTD bridge for
   Adjusted EBITDA (or unadjusted EBITDA when no Adj. figure is disclosed).
   Bridge only — no segment overview.

Arithmetic lives in cell formulas (% of total, the total row, the bridge sums)
so the workbook stays analyst-auditable, matching the cap-table convention.
Each hand-extracted input (a segment's LTM revenue, a bridge component) carries a
``source`` :class:`provenance.FigureSource` — the filing, statement/note and page
it came from. Since Phase G that record is the record: it goes into the run's
provenance ledger, and the amount cell's ``Source: …`` comment is *rendered from*
it (`comment_citations.cite_cell`), so the artefact and the machine-readable
record cannot drift apart. Each bridge's result row is recorded too, as a
*derived* figure whose provenance is its components — which is the chain
`deckcheck` walks from a deck tile back to a filing page.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from comment_citations import cite_cell
from deal_workbook import TAB_LTM_METRICS, TabSpec, write_tab
from provenance import FigureSource, ProvenanceError, ProvenanceLedger

# No template_layout anchors here: this tab is authored from scratch (no shipped
# template to shift), and every downstream reader locates its blocks by label,
# not address (financial_charts.ltm_revenue_overview_range, financial-summary's
# MATCH-keyed links).

# INFOR mid-blue header fill / Palatino body, mirroring the deck brand.
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(name="Palatino Linotype", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Palatino Linotype", size=14, bold=True, color="1F3864")
_SECTION_FONT = Font(name="Palatino Linotype", size=11, bold=True, color="1F3864")
_META_FONT = Font(name="Palatino Linotype", size=10, italic=True, color="595959")
_BODY_FONT = Font(name="Palatino Linotype", size=11)
_TOTAL_FONT = Font(name="Palatino Linotype", size=11, bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_MINUS = "−"  # typographic minus sign used in bridge labels

# ─── The bridge-total row, and how other tabs address it ─────────────────────
#
# A bridge's total lives on a row labelled `(=) <result_label>` in column A with
# the value in column B. That label — not a row number — is the join key every
# consumer uses, because the bridges' row positions move with the segment count.
# The two labels below are the ones the two shipped plans produce, and both the
# `financial-summary` tab and the cap table's LTM valuation cells link off them.

#: Column-A prefix on a bridge's total row. Not a leading "=" — openpyxl would
#: store the label as a formula and Excel would render it as "=@LTM Revenue".
RESULT_ROW_PREFIX = "(=) "

#: The revenue bridge's result label — fixed.
LTM_REVENUE_RESULT_LABEL = "LTM Revenue"

#: The EBITDA bridge's result label, in preference order: Adjusted when the
#: company discloses it, unadjusted when it does not (`ebitda_label`). A consumer
#: linking to "the EBITDA total" passes both and takes whichever the tab has.
LTM_EBITDA_RESULT_LABELS = ("LTM Adj. EBITDA", "LTM EBITDA")


@dataclass(frozen=True)
class RevenueSegment:
    """One row of the LTM revenue overview.

    ``source`` is the :class:`provenance.FigureSource` naming the filing,
    statement/note and page the segment figure came from — e.g.
    ``FigureSource(filing="Q3 2026 10-Q", statement="revenue disaggregation
    note", page=14)``. It is recorded in the run's provenance ledger and the
    amount cell's ``Source: …`` comment is rendered from it.
    """

    name: str
    ltm_revenue: float
    source: FigureSource | None = None


@dataclass(frozen=True)
class BridgeComponent:
    """One additive (or subtractive) line of an LTM bridge.

    ``source`` is the :class:`provenance.FigureSource` naming the filing,
    statement and page the component figure came from — e.g.
    ``FigureSource(filing="FY2025 10-K", statement="Consolidated Statements of
    Operations", page=61)``. It is recorded in the run's provenance ledger and the
    amount cell's ``Source: …`` comment is rendered from it.
    """

    name: str
    value: float
    subtract: bool = False
    source: FigureSource | None = None


@dataclass(frozen=True)
class Bridge:
    """A whole LTM bridge block — its section title, result label, and components.

    Used for the pitch-only ``extra_bridges`` (e.g. an LTM Net Income or LTM Gross
    Profit bridge for the Financial Summary tab). ``result_label`` is the figure
    the bridge derives (e.g. ``"LTM Net Income"``); the result row is written as
    ``(=) <result_label>``, which the financial-summary tab keys its LTM link off.
    """

    section_title: str
    result_label: str
    components: list[BridgeComponent] = field(default_factory=list)


def _coerce_source(owner: str, source) -> FigureSource | None:
    """Validate one figure's source record.

    A citation string used to BE the record; since Phase G it is rendered from
    one, so a string here would silently produce a record with the whole citation
    in ``filing`` and no statement or page — provenance that reads fine and cannot
    be followed. Reject it, naming the fix.
    """
    if source is None or isinstance(source, FigureSource):
        return source
    raise ProvenanceError(
        f"{owner!r} has a source of type {type(source).__name__} ({source!r}); pass "
        f"FigureSource(filing=…, statement=…, page=…) — a citation string is no "
        f"longer a source record."
    )


def _coerce_segments(
    segments: list[RevenueSegment] | list[tuple],
) -> list[RevenueSegment]:
    out: list[RevenueSegment] = []
    for seg in segments:
        if isinstance(seg, RevenueSegment):
            out.append(RevenueSegment(seg.name, seg.ltm_revenue, _coerce_source(seg.name, seg.source)))
        else:
            source = seg[2] if len(seg) > 2 else None
            out.append(RevenueSegment(seg[0], float(seg[1]), _coerce_source(seg[0], source)))
    return out


def _coerce_components(
    components: list[BridgeComponent] | list[tuple] | None,
) -> list[BridgeComponent]:
    if not components:
        return []
    out: list[BridgeComponent] = []
    for c in components:
        if isinstance(c, BridgeComponent):
            out.append(
                BridgeComponent(c.name, c.value, c.subtract, _coerce_source(c.name, c.source))
            )
        else:
            name, value = c[0], float(c[1])
            subtract = bool(c[2]) if len(c) > 2 else False
            source = c[3] if len(c) > 3 else None
            out.append(BridgeComponent(name, value, subtract, _coerce_source(name, source)))
    return out


def _coerce_bridges(bridges: "list[Bridge] | list[dict] | None") -> list[Bridge]:
    if not bridges:
        return []
    out: list[Bridge] = []
    for b in bridges:
        if isinstance(b, Bridge):
            out.append(Bridge(b.section_title, b.result_label, _coerce_components(b.components)))
        else:
            out.append(
                Bridge(
                    section_title=b["section_title"],
                    result_label=b["result_label"],
                    components=_coerce_components(b.get("components")),
                )
            )
    return out


def bridge_total(
    components: list[BridgeComponent] | list[tuple] | None,
) -> float | None:
    """Sum a bridge's components (additive minus subtractive) for the handoff.

    Returns None when no components are supplied. The workbook derives the same
    total via a cell formula; this mirrors it so the typed stage handoff can pass
    the LTM revenue / EBITDA figure to a downstream stage (e.g. the cap table).
    The figure is in the same currency as the bridge components — i.e. the
    filing's reporting currency.
    """
    comps = _coerce_components(components)
    if not comps:
        return None
    return sum(-c.value if c.subtract else c.value for c in comps)


def _quote_sheet(name: str) -> str:
    """Single-quote a sheet name for a formula (the hyphen in `ltm-metrics`)."""
    return "'" + name.replace("'", "''") + "'"


def result_row_labels(ws: Worksheet) -> list[str]:
    """Every bridge result label on an `ltm-metrics` tab, top to bottom.

    The `(=) ` prefix is stripped, so a tab carrying a revenue and an Adjusted
    EBITDA bridge gives `["LTM Revenue", "LTM Adj. EBITDA"]`.
    """
    out: list[str] = []
    for (cell,) in ws.iter_rows(min_col=1, max_col=1):
        value = cell.value
        if isinstance(value, str) and value.startswith(RESULT_ROW_PREFIX):
            label = value[len(RESULT_ROW_PREFIX) :].strip()
            if label:
                out.append(label)
    return out


def ltm_total_formula(
    result_label: str, *, sheet: str = TAB_LTM_METRICS, times: str | None = None
) -> str:
    """The label-keyed lookup of one bridge total — the formula, built once.

    ``=INDEX('ltm-metrics'!$B:$B, MATCH("(=) LTM Revenue", 'ltm-metrics'!$A:$A, 0))``.
    Keyed on the `(=) <result_label>` row rather than an address, because the
    bridges sit below a segment overview whose height varies.

    Both consumers of an LTM total build their formula here, so the two cannot
    drift: `financial_summary_workbook` writes it unconditionally (in the pitch
    plan its stage runs *before* `ltm-metrics`, so there is no tab to check yet),
    while the cap table goes through `ltm_total_link`, which only links to a
    bridge that is really there.
    """
    quoted = _quote_sheet(sheet)
    formula = (
        f"=INDEX({quoted}!$B:$B, "
        f'MATCH("{RESULT_ROW_PREFIX}{result_label}", {quoted}!$A:$A, 0))'
    )
    return f"{formula}*{times}" if times else formula


def ltm_total_link(wb, *result_labels: str, times: str | None = None) -> str | None:
    """A live link to a bridge total on this workbook's `ltm-metrics` tab.

    Returns ``=INDEX('ltm-metrics'!$B:$B, MATCH("(=) <label>", 'ltm-metrics'!$A:$A, 0))``
    for the first of `result_labels` the tab actually carries — keyed on the label
    so it survives the bridge's variable row position — or **None** when this
    workbook has no `ltm-metrics` tab, or has one without any of those bridges.
    A None is the caller's signal to fall back (the cap table restores its CapIQ
    formulas); it is never a reason to write the figure as a literal.

    `times` names a cell to multiply the link by, for a consumer whose column is
    in a different currency — the cap table's `D47`/`D48` pass the FX-rate cell
    resolved through `NAME_FX_RATE`, giving `=INDEX(…)*F7`.

    `wb` is the open `openpyxl` workbook — inside `deal_workbook.write_tab` that
    is the `wb` handed to the `write(wb, ws)` callback, so a stage writing one tab
    can link to a sibling tab without opening the file twice.

    This is the ONE way a figure computed on the `ltm-metrics` tab reaches another
    tab. Embedding the number instead leaves two copies of it in the deal
    workbook: correct the bridge and the linked consumer follows while the
    hardcoded one silently does not, and the deck then ships two different LTM
    revenues with no error anywhere.
    """
    if TAB_LTM_METRICS not in getattr(wb, "sheetnames", ()):
        return None
    present = set(result_row_labels(wb[TAB_LTM_METRICS]))
    label = next((candidate for candidate in result_labels if candidate in present), None)
    if label is None:
        return None
    return ltm_total_formula(label, times=times)


def _section(ws: Worksheet, row: int, text: str) -> None:
    for col in (1, 2, 3):
        cell = ws.cell(row=row, column=col)
        cell.fill = _SECTION_FILL
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT


def _ref(cell) -> str:
    """`"ltm-metrics!B12"` — a provenance record's location for a written cell."""
    return f"{TAB_LTM_METRICS}!{cell.coordinate}"


def _write_bridge(
    ws: Worksheet,
    *,
    start_row: int,
    section_title: str,
    result_label: str,
    currency: str,
    components: list[BridgeComponent],
    ledger: ProvenanceLedger,
) -> int:
    """Write a bridge block and return the row after its result row."""
    _section(ws, start_row, section_title)
    header_row = start_row + 1
    for col, text in enumerate(("Component", f"Amount ({currency})"), start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    first_data = header_row + 1
    for i, comp in enumerate(components):
        r = first_data + i
        prefix = f"({_MINUS}) " if comp.subtract else "(+) "
        ws.cell(row=r, column=1, value=prefix + comp.name).font = _BODY_FONT
        v = ws.cell(row=r, column=2, value=comp.value)
        v.font = _BODY_FONT
        v.number_format = "#,##0.0"
        ws.cell(row=r, column=1).border = _BORDER
        v.border = _BORDER
        if comp.source is not None:
            # The record first, the comment rendered from it — never the reverse.
            cite_cell(
                v,
                ledger.record(
                    f"{result_label} — {comp.name}",
                    sources=comp.source,
                    value=comp.value,
                    units=currency,
                    location=_ref(v),
                ),
            )

    last_data = first_data + len(components) - 1
    result_row = last_data + 1

    terms = "".join(
        f"-B{first_data + i}" if comp.subtract else f"+B{first_data + i}"
        for i, comp in enumerate(components)
    ).lstrip("+")

    # The label is a plain string. A leading "=" makes openpyxl store the cell
    # as a formula (Excel then renders it as "=@LTM Revenue"), so use the "(=)"
    # bridge glyph, mirroring the "(+)" / "(−)" component rows above.
    ws.cell(
        row=result_row, column=1, value=f"{RESULT_ROW_PREFIX}{result_label}"
    ).font = _TOTAL_FONT
    rv = ws.cell(row=result_row, column=2, value=f"={terms}")
    rv.font = _TOTAL_FONT
    rv.number_format = "#,##0.0"
    for col in (1, 2):
        ws.cell(row=result_row, column=col).border = _BORDER

    # The bridge total is DERIVED: its provenance is the components above, each of
    # which carries its own filing record. This is the row the deck's LTM tile and
    # the financial-summary tab's LTM link both read, so it is the join point
    # `deckcheck` walks from a figure on a slide back to a filing page.
    ledger.record(
        result_label,
        value=bridge_total(components),
        units=currency,
        location=_ref(rv),
        derivation=" ".join(
            f"{'−' if comp.subtract else '+'} {comp.name}" for comp in components
        ).lstrip("+ "),
    )

    return result_row + 1


def build_ltm_metrics_workbook(
    *,
    company_name: str,
    period_label: str,
    currency: str,
    segmentation_basis: str,
    segments: list[RevenueSegment] | list[tuple],
    revenue_bridge: list[BridgeComponent] | list[tuple] | None = None,
    ebitda_bridge: list[BridgeComponent] | list[tuple] | None = None,
    ebitda_label: str = LTM_EBITDA_RESULT_LABELS[0],
    extra_bridges: "list[Bridge] | list[dict] | None" = None,
    deal_workbook: Path | str,
    provenance: ProvenanceLedger | None = None,
) -> Path:
    """Write the `ltm-metrics` tab (overview + bridges) into the deal workbook.

    ``provenance`` is filled **in place** with one record per figure written —
    every segment, every bridge component, and each bridge total as a derived
    figure — and each amount cell's ``Source: …`` comment is rendered from its
    record. Pass the stage's ledger and write it afterwards
    (`ledger.write(io.stage_dir)`); pass nothing and the records are still built
    (the comments come from them either way), just not kept.

    Returns the deal workbook's path. Since Phase D there is no standalone LTM
    metrics file: the tab is written straight into the deal's single workbook, so
    the `financial-summary` tab's `=INDEX('ltm-metrics'!…)` links resolve as soon
    as both tabs exist rather than only after an aggregation step.

    `segments` and the bridge components may be dataclasses or plain tuples
    (`(name, value[, source])` for segments; `(name, value[, subtract[, source]])`
    for components). A `source` string is written as a `Source: …` comment on
    that figure's amount cell — in-artefact provenance for every extracted value.
    `segmentation_basis` is a human label such as "Service line" or "Geography".
    `ebitda_label` is "LTM Adj. EBITDA" by default; pass "LTM EBITDA" when no
    Adjusted figure is disclosed. Either bridge may be omitted.

    `extra_bridges` (pitch only) appends one further `FY + YTD − prior-YTD` bridge
    block per `Bridge`, after the revenue and EBITDA bridges, reusing the same
    layout. The `financial-summary` stage passes these for the additional metrics
    it selected (e.g. Net Income, Gross Profit) so its tab can link each metric's
    LTM total off the bridge's `(=) <result_label>` row. Omitted (None) by the
    earnings-update plan, which keeps the revenue + EBITDA bridges unchanged.
    """
    rows = _coerce_segments(segments)
    if not rows:
        raise ValueError("LTM metrics workbook requires at least one revenue segment")
    rev_components = _coerce_components(revenue_bridge)
    ebitda_components = _coerce_components(ebitda_bridge)
    extra = _coerce_bridges(extra_bridges)
    ledger = provenance if provenance is not None else ProvenanceLedger(stage="ltm-metrics")

    def _write(_wb, ws) -> None:
        _fill_ltm_metrics_tab(
            ws,
            company_name=company_name,
            period_label=period_label,
            currency=currency,
            segmentation_basis=segmentation_basis,
            rows=rows,
            rev_components=rev_components,
            ebitda_components=ebitda_components,
            ebitda_label=ebitda_label,
            extra=extra,
            ledger=ledger,
        )

    write_tab(deal_workbook, TAB_LTM_METRICS, TabSpec(create=True, write=_write))
    return Path(deal_workbook)


def _fill_ltm_metrics_tab(
    ws: Worksheet,
    *,
    company_name: str,
    period_label: str,
    currency: str,
    segmentation_basis: str,
    rows: list[RevenueSegment],
    rev_components: list[BridgeComponent],
    ebitda_components: list[BridgeComponent],
    ebitda_label: str,
    extra: list[Bridge],
    ledger: ProvenanceLedger,
) -> None:
    """Write the tab's three blocks. Layout unchanged from the standalone workbook."""
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"{company_name} — LTM Metrics"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Period: {period_label}"
    ws["A3"] = f"Revenue segmentation: {segmentation_basis}"
    ws["A4"] = f"Figures in {currency}"
    for cell in ("A2", "A3", "A4"):
        ws[cell].font = _META_FONT

    # --- LTM Revenue overview (row 5 left blank as a spacer after the meta block) ---
    section_row = 6
    _section(ws, section_row, "LTM Revenue Overview")
    header_row = section_row + 1
    for col, text in enumerate(("Segment", f"LTM Revenue ({currency})", "% of Total"), start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    first_data = header_row + 1
    for i, seg in enumerate(rows):
        r = first_data + i
        ws.cell(row=r, column=1, value=seg.name).font = _BODY_FONT
        v = ws.cell(row=r, column=2, value=seg.ltm_revenue)
        v.font = _BODY_FONT
        v.number_format = "#,##0.0"
        ws.cell(row=r, column=1).border = _BORDER
        v.border = _BORDER
        if seg.source is not None:
            cite_cell(
                v,
                ledger.record(
                    f"LTM Revenue — {seg.name}",
                    sources=seg.source,
                    value=seg.ltm_revenue,
                    units=currency,
                    location=_ref(v),
                ),
            )

    last_data = first_data + len(rows) - 1
    total_row = last_data + 1
    total_ref = f"B{total_row}"

    for i in range(len(rows)):
        r = first_data + i
        p = ws.cell(row=r, column=3, value=f"=B{r}/{total_ref}")
        p.font = _BODY_FONT
        p.number_format = "0.0%"
        p.border = _BORDER
        p.alignment = Alignment(horizontal="center")

    ws.cell(row=total_row, column=1, value="Total").font = _TOTAL_FONT
    tv = ws.cell(row=total_row, column=2, value=f"=SUM(B{first_data}:B{last_data})")
    tv.font = _TOTAL_FONT
    tv.number_format = "#,##0.0"
    tp = ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data}:C{last_data})")
    tp.font = _TOTAL_FONT
    tp.number_format = "0.0%"
    tp.alignment = Alignment(horizontal="center")
    for col in (1, 2, 3):
        ws.cell(row=total_row, column=col).border = _BORDER

    # Derived: the overview total is the sum of the segments above, each of which
    # carries its own filing record.
    ledger.record(
        "LTM Revenue by segment — Total",
        value=sum(seg.ltm_revenue for seg in rows),
        units=currency,
        location=_ref(tv),
        derivation=f"sum of the {len(rows)} segment rows above ({segmentation_basis})",
    )

    # --- Bridges, each preceded by a blank spacer row ---
    next_row = total_row + 1
    if rev_components:
        next_row = _write_bridge(
            ws,
            start_row=next_row + 1,
            section_title="LTM Revenue Bridge",
            result_label=LTM_REVENUE_RESULT_LABEL,
            currency=currency,
            components=rev_components,
            ledger=ledger,
        )
    if ebitda_components:
        next_row = _write_bridge(
            ws,
            start_row=next_row + 1,
            section_title=f"{ebitda_label} Bridge",
            result_label=ebitda_label,
            currency=currency,
            components=ebitda_components,
            ledger=ledger,
        )
    for bridge in extra:
        if not bridge.components:
            continue
        next_row = _write_bridge(
            ws,
            start_row=next_row + 1,
            section_title=bridge.section_title,
            result_label=bridge.result_label,
            currency=currency,
            components=bridge.components,
            ledger=ledger,
        )

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
