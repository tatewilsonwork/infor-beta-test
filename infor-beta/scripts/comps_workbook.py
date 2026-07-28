"""Populate the INFOR comps workbook from selected public comparables.

Companion-workbook builder for the pitch deck's comparable-companies ("public
comps" / trading comps) slide, mirroring ``ltm_metrics.py`` and
``ownership_workbook.py``. Choosing the three verticals, the six public peers
per vertical, their Capital IQ tickers, and the one-line descriptions is the
``comps`` skill's LLM/web work; this module is the deterministic write.

The template's ``Comps`` sheet holds three vertical (peer-group) blocks. Each
block has a label cell, six ticker rows and six description rows; every other
column is a Capital IQ array formula (``_xll.SNL…SPG($B10, …)``) that resolves
off the ticker in column B once the analyst opens the workbook with the Capital
IQ add-in active. This builder writes ONLY the three input fields — the vertical
label, the CapIQ ticker (column B) and the description (column AA) — and never
touches the CapIQ formulas or the group-average / global-statistic rows.

Where those fields ARE comes from the template, not from this file: each
block's label cell and row span are read off its ``infor_comps_groupN_label`` /
``infor_comps_groupN_block`` defined names (``D9`` + rows 10-15, ``D19`` +
20-25, ``D29`` + 30-35 as shipped). A template whose verticals were moved, or
resized to a different number of peers, needs no code change.

Capital IQ cannot be refreshed in this environment, so the workbook ships with
its formulas un-evaluated; the analyst refreshes them in Excel. The shipped
template round-trips cleanly through openpyxl (CapIQ array formulas preserved,
file stays Excel-openable), so — unlike the ownership template — it needs no
defined-name / external-link pre-cleaning; a regression test guards that the
formulas survive the build.
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from openpyxl.utils import range_boundaries

from template_layout import (
    COMPS_BLOCK_ANCHORS,
    COMPS_SHEET,
    COMPS_TEMPLATE,
    NAME_COMPS_GROUP_BLOCKS,
    NAME_COMPS_GROUP_LABELS,
    resolve_name_cell,
    resolve_name_range,
    verify_anchors,
)

_SHEET = COMPS_SHEET
_MAX_VERTICALS = 3
_MAX_DESCRIPTION_CHARS = 50  # column AA width ~50; longer overflows visually in the cell

_COL_TICKER = "B"
_COL_DESCRIPTION = "AA"


def _vertical_slots(ws) -> list[tuple[str, range]]:
    """``(label cell, data rows)`` per vertical, resolved from the defined names.

    Both the label cell and the block extent come from the template rather than
    from constants: the shipped blocks are ``D9`` + rows 10-15, ``D19`` + 20-25,
    ``D29`` + 30-35, but a template whose verticals were moved or resized needs
    no code change, because the row span is read off ``infor_comps_groupN_block``.
    """
    slots: list[tuple[str, range]] = []
    for label_name, block_name in zip(NAME_COMPS_GROUP_LABELS, NAME_COMPS_GROUP_BLOCKS):
        label = resolve_name_cell(ws, label_name, template=COMPS_TEMPLATE)
        block = resolve_name_range(ws, block_name, template=COMPS_TEMPLATE)
        _, first_row, _, last_row = range_boundaries(block)
        slots.append((label, range(first_row, last_row + 1)))
    return slots


@dataclass
class CompCompany:
    """One public comparable: its Capital IQ ticker and a one-line description.

    - ``ticker``: CapIQ ``Exchange:Ticker`` identifier (e.g. ``"NasdaqGS:MSFT"``,
      ``"TSX:RY"``, ``"NYSE:JPM"``) -> column B. Every downstream metric column's
      array formula keys off this, so the exchange prefix is required.
    - ``description``: a <=50-char note on what the company does / sells ->
      column AA. No geography (the exchange prefix already signals it).
    """

    ticker: str
    description: str = ""


@dataclass
class Vertical:
    """One peer group ("vertical"): a label plus up to six comparables.

    - ``name``: the vertical / peer-group label -> ``D9`` / ``D19`` / ``D29``.
    - ``companies``: up to six ``CompCompany`` entries -> the block's rows.
    """

    name: str
    companies: list[CompCompany] = field(default_factory=list)


def _normalize_company(company: "CompCompany | dict") -> CompCompany:
    if isinstance(company, CompCompany):
        return company
    return CompCompany(ticker=company["ticker"], description=company.get("description", ""))


def _normalize_vertical(vertical: "Vertical | dict") -> Vertical:
    if isinstance(vertical, Vertical):
        companies = vertical.companies
        name = vertical.name
    else:
        companies = vertical.get("companies", [])
        name = vertical["name"]
    return Vertical(name=name, companies=[_normalize_company(c) for c in companies])


def build_comps_workbook(
    *,
    template_path: Path | str,
    verticals: "list[Vertical | dict]",
    output_path: Path | str,
) -> Path:
    """Fill the comps template's vertical blocks and return the saved path.

    ``verticals`` lists up to three peer groups, each with up to six public
    comparables. Writes only the vertical labels, the CapIQ tickers (column B)
    and the descriptions (column AA); the CapIQ array formulas and the statistic
    rows are left untouched so they resolve when the analyst refreshes Capital
    IQ. Unfilled verticals keep the template's ``[Group #N]`` placeholder. Raises
    ValueError on too many verticals / companies, an empty ticker, or an
    over-length description.
    """
    verticals = [_normalize_vertical(v) for v in verticals]
    if not verticals:
        raise ValueError("no verticals supplied — expected 1–3 peer groups")
    if len(verticals) > _MAX_VERTICALS:
        raise ValueError(
            f"comps template holds {_MAX_VERTICALS} vertical blocks; got {len(verticals)}"
        )
    for vertical in verticals:
        for company in vertical.companies:
            if not str(company.ticker).strip():
                raise ValueError(f"empty ticker in vertical {vertical.name!r}")
            if len(company.description) > _MAX_DESCRIPTION_CHARS:
                raise ValueError(
                    f"description for {company.ticker!r} is {len(company.description)} "
                    f"chars; max {_MAX_DESCRIPTION_CHARS} (column AA overflows beyond that)"
                )

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"comps template not found: {template}")
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template, out)

    wb = load_workbook(out)  # preserve formulas (no data_only)
    if _SHEET not in wb.sheetnames:
        raise KeyError(f"sheet {_SHEET!r} not found in comps template (have {wb.sheetnames})")
    ws = wb[_SHEET]
    # Verify the vertical blocks' sentinel anchors (row-7 headers + the 'Group
    # Average' row closing each block) AND that each block's defined name still
    # resolves to the region its sentinel pins, before writing anything.
    verify_anchors(ws, COMPS_BLOCK_ANCHORS, template=COMPS_TEMPLATE, require_names=True)

    for vertical, (label_cell, rows) in zip(verticals, _vertical_slots(ws)):
        if len(vertical.companies) > len(rows):
            raise ValueError(
                f"vertical {vertical.name!r} has {len(vertical.companies)} companies; "
                f"the template holds {len(rows)} rows per vertical"
            )
        ws[label_cell] = vertical.name
        for row, company in zip(rows, vertical.companies):
            ws[f"{_COL_TICKER}{row}"] = str(company.ticker).strip()
            description = company.description.strip()
            if description:
                ws[f"{_COL_DESCRIPTION}{row}"] = description

    wb.save(out)
    return out
