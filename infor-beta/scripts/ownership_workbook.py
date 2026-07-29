"""Populate the INFOR ownership workbook from SEDI insider holdings.

Companion-workbook builder for the pitch deck's insider-ownership slide,
mirroring ``ltm_metrics.py``. Parsing a SEDI "Insider Information by Issuer"
report — which insiders are current, their relationship codes, common-share
tranches, dates, and roles — is the ``ownership`` skill's LLM/web work; this
module is the deterministic write. Given the issuer's **current** insiders and
its total basic shares outstanding, it fills the template's Select-Insiders
data block (rows 39-65, columns B/F/G/J) and the % denominator (``F35``),
leaving the formula-driven display block (``B4:G17``), the include flags (col
H), and the adjusted-# formulas (col I) untouched.

When the analyst also attaches a **Bloomberg ownership export** (the BBG Excel
add-in "Ownership" template — a ``Summary View`` sheet with holder rows from
row 14), the builder fills the institutional side too: the export's holder
rows are copied into the template's ``Bloomberg Output`` tab (whose layout
mirrors the BBG Summary View, so the Ownership tab's pre-wired rows 68-185
``XLOOKUP`` link up unchanged), each Bloomberg holder that duplicates a SEDI
insider gets its Include flag (col H) set to 0 — **the SEDI figure always
wins; only the Bloomberg duplicate is excluded** — every populated Bloomberg
row gets an adjusted display name (col J, legal suffixes stripped), and the
unused link rows are neutralised so the Select-Institutions ``LARGE`` block
computes. Without an export the institutional side stays untouched (today's
insider-only behaviour).

The display block ranks the top 12 insiders by common shares via
``LARGE``/``XLOOKUP`` over the I (``=H*F``) helper column, so only an insider's
common-share count drives the slide — options, RSUs, PSUs and DSUs are excluded
by design (they are not common shares). Insiders holding 0 common shares are
still written (the analyst can toggle col H to exclude them).

The template's top-12 display *assumes* at least 12 positive balances, and
``_guard_display_block`` is what makes a shorter list safe: both display blocks
get their rank formulas rewritten so a surplus rank renders as an empty row
instead of repeating the first zero-balance holder. See that function for the
mechanism and for why the fix is here rather than in the template.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from deal_workbook import (
    TAB_BLOOMBERG_OUTPUT,
    TAB_CAPTABLE,
    TAB_OWNERSHIP,
    TabSpec,
    write_tab,
)
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries

from comment_citations import cite_cell
from provenance import FigureRef, FigureSource, ProvenanceError, ProvenanceLedger
from template_layout import (
    CAP_TABLE_SECTION_VII_NAMES,
    CAP_TABLE_SOURCE_SHEET,
    NAME_CAP_SHARE_INPUTS,
    NAME_OWN_BBG_HOLDER_BLOCK,
    NAME_OWN_BBG_LINK_BLOCK,
    NAME_OWN_INSIDER_BLOCK,
    NAME_OWN_INSIDERS_PICTURE,
    NAME_OWN_INSTITUTIONS_PICTURE,
    NAME_OWN_TOTAL_SHARES,
    OWNERSHIP_BBG_HOLDER_NAMES,
    OWNERSHIP_BBG_LINK_NAMES,
    OWNERSHIP_INSIDER_WRITE_NAMES,
    OWNERSHIP_INSIDERS_PICTURE_NAMES,
    OWNERSHIP_INSTITUTIONS_PICTURE_NAMES,
    OWNERSHIP_TEMPLATE,
    TemplateLayoutError,
    defined_name_ref,
    resolve_name_cell,
    resolve_name_range,
    verify_names,
)

# The tabs this module writes live in the DEAL workbook, so they are addressed by
# `deal_workbook.TAB_*` and not by the ownership source template's sheet names
# (identical strings today, but the cap table's are not — see the sheet-name note
# in `template_layout`).
_SHEET = TAB_OWNERSHIP
_DATE_FORMAT = "yyyy-mm-dd"  # SEDI reports dates as ISO 'YYYY-MM-DD'


def _row_span(ws, name: str) -> range:
    """The row span of a named block on ``ws`` (1-based, inclusive)."""
    _, first, _, last = range_boundaries(
        resolve_name_range(ws, name, template=OWNERSHIP_TEMPLATE)
    )
    return range(first, last + 1)

_COL_SEDI_NAME = "B"
_COL_BASIC = "F"
_COL_DATE = "G"
_COL_ADJ_NAME = "J"
_COL_INCLUDE = "H"
# The two display blocks reuse B/F/G as name / shares / %. A block's *rank* rows
# are recognised by their shares cell holding a `LARGE(<pool>, <rank>)` call, so
# no row span is hardcoded and the header / Subtotal / Other-Shareholders rows
# inside the same named range are skipped without being enumerated.
_COL_DISPLAY_PCT = "G"
_LARGE_CALL_RE = re.compile(r"\bLARGE\(\s*([^,()]+?)\s*,\s*([^,()]+?)\s*\)", re.IGNORECASE)

# ── Bloomberg institutional side ─────────────────────────────────────────────
# The template's 'Bloomberg Output' tab mirrors the BBG Excel add-in's
# "Summary View" layout: info cells (E7 name, E9 ticker, I9 view, E11/I11
# sort), row-13 column headers, holder rows from row 14. The Ownership tab's
# rows 68-185 are pre-wired against C14:C131 (name), L (position), N (filing
# date), so writing the export's rows into the same coordinates links the
# Select-Institutions block up with no formula work.
# These two pin the ANALYST'S EXPORT, not our template: they are the Bloomberg
# add-in's own Summary View layout, so no defined name of ours applies. The
# template side (where the rows land, and the Ownership tab's link rows) is
# resolved from `infor_own_bbg_holder_block` / `infor_own_bbg_link_block`.
_BBG_HEADER_ROW = 13
_BBG_FIRST_ROW = 14
# Columns copied per holder row (col B keeps the template's own 1-118 numbering).
_BBG_COPY_COLS = (
    "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q",
    "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC",
)
_BBG_INFO_CELLS = ("E7", "E9", "I9", "E11", "I11")  # name, ticker, view, sort by, order
# Row-13 headers used to recognise / validate the BBG Summary View layout.
_BBG_EXPECTED_HEADERS = {"C": "Holder Name", "L": "Position", "N": "Filing Date", "R": "Insider Status"}


def _blue(cell) -> Font:
    """Blue font for a hardcoded input cell, preserving the template's typeface.

    The hardcoded-value convention is blue text (matches the cap table), but a
    bare ``Font(color="0000FF")`` carries no name/size, so it resets the cell to
    PowerPoint/Excel's Calibri 11 default and drops the template's Palatino. Read
    the cell's existing font and re-emit it blue so name/size/weight survive.
    """
    f = cell.font
    return Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic, color="0000FF")


@dataclass
class InsiderHolding:
    """One current insider's common-share holding from a SEDI report.

    - ``sedi_name``: the raw SEDI name ("Last, First Middle") -> col B.
    - ``adjusted_name``: presentation name + role, e.g.
      ``"Mark Barrenechea (CEO & Director)"`` -> col J (what the slide shows).
    - ``common_shares``: the common-share count, or a list of per-registered-
      holder tranche counts (direct, RRSP, Holdco, ...) which is written as a
      sum formula -> col F. Common shares ONLY — never options/RSU/PSU/DSU.
    - ``most_recent_date``: the latest common-share transaction date -> col G.
    - ``source``: the :class:`provenance.FigureSource` for the share count — the
      SEDI report the analyst attached, and the page this insider's holdings are
      on: ``FigureSource(filing="SEDI Insider Information by Issuer report",
      statement="<insider name>", page=4)``. Recorded in the stage's ledger and
      rendered as the amount cell's ``Source: …`` comment.
    """

    sedi_name: str
    adjusted_name: str
    common_shares: int | list[int]
    most_recent_date: date | str | None = None
    source: "FigureSource | None" = None


def _coerce_date(value: date | str | None) -> date | None:
    if value is None or isinstance(value, date):
        return value
    text = str(value).strip()
    return date.fromisoformat(text) if text else None


def _basic_shares_cell_value(common_shares: int | list[int]):
    """A plain int for a single tranche; an Excel sum formula for several.

    Per the analyst workflow, multiple common-share tranches are summed *in the
    cell* (``=193000+0+0``) rather than pre-totalled, so the workbook documents
    every registered holding and stays auditable.
    """
    if isinstance(common_shares, (list, tuple)):
        tranches = [int(t) for t in common_shares]
        if not tranches:
            return None
        if len(tranches) == 1:
            return tranches[0]
        return "=" + "+".join(str(t) for t in tranches)
    return int(common_shares)


def _coerce_source(owner: str, source):
    """Validate one insider's source record — a citation string is not one.

    Same rule and same reason as `ltm_metrics._coerce_source`: a string would build
    a record with the whole sentence in ``filing`` and no statement or page, which
    reads like provenance and cannot be followed.
    """
    if source is None or isinstance(source, FigureSource):
        return source
    raise ProvenanceError(
        f"{owner!r} has a source of type {type(source).__name__} ({source!r}); pass "
        f"FigureSource(filing=…, statement=…, page=…) — a citation string is no "
        f"longer a source record."
    )


def _normalize(insider: "InsiderHolding | dict") -> InsiderHolding:
    if isinstance(insider, InsiderHolding):
        return InsiderHolding(
            sedi_name=insider.sedi_name,
            adjusted_name=insider.adjusted_name,
            common_shares=insider.common_shares,
            most_recent_date=insider.most_recent_date,
            source=_coerce_source(insider.sedi_name, insider.source),
        )
    return InsiderHolding(
        sedi_name=insider["sedi_name"],
        adjusted_name=insider["adjusted_name"],
        common_shares=insider["common_shares"],
        most_recent_date=insider.get("most_recent_date"),
        source=_coerce_source(insider["sedi_name"], insider.get("source")),
    )


def read_basic_shares_from_cap_table(captable_path: Path | str) -> int | None:
    """Return total basic shares outstanding (full units) from the cap table, or None.

    Sources ``F35`` of the `Ownership` tab from the `captable` tab of the same
    deal workbook — `captable_path` is the deal workbook (Phase D), though a
    standalone cap table with the original ``Cap with Links`` sheet still reads.
    Sums the Section VII basic-share **input** rows (col F, rows 168-185, in
    millions) and converts to full units. The captable skill writes those as
    hardcoded literals (not formulas), so they read reliably with openpyxl
    regardless of recalc state — unlike the Section VII total ``F186`` (a ``SUM``
    formula whose cached value openpyxl may not see). Sub-event rows (e.g. a
    buyback's negative row) net in correctly because they live in the same
    column. Returns None when the file is missing/unreadable or the sum is
    non-positive, so the caller can leave ``F35`` blank for the analyst.
    """
    try:
        wb = load_workbook(Path(captable_path), data_only=False)
    except Exception:
        return None
    for candidate in (TAB_CAPTABLE, CAP_TABLE_SOURCE_SHEET):
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            break
    else:
        ws = wb.active
    # The summing window comes from the cap table's `infor_cap_share_inputs`
    # name (F168:F185 as shipped), so an inserted Section VII row is summed too.
    # Verify it is there first: a cap table that lost the name must raise rather
    # than silently sum a hardcoded window that may no longer be Section VII.
    verify_names(ws, CAP_TABLE_SECTION_VII_NAMES, template=Path(captable_path).name)
    block = resolve_name_range(ws, NAME_CAP_SHARE_INPUTS, template=Path(captable_path).name)
    min_col, min_row, max_col, max_row = range_boundaries(block)
    total_millions = 0.0
    found = False
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                total_millions += float(cell.value)
                found = True
    if not found or total_millions <= 0:
        return None
    return round(total_millions * 1_000_000)


@dataclass
class BloombergHolder:
    """One holder row from a Bloomberg ownership export's Summary View.

    ``values`` / ``number_formats`` map column letters (C..AC) to the source
    cell's value / number format, so the row can be copied into the template's
    ``Bloomberg Output`` tab verbatim.
    """

    name: str
    position: float | int | None = None
    insider_status: str | None = None  # "Y" = insider, "N-P" etc. = institution
    institution_type: str | None = None
    filing_date: date | datetime | None = None
    values: dict[str, object] = field(default_factory=dict)
    number_formats: dict[str, str] = field(default_factory=dict)


@dataclass
class BloombergExport:
    """A parsed Bloomberg ownership export (BBG Excel add-in, Summary View)."""

    company_name: str | None
    ticker: str | None
    holders: list[BloombergHolder]
    info: dict[str, object] = field(default_factory=dict)  # info-cell coordinate -> value


def _find_summary_view(wb):
    """The sheet carrying the BBG Summary View layout ('Holder Name' in C13)."""
    candidates = list(wb.worksheets)
    if "Summary View" in wb.sheetnames:
        candidates.insert(0, wb["Summary View"])
    for ws in candidates:
        if ws.cell(row=_BBG_HEADER_ROW, column=3).value == "Holder Name":
            return ws
    return None


def read_bloomberg_export(path: Path | str) -> BloombergExport:
    """Parse a Bloomberg ownership export (.xlsm/.xlsx from the BBG Excel add-in).

    Locates the Summary View sheet (``'Holder Name'`` in ``C13``), validates the
    column layout the ownership template's link formulas assume, and returns the
    holder rows (values + number formats for columns C..AC). Raises with a clear
    message when the file does not look like a BBG ownership Summary View, so the
    skill can tell the analyst what to re-attach.
    """
    src = Path(path)
    if not src.exists():
        raise FileNotFoundError(f"Bloomberg ownership export not found: {src}")
    wb = load_workbook(src, data_only=True)  # cached values for any add-in formulas
    ws = _find_summary_view(wb)
    if ws is None:
        raise ValueError(
            f"{src.name} does not contain a Bloomberg ownership Summary View "
            "('Holder Name' expected in C13) — attach the BBG Excel add-in "
            "Ownership export (Summary View)"
        )
    mismatched = {
        col: ws[f"{col}{_BBG_HEADER_ROW}"].value
        for col, expected in _BBG_EXPECTED_HEADERS.items()
        if ws[f"{col}{_BBG_HEADER_ROW}"].value != expected
    }
    if mismatched:
        raise ValueError(
            f"{src.name} Summary View columns differ from the layout the ownership "
            f"template links against (row {_BBG_HEADER_ROW}: expected "
            f"{_BBG_EXPECTED_HEADERS}, found {mismatched})"
        )

    holders: list[BloombergHolder] = []
    row = _BBG_FIRST_ROW
    while True:
        name = ws[f"C{row}"].value
        if name is None or str(name).strip() == "":
            break
        values: dict[str, object] = {}
        formats: dict[str, str] = {}
        for col in _BBG_COPY_COLS:
            cell = ws[f"{col}{row}"]
            if cell.value is not None:
                values[col] = cell.value
                formats[col] = cell.number_format
        holders.append(
            BloombergHolder(
                name=str(name).strip(),
                position=values.get("L"),
                insider_status=values.get("R"),
                institution_type=values.get("S"),
                filing_date=values.get("N"),
                values=values,
                number_formats=formats,
            )
        )
        row += 1

    info = {coord: ws[coord].value for coord in _BBG_INFO_CELLS if ws[coord].value is not None}
    return BloombergExport(
        company_name=ws["E7"].value,
        ticker=ws["E9"].value,
        holders=holders,
        info=info,
    )


# Trailing legal-form tokens stripped from an institution's display name. A
# token preceded by '&' is part of the brand, not a suffix ("Kelso & Co LP" ->
# "Kelso & Co", never "Kelso"). Stripping repeats ("Vanguard Group Inc" ->
# "Vanguard") but never empties the name.
_LEGAL_SUFFIX_TOKENS = {
    "inc", "incorporated", "corp", "corporation", "co", "company",
    "ltd", "limited", "llc", "lp", "llp", "ulc", "plc",
    "sa", "nv", "ag", "se", "spa",
    "group", "partners", "holdings", "holding",
}


def strip_legal_suffixes(name: str) -> str:
    """Default adjusted name: 'T Rowe Price Group Inc' -> 'T Rowe Price'.

    The skill agent may override individual names where house style differs
    (e.g. restoring 'T. Rowe Price' punctuation); this is the deterministic
    baseline. Person names pass through unchanged (their tokens are not legal
    suffixes).
    """
    tokens = str(name).split()
    while len(tokens) > 1:
        tail = tokens[-1].rstrip(".,").lower()
        if tail in _LEGAL_SUFFIX_TOKENS and tokens[-2] != "&":
            tokens.pop()
        else:
            break
    return " ".join(tokens) if tokens else str(name)


_NON_WORD_RE = re.compile(r"[^\w\s]", re.UNICODE)


def _name_tokens(name: str) -> list[str]:
    return [t for t in _NON_WORD_RE.sub(" ", str(name)).lower().split() if t]


def _token_matches(a: str, b: str) -> bool:
    """Exact token match, or an initial matching a full token's first letter."""
    if a == b:
        return True
    if len(a) == 1 or len(b) == 1:
        return a[0] == b[0]
    return False


def _givens_compatible(a: list[str], b: list[str]) -> bool:
    """Given-name token lists agree (order-free, initials tolerated).

    Both sides must carry *something* (surname-only never matches), full names
    must share at least one exact token when both sides have full names
    ("Mary" vs "M Christine" is rejected; "Eric" vs "Eric D" accepted), and
    every token of the shorter list must find a distinct flexible match in the
    longer ("Smith M Christine" matches SEDI "Smith, M Christine").
    """
    if not a or not b:
        return False
    a_full = {t for t in a if len(t) > 1}
    b_full = {t for t in b if len(t) > 1}
    if a_full and b_full and not (a_full & b_full):
        return False
    short, rest = (list(a), list(b)) if len(a) <= len(b) else (list(b), list(a))
    for t in short:
        matched = next((u for u in rest if _token_matches(t, u)), None)
        if matched is None:
            return False
        rest.remove(matched)
    return True


def bloomberg_matches_sedi(bbg_name: str, sedi_name: str) -> bool:
    """True when a Bloomberg holder duplicates a SEDI insider.

    Two paths: (1) corporate / exact — the suffix-stripped, punctuation-free
    token lists are identical (covers holdco insiders and letter-perfect person
    matches); (2) person — SEDI prints ``Last, First Middle`` and Bloomberg
    prints ``Last First Middle``, so the SEDI surname must prefix the Bloomberg
    tokens and the given names must agree per ``_givens_compatible``.
    """
    if _name_tokens(strip_legal_suffixes(bbg_name)) == _name_tokens(
        strip_legal_suffixes(sedi_name.replace(",", " "))
    ):
        return True
    if "," not in sedi_name:
        return False
    last, _, given = sedi_name.partition(",")
    surname = _name_tokens(last)
    btoks = _name_tokens(bbg_name)
    k = len(surname)
    if k == 0 or len(btoks) <= k or btoks[:k] != surname:
        return False
    return _givens_compatible(_name_tokens(given), btoks[k:])


def match_bloomberg_to_sedi(
    bbg_names: "list[str]", sedi_names: "list[str]"
) -> dict[str, str]:
    """Map each Bloomberg holder name to the SEDI insider it duplicates.

    Deterministic and conservative — the skill reviews this report (and any
    Bloomberg rows flagged Insider Status "Y" that did NOT match) and can
    correct individual rows via ``build_ownership_workbook``'s
    ``bloomberg_include_overrides``.
    """
    matches: dict[str, str] = {}
    for bbg in bbg_names:
        for sedi in sedi_names:
            if bloomberg_matches_sedi(bbg, sedi):
                matches[bbg] = sedi
                break
    return matches


def _write_bloomberg_side(
    wb,
    export: BloombergExport,
    *,
    sedi_names: "list[str]",
    adjusted_overrides: dict[str, str],
    include_overrides: dict[str, int],
    ledger: "ProvenanceLedger | None" = None,
    source_name: str | None = None,
) -> None:
    """Fill 'Bloomberg Output' + the Ownership tab's institutional link rows."""
    if TAB_BLOOMBERG_OUTPUT not in wb.sheetnames:
        raise KeyError(
            f"tab {TAB_BLOOMBERG_OUTPUT!r} not found in the deal workbook "
            f"(have {wb.sheetnames})"
        )
    ws_bbg = wb[TAB_BLOOMBERG_OUTPUT]
    ws_own = wb[_SHEET]
    # The export side is validated in read_bloomberg_export; this validates the
    # template side the rows land on — the 'Bloomberg Output' holder block the
    # holder rows are copied into, and the Ownership tab's pre-wired link rows
    # (whose H/J columns are written and B/F/G neutralised below).
    verify_names(ws_bbg, OWNERSHIP_BBG_HOLDER_NAMES, template=OWNERSHIP_TEMPLATE)
    verify_names(ws_own, OWNERSHIP_BBG_LINK_NAMES, template=OWNERSHIP_TEMPLATE)

    # Both spans come from the template. They must stay the same length: the
    # Ownership tab's link rows are pre-wired one-to-one against the holder
    # rows, so a template that resized one and not the other is a layout error.
    bbg_rows = _row_span(ws_bbg, NAME_OWN_BBG_HOLDER_BLOCK)
    own_rows = _row_span(ws_own, NAME_OWN_BBG_LINK_BLOCK)
    if len(bbg_rows) != len(own_rows):
        raise TemplateLayoutError(
            f"{OWNERSHIP_TEMPLATE}: the Bloomberg holder block holds {len(bbg_rows)} rows "
            f"but the Ownership tab's link block holds {len(own_rows)} — they are wired "
            f"one-to-one and must match."
        )

    holders = export.holders
    if len(holders) > len(bbg_rows):
        print(
            f"ownership: Bloomberg export has {len(holders)} holders; the template "
            f"holds {len(bbg_rows)} (rows {bbg_rows[0]}-{bbg_rows[-1]}) — writing the "
            f"first {len(bbg_rows)} (the Summary View is sorted by position, so the "
            "tail is smallest)",
            file=sys.stderr,
        )
        holders = holders[: len(bbg_rows)]

    for coord, value in export.info.items():
        ws_bbg[coord] = value

    matches = match_bloomberg_to_sedi([h.name for h in holders], sedi_names)

    position_col = next(
        (col for col, header in _BBG_EXPECTED_HEADERS.items() if header == "Position"), None
    )

    for offset, holder in enumerate(holders):
        bbg_row = bbg_rows[offset]
        for col, value in holder.values.items():
            cell = ws_bbg[f"{col}{bbg_row}"]
            cell.value = value
            fmt = holder.number_formats.get(col)
            if fmt and fmt != "General":
                cell.number_format = fmt

        if ledger is not None and position_col is not None:
            # The institutional side's source is the attached export, not a filing —
            # Bloomberg is where the position was read, and the export is the
            # document in the deal directory that says so.
            ledger.record(
                f"Institutional position — {holder.name}",
                sources=FigureSource(
                    filing=source_name or "Bloomberg ownership export",
                    statement="Summary View",
                ),
                value=holder.values.get(position_col),
                units="shares",
                location=f"{TAB_BLOOMBERG_OUTPUT}!{position_col}{bbg_row}",
            )

        own_row = own_rows[offset]
        include = include_overrides.get(holder.name, 0 if holder.name in matches else 1)
        if include != 1:
            include_cell = ws_own[f"{_COL_INCLUDE}{own_row}"]
            include_cell.value = 0
            include_cell.font = _blue(include_cell)
            reason = (
                f"Excluded - duplicate of SEDI insider '{matches[holder.name]}'; "
                "the SEDI figure (Select Insiders block) is kept."
                if holder.name in matches
                else "Excluded by the ownership skill (analyst/agent override)."
            )
            include_cell.comment = Comment(reason, "INFOR")

        adj_cell = ws_own[f"{_COL_ADJ_NAME}{own_row}"]
        adj_cell.value = adjusted_overrides.get(holder.name) or strip_legal_suffixes(holder.name)
        adj_cell.font = _blue(adj_cell)

    # Neutralise the unused link rows so the Select-Institutions LARGE range
    # stays numeric: an un-fed 'Bloomberg Output' XLOOKUP evaluates to #N/A,
    # which would poison LARGE($I$68:$I$185, k). Clearing B/F/G and zeroing H
    # leaves I (=H*F) at 0.
    for own_row in own_rows[len(holders) :]:
        for col in (_COL_SEDI_NAME, _COL_BASIC, _COL_DATE):
            ws_own[f"{col}{own_row}"] = None
        ws_own[f"{_COL_INCLUDE}{own_row}"] = 0


def _guard_display_block(ws, picture_name: str) -> None:
    """Make a top-12 display block survive fewer than 12 positive balances.

    A display row is ``F=LARGE(<pool>, <rank>)`` for the share count,
    ``B=XLOOKUP(F<row>, ...)`` for the holder name and ``G=F<row>/$F$35`` for the
    percentage. The pool is the ``I`` (``=H*F``) helper column over *every* data
    row, and an unfed data row computes to **0** rather than being absent — so
    when the pool holds fewer positive balances than there are ranks, ``LARGE``
    returns 0 for each surplus rank and ``XLOOKUP(0, ...)`` matches the first
    zero-balance row. The slide then prints that one holder's name once per
    surplus rank (the shipped symptom: "Ayman Antoun (CEO & Director) -- <0.1%"
    twice).

    Each rank is therefore guarded by a ``COUNTIF`` over the same pool it draws
    from, and the name and percentage follow the share cell: a surplus rank
    renders as an empty row. The arithmetic stays in Excel, and the block's
    ``Subtotal`` is untouched — ``SUM`` ignores the ``""`` a guarded rank yields.

    Raises ``TemplateLayoutError`` when the block holds no recognisable rank row,
    rather than leaving the guard a silent no-op.

    Written from the stage rather than baked into the template on purpose: a
    template edit forces an `add_template_named_ranges.py` +
    `build_deal_workbook_template.py` rebuild. Idempotent — a block whose share
    cells already carry the ``COUNTIF`` guard is left alone, so a re-run of the
    stage does not wrap twice.
    """
    _, first, _, last = range_boundaries(
        resolve_name_range(ws, picture_name, template=OWNERSHIP_TEMPLATE)
    )
    rank_rows = 0
    for row in range(first, last + 1):
        basic = ws[f"{_COL_BASIC}{row}"]
        formula = basic.value
        if not isinstance(formula, str):
            continue
        match = _LARGE_CALL_RE.search(formula)
        if match is None:
            continue  # a header, the Subtotal, Other Shareholders — not a rank
        rank_rows += 1
        if "COUNTIF(" in formula.upper():
            continue  # already guarded (the stage ran before on this workbook)
        pool, rank = match.group(1), match.group(2)
        basic.value = f'=IF({rank}<=COUNTIF({pool},">0"),LARGE({pool},{rank}),"")'
        # The name and the percentage read the share cell, so both go blank with
        # it — otherwise B repeats a name and G divides "" and shows #VALUE!.
        for col in (_COL_SEDI_NAME, _COL_DISPLAY_PCT):
            cell = ws[f"{col}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = f'=IF({_COL_BASIC}{row}="","",{cell.value[1:]})'
    if not rank_rows:
        # A block whose rank formulas stopped looking like `LARGE(<pool>, <rank>)`
        # would leave this a silent no-op, and the padding duplicates would come
        # straight back — so halt instead of writing a workbook that looks guarded.
        raise TemplateLayoutError(
            f"{OWNERSHIP_TEMPLATE}: found no rank row in the block named "
            f"{picture_name!r} (rows {first}-{last}). A display row is recognised by "
            f"its {_COL_BASIC} cell holding a LARGE(<pool>, <rank>) call; if the "
            f"template's ranking changed shape, update _guard_display_block with it."
        )


def build_ownership_workbook(
    *,
    insiders: "list[InsiderHolding | dict]",
    total_shares_outstanding: int | None,
    deal_workbook: Path | str,
    bloomberg_export_path: Path | str | None = None,
    bloomberg_adjusted_names: dict[str, str] | None = None,
    bloomberg_include_overrides: dict[str, int] | None = None,
    provenance: "ProvenanceLedger | None" = None,
) -> Path:
    """Fill the deal workbook's `Ownership` tab; return the workbook path.

    Since Phase D there is no standalone ownership file and no template to copy:
    the `Ownership` and `Bloomberg Output` tabs are already in the deal workbook,
    carried in from `INFOR Deal Workbook Template.xlsx` at deal-init — as a PAIR,
    so the Ownership tab's `XLOOKUP` rows against `'Bloomberg Output'` are an
    internal reference rather than a link back to the template.

    ``insiders`` lists the issuer's **current** insiders only — those whose
    SEDI "Ceased to be Insider" is "Not Applicable". ``total_shares_outstanding``
    is the issuer's basic shares outstanding in **full units** (not millions),
    typically sourced from the companion cap table; it drives the % column via
    ``F35``. Pass ``None`` to leave ``F35`` for the analyst to fill.

    ``bloomberg_export_path`` (optional) is the analyst-attached Bloomberg
    ownership export; when supplied the builder also fills the ``Bloomberg
    Output`` tab and the Ownership tab's institutional H/J columns (see module
    docstring). ``bloomberg_adjusted_names`` overrides the default
    suffix-stripped display name per Bloomberg holder name;
    ``bloomberg_include_overrides`` (holder name -> 0/1) overrides the
    computed SEDI-duplicate exclusion for individual rows.

    ``provenance`` is filled **in place**: one record per insider share count (from
    that insider's ``source``), one for each institutional position (the attached
    Bloomberg export), and one for the ``F35`` denominator as a **derived** figure
    referring to the cap table's basic shares — the cross-tab link that makes the
    ownership slide's percentages traceable to the filing the share count came
    from. Pass the stage's ledger and write it afterwards
    (`ledger.write(io.stage_dir)`).
    """
    insiders = [_normalize(i) for i in insiders]

    def _write(wb, ws) -> None:
        _fill_ownership_tab(
            wb,
            ws,
            insiders=insiders,
            total_shares_outstanding=total_shares_outstanding,
            bloomberg_export_path=bloomberg_export_path,
            bloomberg_adjusted_names=bloomberg_adjusted_names,
            bloomberg_include_overrides=bloomberg_include_overrides,
            ledger=provenance,
        )

    write_tab(
        deal_workbook,
        TAB_OWNERSHIP,
        TabSpec(write=_write, verify_names=(NAME_OWN_INSIDER_BLOCK, NAME_OWN_TOTAL_SHARES)),
    )
    return Path(deal_workbook)


def _fill_ownership_tab(
    wb,
    ws,
    *,
    insiders: "list[InsiderHolding]",
    total_shares_outstanding: int | None,
    bloomberg_export_path: Path | str | None,
    bloomberg_adjusted_names: dict[str, str] | None,
    bloomberg_include_overrides: dict[str, int] | None,
    ledger: "ProvenanceLedger | None" = None,
) -> None:
    """Write the insider block, the % denominator and (optionally) the BBG side."""
    # Verify the names the writes resolve through — the insider block, the %
    # denominator, and the two display blocks whose rank formulas get guarded —
    # before touching a cell, so a tab that lost one reports every miss at once.
    verify_names(
        ws,
        OWNERSHIP_INSIDER_WRITE_NAMES
        + OWNERSHIP_INSIDERS_PICTURE_NAMES
        + OWNERSHIP_INSTITUTIONS_PICTURE_NAMES,
        template=OWNERSHIP_TEMPLATE,
    )

    insider_rows = _row_span(ws, NAME_OWN_INSIDER_BLOCK)
    if len(insiders) > len(insider_rows):
        raise ValueError(
            f"ownership template holds {len(insider_rows)} insider rows "
            f"({insider_rows[0]}-{insider_rows[-1]}); got {len(insiders)} current insiders"
        )

    for row, insider in zip(insider_rows, insiders):
        name_cell = ws[f"{_COL_SEDI_NAME}{row}"]
        name_cell.value = insider.sedi_name
        name_cell.font = _blue(name_cell)

        basic_cell = ws[f"{_COL_BASIC}{row}"]
        basic_cell.value = _basic_shares_cell_value(insider.common_shares)
        basic_cell.font = _blue(basic_cell)
        if ledger is not None and insider.source is not None:
            # The record first, the comment rendered from it — never the reverse.
            cite_cell(
                basic_cell,
                ledger.record(
                    f"Insider holding — {insider.adjusted_name}",
                    sources=insider.source,
                    value=basic_cell.value,
                    units="shares",
                    location=f"{TAB_OWNERSHIP}!{_COL_BASIC}{row}",
                ),
            )

        date_cell = ws[f"{_COL_DATE}{row}"]
        coerced = _coerce_date(insider.most_recent_date)
        if coerced is not None:
            date_cell.value = coerced
            date_cell.number_format = _DATE_FORMAT
            date_cell.font = _blue(date_cell)

        adj_cell = ws[f"{_COL_ADJ_NAME}{row}"]
        adj_cell.value = insider.adjusted_name
        adj_cell.font = _blue(adj_cell)

    if total_shares_outstanding is not None:
        total_ref = resolve_name_cell(ws, NAME_OWN_TOTAL_SHARES, template=OWNERSHIP_TEMPLATE)
        total_cell = ws[total_ref]
        total_cell.value = int(total_shares_outstanding)
        # Leave the cell's font untouched — the template ships it Palatino (bold), and
        # the aggregator later relinks it to the cap table's basic shares. Setting
        # a bare Font here would reset it to Calibri 11.
        total_cell.comment = Comment(
            "Total basic shares outstanding (full units) - from the companion cap table.",
            "INFOR",
        )
        if ledger is not None:
            # Every percentage on the ownership slide divides by this cell, and its
            # provenance is the cap table's — `read_basic_shares_from_cap_table`
            # summed the Section VII input rows, which the captable stage recorded
            # against the filing they came from. The ref is by figure name, not by
            # cell: it crosses stages, so it resolves in the run merge.
            ledger.record(
                "Total basic shares outstanding (% denominator)",
                value=int(total_shares_outstanding),
                units="shares",
                location=f"{TAB_OWNERSHIP}!{total_ref}",
                derivation=(
                    "the cap table's Section VII basic-share input rows, converted "
                    "from millions to full units"
                ),
                derived_from=[FigureRef(figure="Total Basic Shares Outstanding")],
            )

    if bloomberg_export_path is not None:
        _write_bloomberg_side(
            wb,
            read_bloomberg_export(bloomberg_export_path),
            sedi_names=[i.sedi_name for i in insiders],
            adjusted_overrides=bloomberg_adjusted_names or {},
            include_overrides=bloomberg_include_overrides or {},
            ledger=ledger,
            source_name=Path(bloomberg_export_path).name,
        )

    # Both display blocks, unconditionally: an issuer with fewer than 12
    # positive-balance insiders is ordinary, and the institutional block pads the
    # same way (an excluded SEDI duplicate or an unused link row is a 0 in the
    # pool). Without a Bloomberg export the institutional pool is #N/A rather
    # than 0, which COUNTIF ignores — so the guard blanks those rows too.
    for picture_name in (NAME_OWN_INSIDERS_PICTURE, NAME_OWN_INSTITUTIONS_PICTURE):
        _guard_display_block(ws, picture_name)
