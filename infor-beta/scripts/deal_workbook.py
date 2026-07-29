"""The deal's single workbook — one file, owned from stage one.

Phase D's centrepiece. Before this, every producing skill wrote a standalone
`.xlsx` and a final `workbook-aggregation` stage merged them: on Windows through
Excel COM, off-Windows through a best-effort openpyxl copy plus a LibreOffice
recalc. That merge is what produced the phase's whole tail of defects — the
LibreOffice `~` union-operator rewrite that made Excel repair-strip 44 formulas
on open, external-reference rebinding, order-dependent sheet renames, a
source-deletion gate to decide when the merge could be trusted, comment and
hyperlink carry-over, and the rule that a `financial-summary` LTM link resolves
*only* in the combined workbook. None of it exists here, because there is
nothing to merge: `pitch-<codename>.xlsx` is created at deal-init with its
template tabs already in place, and each stage writes its own tab.

Three consequences worth stating, because code elsewhere used to work around
their absence:

1. **Cross-tab links resolve immediately.** `financial-summary`'s
   `=INDEX('ltm-metrics'!…)` LTM lookup points at a tab in the same file the
   moment it is written, so nothing has to run "after aggregation".
2. **Tab names are fixed, not derived.** They are the constants below, matching
   the names the merge used to produce, so the deck assemblers and
   `financial_charts` address the same tabs as before.
3. **Writes are serialized.** The conductor dispatches a wave's stages
   concurrently, so `comps`, `precedents` and `financial-summary` can all reach
   for this file at once. `write_tab` takes an exclusive lock for the whole
   load → mutate → save cycle; two stages cannot interleave and lose one
   another's tab.

The lock releases without deleting anything
-------------------------------------------
The deal directory is the analyst's, which means it can be a cloud-synced mount
that refuses `unlink()`. So the lock's state is its file's **content** and release
truncates when it cannot delete — see `_workbook_lock`. **No stage should ever
have a reason to patch the locking**, and two did once (one no-op'd
`_workbook_lock`, the other wrote to a `/tmp` copy and copied back) because a
denied unlink wedged the workbook for 120 s per write and the error text asked for
a deletion the filesystem forbade. Serialization is not optional here — without it
a wave's last save wins and the other tabs are gone — so if the lock ever looks
like the thing standing between a stage and forward progress, that is a defect in
the lock, not a case for working around it.

Template fidelity
-----------------
The five template tabs are copied in at init by copying the shipped
`INFOR Deal Workbook Template.xlsx`, which `tools/build_deal_workbook_template.py`
assembles from the four source templates. That template is a *pre-assembled*
artefact for a reason: openpyxl cannot move a sheet between workbooks, and a
cell-by-cell copy is precisely the lossy merge this phase deletes. So init is a
`shutil.copyfile` and nothing is reconstructed at runtime.

The Phase C `infor_` defined names are what make that safe. They are
**worksheet-scoped**, so they travel with their sheet and keep resolving after
the copy — `resolve_name_cell(ws, NAME_FX_RATE)` works on the `captable` tab of
a deal workbook exactly as it did on a standalone cap table. `write_tab` verifies
that on every write (see `_verify_tab_names`) rather than trusting it, because
`resolve_name_range` raises loudly and a stage failing three waves in is a worse
diagnostic than a stage failing at its first write.
"""

from __future__ import annotations

import os
import shutil
import sys
import time
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from naming import safe_filename

# ─── Tab names ───────────────────────────────────────────────────────────────
# Verbatim the names the aggregator's merge produced (a single-sheet source took
# its skill key; the ownership pair kept its own sheet names), so every
# downstream reader — both deck assemblers, `financial_charts` — addresses the
# same tabs it always did.
TAB_CAPTABLE = "captable"
TAB_COMPS = "comps"
TAB_FINANCIAL_SUMMARY = "financial-summary"
TAB_LTM_METRICS = "ltm-metrics"
TAB_OWNERSHIP = "Ownership"
TAB_BLOOMBERG_OUTPUT = "Bloomberg Output"
TAB_PRECEDENTS = "precedents"

#: Canonical left-to-right order, matching the frozen production fixture
#: (`scripts/tests/fixtures/pitch-workbook.xlsx`) so a Phase D workbook opens
#: looking like the ones analysts have already been handed.
TAB_ORDER: tuple[str, ...] = (
    TAB_CAPTABLE,
    TAB_COMPS,
    TAB_FINANCIAL_SUMMARY,
    TAB_LTM_METRICS,
    TAB_OWNERSHIP,
    TAB_BLOOMBERG_OUTPUT,
    TAB_PRECEDENTS,
)

#: Tabs the shipped deal-workbook template carries (the rest are authored from
#: scratch by their producer, so they are created on first write).
TEMPLATE_TABS: frozenset[str] = frozenset(
    {TAB_CAPTABLE, TAB_COMPS, TAB_OWNERSHIP, TAB_BLOOMBERG_OUTPUT, TAB_PRECEDENTS}
)

#: Which tabs each deliverable needs. An earnings update has no comps /
#: precedents / ownership section, so those template tabs are dropped at init
#: rather than shipped empty in a client-facing workbook.
DELIVERABLE_TABS: dict[str, tuple[str, ...]] = {
    "pitch": TAB_ORDER,
    "earnings-update": (TAB_CAPTABLE, TAB_LTM_METRICS),
}

DEAL_WORKBOOK_TEMPLATE = "INFOR Deal Workbook Template.xlsx"


class DealWorkbookError(RuntimeError):
    """The deal workbook is missing, locked by a stuck writer, or malformed."""


# ─── Naming ──────────────────────────────────────────────────────────────────


def workbook_filename(deliverable_type: str, deal_name: str) -> str:
    """Return `<deliverable>-<deal name>.xlsx`.

    Unchanged from the aggregator's `combined_filename`, so a deal directory's
    workbook keeps the name analysts know: the deliverable prefix drops hyphens
    (`earnings-update` -> `earningsupdate`) while `pitch` stays `pitch`.
    """
    prefix = deliverable_type.replace("-", "").strip() or "deliverable"
    return f"{prefix}-{safe_filename(deal_name, default='Deal')}.xlsx"


def deal_workbook_path(
    deal_dir: Path | str, deliverable_type: str, deal_name: str
) -> Path:
    """Where the deal's workbook lives. Does not create it — see `init_deal_workbook`."""
    return Path(deal_dir).expanduser() / workbook_filename(deliverable_type, deal_name)


# ─── Init ────────────────────────────────────────────────────────────────────


def _template_path(template_name: str = DEAL_WORKBOOK_TEMPLATE) -> Path:
    """Resolve a shipped template through `CLAUDE_PLUGIN_ROOT`, as the skills do."""
    root = Path(os.environ.get("CLAUDE_PLUGIN_ROOT", Path(__file__).resolve().parents[1]))
    candidate = root / "templates" / template_name
    if candidate.is_file():
        return candidate
    # A worktree / test run whose CLAUDE_PLUGIN_ROOT is unset or points elsewhere.
    fallback = Path(__file__).resolve().parents[1] / "templates" / template_name
    if fallback.is_file():
        return fallback
    raise DealWorkbookError(
        f"{template_name} not found under {candidate.parent} or {fallback.parent}. "
        f"It is assembled by tools/build_deal_workbook_template.py — re-run that "
        f"if a source template was re-saved."
    )


# ─── Filesystem capability ───────────────────────────────────────────────────
#: Whether a deal directory permits `unlink()`, probed once per directory. The
#: lock does NOT need unlink — release falls back to truncation — so this exists
#: to say out loud that a directory denies it, not to change how the lock behaves.
#: There is deliberately no lock-free degraded mode: going unserialized would
#: trade a 120 s hang for a lost tab, which is the worse of the two defects, and
#: `write_tab` is the only mutation path precisely so a wave cannot lose one.
_UNLINK_PROBED: dict[str, bool] = {}

#: The probe file's name. **Fixed, not unique per process**, because on the
#: filesystem this exists to detect the probe file cannot be deleted either — a
#: unique name would accumulate one permanent scrap per deal-init, littering the
#: analyst's deal directory to report that littering it is unavoidable. One file,
#: overwritten by every later probe, whose surviving presence is itself the finding.
_UNLINK_PROBE_NAME = ".infor-fs-probe"
_UNLINK_PROBE_BODY = (
    "Written and immediately deleted by infor-beta to check whether this directory "
    "permits file deletion. If you are reading it, the answer was no — the deal "
    "workbook's lock releases by emptying its .lock file here instead of removing "
    "it. Safe to delete if your filesystem ever lets you.\n"
)


def deal_dir_permits_unlink(deal_dir: Path | str, *, refresh: bool = False) -> bool:
    """Whether this directory lets us delete a file we just created.

    A real create + unlink, because the answer is a property of the mount (a
    cloud-sync mount can refuse) and not of the platform, and cached per directory
    because the answer cannot change under a run. An inconclusive probe — one that
    could not even create the file — reports True and is not cached: the deal
    directory is unwritable, which `init_deal_workbook` is about to fail on for a
    much more direct reason.
    """
    key = str(Path(deal_dir).expanduser())
    if refresh:
        _UNLINK_PROBED.pop(key, None)
    if key in _UNLINK_PROBED:
        return _UNLINK_PROBED[key]

    probe = Path(key) / _UNLINK_PROBE_NAME
    try:
        probe.write_text(_UNLINK_PROBE_BODY, encoding="utf-8")
    except OSError:
        return True
    try:
        probe.unlink()
    except FileNotFoundError:
        pass  # a concurrent probe removed it, which answers the question
    except OSError:
        _UNLINK_PROBED[key] = False
        return False
    _UNLINK_PROBED[key] = True
    return True


def _report_unlink_capability(path: Path) -> None:
    """Say so, once per deal directory, when the lock will release by truncation.

    Named at deal-init rather than discovered mid-wave, because a 0-byte `.lock`
    file surviving in the deal directory is otherwise the kind of thing a reader
    diagnoses as the bug instead of the accommodation.
    """
    if deal_dir_permits_unlink(path.parent):
        return
    print(
        f"deal_workbook: {path.parent} denies unlink(). The workbook lock will "
        f"release by emptying {path.name}.lock instead of deleting it, so a 0-byte "
        f"lock file (and a {_UNLINK_PROBE_NAME} explaining itself) may stay in the "
        f"deal directory — an empty lock is a free lock. Writes stay serialized; "
        f"nothing degrades.",
        file=sys.stderr,
    )


def init_deal_workbook(
    *,
    deal_dir: Path | str,
    deliverable_type: str,
    deal_name: str,
    overwrite: bool = False,
) -> Path:
    """Create the deal's workbook from the shipped template. Returns its path.

    Idempotent by default: an existing workbook is left exactly as it is, so
    re-running deal-init over a deal directory mid-run cannot discard the tabs
    stages have already written. Pass `overwrite=True` to start clean.

    Tabs the deliverable does not use are dropped (see `DELIVERABLE_TABS`), so
    an earnings-update workbook does not carry an empty comps section.
    """
    path = deal_workbook_path(deal_dir, deliverable_type, deal_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    _report_unlink_capability(path)
    if path.exists() and not overwrite:
        return path

    shutil.copyfile(_template_path(), path)

    wanted = DELIVERABLE_TABS.get(deliverable_type)
    if wanted is not None:
        drop = [tab for tab in TEMPLATE_TABS if tab not in wanted]
        if drop:
            with _workbook_lock(path):
                wb = load_workbook(path)
                try:
                    for tab in drop:
                        if tab in wb.sheetnames:
                            del wb[tab]
                    wb.save(path)
                finally:
                    wb.close()
    return path


# ─── Serialized tab writes ───────────────────────────────────────────────────


@dataclass(frozen=True)
class TabSpec:
    """One tab's write.

    - ``write``: called as ``write(workbook, worksheet)`` with the deal workbook
      open and the target tab resolved. Everything cell-level stays in the
      producer module; this module owns only the file, the lock and the tab.
    - ``create``: the tab is authored from scratch rather than shipped by the
      template (`ltm-metrics`, `financial-summary`). A re-run replaces it, so a
      retried stage does not append a second copy or write over half of one.
    - ``verify_names``: `infor_` defined names that must resolve on the tab
      before ``write`` is called. Producers resolving through names pass theirs,
      turning a template that lost its names into an immediate, named failure.
    """

    write: Callable[[object, object], None]
    create: bool = False
    verify_names: tuple[str, ...] = ()


_LOCK_TIMEOUT_S = 120.0
_LOCK_POLL_S = 0.1
#: A lock older than this is from a writer that died (the conductor's stages are
#: single writes, not long sessions), so it is broken rather than waited out.
_LOCK_STALE_S = 300.0
#: Pause between claiming a free lock file and confirming the claim survived. Two
#: writers that both found it empty both write their token; whichever token is
#: still there after this owns the lock and the other retries, so exactly one
#: proceeds. It only has to cover a rival's gap between reading "empty" and
#: writing — two adjacent syscalls — so 50 ms is generous.
_LOCK_CONFIRM_S = 0.05

#: `_read_state` outcomes that are not a holder token.
_VANISHED = object()  # the file is gone: free, take the fast path again
_UNREADABLE = object()  # a transient read error: assume held, retry


class _workbook_lock:
    """Exclusive lock on one workbook, held for a whole load → mutate → save.

    A sibling `.lock` file — uniform across Windows and Linux, unlike
    `msvcrt.locking` / `fcntl.flock`, and visible to a developer wondering why a
    stage is waiting. Waits up to `_LOCK_TIMEOUT_S`, breaks a lock left behind by
    a killed writer after `_LOCK_STALE_S`.

    **The lock's state is the file's CONTENT, not its existence.** A non-empty
    lock is held by the token it names; an empty one is free. That is what makes
    it survive a filesystem that refuses `unlink()`, which a cloud-synced deal
    directory does. The previous revision released by unlinking inside
    `except OSError: pass`, so a denied unlink left the *holder's own* lock behind
    silently, and `_break_if_stale` could not recover it either — it returned
    False when unlink raised. Every later `write_tab` then waited the full 120 s
    and failed with a message telling the analyst to delete a file nothing on that
    mount could delete. In one pitch run two sub-agents reached the same
    conclusion independently and monkeypatched this class — one to a no-op, the
    other onto a `/tmp` copy — and patching the serialization out is exactly how a
    wave loses a tab. So the lock has to work here, and it does not need to delete
    anything to do it.

    Release tries `unlink()` first (identical to the old behaviour wherever it is
    allowed: no leftover file) and falls back to truncating in place. Truncation
    is a *write*, and a directory we just saved a workbook into always permits
    one. Exactly one of the two makes the lock free, never both — truncating and
    *then* unlinking would let a rival claim the emptied file and have its claim
    deleted out from under it.
    """

    def __init__(self, path: Path):
        self._lock_path = Path(path).with_suffix(Path(path).suffix + ".lock")
        # pid for a human reading the file; the uuid makes the claim confirmable
        # even against another writer in this same process (a conductor wave's
        # stages are threads as often as processes).
        self._token = f"{os.getpid()}:{uuid4().hex}"
        self._held = False

    def __enter__(self) -> "_workbook_lock":
        deadline = time.monotonic() + _LOCK_TIMEOUT_S
        while True:
            if self._claim():
                self._held = True
                return self
            if time.monotonic() >= deadline:
                raise DealWorkbookError(self._timeout_message()) from None
            time.sleep(_LOCK_POLL_S)

    # ── acquire ──────────────────────────────────────────────────────────────

    def _claim(self) -> bool:
        """One attempt. True when this instance now holds the lock."""
        if self._create_exclusive():
            return True
        state = self._read_state()
        if state is _UNREADABLE:
            return False
        if state is _VANISHED:
            return self._create_exclusive()  # released between our two calls
        if state and not self._is_stale():
            return False  # genuinely held by another writer
        return self._claim_free()  # empty, or a dead writer's token

    def _create_exclusive(self) -> bool:
        """The fast path: no lock file at all, so `O_EXCL` settles it atomically."""
        try:
            fd = os.open(str(self._lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            return False
        except OSError as exc:
            raise DealWorkbookError(
                f"cannot create the deal-workbook lock {self._lock_path}: {exc}. The "
                f"deal directory has to be writable — the workbook is saved into it."
            ) from exc
        try:
            os.write(fd, self._token.encode())
        finally:
            os.close(fd)
        return True

    def _claim_free(self) -> bool:
        """Write our token over a free (empty or stale) lock, and confirm it stuck.

        Two writers can both find the lock free and both write. Whichever token is
        still there after `_LOCK_CONFIRM_S` owns it; the loser reads someone
        else's token back and retries.
        """
        try:
            self._lock_path.write_text(self._token, encoding="utf-8")
        except OSError:
            return False
        time.sleep(_LOCK_CONFIRM_S)
        return self._read_state() == self._token

    def _read_state(self):
        """The holder's token, `""` when free, or a sentinel. Never raises."""
        try:
            return self._lock_path.read_text(encoding="utf-8", errors="replace").strip()
        except FileNotFoundError:
            return _VANISHED
        except OSError:
            return _UNREADABLE

    def _is_stale(self) -> bool:
        try:
            return (time.time() - self._lock_path.stat().st_mtime) > _LOCK_STALE_S
        except OSError:
            return True  # cannot stat it, so nobody is holding it

    def _timeout_message(self) -> str:
        state = self._read_state()
        holder = state if isinstance(state, str) and state else "unknown"
        return (
            f"timed out after {_LOCK_TIMEOUT_S:.0f}s waiting for "
            f"{self._lock_path.name} (held by {holder}). Another stage is writing "
            f"the deal workbook. If none is running, a killed writer left the lock "
            f"held: EMPTY the file to clear it — an empty lock file is a free lock, "
            f"and emptying it works even on a mount that refuses to remove files."
        )

    # ── release ──────────────────────────────────────────────────────────────

    def __exit__(self, *_exc) -> None:
        if not self._held:
            return
        self._held = False
        if self._unlink() or self._truncate():
            return
        print(
            f"deal_workbook: WARNING — could not release {self._lock_path}: neither "
            f"unlink nor truncation was permitted. Later writes will wait for it to "
            f"age out after {_LOCK_STALE_S:.0f}s. Empty the file to clear it now.",
            file=sys.stderr,
        )

    def _unlink(self) -> bool:
        try:
            self._lock_path.unlink()
        except FileNotFoundError:
            return True  # already gone — free either way
        except OSError:
            return False
        return True

    def _truncate(self) -> bool:
        try:
            self._lock_path.write_text("", encoding="utf-8")
        except OSError:
            return False
        return True


def _tab_index(tab_name: str, present: list[str]) -> int:
    """Where a newly created tab belongs, to keep `TAB_ORDER` left to right."""
    if tab_name not in TAB_ORDER:
        return len(present)
    wanted = TAB_ORDER.index(tab_name)
    for index, existing in enumerate(present):
        if existing in TAB_ORDER and TAB_ORDER.index(existing) > wanted:
            return index
    return len(present)


def _verify_tab_names(ws, tab_name: str, names: tuple[str, ...]) -> None:
    """Every named range the producer relies on must resolve on this tab."""
    if not names:
        return
    from template_layout import TemplateLayoutError, defined_name_ref

    missing = [name for name in names if defined_name_ref(ws, name) is None]
    if missing:
        raise TemplateLayoutError(
            f"deal workbook, tab {tab_name!r}: defined name(s) {', '.join(missing)} "
            f"do not resolve. The tab should carry them from "
            f"{DEAL_WORKBOOK_TEMPLATE}; re-run tools/build_deal_workbook_template.py "
            f"if a source template was re-saved."
        )


def write_tab(workbook_path: Path | str, tab_name: str, spec: TabSpec) -> None:
    """Apply `spec` to `tab_name` of the deal workbook, under an exclusive lock.

    The single mutation path. Serialized, so concurrent stages in one wave queue
    instead of clobbering each other; the whole load → mutate → save happens
    inside the lock, because openpyxl rewrites the entire file on save and a
    read that straddles another stage's save would drop that stage's tab.
    """
    path = Path(workbook_path)
    if not path.is_file():
        raise DealWorkbookError(
            f"{path} does not exist. The conductor creates the deal workbook at "
            f"deal-init via init_deal_workbook(); a stage must not create it."
        )

    with _workbook_lock(path):
        wb = load_workbook(path)
        try:
            if tab_name in wb.sheetnames:
                if spec.create:
                    del wb[tab_name]
                    ws = wb.create_sheet(tab_name, _tab_index(tab_name, wb.sheetnames))
                else:
                    ws = wb[tab_name]
            elif spec.create:
                ws = wb.create_sheet(tab_name, _tab_index(tab_name, wb.sheetnames))
            else:
                raise DealWorkbookError(
                    f"tab {tab_name!r} is not in {path.name} (has: "
                    f"{', '.join(wb.sheetnames)}). Template tabs are created at "
                    f"deal-init; pass TabSpec(create=True) for a tab authored "
                    f"from scratch."
                )

            _verify_tab_names(ws, tab_name, spec.verify_names)
            spec.write(wb, ws)
            wb.save(path)
        finally:
            wb.close()


def tab_names(workbook_path: Path | str) -> list[str]:
    """The workbook's tabs, in order. A cheap read for callers checking presence."""
    wb = load_workbook(Path(workbook_path), read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
