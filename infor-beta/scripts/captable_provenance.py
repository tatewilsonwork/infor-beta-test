"""Provenance for the cap table's DERIVED figures — read off its own formulas.

The captable skill hand-types the leaves: the debt tranches, the lease and cash
balances, the share-count rows, the FX rate and the share price. Every figure the
deck actually shows is computed from them by the template — and the block the
pitch deck pastes onto the overview slide (``infor_cap_picture_range``) is almost
entirely those computed cells:

    Basic Market Cap        =F16*F17
    Add: Debt               =F122*F7
    Less: Cash              =-F164*F7
    Net Debt                =F24+F25+F26+F27
    Enterprise Value        =F22+F28+F29+F30

None of them had a record. Enterprise Value is the largest number on the slide,
and a reviewer reading the pasted range could not get from it to a filing page —
not because the chain was missing but because it was only ever *in the workbook*,
as formulas nothing outside Excel followed.

So this module records it, and records it **from the workbook**: for every formula
cell in the pasted range it writes a derived :class:`provenance.FigureProvenance`
whose ``derived_from`` refs are the cells the formula reads, then follows those
refs and does the same again. The closure stops at the hand-typed leaves, which
are the skill's own records (Step 6), so the graph joins up: Enterprise Value →
Net Debt → Add: Debt → Total Debt → each debt tranche → the filing note it was
read from.

Two properties fall out of deriving it rather than writing it down.

- **No addresses.** The only address in this module is resolved from a defined
  name (`NAME_CAP_PICTURE_RANGE`), exactly as Phase C requires; the rest of the
  graph is whatever the template's formulas say it is. A re-saved template that
  moved the EV cascade needs no change here — which is the point, because the
  alternative is the table of hardcoded rows that Phase C deleted three times.
- **The labels are the template's.** A figure is named by the label in its row's
  first column, so the record reads like the tab: "Enterprise Value", not
  "captable!F31". A row whose label is itself a formula (the template concatenates
  two of them) has no readable label, so the record falls back to its cell
  reference — and, because the refs use the same function, the graph still joins.

What is NOT recorded: a formula cell whose value cannot be traced to anything —
nothing here evaluates arithmetic (Excel does the math, not us), so a derived
record carries the **formula** as its value. `deckcheck` cannot auto-match a
formula against a deck figure, and does not pretend to: these records exist so a
reviewer reading the rasterised range has a chain to walk.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl.utils import get_column_letter, range_boundaries

from deal_workbook import TAB_CAPTABLE
from provenance import FigureProvenance, FigureRef, ProvenanceLedger
from template_layout import (
    CAP_TABLE_PICTURE_RANGE,
    CAP_TABLE_TEMPLATE,
    NAME_CAP_PICTURE_RANGE,
    defined_name_ref,
    resolve_name_range,
)

#: A single cell or a range inside a formula, with optional `$` anchors and an
#: optional sheet qualifier. A qualified reference to another sheet is skipped —
#: this walks one tab, and a cross-tab reference is the other tab's record.
_REF_RE = re.compile(
    r"""
    (?P<sheet>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?
    \$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d{1,7})
    (?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d{1,7}))?
    """,
    re.VERBOSE,
)

#: Excel functions and the boolean/text literals inside a formula look like cell
#: references to a regex (`IF`, `SUM`, `LOG10`). A match immediately followed by
#: `(` is a function call, and a match preceded by a letter is part of a name.
_FUNCTION_AHEAD = re.compile(r"\s*\(")

#: How deep to follow the formula graph. The cap table's deepest real chain is
#: EV → Net Debt → Add: Debt → Total Debt → a tranche row: five. The cap is a
#: backstop against a circular reference an analyst introduced by hand, which
#: Excel would flag and openpyxl will not.
_MAX_DEPTH = 12


def _formula(cell) -> str | None:
    """The cell's formula, or None when it holds a literal (or an array formula).

    An `ArrayFormula` (the CapIQ estimate cells) is not walked: its text is a
    CapIQ call whose inputs are a ticker and a date, not other cells on this tab,
    and its value is un-evaluable here by design (`deckcheck.EXPECTED_ERROR_CONTEXTS`).
    """
    value = cell.value
    return value if isinstance(value, str) and value.startswith("=") else None


def _cells_in(formula: str) -> list[tuple[str, bool]]:
    """Every same-sheet cell a formula reads, as ``(coordinate, came_from_a_range)``.

    Order is left-to-right through the formula, deduplicated, so a bridge's refs
    read in the order the arithmetic does. The range flag matters downstream: a
    range spans a whole input section, most of whose rows a given deal leaves
    empty, while a direct reference is always deliberate.
    """
    out: dict[str, bool] = {}
    for match in _REF_RE.finditer(formula):
        if match.group("sheet"):
            continue  # another tab's figure — that tab's record, not this one's
        start, end = match.span()
        if _FUNCTION_AHEAD.match(formula, end):
            continue  # `IF(`, `SUM(` — a function name, not a reference
        if start and (formula[start - 1].isalnum() or formula[start - 1] in "_$'\""):
            continue  # inside a longer identifier or a quoted string
        c1, r1 = match.group("c1"), int(match.group("r1"))
        if match.group("c2") is None:
            out.setdefault(f"{c1}{r1}", False)
            continue
        left, top, right, bottom = range_boundaries(
            f"{c1}{r1}:{match.group('c2')}{match.group('r2')}"
        )
        for row in range(top, bottom + 1):
            for col in range(left, right + 1):
                out.setdefault(f"{get_column_letter(col)}{row}", True)
    return list(out.items())


def _is_numeric_format(number_format: str | None) -> bool:
    """Whether a cell's number format is a **figure**'s format.

    The discriminator is the workbook's own, not a guess about what a cell means:
    a figure carries digit placeholders (``#,##0.0``, ``"$"#,##0.00``, ``0.0\\x``,
    ``#,##0.0%``), while the cap table's non-figures carry ``General`` (the ticker
    echo, the two `CONCATENATE`d row captions) or a date format (``d-mmm-yy`` —
    the as-of date and the Treasury-Stock switch, which shares its format).
    Quoted literals, escaped characters and bracketed locale sections are notation
    and are stripped before looking for date letters.
    """
    text = number_format or ""
    text = re.sub(r'"[^"]*"', "", text)
    text = re.sub(r"\\.", "", text)
    text = re.sub(r"\[[^\]]*\]", "", text)
    positive = text.split(";")[0]  # the positive section decides the kind
    if re.search(r"[ymdhs]", positive, re.IGNORECASE):
        return False
    return any(ch in positive for ch in "0#?")


def _is_figure_cell(ws, coordinate: str) -> bool:
    """Whether a cell holds a figure: a numeric format, and something in it.

    Both halves earn their keep. Without the format test the walk records the
    ticker echo and the two formula-built row captions as though they were
    figures; without the emptiness test it records every blank row of every input
    section, because the template pre-fills their ITM formulas.
    """
    cell = ws[coordinate]
    if not _is_numeric_format(cell.number_format):
        return False
    value = cell.value
    if value is None:
        return False
    if isinstance(value, str):
        return value.startswith("=")
    return not isinstance(value, bool)


def _row_is_in_use(ws, coordinate: str, *, label_column: str) -> bool:
    """Whether a row inside an expanded range is one this deal actually filled.

    Every populated data row carries a label — the facility, the lease type, the
    cash line, the share description — because the skill writes one (SKILL.md
    Step 6, and the ``Sub-Event:`` prefix rule depends on it). An unfilled row
    carries none, and its pre-filled ITM formula computes to zero. So the label is
    the "in use" signal, and it is the row's own, not a count kept anywhere.
    """
    return _label_for(ws, coordinate, label_column=label_column) is not None


def _label_for(ws, coordinate: str, *, label_column: str) -> str | None:
    """The row label beside a cell, or None when it is missing or a formula.

    A formula label (the template builds two by `CONCATENATE`/`IF`) renders to
    text only in Excel, and guessing what it renders to is exactly the kind of
    estimate this repo does not make — so the caller falls back to the cell
    reference, which is unambiguous and joins the same way.
    """
    row = int(re.sub(r"[^0-9]", "", coordinate))
    value = ws[f"{label_column}{row}"].value
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text or text.startswith("="):
        return None
    return text


def _ref(coordinate: str) -> str:
    """`"captable!F31"` — a record's location for a cell on this tab."""
    return f"{TAB_CAPTABLE}!{coordinate}"


def cap_table_figure_name(ws, coordinate: str, *, label_column: str = "B") -> str:
    """What to call the figure in a cell: its row label, else its cell reference.

    Used for both the record's ``figure`` and (indirectly) its refs, so a name and
    the thing that points at it are produced by one function and cannot disagree.
    When several cells in one row hold figures — the template's Section II total
    row totals ITM shares in one column and proceeds in another, both labelled
    "Total" — the column letter disambiguates.
    """
    label = _label_for(ws, coordinate, label_column=label_column)
    if label is None:
        return _ref(coordinate)
    column = re.sub(r"[0-9]", "", coordinate)
    row = int(re.sub(r"[^0-9]", "", coordinate))
    siblings = [
        get_column_letter(col)
        for col in range(1, ws.max_column + 1)
        if _formula(ws.cell(row=row, column=col)) is not None
    ]
    return f"{label} ({column})" if len(siblings) > 1 else label


def _picture_range(ws) -> str:
    """The block the deck pastes, resolved by name (never by address).

    Falls back to the shipped address only for a tab that lost the name, which
    `verify_cap_table_before_write` would already have refused to write to —
    recording provenance is not the place to start halting a finished run.
    """
    if defined_name_ref(ws, NAME_CAP_PICTURE_RANGE) is None:
        return CAP_TABLE_PICTURE_RANGE
    return resolve_name_range(ws, NAME_CAP_PICTURE_RANGE, template=CAP_TABLE_TEMPLATE)


def record_cap_table_derived_figures(
    ws,
    ledger: ProvenanceLedger,
    *,
    label_column: str | None = None,
) -> list[FigureProvenance]:
    """Record every derived figure the deck's cap-table block depends on.

    Walks the formula graph from the pasted picture range outwards, one record per
    formula cell, each naming the cells it was computed from. Returns the records
    in the order they were added (deepest components first, so the ledger reads
    bottom-up like the arithmetic does).

    Call it **after** writing every hand-typed row (captable SKILL.md Steps 3b–6b),
    because the leaves' records are the skill's own and the refs land on them. A
    figure already recorded — F7 and F16, whose records carry the URL they were
    read from — is not recorded twice.
    """
    left, top, right, bottom = range_boundaries(_picture_range(ws))
    # The label column is the pasted block's left edge — the template puts every
    # row's caption there, and so does the skill for the rows it adds.
    label_col = label_column or get_column_letter(left)

    recorded: list[FigureProvenance] = []
    visiting: set[str] = set()
    done = {
        entry.location.split("!")[-1].replace("$", "").upper()
        for entry in ledger.figures
        if entry.location
    }

    def visit(coordinate: str, depth: int) -> None:
        if coordinate in done or coordinate in visiting or depth > _MAX_DEPTH:
            return
        if not _is_figure_cell(ws, coordinate):
            return  # a caption, the ticker echo, the as-of date, an empty row
        formula = _formula(ws[coordinate])
        if formula is None:
            return  # a hand-typed leaf — the skill's own record covers it
        visiting.add(coordinate)
        components = [
            c
            for c, from_range in _cells_in(formula)
            if c != coordinate
            and _is_figure_cell(ws, c)
            and (not from_range or _row_is_in_use(ws, c, label_column=label_col))
        ]
        for component in components:
            visit(component, depth + 1)
        visiting.discard(coordinate)
        if coordinate in done:
            return
        done.add(coordinate)
        recorded.append(
            ledger.record(
                cap_table_figure_name(ws, coordinate, label_column=label_col),
                value=formula,
                location=_ref(coordinate),
                derivation=f"cap-table formula {formula}",
                derived_from=[
                    FigureRef(
                        figure=cap_table_figure_name(ws, c, label_column=label_col),
                        location=_ref(c),
                    )
                    for c in components
                ]
                or None,
            )
        )

    for row in range(top, bottom + 1):
        for col in range(left, right + 1):
            visit(f"{get_column_letter(col)}{row}", 0)
    return recorded


def record_cap_table_derived_figures_in_workbook(
    deal_workbook: Path | str,
    ledger: ProvenanceLedger,
) -> list[FigureProvenance]:
    """Same, for a caller holding the deal workbook path rather than a sheet.

    Read-only: it opens its own handle, records, and closes. Never used to write —
    every cap-table mutation goes through `deal_workbook.write_tab`, which holds
    the lock for the whole load → mutate → save cycle.
    """
    from openpyxl import load_workbook

    wb = load_workbook(Path(deal_workbook), data_only=False)
    try:
        return record_cap_table_derived_figures(wb[TAB_CAPTABLE], ledger)
    finally:
        wb.close()
