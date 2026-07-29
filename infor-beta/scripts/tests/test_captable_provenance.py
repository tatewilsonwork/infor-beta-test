"""The cap table's derived figures — read off its own formulas.

The gap this closes was reported from a real pitch run: Enterprise Value
``C$15,796.0``, the largest number on the overview slide, had **no provenance
record anywhere**, and neither did ``captable!F17``. Not because the chain was
missing — the template computes EV from market cap, net debt, preferreds and NCI,
each of those from a section total, each of those from a hand-typed row — but
because the chain only ever existed as formulas inside the workbook.

Two properties are asserted here, and the second is why this is derived from the
workbook rather than written down:

1. **The graph joins up.** A record exists for every formula cell the pasted
   picture range depends on, its ``derived_from`` refs resolve, and the walk from
   Enterprise Value reaches the filing pages the leaves were read from.
2. **Nothing is addressed by position.** The scan is seeded from
   ``infor_cap_picture_range`` and follows formulas, so a template whose rows moved
   still records the same figures — the failure mode Phase C's deleted
   ``_KEEP_LIBRARY_INDICES`` demonstrated three times.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from captable_provenance import (
    _is_numeric_format,
    cap_table_figure_name,
    record_cap_table_derived_figures,
)
from tests.conftest import stamp_defined_names
from deal_workbook import TAB_CAPTABLE
from provenance import FigureRef, FigureSource, ProvenanceLedger
from template_layout import NAME_CAP_PICTURE_RANGE

_CURRENCY = '"$"#,##0.0_);\\("$"#,##0.0\\);"-- "'
_SHARES = '#,##0.0_);\\(#,##0.0\\);"-- "'
_DATE = "d-mmm-yy"
_TEXT = "General"


def _cap_tab():
    """A miniature cap table with the shipped template's shape, not its addresses.

    Rows are deliberately *not* the shipped ones: the module must find the graph
    through the picture-range name and the formulas, so a fixture at different
    addresses proves it does.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = TAB_CAPTABLE

    def put(coord, value, fmt=_CURRENCY):
        ws[coord] = value
        ws[coord].number_format = fmt

    # Header inputs (hand-typed; the skill records these itself).
    ws["B2"] = "Share Price"
    put("C2", 21.2)
    ws["B3"] = "FX Rate"
    put("C3", 1.37, _SHARES)
    ws["B4"] = "Date:"
    put("C4", "=TODAY()-1", _DATE)          # a date, not a figure
    ws["B5"] = "Treasury Method"
    put("C5", "Treasury Stock", _TEXT)      # a switch, not a figure

    # The summary block the deck pastes (C6:C10 hold the cascade).
    ws["B6"] = "Basic Shares Outstanding"
    put("C6", "=C21", _SHARES)
    ws["B7"] = "Basic Market Cap"
    put("C7", "=C2*C6")
    ws["B8"] = "Add: Debt"
    put("C8", "=C15*C3")
    ws["B9"] = "Less: Cash"
    put("C9", "=-C18*C3")
    ws["B10"] = "Enterprise Value"
    put("C10", "=C7+C8+C9")

    # Section: debt (two filled rows, one blank).
    ws["B13"] = "Senior Notes due 2034"
    put("C13", 500.0)
    ws["B14"] = "Revolving Credit Facility (undrawn)"
    put("C14", 0.0)
    put("C15", "=SUM(C13:C14)")
    ws["B15"] = "Total Debt"

    # Section: cash.
    ws["B17"] = "Cash and equivalents"
    put("C17", 120.0)
    ws["B18"] = "Total Cash"
    put("C18", "=SUM(C17:C17)")

    # Section: basic shares.
    ws["B20"] = "Common shares"
    put("C20", 22.8, _SHARES)
    ws["B21"] = "Total Basic Shares Outstanding"
    put("C21", "=SUM(C20:C20)", _SHARES)

    stamp_defined_names(ws, {NAME_CAP_PICTURE_RANGE: "B6:C10"})
    return ws


def _with_ltm_link(ws, formula: str):
    """Add the LTM valuation cell to the pasted block — `D47` on the real template.

    Inside the block because the deck shows it: `EV / Revenue` divides by it, and the
    whole point of following the link is that a figure on the overview slide reaches
    the bridge total that computed it.
    """
    ws["B11"] = "LTM Revenue (valuation)"
    ws["C11"] = formula
    ws["C11"].number_format = _CURRENCY
    stamp_defined_names(ws, {NAME_CAP_PICTURE_RANGE: "B6:C11"})
    return ws


def _leaf_ledger() -> ProvenanceLedger:
    """What the captable skill records itself: the hand-typed leaves."""
    ledger = ProvenanceLedger(stage="captable")
    ledger.record("Share price", value=21.2, location=f"{TAB_CAPTABLE}!C2",
                  sources=FigureSource(url="https://example.com/quote", retrieved="2026-07-29"))
    ledger.record("FX rate", value=1.37, location=f"{TAB_CAPTABLE}!C3",
                  sources=FigureSource(url="https://example.com/fx", retrieved="2026-07-29"))
    ledger.record("Senior Notes due 2034", value=500.0, location=f"{TAB_CAPTABLE}!C13",
                  sources=FigureSource(filing="FY2025 10-K", statement="Note 12: Long-Term Debt",
                                       page=99))
    ledger.record("Revolving Credit Facility (undrawn)", value=0.0,
                  location=f"{TAB_CAPTABLE}!C14",
                  sources=FigureSource(filing="FY2025 10-K", statement="Note 12: Long-Term Debt",
                                       page=99))
    ledger.record("Cash and equivalents", value=120.0, location=f"{TAB_CAPTABLE}!C17",
                  sources=FigureSource(filing="FY2025 10-K", statement="Consolidated Balance Sheets",
                                       page=88))
    ledger.record("Common shares", value=22.8, location=f"{TAB_CAPTABLE}!C20",
                  sources=FigureSource(filing="FY2025 10-K", statement="Capital stock note",
                                       page=95))
    return ledger


def _recorded():
    ws = _cap_tab()
    ledger = _leaf_ledger()
    record_cap_table_derived_figures(ws, ledger)
    return ledger


def test_the_ev_bridge_total_now_has_a_record():
    by_location = {f.location: f for f in _recorded().figures}
    ev = by_location[f"{TAB_CAPTABLE}!C10"]
    assert ev.figure == "Enterprise Value"
    assert ev.derived and ev.derivation_line
    assert [r.render() for r in ev.derived_from] == [
        f"{TAB_CAPTABLE}!C7", f"{TAB_CAPTABLE}!C8", f"{TAB_CAPTABLE}!C9"
    ]


def test_basic_shares_outstanding_traces_to_the_capital_stock_note():
    # `captable!F17` on the shipped template: a one-cell formula (`=F186`) whose
    # own component is the section total of the rows the skill typed in.
    ledger = _recorded()
    shares = next(f for f in ledger.figures if f.location == f"{TAB_CAPTABLE}!C6")
    trace = ledger.trace(shares)
    assert trace.resolved
    assert [c.figure for c in trace.components] == ["Total Basic Shares Outstanding"]
    assert [s.render() for s in trace.root_sources] == [
        "FY2025 10-K, Capital stock note, p. 95"
    ]


def test_enterprise_value_walks_all_the_way_to_the_filings():
    ledger = _recorded()
    ev = next(f for f in ledger.figures if f.figure == "Enterprise Value")
    trace = ledger.trace(ev)

    assert trace.resolved, f"unresolved refs: {[r.render() for r in trace.unresolved]}"
    assert {s.render() for s in trace.root_sources} == {
        "https://example.com/quote — retrieved 2026-07-29",
        "https://example.com/fx — retrieved 2026-07-29",
        "FY2025 10-K, Note 12: Long-Term Debt, p. 99",
        "FY2025 10-K, Consolidated Balance Sheets, p. 88",
        "FY2025 10-K, Capital stock note, p. 95",
    }


def test_a_hand_typed_leaf_is_not_re_recorded():
    # The skill's own records carry the URL / filing; a second record for the same
    # cell would be a duplicate claiming to be derived.
    locations = [f.location for f in _recorded().figures]
    assert len(locations) == len(set(locations))
    assert locations.count(f"{TAB_CAPTABLE}!C2") == 1


def test_a_switch_a_date_and_a_caption_are_not_figures():
    recorded = {f.location for f in _recorded().figures}
    assert f"{TAB_CAPTABLE}!C4" not in recorded, "the =TODAY()-1 as-of date is not a figure"
    assert f"{TAB_CAPTABLE}!C5" not in recorded, "the treasury-method switch is not a figure"
    assert not any(loc.endswith("!B10") for loc in recorded), "a row caption is not a figure"


def test_an_unfilled_input_row_is_not_recorded():
    # `SUM` spans a whole input section; a deal fills a few rows of it. The template
    # pre-fills the rest with formulas, and recording them would bury the real
    # figures — an agenda nobody reads is the same as no review.
    ws = _cap_tab()
    ws["C25"] = "=SUM(C13:C14)"     # a formula row with no label
    ws["C25"].number_format = _CURRENCY
    ledger = _leaf_ledger()
    record_cap_table_derived_figures(ws, ledger)
    assert f"{TAB_CAPTABLE}!C25" not in {f.location for f in ledger.figures}


def test_a_row_the_skill_filled_inside_a_sum_range_is_recorded():
    ws = _cap_tab()
    ws["B14"] = "Sub-Event: Feb 2026 term loan"   # a labelled row is a filled row
    ledger = _leaf_ledger()
    record_cap_table_derived_figures(ws, ledger)
    total = next(f for f in ledger.figures if f.figure == "Total Debt")
    assert [r.render() for r in total.derived_from] == [
        f"{TAB_CAPTABLE}!C13", f"{TAB_CAPTABLE}!C14"
    ]


def test_the_scan_follows_the_picture_range_name_not_an_address():
    # Move the block; the same figures are recorded, because the name moved with it.
    ws = _cap_tab()
    stamp_defined_names(ws, {NAME_CAP_PICTURE_RANGE: "B6:C7"})
    ledger = _leaf_ledger()
    record_cap_table_derived_figures(ws, ledger)
    names = {f.figure for f in ledger.figures}
    assert {"Basic Market Cap", "Total Basic Shares Outstanding"} <= names
    assert "Enterprise Value" not in names, "C10 is outside the narrowed block"


def test_a_row_whose_label_is_itself_a_formula_falls_back_to_its_cell():
    ws = _cap_tab()
    ws["B7"] = '=CONCATENATE("Basic ","Market Cap")'
    assert cap_table_figure_name(ws, "C7") == f"{TAB_CAPTABLE}!C7"


def test_two_figures_in_one_row_are_named_apart():
    ws = _cap_tab()
    ws["D15"] = "=C15*2"
    ws["D15"].number_format = _CURRENCY
    assert cap_table_figure_name(ws, "C15") == "Total Debt (C)"
    assert cap_table_figure_name(ws, "D15") == "Total Debt (D)"


@pytest.mark.parametrize(
    "number_format, is_figure",
    [
        ('#,##0.0_);\\(#,##0.0\\);"-- "', True),
        ('"$"#,##0.00_);\\("$"#,##0.00\\)', True),
        ("#,##0.0%", True),
        ("0.0\\x_);\\(0.0\\x\\)", True),
        ("General", False),
        ("d-mmm-yy", False),
        ("[$-1009]d\\-mmm\\-yy;@", False),
        ("@", False),
    ],
)
def test_the_format_decides_what_counts_as_a_figure(number_format, is_figure):
    # The discriminator is the workbook's own formatting, not a guess about meaning.
    assert _is_numeric_format(number_format) is is_figure


def test_recording_twice_is_idempotent():
    ws = _cap_tab()
    ledger = _leaf_ledger()
    first = record_cap_table_derived_figures(ws, ledger)
    second = record_cap_table_derived_figures(ws, ledger)
    assert first and second == []
    assert len({f.location for f in ledger.figures}) == len(ledger.figures)


# ─── The LTM valuation cells link ANOTHER tab ─────────────────────────────────
#
# `D47` / `D48` are label-keyed links into the `ltm-metrics` bridge totals
# (`ltm_metrics.ltm_total_link`), which is the only copy of each figure. The link
# names its upstream figure in the formula, so the chain continues into the other
# stage's fragment instead of stopping at "a cap-table formula".


def test_a_cross_tab_link_names_the_bridge_total_it_depends_on():
    from ltm_metrics import ltm_total_formula

    ws = _with_ltm_link(_cap_tab(), ltm_total_formula("LTM Revenue", times="C3"))
    ledger = _leaf_ledger()
    record_cap_table_derived_figures(ws, ledger)
    linked = next(f for f in ledger.figures if f.location == f"{TAB_CAPTABLE}!C11")
    assert [r.render() for r in linked.derived_from] == [f"{TAB_CAPTABLE}!C3", "LTM Revenue"]


def test_the_link_resolves_against_the_ltm_metrics_fragment(tmp_path):
    # The whole point: EV / Revenue on the overview slide reaches the bridge total,
    # and the bridge total reaches the filing — across two stages' fragments.
    from ltm_metrics import ltm_total_formula
    from provenance import read_run_provenance

    ltm = ProvenanceLedger(stage="ltm-metrics")
    ltm.record("LTM Revenue — FY2025 revenue", value=5168.4, units="US$MM",
               location="ltm-metrics!B21",
               sources=FigureSource(filing="FY2025 10-K",
                                    statement="Consolidated Statements of Operations", page=142))
    ltm.record("LTM Revenue", value=5196.1, units="US$MM", location="ltm-metrics!B23",
               derivation="FY2025 + Q1 2026 YTD − Q1 2025 YTD",
               derived_from=[FigureRef(location="ltm-metrics!B21")])
    ltm.write(tmp_path / "stages" / "ltm-metrics")

    ws = _with_ltm_link(_cap_tab(), ltm_total_formula("LTM Revenue", times="C3"))
    cap = _leaf_ledger()
    record_cap_table_derived_figures(ws, cap)
    cap.write(tmp_path / "stages" / "captable")

    merged = read_run_provenance(tmp_path)
    linked = next(f for f in merged.figures if f.location == f"{TAB_CAPTABLE}!C11")
    trace = merged.trace(linked)
    assert trace.resolved, [r.render() for r in trace.unresolved]
    assert "LTM Revenue" in [c.figure for c in trace.components]
    assert any("p. 142" in s.render() for s in trace.root_sources)


def test_the_capiq_fallback_formula_names_no_upstream_figure():
    # Case B: no `ltm-metrics` tab, so the cell holds a CapIQ call whose inputs are
    # a ticker and a date — not another tab's figure.
    ws = _with_ltm_link(
        _cap_tab(),
        '=_xll.SNL.Clients.Office.Excel.Functions.SPG($F$3,"IQ_REV",D$33,$F$6,'
        '"Options:Mag=Millions,NA=NA,Curr="&$F$5)',
    )
    ledger = _leaf_ledger()
    record_cap_table_derived_figures(ws, ledger)
    linked = next(f for f in ledger.figures if f.location == f"{TAB_CAPTABLE}!C11")
    assert all(r.figure != "LTM Revenue" for r in linked.derived_from)
