"""Unit tests for the deal workbook's `ltm-metrics` tab."""

import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from deal_workbook import (
    TAB_CAPTABLE,
    TAB_LTM_METRICS,
    TabSpec,
    init_deal_workbook,
    write_tab,
)
from ltm_metrics import (
    LTM_EBITDA_RESULT_LABELS,
    LTM_REVENUE_RESULT_LABEL,
    Bridge,
    BridgeComponent,
    RevenueSegment,
    build_ltm_metrics_workbook,
    ltm_total_link,
    result_row_labels,
)
from provenance import FigureSource, ProvenanceError, ProvenanceLedger
from template_layout import (
    CAP_TABLE_NAMED_RANGES,
    NAME_FX_RATE,
    NAME_LTM_EBITDA_VALUATION,
    NAME_LTM_REVENUE_VALUATION,
    resolve_name_cell,
)


def _build(tmp_path: Path, **overrides) -> Path:
    kwargs = dict(
        company_name="SampleCo",
        period_label="LTM ended March 31, 2026",
        currency="US$MM",
        segmentation_basis="Service line",
        segments=[
            RevenueSegment("Cloud Services & Subscriptions", 1932.0),
            RevenueSegment("Customer Support", 1480.0),
            RevenueSegment("License", 360.0),
            RevenueSegment("Professional Service & Other", 290.0),
        ],
        revenue_bridge=[
            BridgeComponent("FY2025 Revenue", 5400.0),
            BridgeComponent("Q3 2026 YTD Revenue", 3050.0),
            BridgeComponent("Q3 2025 YTD Revenue", 2388.0, subtract=True),
        ],
        ebitda_bridge=[
            BridgeComponent("FY2025 Adj. EBITDA", 1820.0),
            BridgeComponent("Q3 2026 YTD Adj. EBITDA", 1040.0),
            BridgeComponent("Q3 2025 YTD Adj. EBITDA", 815.0, subtract=True),
        ],
        deal_workbook=init_deal_workbook(
            deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
        ),
    )
    kwargs.update(overrides)
    return build_ltm_metrics_workbook(**kwargs)


def _ws(path: Path):
    """The deal workbook's `ltm-metrics` tab — formulas, not computed values."""
    return load_workbook(path)[TAB_LTM_METRICS]


def test_build_ltm_metrics_writes_the_tab_into_the_deal_workbook(tmp_path: Path):
    path = _build(tmp_path)
    assert path.exists()
    assert path.name == "pitch-Project Test.xlsx", "the deal owns one workbook"
    assert TAB_LTM_METRICS in load_workbook(path).sheetnames
    assert _ws(path).title == TAB_LTM_METRICS


def test_revenue_overview_uses_formulas_for_totals_and_percentages(tmp_path: Path):
    path = _build(tmp_path)
    ws = _ws(path)  # formulas, not computed values

    # Section row 6, header row 7, data rows 8-11, total row 12.
    assert ws["A6"].value == "LTM Revenue Overview"
    assert ws["A7"].value == "Segment"
    assert ws["B8"].value == 1932.0
    # % of total references the total cell so Excel recomputes.
    assert ws["C8"].value == "=B8/B12"
    assert ws["B12"].value == "=SUM(B8:B11)"
    assert ws["C12"].value == "=SUM(C8:C11)"
    assert ws["A12"].value == "Total"


def test_revenue_bridge_sums_fy_plus_ytd_minus_prior(tmp_path: Path):
    path = _build(tmp_path)
    ws = _ws(path)

    # Overview total is row 12; row 13 is a blank spacer, bridge starts at 14.
    assert ws["A14"].value == "LTM Revenue Bridge"
    assert ws["A15"].value == "Component"
    assert ws["A16"].value == "(+) FY2025 Revenue"
    assert ws["B16"].value == 5400.0
    assert ws["A18"].value == "(−) Q3 2025 YTD Revenue"
    # LTM = FY + current YTD − prior YTD, referencing the component cells.
    assert ws["A19"].value == "(=) LTM Revenue"
    assert ws["B19"].value == "=B16+B17-B18"


def test_ebitda_bridge_only_and_label(tmp_path: Path):
    path = _build(tmp_path)
    ws = _ws(path)

    # Revenue result is row 19; row 20 is a blank spacer, EBITDA starts at 21.
    assert ws["A21"].value == "LTM Adj. EBITDA Bridge"
    assert ws["A26"].value == "(=) LTM Adj. EBITDA"
    assert ws["B26"].value == "=B23+B24-B25"


def test_ebitda_label_falls_back_to_unadjusted(tmp_path: Path):
    path = _build(tmp_path, ebitda_label="LTM EBITDA")
    ws = _ws(path)
    assert ws["A21"].value == "LTM EBITDA Bridge"
    assert ws["A26"].value == "(=) LTM EBITDA"


def test_segments_and_bridges_accept_plain_tuples(tmp_path: Path):
    path = _build(
        tmp_path,
        segmentation_basis="Geography",
        segments=[("Americas", 2500.0), ("EMEA", 1100.0), ("Asia-Pacific", 462.0)],
        revenue_bridge=[
            ("FY2025 Revenue", 3500.0),
            ("Q3 2026 YTD Revenue", 2700.0),
            ("Q3 2025 YTD Revenue", 2138.0, True),
        ],
    )
    ws = _ws(path)
    assert ws["A3"].value == "Revenue segmentation: Geography"
    # 3 segments -> data rows 8-10, total row 11.
    assert ws["B11"].value == "=SUM(B8:B10)"
    # Spacer row 12, then bridge: section 13, header 14, data 15-17, result 18.
    assert ws["A15"].value == "(+) FY2025 Revenue"
    assert ws["B18"].value == "=B15+B16-B17"


def test_bridges_optional(tmp_path: Path):
    path = _build(tmp_path, revenue_bridge=None, ebitda_bridge=None)
    ws = _ws(path)
    # Only the overview block is present; no bridge section rows.
    assert ws["A6"].value == "LTM Revenue Overview"
    assert ws["A14"].value is None


def test_empty_segments_raise(tmp_path: Path):
    with pytest.raises(ValueError):
        _build(tmp_path, segments=[])


def test_extra_bridges_append_after_ebitda(tmp_path: Path):
    # EBITDA bridge result is row 26; row 27 is a blank spacer, the extra bridge
    # starts at 28 (section 28, header 29, data 30-32, result 33).
    path = _build(
        tmp_path,
        extra_bridges=[
            Bridge(
                "LTM Net Income Bridge",
                "LTM Net Income",
                [
                    BridgeComponent("FY2025 Net Income", 700.0),
                    BridgeComponent("Q3 2026 YTD Net Income", 420.0),
                    BridgeComponent("Q3 2025 YTD Net Income", 330.0, subtract=True),
                ],
            ),
        ],
    )
    ws = _ws(path)
    assert ws["A28"].value == "LTM Net Income Bridge"
    assert ws["A33"].value == "(=) LTM Net Income"
    assert ws["B33"].value == "=B30+B31-B32"


def test_extra_bridges_accept_dicts(tmp_path: Path):
    path = _build(
        tmp_path,
        extra_bridges=[
            {
                "section_title": "LTM Gross Profit Bridge",
                "result_label": "LTM Gross Profit",
                "components": [
                    ("FY2025 GP", 1000.0),
                    ("Q3 2026 YTD GP", 600.0),
                    ("Q3 2025 YTD GP", 470.0, True),
                ],
            },
        ],
    )
    ws = _ws(path)
    assert ws["A28"].value == "LTM Gross Profit Bridge"
    assert ws["A33"].value == "(=) LTM Gross Profit"
    assert ws["B33"].value == "=B30+B31-B32"


def test_extra_bridges_default_none_leaves_workbook_unchanged(tmp_path: Path):
    # Without extra_bridges the workbook ends at the EBITDA bridge (result row 26).
    ws = _ws(_build(tmp_path))
    assert ws["A28"].value is None


# ─── Provenance records + the comments rendered from them (Phase G) ──────────
#
# v0.5.34 wrote the citation string straight onto the cell and that string WAS the
# record. Phase G inverts it: a `FigureSource` goes in, a record comes out into the
# run's ledger, and the comment is a view of the record. These tests assert both
# halves — the comment text is unchanged, and a record now exists for it.

_SEG_SRC = FigureSource(filing="Q3 2026 10-Q", statement="revenue disaggregation note")
_FY_SRC = FigureSource(filing="FY2025 10-K", statement="income statement")
_PRIOR_SRC = FigureSource(filing="Q3 2026 10-Q", statement="comparative prior-year period")


def test_segment_source_written_as_cell_comment(tmp_path: Path):
    path = _build(
        tmp_path,
        segments=[
            RevenueSegment("Segment A", 999.0, source=_SEG_SRC),
            RevenueSegment("Segment B", 888.0),  # no source -> no comment
        ],
    )
    ws = _ws(path)
    assert ws["B8"].comment is not None
    assert ws["B8"].comment.text == "Source: Q3 2026 10-Q, revenue disaggregation note"
    assert ws["B9"].comment is None


def test_bridge_component_source_written_as_cell_comment(tmp_path: Path):
    path = _build(
        tmp_path,
        revenue_bridge=[
            BridgeComponent("FY2025 Revenue", 9999.0, source=_FY_SRC),
            BridgeComponent("Q3 2026 YTD Revenue", 999.0),
            BridgeComponent("Q3 2025 YTD Revenue", 888.0, subtract=True, source=_PRIOR_SRC),
        ],
    )
    ws = _ws(path)
    # Default 4-segment fixture geometry: bridge section 14, header 15, data 16-18.
    assert ws["B16"].comment is not None
    assert ws["B16"].comment.text == "Source: FY2025 10-K, income statement"
    assert ws["B17"].comment is None
    assert ws["B18"].comment.text == "Source: Q3 2026 10-Q, comparative prior-year period"


def test_tuple_forms_accept_trailing_source(tmp_path: Path):
    path = _build(
        tmp_path,
        segments=[("Americas", 2500.0, _SEG_SRC), ("EMEA", 1100.0)],
        revenue_bridge=[
            ("FY2025 Revenue", 3500.0, False, _FY_SRC),
            ("Q3 2026 YTD Revenue", 2700.0),
            ("Q3 2025 YTD Revenue", 2138.0, True),
        ],
    )
    ws = _ws(path)
    assert ws["B8"].comment.text == "Source: Q3 2026 10-Q, revenue disaggregation note"
    # 2 segments -> total row 10, spacer 11, bridge section 12, header 13, data 14-16.
    assert ws["B14"].comment.text == "Source: FY2025 10-K, income statement"
    assert ws["B15"].comment is None


def test_page_number_reaches_the_comment(tmp_path: Path):
    # The field the old string convention could not enforce: a page.
    path = _build(
        tmp_path,
        segments=[("Americas", 2500.0, FigureSource(filing="FY2025 10-K",
                                                    statement="Note 23: Segments", page=112))],
    )
    assert _ws(path)["B8"].comment.text == (
        "Source: FY2025 10-K, Note 23: Segments, p. 112"
    )


def test_a_citation_string_is_rejected(tmp_path: Path):
    # A string used to be the record. Accepting it now would build a record whose
    # whole citation sits in `filing` with no statement or page — provenance that
    # reads fine and cannot be followed.
    with pytest.raises(ProvenanceError, match="FigureSource"):
        _build(tmp_path, segments=[("Americas", 2500.0, "Q3 10-Q, segment note")])
    with pytest.raises(ProvenanceError, match="FigureSource"):
        _build(tmp_path, revenue_bridge=[("FY2025 Revenue", 3500.0, False, "FY2025 10-K")])


def test_every_extracted_figure_lands_in_the_ledger(tmp_path: Path):
    ledger = ProvenanceLedger(stage="ltm-metrics")
    _build(
        tmp_path,
        segments=[
            RevenueSegment("Segment A", 999.0, source=_SEG_SRC),
            RevenueSegment("Segment B", 888.0, source=_SEG_SRC),
        ],
        revenue_bridge=[
            BridgeComponent("FY2025 Revenue", 9999.0, source=_FY_SRC),
            BridgeComponent("Q3 2026 YTD Revenue", 999.0, source=_FY_SRC),
            BridgeComponent("Q3 2025 YTD Revenue", 888.0, subtract=True, source=_PRIOR_SRC),
        ],
        ebitda_bridge=None,
        provenance=ledger,
    )
    by_figure = {f.figure: f for f in ledger.figures}
    assert by_figure["LTM Revenue — Segment A"].value == 999.0
    assert by_figure["LTM Revenue — Segment A"].location == "ltm-metrics!B8"
    assert by_figure["LTM Revenue — Segment A"].units == "US$MM"
    assert by_figure["LTM Revenue — Segment A"].sources == (_SEG_SRC,)
    assert by_figure["LTM Revenue — FY2025 Revenue"].sources == (_FY_SRC,)
    assert all(f.stage == "ltm-metrics" for f in ledger.figures)


def test_bridge_total_is_recorded_as_a_derived_figure(tmp_path: Path):
    # The join point `deckcheck` walks from a deck tile back to a filing: the LTM
    # total has no source of its own, only the components that do.
    ledger = ProvenanceLedger(stage="ltm-metrics")
    _build(
        tmp_path,
        revenue_bridge=[
            BridgeComponent("FY2025 Revenue", 5400.0, source=_FY_SRC),
            BridgeComponent("Q3 2026 YTD Revenue", 3050.0, source=_FY_SRC),
            BridgeComponent("Q3 2025 YTD Revenue", 2388.0, subtract=True, source=_PRIOR_SRC),
        ],
        provenance=ledger,
    )
    total = next(f for f in ledger.figures if f.figure == "LTM Revenue")
    assert total.sources == ()
    assert total.value == pytest.approx(5400.0 + 3050.0 - 2388.0)
    assert total.derivation == (
        "FY2025 Revenue + Q3 2026 YTD Revenue − Q3 2025 YTD Revenue"
    )
    assert total.location == "ltm-metrics!B19"


def test_ledger_survives_a_round_trip_through_its_fragment(tmp_path: Path):
    ledger = ProvenanceLedger(stage="ltm-metrics")
    _build(
        tmp_path,
        segments=[RevenueSegment("Segment A", 999.0,
                                 source=FigureSource(filing="FY2025 10-K",
                                                     statement="Note 23", page="F-12"))],
        provenance=ledger,
    )
    fragment = ledger.write(tmp_path / "stages" / "ltm-metrics")
    reloaded = ProvenanceLedger.read(fragment)
    assert reloaded.stage == "ltm-metrics"
    assert [f.to_dict() for f in reloaded.figures] == [f.to_dict() for f in ledger.figures]


# ─── The LTM total is linked, never copied (B11) ──────────────────────────────


def _fresh_workbook(tmp_path: Path) -> Path:
    """A deal workbook with its template tabs but no `ltm-metrics` tab yet."""
    return init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )


def test_result_row_labels_lists_every_bridge_total(tmp_path: Path):
    labels = result_row_labels(_ws(_build(tmp_path)))
    assert labels == ["LTM Revenue", "LTM Adj. EBITDA"]


def test_ltm_total_link_is_keyed_on_the_result_row_label(tmp_path: Path):
    wb = load_workbook(_build(tmp_path))
    link = ltm_total_link(wb, LTM_REVENUE_RESULT_LABEL)
    assert link == (
        "=INDEX('ltm-metrics'!$B:$B, "
        'MATCH("(=) LTM Revenue", \'ltm-metrics\'!$A:$A, 0))'
    )


def test_ltm_total_link_multiplies_by_the_named_fx_cell(tmp_path: Path):
    wb = load_workbook(_build(tmp_path))
    assert ltm_total_link(wb, LTM_REVENUE_RESULT_LABEL, times="F7").endswith(")*F7")


def test_ltm_total_link_takes_the_ebitda_label_the_tab_actually_carries(tmp_path: Path):
    """`ebitda_label` varies with disclosure, so the consumer passes both."""
    adjusted = load_workbook(_build(tmp_path))
    assert '"(=) LTM Adj. EBITDA"' in ltm_total_link(adjusted, *LTM_EBITDA_RESULT_LABELS)

    unadjusted = load_workbook(_build(tmp_path, ebitda_label="LTM EBITDA"))
    assert '"(=) LTM EBITDA"' in ltm_total_link(unadjusted, *LTM_EBITDA_RESULT_LABELS)


def test_ltm_total_link_is_none_without_an_ltm_metrics_tab(tmp_path: Path):
    """Direct /captable, or a plan with no `ltm-metrics` stage — the fallback."""
    wb = load_workbook(_fresh_workbook(tmp_path))
    assert TAB_LTM_METRICS not in wb.sheetnames
    assert ltm_total_link(wb, LTM_REVENUE_RESULT_LABEL) is None
    assert ltm_total_link(wb, *LTM_EBITDA_RESULT_LABELS) is None


def test_ltm_total_link_is_none_for_a_bridge_the_tab_does_not_carry(tmp_path: Path):
    """Each cell falls back independently: a revenue-only tab links D47, not D48."""
    wb = load_workbook(_build(tmp_path, ebitda_bridge=None))
    assert ltm_total_link(wb, LTM_REVENUE_RESULT_LABEL) is not None
    assert ltm_total_link(wb, *LTM_EBITDA_RESULT_LABELS) is None


def _write_cap_table_ltm_cells(workbook_path: Path) -> dict[str, object]:
    """Captable SKILL.md Step 6b, verbatim — returns the two cells' values."""
    written: dict[str, object] = {}

    def fill(wb, ws) -> None:
        fx = resolve_name_cell(ws, NAME_FX_RATE)
        rev_link = ltm_total_link(wb, LTM_REVENUE_RESULT_LABEL, times=fx)
        ebitda_link = ltm_total_link(wb, *LTM_EBITDA_RESULT_LABELS, times=fx)
        cells = {
            "revenue": ws[resolve_name_cell(ws, NAME_LTM_REVENUE_VALUATION)],
            "ebitda": ws[resolve_name_cell(ws, NAME_LTM_EBITDA_VALUATION)],
        }
        for key, link in (("revenue", rev_link), ("ebitda", ebitda_link)):
            if link is not None:
                cells[key].value = link
            else:  # Case B — restore the CapIQ formulas
                field = "IQ_REV" if key == "revenue" else "SP_EBITDA"
                cells[key].value = (
                    f'=_xll.SNL.Clients.Office.Excel.Functions.SPG($F$3,"{field}",'
                    'D$33,$F$6,"Options:Mag=Millions,NA=NA,Curr="&$F$5)'
                )
            written[key] = cells[key].value

    write_tab(workbook_path, TAB_CAPTABLE, TabSpec(write=fill))
    return written


def test_a_cap_table_built_with_ltm_metrics_present_links_instead_of_hardcoding(
    tmp_path: Path,
):
    """The delivered workbook had `captable!D47 = 5207.912*F7` — a second copy of
    a figure the `ltm-metrics` tab already owns. Correct the bridge and the
    Financial Summary slide follows while the cap-table picture silently does not,
    so the deck ships two different LTM revenues with no error value anywhere.
    """
    workbook = _build(tmp_path)  # writes the `ltm-metrics` tab into the deal workbook
    written = _write_cap_table_ltm_cells(workbook)

    for key in ("revenue", "ebitda"):
        formula = written[key]
        assert isinstance(formula, str) and formula.startswith("=INDEX(")
        assert "'ltm-metrics'" in formula, f"{key} must LINK to the tab that computed it"
        assert formula.endswith("*F7"), "converted to the Output currency (F5)"
        # No literal: the only digits in the formula are the MATCH's exact-match
        # flag and the FX cell's row.
        assert not re.search(r"\d{3,}", formula), f"{key} embeds a figure: {formula}"

    # Round-tripped through the file, and reading the sibling tab found its bridges.
    saved = load_workbook(workbook)[TAB_CAPTABLE]
    assert saved[CAP_TABLE_NAMED_RANGES[NAME_LTM_REVENUE_VALUATION]].value == written["revenue"]
    assert '"(=) LTM Adj. EBITDA"' in saved[
        CAP_TABLE_NAMED_RANGES[NAME_LTM_EBITDA_VALUATION]
    ].value


def test_a_cap_table_built_without_ltm_metrics_restores_the_capiq_formulas(
    tmp_path: Path,
):
    """Case B, preserved exactly: no tab, no link — the CapIQ fallback, not a literal."""
    written = _write_cap_table_ltm_cells(_fresh_workbook(tmp_path))
    assert "IQ_REV" in written["revenue"]
    assert "SP_EBITDA" in written["ebitda"]
    assert "'ltm-metrics'" not in written["revenue"]


def test_the_financial_summary_tab_and_the_cap_table_build_the_same_link(tmp_path: Path):
    """Two consumers of one figure, one formula builder — they cannot drift."""
    from financial_summary_workbook import _ltm_link

    wb = load_workbook(_build(tmp_path))
    assert _ltm_link(TAB_LTM_METRICS, LTM_REVENUE_RESULT_LABEL) == ltm_total_link(
        wb, LTM_REVENUE_RESULT_LABEL
    )


def test_the_captable_skill_doc_forbids_a_hardcoded_ltm_figure():
    """The drift lock. The offender was Step 6b Case A writing

        c.value = f"={ltm_revenue}*{fx}"

    which no test could see, because the rule lived only in prose. Prose is still
    where it has to live — this stage is a sub-agent, not a function — so the lock
    is on the prose.
    """
    doc = (
        Path(__file__).resolve().parents[2] / "skills" / "captable" / "SKILL.md"
    ).read_text(encoding="utf-8")

    assert "ltm_total_link" in doc, "Step 6b must build the link through the helper"
    assert "LTM_EBITDA_RESULT_LABELS" in doc, "and pass the label preference order"
    assert not re.search(r"=\{ltm_(revenue|adj_ebitda)\}", doc), (
        "Step 6b must not interpolate an LTM figure into a formula"
    )
    # The fallback stays, and stays keyed on the absent tab rather than on absent
    # input values.
    assert "IQ_REV" in doc and "SP_EBITDA" in doc
    assert "no `ltm-metrics` tab" in doc
