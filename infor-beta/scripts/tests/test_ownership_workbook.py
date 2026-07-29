"""Tests for the ownership workbook builder (deterministic; no Excel needed).

The picture render (Excel COM / LibreOffice) is covered separately and skips
where unavailable; these tests pin the openpyxl write that fills the template's
Select-Insiders block and the vestigial-cruft strip that keeps the output
openable by the render step.
"""

import re
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from deal_workbook import init_deal_workbook
from ownership_workbook import (
    InsiderHolding,
    bloomberg_matches_sedi,
    build_ownership_workbook,
    match_bloomberg_to_sedi,
    read_basic_shares_from_cap_table,
    read_bloomberg_export,
    strip_legal_suffixes,
)

PLUGIN_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = PLUGIN_ROOT / "templates" / "INFOR Ownership Template.xlsx"


def _deal(tmp_path: Path) -> Path:
    """A fresh deal workbook — the Ownership + Bloomberg Output tabs come with it."""
    return init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )


def _build(tmp_path, insiders, total=261_000_000):
    return build_ownership_workbook(
        deal_workbook=_deal(tmp_path),
        insiders=insiders,
        total_shares_outstanding=total,
    )


def test_writes_insider_block_with_names_shares_dates_and_roles(tmp_path: Path):
    out = _build(
        tmp_path,
        [
            InsiderHolding("Barrenechea, Mark James", "Mark Barrenechea (CEO & Director)", 1219092, "2025-03-31"),
            InsiderHolding("Fowlie, Randy", "Randy Fowlie (Director)", [193000, 0, 0], "2025-12-01"),
        ],
    )
    ws = load_workbook(out)["Ownership"]  # formulas preserved (no data_only)

    # Row 39 — single tranche -> plain int.
    assert ws["B39"].value == "Barrenechea, Mark James"
    assert ws["F39"].value == 1219092
    # openpyxl reads a date-formatted cell back as datetime; compare the date part.
    g39 = ws["G39"].value
    assert (g39.date() if isinstance(g39, datetime) else g39) == date(2025, 3, 31)
    assert ws["G39"].number_format == "yyyy-mm-dd"
    assert ws["J39"].value == "Mark Barrenechea (CEO & Director)"
    # Hardcoded values are blue; H/I template formulas are preserved.
    assert str(ws["F39"].font.color.rgb).endswith("0000FF")
    assert ws["H39"].value == 1
    assert ws["I39"].value == "=H39*F39"

    # Row 40 — multiple tranches -> in-cell sum formula (not pre-computed).
    assert ws["F40"].value == "=193000+0+0"
    assert ws["J40"].value == "Randy Fowlie (Director)"

    # F35 (% denominator) is written with a source comment.
    assert ws["F35"].value == 261_000_000
    assert ws["F35"].comment is not None and "cap table" in ws["F35"].comment.text


def test_accepts_plain_dicts_and_handles_missing_date(tmp_path: Path):
    out = _build(
        tmp_path,
        [{"sedi_name": "Doe, Jane", "adjusted_name": "Jane Doe (CFO)", "common_shares": 500}],
    )
    ws = load_workbook(out)["Ownership"]
    assert ws["B39"].value == "Doe, Jane"
    assert ws["F39"].value == 500
    assert ws["G39"].value is None  # no date supplied


def test_too_many_insiders_raises(tmp_path: Path):
    too_many = [
        InsiderHolding(f"Name {i}", f"Name {i} (Director)", i + 1, "2025-01-01") for i in range(28)
    ]
    with pytest.raises(ValueError, match="insider rows"):
        _build(tmp_path, too_many)


def test_template_and_output_carry_no_external_links_or_invented_names(tmp_path: Path):
    """The write must not re-dirty the file, or the render step cannot open it.

    The source template is still pre-cleaned of dead external links and legacy
    defined names. The *output* is now the deal workbook, which legitimately
    inherits the comps template's 1,246 legacy artefacts and the 27 `infor_`
    names — so the assertion is that this build ADDS nothing and loses nothing,
    compared against a freshly initialised deal workbook.
    """
    src = load_workbook(TEMPLATE)
    assert list(getattr(src, "_external_links", [])) == [], "shipped template must carry no external links"
    assert len(src.defined_names) == 0, "shipped template must carry no defined names"

    baseline = set(load_workbook(_deal(tmp_path / "baseline")).defined_names)
    out = _build(tmp_path, [InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")])
    cleaned = load_workbook(out)
    assert list(getattr(cleaned, "_external_links", [])) == []
    assert set(cleaned.defined_names) == baseline


def test_f35_keeps_template_palatino_font(tmp_path: Path):
    """F35 keeps the template's Palatino typeface — a bare Font() would reset it to
    Calibri 11. The aggregator later relinks F35 to the cap table, preserving font."""
    out = _build(tmp_path, [InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")])
    f35 = load_workbook(out)["Ownership"]["F35"].font
    assert f35.name == "Palatino Linotype", "F35 must keep the template's Palatino font"
    # The hardcoded insider data cells stay blue but also keep Palatino (not Calibri 11).
    f39 = load_workbook(out)["Ownership"]["F39"].font
    assert f39.name == "Palatino Linotype", "insider data cells must keep Palatino"


def test_total_shares_none_leaves_f35_blank(tmp_path: Path):
    out = build_ownership_workbook(
        deal_workbook=_deal(tmp_path),
        insiders=[InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")],
        total_shares_outstanding=None,
    )
    assert load_workbook(out)["Ownership"]["F35"].value is None


# ─── The padding top-12 display (the duplicate-name bug) ─────────────────────
#
# `Ownership!F5:F16` is `=LARGE($I$39:$I$65, A<row>)` and `B5:B16` is
# `=XLOOKUP(F<row>, $I$39:$I$178, $J$39:$J$178)`. The pool is `=H*F` over every
# data row, and an unfed row computes to 0 rather than being absent — so with
# fewer than 12 positive balances `LARGE` returns 0 for each surplus rank and
# `XLOOKUP(0, ...)` matches the first zero-balance holder, printing that one
# name once per surplus rank. Renderer-free by design: these assert the guard the
# stage writes, not LibreOffice's arithmetic.

_RANK_ROWS = {"insiders": range(5, 17), "institutions": range(20, 32)}
_POOL = {"insiders": "$I$39:$I$65", "institutions": "$I$68:$I$185"}


def _guard_of(ws, row: int) -> tuple[str, str]:
    """``(rank ref, pool ref)`` from a guarded share formula, or fail loudly."""
    formula = ws[f"F{row}"].value
    match = re.fullmatch(
        r'=IF\((\S+?)<=COUNTIF\((\S+?),">0"\),LARGE\(\2,\1\),""\)', formula
    )
    assert match, (
        f"F{row} is not a guarded rank formula: {formula!r}. Without the COUNTIF "
        f"guard a surplus rank resolves to LARGE(...)=0 and repeats the first "
        f"zero-balance holder."
    )
    return match.group(1), match.group(2)


@pytest.mark.parametrize("block", ("insiders", "institutions"))
def test_every_display_rank_is_guarded_by_a_countif_over_its_own_pool(
    tmp_path: Path, block: str
):
    out = _build(tmp_path, [InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")])
    ws = load_workbook(out)["Ownership"]

    for row in _RANK_ROWS[block]:
        rank, pool = _guard_of(ws, row)
        assert rank == f"A{row}", f"F{row} must rank by its own row's A cell"
        assert pool == _POOL[block], f"F{row} counts a different pool than it ranks"
        # The name and the percentage follow the share cell: otherwise B repeats a
        # name and G divides "" into #VALUE!.
        for col in ("B", "G"):
            assert ws[f"{col}{row}"].value.startswith(f'=IF(F{row}="","",'), (
                f"{col}{row} is not gated on F{row} being blank: "
                f"{ws[f'{col}{row}'].value!r}"
            )


def test_the_guard_leaves_the_subtotals_summing_the_display_block(tmp_path: Path):
    """The arithmetic stays in Excel: `SUM` ignores the `""` a guarded rank yields,
    so the Subtotal needs no rewrite — and must not get one."""
    ws = load_workbook(
        _build(tmp_path, [InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")])
    )["Ownership"]

    assert ws["F17"].value == "=SUM(F5:F16)"
    assert ws["F32"].value == "=SUM(F20:F31)"
    assert ws["F33"].value == "=+F35-F32-F17"


def test_fewer_than_twelve_positive_insiders_renders_no_duplicate_name_or_row(
    tmp_path: Path,
):
    """The shipped symptom: "Ayman Antoun (CEO & Director) -- <0.1%" printed twice.

    Three insiders, two of them holding common shares. The pool therefore holds two
    positive balances, so ranks 3-12 are guarded off and the display can only show
    the two distinct holders.
    """
    insiders = [
        InsiderHolding("Antoun, Ayman", "Ayman Antoun (CEO & Director)", 150_000, "2025-03-31"),
        InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 90_000, "2025-04-01"),
        InsiderHolding("Zero, Zed", "Zed Zero (Director)", 0, "2025-04-02"),
    ]
    ws = load_workbook(_build(tmp_path, insiders))["Ownership"]

    # The pool is `=H<row>*F<row>`; H ships as 1 and F holds what the stage wrote,
    # so counting the positive balances needs no formula evaluation.
    positive = [
        row
        for row in range(39, 66)
        if isinstance(ws[f"F{row}"].value, (int, float)) and ws[f"F{row}"].value > 0
    ]
    assert len(positive) == 2

    rendered, blank = [], []
    for row in _RANK_ROWS["insiders"]:
        rank, _ = _guard_of(ws, row)
        (rendered if ws[rank].value <= len(positive) else blank).append(row)

    assert rendered == [5, 6], "only the ranks with a positive balance may render"
    assert blank == list(range(7, 17)), "every surplus rank must render as an empty row"
    # `LARGE` returns the k-th largest, so distinct balances give distinct names —
    # and the pre-fix formulas would have put rank 3-12's XLOOKUP(0, ...) hit here
    # ten more times.
    assert len({ws[f"F{row}"].value for row in rendered}) == len(rendered)


def test_a_block_with_no_recognisable_rank_row_halts(tmp_path: Path):
    """A silent no-op would put the duplicates straight back."""
    from ownership_workbook import _guard_display_block
    from template_layout import NAME_OWN_INSIDERS_PICTURE, TemplateLayoutError

    ws = load_workbook(
        _build(tmp_path, [InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")])
    )["Ownership"]
    for row in _RANK_ROWS["insiders"]:
        ws[f"F{row}"] = None

    with pytest.raises(TemplateLayoutError, match="no rank row"):
        _guard_display_block(ws, NAME_OWN_INSIDERS_PICTURE)


def test_the_guard_is_idempotent(tmp_path: Path):
    """The stage may run twice over one deal workbook; the wrap must not nest."""
    deal = _deal(tmp_path)
    insiders = [InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")]
    for _ in range(2):
        build_ownership_workbook(
            deal_workbook=deal, insiders=insiders, total_shares_outstanding=261_000_000
        )
    ws = load_workbook(deal)["Ownership"]

    for block in ("insiders", "institutions"):
        for row in _RANK_ROWS[block]:
            _guard_of(ws, row)  # still exactly one guard, not a nested pair
            assert ws[f"B{row}"].value.count("=IF(") == 1
            assert ws[f"G{row}"].value.count("=IF(") == 1


def _cap_table(path: Path, rows: dict[int, float]) -> Path:
    """A pre-Phase-D STANDALONE cap table, i.e. the legacy branch of the reader.

    `read_basic_shares_from_cap_table` accepts both spellings on purpose (it tries
    `TAB_CAPTABLE` then `CAP_TABLE_SOURCE_SHEET`), so this covers the fallback.
    The branch production actually takes is covered by
    `test_read_basic_shares_reads_the_deal_workbook_tab` below — which is the one
    that was missing, and is the same fixture/artefact mismatch that hid v0.5.45.
    """
    from openpyxl.workbook.defined_name import DefinedName

    from template_layout import CAP_TABLE_SOURCE_SHEET, NAME_CAP_SHARE_INPUTS

    wb = Workbook()
    ws = wb.active
    ws.title = CAP_TABLE_SOURCE_SHEET
    # read_basic_shares_from_cap_table resolves the summing window through this
    # name, and verifies it is present first — so a synthetic cap table has to
    # carry it, exactly as the shipped template and the deal workbook do.
    ws.defined_names.add(
        DefinedName(NAME_CAP_SHARE_INPUTS, attr_text=f"'{CAP_TABLE_SOURCE_SHEET}'!$F$168:$F$185")
    )
    for row, value in rows.items():
        ws.cell(row=row, column=6, value=value)  # column F
    wb.save(path)
    return path


def test_read_basic_shares_sums_section_vii_to_full_units(tmp_path: Path):
    # Section VII basic-share inputs are in millions; reader returns full units.
    cap = _cap_table(tmp_path / "cap.xlsx", {168: 100.0, 169: 72.34})
    assert read_basic_shares_from_cap_table(cap) == 172_340_000


def test_read_basic_shares_missing_or_empty_returns_none(tmp_path: Path):
    assert read_basic_shares_from_cap_table(tmp_path / "nope.xlsx") is None
    empty = _cap_table(tmp_path / "empty.xlsx", {})
    assert read_basic_shares_from_cap_table(empty) is None


def test_read_basic_shares_reads_the_deal_workbook_tab(tmp_path: Path):
    """The branch the pitch plan actually takes: the deal workbook's `captable`.

    The reader's other tests all build a standalone `Cap with Links` sheet, so
    only its legacy fallback was covered. `ownership.yaml`/`pitch.yaml` pass the
    deal workbook, whose tab is `captable` — the same fixture-vs-artefact gap
    that hid the v0.5.45 assembler bug, on a different reader.
    """
    from deal_workbook import TAB_CAPTABLE, TabSpec, init_deal_workbook, write_tab
    from template_layout import NAME_CAP_SHARE_INPUTS, resolve_name_range

    deal = init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )

    def _fill(_wb, ws):
        # Write into the top of the real `infor_cap_share_inputs` block, wherever
        # the template now puts it, rather than at a hardcoded F168.
        from openpyxl.utils import range_boundaries

        min_col, min_row, _, _ = range_boundaries(
            resolve_name_range(ws, NAME_CAP_SHARE_INPUTS)
        )
        ws.cell(row=min_row, column=min_col, value=100.0)
        ws.cell(row=min_row + 1, column=min_col, value=72.34)

    write_tab(
        deal, TAB_CAPTABLE, TabSpec(write=_fill, verify_names=(NAME_CAP_SHARE_INPUTS,))
    )
    assert read_basic_shares_from_cap_table(deal) == 172_340_000


# ─── Bloomberg institutional side ─────────────────────────────────────────────

def _bbg_export(path: Path, holders, sheet_title: str = "Summary View") -> Path:
    """A minimal BBG Excel add-in ownership export (Summary View layout).

    ``holders`` is a list of ``(name, position, insider_status)`` tuples.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws["C7"] = "Name"
    ws["E7"] = "Test Co Inc."
    ws["C9"] = "Ticker"
    ws["E9"] = "TST CN EQUITY"
    ws["G9"] = "View"
    ws["I9"] = "Summary"
    ws["C11"] = "Sort By"
    ws["E11"] = "Position"
    ws["G11"] = "Order"
    ws["I11"] = "Descending"
    for col, header in {
        "C": "Holder Name", "G": "Portfolio Name", "L": "Position", "M": "Latest Chg",
        "N": "Filing Date", "O": "Source", "P": "% Out", "R": "Insider Status",
        "S": "Institution Type", "U": "Country", "V": "Market Value",
    }.items():
        ws[f"{col}13"] = header
    for i, (name, position, status) in enumerate(holders):
        row = 14 + i
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = name
        ws[f"L{row}"] = position
        ws[f"N{row}"] = datetime(2026, 5, 21)
        ws[f"N{row}"].number_format = "m/d/yyyy"
        ws[f"O{row}"] = "Form 4" if status == "Y" else "ULT-AGG"
        ws[f"R{row}"] = status
        ws[f"S{row}"] = "Unclassified" if status == "Y" else "Investment Advisor"
    wb.save(path)
    return path


_SAMPLE_HOLDERS = [
    ("Malchow Eric D", 4_972_500, "Y"),
    ("T Rowe Price Group Inc", 150_311, "N-P"),
    ("Kelso & Co LP", 14_983, "N-P"),
]


@pytest.mark.parametrize(
    ("raw", "adjusted"),
    [
        ("T Rowe Price Group Inc", "T Rowe Price"),
        ("Kelso & Co LP", "Kelso & Co"),  # '& Co' is the brand, not a suffix
        ("Vanguard Group Inc", "Vanguard"),
        ("BlackRock Inc.", "BlackRock"),
        ("Malchow Eric D", "Malchow Eric D"),  # person names pass through
        ("Royal Bank of Canada", "Royal Bank of Canada"),
        ("Inc", "Inc"),  # never stripped to nothing
    ],
)
def test_strip_legal_suffixes(raw: str, adjusted: str):
    assert strip_legal_suffixes(raw) == adjusted


@pytest.mark.parametrize(
    ("bbg", "sedi", "expected"),
    [
        ("Malchow Eric D", "Malchow, Eric D", True),
        ("Malchow Eric", "Malchow, Eric D", True),  # middle initial missing on one side
        ("Weber Mary R.", "Weber, Mary R", True),  # punctuation-insensitive
        ("Smith M Christine", "Smith, M Christine", True),
        ("Smith J", "Smith, Jane", True),  # initial-only given name
        ("Van Der Berg Jan", "Van Der Berg, Jan", True),  # multi-token surname
        ("Fairfax Financial Holdings Ltd", "Fairfax Financial Holdings Limited", True),  # corporate insider
        ("Smith M Christine", "Smith, Mary", False),  # full given names disagree
        ("Brown Robert T", "Brown, Richard", False),
        ("T Rowe Price Group Inc", "Barrenechea, Mark", False),
        ("Malchow Eric D", "Malchowski, Eric", False),  # surname must match exactly
    ],
)
def test_bloomberg_matches_sedi(bbg: str, sedi: str, expected: bool):
    assert bloomberg_matches_sedi(bbg, sedi) is expected


def test_match_bloomberg_to_sedi_maps_first_match_only():
    matches = match_bloomberg_to_sedi(
        [h[0] for h in _SAMPLE_HOLDERS], ["Malchow, Eric D", "Brown, Robert T"]
    )
    assert matches == {"Malchow Eric D": "Malchow, Eric D"}


def test_read_bloomberg_export_parses_summary_view(tmp_path: Path):
    src = _bbg_export(tmp_path / "bbg.xlsx", _SAMPLE_HOLDERS)
    export = read_bloomberg_export(src)
    assert export.company_name == "Test Co Inc."
    assert export.ticker == "TST CN EQUITY"
    assert [h.name for h in export.holders] == [h[0] for h in _SAMPLE_HOLDERS]
    assert export.holders[0].position == 4_972_500
    assert export.holders[0].insider_status == "Y"
    assert export.holders[1].institution_type == "Investment Advisor"
    assert export.holders[0].number_formats["N"] == "m/d/yyyy"
    assert export.info["E7"] == "Test Co Inc."


def test_read_bloomberg_export_finds_sheet_by_header_row(tmp_path: Path):
    src = _bbg_export(tmp_path / "bbg.xlsx", _SAMPLE_HOLDERS, sheet_title="My Export")
    assert [h.name for h in read_bloomberg_export(src).holders] == [h[0] for h in _SAMPLE_HOLDERS]


def test_read_bloomberg_export_rejects_non_bbg_workbook(tmp_path: Path):
    wb = Workbook()
    wb.active["A1"] = "not a bloomberg export"
    wb.save(tmp_path / "other.xlsx")
    with pytest.raises(ValueError, match="Summary View"):
        read_bloomberg_export(tmp_path / "other.xlsx")
    with pytest.raises(FileNotFoundError):
        read_bloomberg_export(tmp_path / "missing.xlsx")


def _build_with_bbg(tmp_path: Path, holders=None, **kwargs):
    src = _bbg_export(tmp_path / "bbg.xlsx", holders if holders is not None else _SAMPLE_HOLDERS)
    return build_ownership_workbook(
        deal_workbook=_deal(tmp_path),
        insiders=[
            InsiderHolding("Malchow, Eric D", "Eric Malchow (CEO & Director)", 4_972_500, "2026-05-21"),
            InsiderHolding("Brown, Robert T", "Robert Brown (Director)", 1_665_600, "2026-05-21"),
        ],
        total_shares_outstanding=31_650_000,
        bloomberg_export_path=src,
        **kwargs,
    )


def test_build_with_bloomberg_fills_output_tab_and_links(tmp_path: Path):
    out = _build_with_bbg(tmp_path)
    wb = load_workbook(out)
    bbg = wb["Bloomberg Output"]
    own = wb["Ownership"]

    # Export copied into the Bloomberg Output tab (values + formats + info cells).
    assert bbg["E7"].value == "Test Co Inc."
    assert bbg["E9"].value == "TST CN EQUITY"
    assert bbg["C14"].value == "Malchow Eric D"
    assert bbg["L14"].value == 4_972_500
    assert bbg["R14"].value == "Y"
    assert bbg["N14"].number_format == "m/d/yyyy"
    assert bbg["C16"].value == "Kelso & Co LP"

    # The Ownership tab's pre-wired link formulas are untouched.
    assert own["B68"].value == "='Bloomberg Output'!C14"
    assert "XLOOKUP" in own["F68"].value

    # SEDI duplicate excluded (H=0 + audit comment); institutions stay included.
    assert own["H68"].value == 0
    assert str(own["H68"].font.color.rgb).endswith("0000FF")
    assert "Malchow, Eric D" in own["H68"].comment.text
    assert own["H69"].value == 1
    assert own["H70"].value == 1

    # Adjusted display names (col J) for every populated Bloomberg row.
    assert own["J68"].value == "Malchow Eric D"
    assert own["J69"].value == "T Rowe Price"
    assert own["J70"].value == "Kelso & Co"

    # Unused link rows are neutralised so LARGE($I$68:$I$185) stays numeric.
    for row in (71, 120, 185):
        assert own[f"B{row}"].value is None
        assert own[f"F{row}"].value is None
        assert own[f"G{row}"].value is None
        assert own[f"H{row}"].value == 0
        assert own[f"I{row}"].value == f"=H{row}*F{row}"

    # Insider block untouched by the Bloomberg write.
    assert own["B39"].value == "Malchow, Eric D"
    assert own["F35"].value == 31_650_000


def test_build_with_bloomberg_honours_overrides(tmp_path: Path):
    out = _build_with_bbg(
        tmp_path,
        bloomberg_adjusted_names={"T Rowe Price Group Inc": "T. Rowe Price"},
        bloomberg_include_overrides={"Kelso & Co LP": 0, "Malchow Eric D": 1},
    )
    own = load_workbook(out)["Ownership"]
    assert own["J69"].value == "T. Rowe Price"
    assert own["H70"].value == 0
    assert "override" in own["H70"].comment.text
    # Forced include on a matched duplicate leaves the template's 1 in place.
    assert own["H68"].value == 1
    assert own["H68"].comment is None


def test_build_without_bloomberg_leaves_institutional_side_untouched(tmp_path: Path):
    out = _build(tmp_path, [InsiderHolding("Doe, Jane", "Jane Doe (CFO)", 500, "2025-01-01")])
    wb = load_workbook(out)
    own = wb["Ownership"]
    assert wb["Bloomberg Output"]["C14"].value is None
    assert own["B68"].value == "='Bloomberg Output'!C14"
    assert own["H68"].value == 1
    assert own["H185"].value == 1
    assert own["J68"].value is None


def test_build_with_bloomberg_truncates_at_118_holders(tmp_path: Path):
    holders = [(f"Holder {i}", 1_000 - i, "N-P") for i in range(1, 121)]  # 120 holders
    out = _build_with_bbg(tmp_path, holders=holders)
    wb = load_workbook(out)
    bbg = wb["Bloomberg Output"]
    assert bbg["C131"].value == "Holder 118"
    assert bbg["C132"].value is None
    own = wb["Ownership"]
    assert own["H185"].value == 1  # all 118 slots used, none neutralised
    assert own["J185"].value == "Holder 118"


def test_build_with_bloomberg_stays_excel_clean(tmp_path: Path):
    baseline = set(load_workbook(_deal(tmp_path / "baseline")).defined_names)
    out = _build_with_bbg(tmp_path)
    cleaned = load_workbook(out)
    assert list(getattr(cleaned, "_external_links", [])) == []
    assert set(cleaned.defined_names) == baseline


# ─── Provenance: the SEDI report, the export, and the cap-table denominator ──
#
# Every share count here is a figure read off an attached document, and the tab
# recorded none of them — a run's ledger held 70 records with `ownership`
# contributing zero, so every percentage on the ownership slide was untraceable.
# F35 is the interesting one: it is DERIVED from the cap table, in another stage's
# fragment, so its ref resolves only in the run merge.

_SEDI = "SEDI Insider Information by Issuer report"


def _sedi_source(name: str, page: int = 3):
    from provenance import FigureSource

    return FigureSource(filing=_SEDI, statement=name, page=page)


def test_each_insider_holding_is_recorded_against_the_sedi_report(tmp_path: Path):
    from provenance import ProvenanceLedger

    ledger = ProvenanceLedger(stage="ownership")
    out = build_ownership_workbook(
        deal_workbook=_deal(tmp_path),
        insiders=[
            InsiderHolding("Barrenechea, Mark James", "Mark Barrenechea (CEO & Director)",
                           1_219_092, "2026-03-31",
                           source=_sedi_source("Barrenechea, Mark James")),
            InsiderHolding("Fowlie, Randy", "Randy Fowlie (Director)", [193_000, 5_000],
                           "2026-12-01", source=_sedi_source("Fowlie, Randy", page=4)),
        ],
        total_shares_outstanding=261_000_000,
        provenance=ledger,
    )

    holdings = [f for f in ledger.figures if f.figure.startswith("Insider holding")]
    assert [f.figure for f in holdings] == [
        "Insider holding — Mark Barrenechea (CEO & Director)",
        "Insider holding — Randy Fowlie (Director)",
    ]
    assert holdings[0].value == 1_219_092
    # A multi-tranche holding is a sum FORMULA in the cell, and the record holds
    # exactly what the cell holds — Excel does the math, not the record.
    assert holdings[1].value == "=193000+5000"
    assert holdings[1].citation_lines == (f"{_SEDI}, Fowlie, Randy, p. 4",)
    assert holdings[0].location == "Ownership!F39"

    # The comment on the share cell is rendered FROM the record.
    comment = load_workbook(out)["Ownership"]["F39"].comment
    assert comment is not None and comment.text.endswith(
        f"Source: {_SEDI}, Barrenechea, Mark James, p. 3"
    )


def test_the_percentage_denominator_is_recorded_as_derived_from_the_cap_table(tmp_path: Path):
    from provenance import ProvenanceLedger

    ledger = ProvenanceLedger(stage="ownership")
    build_ownership_workbook(
        deal_workbook=_deal(tmp_path),
        insiders=[InsiderHolding("A, B", "B A (Director)", 1, "2026-01-01",
                                 source=_sedi_source("A, B"))],
        total_shares_outstanding=261_000_000,
        provenance=ledger,
    )

    total = next(f for f in ledger.figures if f.figure.startswith("Total basic shares"))
    assert total.sources == (), "the denominator is not read off a filing here"
    assert total.derived and "cap table" in total.derivation
    assert [r.render() for r in total.derived_from] == ["Total Basic Shares Outstanding"]
    assert total.location == "Ownership!F35"


def test_the_denominators_ref_resolves_against_the_cap_table_stages_fragment(tmp_path: Path):
    # The cross-stage join: on its own the ownership fragment cannot see the cap
    # table's record, and in the merge it can — which is what lets a reviewer walk
    # an ownership percentage back to the capital-stock note.
    from provenance import FigureRef, FigureSource, ProvenanceLedger, read_run_provenance

    captable = ProvenanceLedger(stage="captable")
    captable.record("Common shares", value=261.0, location="captable!F168",
                    sources=FigureSource(filing="FY2025 10-K",
                                         statement="Capital stock note", page=95))
    captable.record("Total Basic Shares Outstanding", value="=SUM(F168:F185)",
                    location="captable!F186", derivation="cap-table formula =SUM(F168:F185)",
                    derived_from=[FigureRef(location="captable!F168")])
    captable.write(tmp_path / "run" / "stages" / "captable")

    ownership = ProvenanceLedger(stage="ownership")
    build_ownership_workbook(
        deal_workbook=_deal(tmp_path),
        insiders=[InsiderHolding("A, B", "B A (Director)", 1, "2026-01-01",
                                 source=_sedi_source("A, B"))],
        total_shares_outstanding=261_000_000,
        provenance=ownership,
    )
    ownership.write(tmp_path / "run" / "stages" / "ownership")

    merged = read_run_provenance(tmp_path / "run")
    total = next(f for f in merged.figures if f.figure.startswith("Total basic shares"))
    trace = merged.trace(total)
    assert trace.resolved
    assert any("Capital stock note" in s.render() for s in trace.root_sources)


def test_each_institutional_position_is_recorded_against_the_attached_export(tmp_path: Path):
    from provenance import ProvenanceLedger

    ledger = ProvenanceLedger(stage="ownership")
    _build_with_bbg(tmp_path, provenance=ledger)

    positions = [f for f in ledger.figures if f.figure.startswith("Institutional position")]
    assert positions, "the Bloomberg side recorded nothing"
    assert all(f.citation_lines == ("bbg.xlsx, Summary View",) for f in positions)
    assert all(f.location.startswith("Bloomberg Output!L") for f in positions)


def test_a_citation_string_instead_of_a_source_record_raises(tmp_path: Path):
    from provenance import ProvenanceError

    with pytest.raises(ProvenanceError, match="no longer a source record"):
        _build(tmp_path, [InsiderHolding("A, B", "B A (Director)", 1, "2026-01-01",
                                         source="the SEDI report, page 3")])
