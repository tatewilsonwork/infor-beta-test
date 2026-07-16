"""Unit tests for the workbook aggregator helper.

These exercise the platform-independent pieces — filename normalization,
Excel-safe sheet naming, theme resolution, and the openpyxl merge backend — so
they run on any platform without Microsoft Excel. The COM backend is covered
only on Windows-with-Excel runtimes and is not unit-tested here.
"""

import re
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import workbook_aggregator
from workbook_aggregator import (
    _combine_via_openpyxl,
    _default_theme_path,
    _excel_safe_sheet_name,
    _extract_theme_xml,
    _recalc_with_libreoffice,
    _unique_sheet_name,
    combine_workbooks,
    combined_filename,
)


@pytest.fixture(autouse=True)
def _stub_recalc(monkeypatch):
    """Keep the merge/relink/theme tests deterministic regardless of whether the
    test machine has LibreOffice: stub combine_workbooks' recalc step to a no-op so
    it never shells out to soffice (which would re-save and could rewrite formulas /
    the theme XML). The recalc-specific tests below call the *real* imported
    `_recalc_with_libreoffice` (a stable reference the module-attr patch doesn't
    rebind) or override this stub with a recorder."""
    monkeypatch.setattr(workbook_aggregator, "_recalc_with_libreoffice", lambda p: False)


def _make_workbook(path: Path, sheets: dict[str, list[list]]) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title=title)
        for r, row in enumerate(rows, start=1):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
    wb.save(path)
    return path


def _make_anchored_captable(path: Path, sheet: str = "Cap with Links") -> Path:
    """A minimal cap-table source carrying the real template's sentinel labels.

    combine_workbooks verifies the relinked cells' anchors (template_layout)
    before merging whenever a relink partner tab is present, so a synthetic
    captable used alongside ltm-metrics/ownership/comps/precedents must look
    like the shipped template around those cells.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws["B5"] = "Output Currency:"
    ws["B7"] = "FX Rate:"
    ws["B17"] = "Basic Shares Outstanding"
    ws["D33"] = "LTM"
    ws["B47"] = "Revenue"
    ws["B48"] = "Adj. EBITDA"
    wb.save(path)
    return path


def _make_anchored_ownership(path: Path) -> Path:
    """A minimal ownership source carrying the F35 sentinel the relink
    pre-flight verifies."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ownership"
    ws["B35"] = "Total Basic Shares Outstanding"
    wb.save(path)
    return path


# --- combined_filename -------------------------------------------------------


def test_combined_filename_drops_hyphens_from_deliverable():
    assert combined_filename("earnings-update", "Project Atlas") == "earningsupdate-Project Atlas.xlsx"


def test_combined_filename_keeps_plain_deliverable():
    assert combined_filename("pitch", "Project Atlas") == "pitch-Project Atlas.xlsx"


def test_combined_filename_sanitizes_deal_name():
    name = combined_filename("pitch", "Project: A/B?")
    assert name == "pitch-Project- A-B-.xlsx"


# --- sheet-name rules --------------------------------------------------------


def test_excel_safe_sheet_name_strips_forbidden_and_truncates():
    assert _excel_safe_sheet_name("Cap:With/Links") == "Cap-With-Links"
    assert len(_excel_safe_sheet_name("x" * 50)) == 31


def test_unique_sheet_name_disambiguates_collisions():
    used: set[str] = set()
    assert _unique_sheet_name("captable", used) == "captable"
    assert _unique_sheet_name("captable", used) == "captable (2)"
    assert _unique_sheet_name("CAPTABLE", used) == "CAPTABLE (3)"


# --- openpyxl merge backend --------------------------------------------------


def test_single_sheet_source_tab_named_after_skill(tmp_path: Path):
    a = _make_workbook(tmp_path / "a.xlsx", {"LTM Metrics": [["x", 1]]})
    out = tmp_path / "out.xlsx"
    _combine_via_openpyxl([("ltm-metrics", a)], out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["ltm-metrics"]


def test_multi_sheet_source_keeps_original_sheet_names(tmp_path: Path):
    # A multi-sheet source keeps its (self-describing) sheet names unprefixed —
    # the skill prefix is only used to label a single-sheet source.
    a = _make_workbook(tmp_path / "a.xlsx", {"Ownership": [["a"]], "Bloomberg Output": [["b"]]})
    out = tmp_path / "out.xlsx"
    _combine_via_openpyxl([("ownership", a)], out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Ownership", "Bloomberg Output"]


def test_multi_sheet_source_preserves_cross_sheet_refs(tmp_path: Path):
    # Because multi-sheet tabs keep their original names, a sheet's formula that
    # references a sibling sheet by name stays valid (no rename -> no #REF). This
    # is the ownership `Ownership` -> `Bloomberg Output` case in miniature.
    a = _make_workbook(
        tmp_path / "a.xlsx",
        {"Ownership": [["=+'Bloomberg Output'!C14"]], "Bloomberg Output": [["x"]]},
    )
    out = tmp_path / "out.xlsx"
    _combine_via_openpyxl([("ownership", a)], out)
    wb = load_workbook(out)
    assert "Bloomberg Output" in wb.sheetnames
    assert wb["Ownership"]["A1"].value == "=+'Bloomberg Output'!C14"


def test_capiq_helper_sheet_is_dropped(tmp_path: Path):
    # The CapIQ add-in's "__snloffice" metadata sheet must not surface as a tab;
    # with it filtered out the lone content sheet collapses to the skill name.
    a = _make_workbook(
        tmp_path / "a.xlsx",
        {"__snloffice": [["garbage"]], "Cap with Links": [["a"]]},
    )
    out = tmp_path / "out.xlsx"
    _combine_via_openpyxl([("captable", a)], out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["captable"]


def test_merge_preserves_values_and_formulas(tmp_path: Path):
    a = _make_workbook(tmp_path / "a.xlsx", {"S": [["Total", 10], [None, "=B1*2"]]})
    out = tmp_path / "out.xlsx"
    _combine_via_openpyxl([("comps", a)], out)
    ws = load_workbook(out)["comps"]
    assert ws["A1"].value == "Total"
    assert ws["B1"].value == 10
    assert ws["B2"].value == "=B1*2"


def test_merge_preserves_comments_and_hyperlinks(tmp_path: Path):
    # v0.5.21: the cap-table template stores the analyst's CapIQ refresh formulas
    # as cell COMMENTS (F7/F16), and openpyxl copies neither comments nor
    # hyperlinks with the style — both must be carried across explicitly or the
    # documented refresh workflow silently breaks on the off-Windows merge.
    # Since v0.5.31 the F7/F16 comments are multi-line: the captable skill appends
    # a "Source: <url> — retrieved <date>" citation below the CapIQ formula, and
    # the whole comment (both lines) must survive the merge verbatim.
    from openpyxl.comments import Comment

    f7_comment = "=IQ_FX_RATE(...)\nSource: https://example.com/fx — retrieved 2026-07-15"

    a = tmp_path / "cap.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Cap"
    ws["F7"] = 1.36
    ws["F7"].comment = Comment(f7_comment, "INFOR")
    ws["B2"] = "Source"
    ws["B2"].hyperlink = "https://example.com/filing"
    wb.save(a)

    out = tmp_path / "out.xlsx"
    _combine_via_openpyxl([("captable", a)], out)
    ws2 = load_workbook(out)["captable"]
    assert ws2["F7"].comment is not None
    assert ws2["F7"].comment.text == f7_comment
    assert ws2["F7"].comment.author == "INFOR"
    assert ws2["B2"].hyperlink is not None
    assert ws2["B2"].hyperlink.target == "https://example.com/filing"


def test_builder_source_comments_survive_openpyxl_merge(tmp_path: Path):
    # v0.5.34: financial-summary / ltm-metrics cite each extracted figure as a
    # "Source: …" comment on the value cell; the openpyxl merge must carry them
    # into the combined workbook like the cap table's F7/F16 citations.
    from financial_summary_workbook import MetricSeries, build_financial_summary_workbook
    from ltm_metrics import BridgeComponent, RevenueSegment, build_ltm_metrics_workbook

    fs = build_financial_summary_workbook(
        company_name="SampleCo",
        currency_note="Figures in US$MM unless noted",
        period_note="FY = fiscal year; LTM = trailing twelve months as of Q3 2026",
        fiscal_labels=["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"],
        metrics=[
            MetricSeries("Revenue", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Revenue",
                         sources=["FY 10-K, income statement"] * 5),
            MetricSeries("Gross Profit", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Gross Profit"),
            MetricSeries("Adjusted EBITDA", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Adj. EBITDA"),
            MetricSeries("Net Income", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Net Income"),
        ],
        output_dir=tmp_path,
    )
    ltm = build_ltm_metrics_workbook(
        company_name="SampleCo",
        period_label="LTM ended March 31, 2026",
        currency="US$MM",
        segmentation_basis="Service line",
        segments=[RevenueSegment("Segment A", 9.0, source="Q3 10-Q, segment note")],
        revenue_bridge=[BridgeComponent("FY Revenue", 9.0, source="FY 10-K, income statement")],
        output_dir=tmp_path,
    )

    out = tmp_path / "out.xlsx"
    _combine_via_openpyxl([("financial-summary", fs), ("ltm-metrics", ltm)], out)
    wb = load_workbook(out)
    assert wb["financial-summary"]["B6"].comment is not None
    assert wb["financial-summary"]["B6"].comment.text == "Source: FY 10-K, income statement"
    # One segment -> overview data row 8; bridge data row 13 (spacer 10,
    # section 11, header 12).
    assert wb["ltm-metrics"]["B8"].comment.text == "Source: Q3 10-Q, segment note"
    assert wb["ltm-metrics"]["B13"].comment.text == "Source: FY 10-K, income statement"


# --- combine_workbooks end-to-end (openpyxl path off-Windows) ---------------


def test_combine_writes_named_file_and_deletes_sources(tmp_path: Path, monkeypatch):
    # Force the openpyxl backend so the test never shells out to Excel.
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    a = _make_anchored_captable(tmp_path / "cap.xlsx", sheet="Cap")
    b = _make_workbook(tmp_path / "ltm.xlsx", {"LTM": [[2]]})

    result = combine_workbooks(
        sources={"captable": a, "ltm-metrics": b},
        output_dir=tmp_path,
        deliverable_type="earnings-update",
        deal_name="Project Atlas",
    )

    out = result.output_path
    assert out.name == "earningsupdate-Project Atlas.xlsx"
    assert out.exists()
    assert load_workbook(out).sheetnames == ["captable", "ltm-metrics"]
    # Verified merge — sources replaced by the combined workbook, no warnings.
    assert result.backend == "openpyxl"
    assert result.degraded is False  # off-Windows, openpyxl IS the intended backend
    assert result.relink_ok is True
    assert result.external_refs == ()
    assert result.warnings == ()
    assert result.sources_deleted is True
    assert result.kept_sources == ()
    assert not a.exists()
    assert not b.exists()


def test_combine_skips_none_and_missing_sources(tmp_path: Path, monkeypatch):
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    a = _make_workbook(tmp_path / "cap.xlsx", {"Cap": [[1]]})

    result = combine_workbooks(
        sources={
            "captable": a,
            "comps": None,
            "precedents": tmp_path / "does-not-exist.xlsx",
        },
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Project Atlas",
    )
    assert load_workbook(result.output_path).sheetnames == ["captable"]


def test_combine_keeps_sources_when_requested(tmp_path: Path, monkeypatch):
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    a = _make_workbook(tmp_path / "cap.xlsx", {"Cap": [[1]]})
    result = combine_workbooks(
        sources={"captable": a},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
        delete_sources=False,
    )
    assert result.output_path.exists()
    assert a.exists()
    # The opt-out is not a warning condition — the merge itself was clean.
    assert result.sources_deleted is False
    assert result.warnings == ()
    assert [p.name for p in result.kept_sources] == ["cap.xlsx"]


def test_combine_raises_when_no_valid_sources(tmp_path: Path, monkeypatch):
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    with pytest.raises(ValueError):
        combine_workbooks(
            sources={"comps": None},
            output_dir=tmp_path,
            deliverable_type="pitch",
            deal_name="Atlas",
        )


# --- source-deletion gate (data-integrity) -----------------------------------


def test_relink_failure_keeps_sources(tmp_path: Path, monkeypatch, capsys):
    # A failed cross-tab relink leaves the combined workbook with broken /
    # externally-bound links; deleting the sources would make that permanent.
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    monkeypatch.setattr(
        workbook_aggregator, "_relink_cross_tab_openpyxl", lambda wb, tabs: False
    )
    a = _make_workbook(tmp_path / "fs.xlsx", {"FS": [[1]]})
    b = _make_workbook(tmp_path / "ltm.xlsx", {"LTM": [[2]]})

    result = combine_workbooks(
        sources={"financial-summary": a, "ltm-metrics": b},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )

    assert result.output_path.exists()  # the merged workbook is not lost
    assert result.relink_ok is False
    assert result.sources_deleted is False
    assert a.exists() and b.exists()
    assert {p.resolve() for p in result.kept_sources} == {a.resolve(), b.resolve()}
    assert any("relink FAILED" in w for w in result.warnings)
    assert "PRESERVED" in capsys.readouterr().err


def test_relink_cross_tab_openpyxl_reports_failure_instead_of_raising(monkeypatch, capsys):
    # The relink pass itself must never raise (the merged workbook would be
    # lost) — it traces to stderr and reports False to the caller.
    def boom(combined, skill_to_tab):
        raise KeyError("Worksheet ltm-metrics does not exist")

    monkeypatch.setattr(workbook_aggregator, "_relink_financial_summary_openpyxl", boom)
    assert workbook_aggregator._relink_cross_tab_openpyxl(Workbook(), {}) is False
    assert "cross-tab relink failed" in capsys.readouterr().err


def test_relink_cross_tab_openpyxl_reports_success():
    assert workbook_aggregator._relink_cross_tab_openpyxl(Workbook(), {}) is True


def test_win32_com_failure_fallback_keeps_sources(tmp_path: Path, monkeypatch, capsys):
    # On Windows the intended backend is COM. When it fails, the documented-lossy
    # openpyxl fallback still produces a combined workbook (CapIQ links / charts
    # gone), but the full-fidelity sources must survive for a retry.
    monkeypatch.setattr("workbook_aggregator.sys.platform", "win32")

    def com_boom(sources, output_path, theme):
        raise RuntimeError("Excel COM workbook aggregation failed: no Excel")

    monkeypatch.setattr(workbook_aggregator, "_combine_via_com", com_boom)
    a = _make_workbook(tmp_path / "cap.xlsx", {"Cap": [[1]]})

    result = combine_workbooks(
        sources={"captable": a},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )

    assert result.output_path.exists()
    assert result.backend == "openpyxl"
    assert result.degraded is True
    assert result.sources_deleted is False
    assert a.exists()
    assert any("DEGRADED" in w for w in result.warnings)
    err = capsys.readouterr().err
    assert "falling back" in err and "PRESERVED" in err


def test_external_refs_block_deletion(tmp_path: Path, monkeypatch):
    # A '[n]'-indexed formula in the combined workbook is a live binding to an
    # external workbook — exactly what a failed relink leaves behind. Deleting
    # the sources would turn it into a permanent #REF!/#N/A.
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    a = _make_workbook(
        tmp_path / "fs.xlsx",
        {"FS": [["=INDEX('[1]ltm-metrics'!$B:$B, MATCH(\"k\", '[1]ltm-metrics'!$A:$A, 0))"]]},
    )

    result = combine_workbooks(
        sources={"financial-summary": a},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )

    assert result.external_refs  # the external binding was detected post-merge
    assert result.sources_deleted is False
    assert a.exists()
    assert any("EXTERNAL workbook references" in w for w in result.warnings)


def test_find_external_workbook_refs_detects_bindings_and_link_parts(tmp_path: Path):
    from workbook_aggregator import _find_external_workbook_refs

    # Formula-level '[n]' binding.
    p = _make_workbook(tmp_path / "a.xlsx", {"S": [["='[1]ltm-metrics'!B10"]]})
    refs = _find_external_workbook_refs(p)
    assert len(refs) == 1 and "ltm-metrics" in refs[0]

    # Workbook-level xl/externalLinks part (no referencing formula needed).
    q = _make_workbook(tmp_path / "b.xlsx", {"S": [[1]]})
    with zipfile.ZipFile(q, "a") as z:
        z.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")
    assert "xl/externalLinks/externalLink1.xml" in _find_external_workbook_refs(q)


def test_find_external_workbook_refs_ignores_internal_formulas_and_strings(tmp_path: Path):
    from workbook_aggregator import _find_external_workbook_refs

    # Internal formulas, a '[1]' inside a string literal, and a bracketed cell
    # VALUE are all clean — no false positives on a self-contained workbook.
    p = _make_workbook(
        tmp_path / "c.xlsx",
        {"S": [["=SUM(A1:A2)", '=COUNTIF(A:A,"[1]x")', "literal [1] text"]]},
    )
    assert _find_external_workbook_refs(p) == []


def test_find_external_workbook_refs_fails_closed_on_unreadable_file(tmp_path: Path):
    from workbook_aggregator import _find_external_workbook_refs

    bad = tmp_path / "bad.xlsx"
    bad.write_text("not a zip")
    refs = _find_external_workbook_refs(bad)
    assert refs and "scan failed" in refs[0]  # a finding, not a green light


# --- cross-tab relink --------------------------------------------------------


def test_relink_wires_cross_tab_formulas(tmp_path: Path, monkeypatch):
    """In the combined workbook, the cap table's LTM cells link to the ltm-metrics
    bridge totals (found by label, since their rows are dynamic) and the ownership
    denominator links to the cap table's basic shares (millions -> full units)."""
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    cap = _make_anchored_captable(tmp_path / "cap.xlsx")
    ltm = _make_workbook(
        tmp_path / "ltm.xlsx",
        {
            "LTM Metrics": [
                ["LTM Revenue Overview", None],
                ["Segment A", 100],
                ["Total", 100],
                [None, None],
                ["LTM Revenue Bridge", None],
                ["Component", "Amount"],
                ["(+) FY", 90],
                ["(+) YTD", 30],
                ["(−) Prior YTD", 20],
                ["(=) LTM Revenue", "=B7+B8-B9"],      # row 10
                [None, None],
                ["LTM Adj. EBITDA Bridge", None],
                ["Component", "Amount"],
                ["(+) FY EBITDA", 40],
                ["(=) LTM Adj. EBITDA", "=B14"],        # row 15
            ]
        },
    )
    own = _make_anchored_ownership(tmp_path / "own.xlsx")

    result = combine_workbooks(
        sources={"captable": cap, "ltm-metrics": ltm, "ownership": own},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )
    wb = load_workbook(result.output_path)
    assert wb["captable"]["D47"].value == "='ltm-metrics'!B10*F7"
    assert wb["captable"]["D48"].value == "='ltm-metrics'!B15*F7"
    assert wb["ownership"]["F35"].value == "='captable'!F17*1000000"


def test_relink_is_noop_without_ltm_bridge_labels(tmp_path: Path, monkeypatch):
    # No "(=) LTM ..." labels in the ltm tab -> the cap table's D47/D48 are left
    # alone (the captable skill's own scalar/CapIQ values stand).
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    cap = _make_anchored_captable(tmp_path / "cap.xlsx")
    ltm = _make_workbook(tmp_path / "ltm.xlsx", {"LTM Metrics": [["Total", 100]]})
    result = combine_workbooks(
        sources={"captable": cap, "ltm-metrics": ltm},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )
    wb = load_workbook(result.output_path)
    assert wb["captable"]["D47"].value is None
    assert wb["captable"]["D48"].value is None


def test_financial_summary_ltm_link_resolves_after_merge(tmp_path: Path, monkeypatch):
    """The financial-summary tab's label-keyed LTM links (written before the
    ltm-metrics tab exists) survive the merge and target the renamed 'ltm-metrics'
    tab — so they resolve in the combined workbook, like the cap table's CapIQ
    formulas."""
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    from financial_summary_workbook import MetricSeries, build_financial_summary_workbook

    fs = build_financial_summary_workbook(
        company_name="SampleCo",
        currency_note="Figures in US$MM",
        period_note="FY = fiscal year; LTM = trailing twelve months",
        fiscal_labels=["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"],
        metrics=[
            MetricSeries("Revenue", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Revenue"),
            MetricSeries("Gross Profit", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Gross Profit"),
            MetricSeries("Adjusted EBITDA", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Adj. EBITDA"),
            MetricSeries("Net Income", "US$MM", [1, 2, 3, 4, 5], result_label="LTM Net Income"),
        ],
        output_dir=tmp_path,
        file_stem="fs",
    )
    ltm = _make_workbook(tmp_path / "ltm.xlsx", {"LTM Metrics": [["(=) LTM Revenue", 4520.0]]})

    result = combine_workbooks(
        sources={"financial-summary": fs, "ltm-metrics": ltm},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )
    wb = load_workbook(result.output_path)
    assert "financial-summary" in wb.sheetnames
    assert "ltm-metrics" in wb.sheetnames  # the link's target tab is present post-merge
    assert wb["financial-summary"]["G6"].value == (
        "=INDEX('ltm-metrics'!$B:$B, MATCH(\"(=) LTM Revenue\", 'ltm-metrics'!$A:$A, 0))"
    )


def test_internalize_external_sheet_ref_rewrites_workbook_index_and_path_forms():
    from workbook_aggregator import _internalize_external_sheet_ref

    idx_form = "=INDEX('[1]ltm-metrics'!$B:$B, MATCH(\"(=) LTM Revenue\", '[1]ltm-metrics'!$A:$A, 0))"
    assert _internalize_external_sheet_ref(idx_form, "ltm-metrics", "ltm-metrics") == (
        "=INDEX('ltm-metrics'!$B:$B, MATCH(\"(=) LTM Revenue\", 'ltm-metrics'!$A:$A, 0))"
    )
    # Full-path external form (Excel emits this once the source workbook is closed).
    path_form = "=INDEX('C:\\d\\[SampleCo - LTM Metrics.xlsx]ltm-metrics'!$B:$B, MATCH(\"k\", 'C:\\d\\[SampleCo - LTM Metrics.xlsx]ltm-metrics'!$A:$A, 0))"
    assert _internalize_external_sheet_ref(path_form, "ltm-metrics", "ltm-metrics") == (
        "=INDEX('ltm-metrics'!$B:$B, MATCH(\"k\", 'ltm-metrics'!$A:$A, 0))"
    )
    # A purely-internal ref (the openpyxl path) is left untouched.
    internal = "=INDEX('ltm-metrics'!$B:$B, MATCH(\"k\", 'ltm-metrics'!$A:$A, 0))"
    assert _internalize_external_sheet_ref(internal, "ltm-metrics", "ltm-metrics") == internal


def test_relink_financial_summary_openpyxl_internalizes_external_links(tmp_path: Path):
    from workbook_aggregator import _relink_financial_summary_openpyxl

    wb = Workbook()
    wb.remove(wb.active)
    fs = wb.create_sheet("financial-summary")
    wb.create_sheet("ltm-metrics")
    # Simulate the external link a COM copy would leave behind.
    fs["G6"] = "=INDEX('[1]ltm-metrics'!$B:$B, MATCH(\"(=) LTM Revenue\", '[1]ltm-metrics'!$A:$A, 0))"
    fs["B6"] = 5.0  # a non-LTM numeric cell is left alone
    _relink_financial_summary_openpyxl(wb, {"financial-summary": "financial-summary", "ltm-metrics": "ltm-metrics"})
    assert fs["G6"].value == (
        "=INDEX('ltm-metrics'!$B:$B, MATCH(\"(=) LTM Revenue\", 'ltm-metrics'!$A:$A, 0))"
    )
    assert fs["B6"].value == 5.0


# --- brand theme -------------------------------------------------------------

_INFOR_ACCENT1 = "0E213F"  # INFOR (New) accent1 — distinguishes it from Office's 4F81BD


def test_infor_theme_is_shipped():
    # The aggregator stamps templates/INFORFG.thmx on the combined workbook, so
    # the file must ship and parse to the INFOR colour scheme.
    theme = _default_theme_path()
    assert theme.exists(), f"shipped theme missing: {theme}"
    xml = _extract_theme_xml(theme)
    assert xml is not None
    assert _INFOR_ACCENT1.encode() in xml


def test_extract_theme_xml_returns_none_for_bad_file(tmp_path: Path):
    bad = tmp_path / "not-a-theme.thmx"
    bad.write_text("not a zip")
    assert _extract_theme_xml(bad) is None
    assert _extract_theme_xml(tmp_path / "missing.thmx") is None


def test_combined_workbook_carries_infor_theme(tmp_path: Path, monkeypatch):
    # End-to-end (openpyxl backend): a fresh openpyxl workbook would carry the
    # default Office theme; the aggregator must inject the INFOR theme instead.
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    a = _make_workbook(tmp_path / "cap.xlsx", {"Cap with Links": [["a"]]})
    result = combine_workbooks(
        sources={"captable": a},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )
    with zipfile.ZipFile(result.output_path) as z:
        theme_xml = z.read("xl/theme/theme1.xml").decode("utf-8", "replace")
    accent1 = re.search(r"<a:accent1>\s*<a:srgbClr val=\"([0-9A-Fa-f]{6})\"", theme_xml)
    assert accent1 is not None and accent1.group(1).upper() == _INFOR_ACCENT1


def test_theme_override_can_be_disabled(tmp_path: Path, monkeypatch):
    # Passing a non-existent theme path leaves the default Office theme in place
    # (no INFOR injection) — proving the stamp is driven by theme_path.
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    a = _make_workbook(tmp_path / "cap.xlsx", {"Cap with Links": [["a"]]})
    result = combine_workbooks(
        sources={"captable": a},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
        theme_path=tmp_path / "does-not-exist.thmx",
    )
    with zipfile.ZipFile(result.output_path) as z:
        theme_xml = z.read("xl/theme/theme1.xml").decode("utf-8", "replace")
    accent1 = re.search(r"<a:accent1>\s*<a:srgbClr val=\"([0-9A-Fa-f]{6})\"", theme_xml)
    assert accent1 is not None and accent1.group(1).upper() != _INFOR_ACCENT1


# --- LibreOffice recalc (P3.2) ----------------------------------------------


def test_combine_invokes_recalc_on_openpyxl_path(tmp_path: Path, monkeypatch):
    # The openpyxl merge path must hand the merged file to the LibreOffice recalc
    # step (so cross-tab links carry evaluated values for downstream stages).
    monkeypatch.setattr("workbook_aggregator.sys.platform", "linux")
    seen: list[Path] = []
    monkeypatch.setattr(
        workbook_aggregator,
        "_recalc_with_libreoffice",
        lambda p: seen.append(Path(p)) or True,
    )
    a = _make_workbook(tmp_path / "cap.xlsx", {"Cap": [[1]]})
    result = combine_workbooks(
        sources={"captable": a},
        output_dir=tmp_path,
        deliverable_type="pitch",
        deal_name="Atlas",
    )
    assert seen == [result.output_path]


def test_recalc_returns_false_when_libreoffice_absent(tmp_path: Path, monkeypatch):
    # No soffice/libreoffice on PATH -> graceful no-op; formulas left untouched.
    wb_path = _make_workbook(tmp_path / "c.xlsx", {"S": [["Total", "=1+1"]]})
    monkeypatch.setattr(workbook_aggregator.shutil, "which", lambda name: None)
    assert _recalc_with_libreoffice(wb_path) is False
    # The merged file is unchanged — the formula is preserved un-evaluated.
    assert load_workbook(wb_path)["S"]["B1"].value == "=1+1"


def test_recalc_replaces_file_and_preserves_formula_when_present(tmp_path: Path, monkeypatch):
    # With LibreOffice present, the recalced workbook replaces the merged file and
    # the formula survives (LibreOffice's xlsx export keeps formula strings).
    wb_path = _make_workbook(tmp_path / "c.xlsx", {"S": [["Total", "=1+1"]]})
    monkeypatch.setattr(
        workbook_aggregator.shutil,
        "which",
        lambda name: "/usr/bin/soffice" if name in ("soffice", "libreoffice") else None,
    )

    def fake_convert(soffice, src, out_fmt, out_dir):
        # Emulate LibreOffice recalc-on-load: keep the formula, add a sentinel cell
        # proving the recalced file is what ends up replacing the original.
        wb = load_workbook(src)
        wb["S"]["C1"] = "RECALCED"
        wb.save(Path(out_dir) / f"{Path(src).stem}.xlsx")

    monkeypatch.setattr("excel_to_powerpoint._soffice_convert", fake_convert)
    assert _recalc_with_libreoffice(wb_path) is True
    merged = load_workbook(wb_path)["S"]
    assert merged["C1"].value == "RECALCED"   # the recalced file replaced the original
    assert merged["B1"].value == "=1+1"        # formula preserved (auditability)


def test_recalc_returns_false_when_conversion_yields_nothing(tmp_path: Path, monkeypatch):
    # soffice present but produces no output file -> graceful False, file untouched.
    wb_path = _make_workbook(tmp_path / "c.xlsx", {"S": [["Total", "=1+1"]]})
    monkeypatch.setattr(workbook_aggregator.shutil, "which", lambda name: "/usr/bin/soffice")
    monkeypatch.setattr("excel_to_powerpoint._soffice_convert", lambda *a, **k: None)
    assert _recalc_with_libreoffice(wb_path) is False
    assert load_workbook(wb_path)["S"]["B1"].value == "=1+1"


@pytest.mark.parametrize(
    "strip_exc",
    [
        zipfile.BadZipFile("File is not a zip file"),
        OSError("disk full"),
        UnicodeDecodeError("utf-8", b"\x9c", 0, 1, "invalid start byte"),
    ],
    ids=["BadZipFile", "OSError", "UnicodeDecodeError"],
)
def test_recalc_never_raises_when_strip_step_fails(tmp_path: Path, monkeypatch, capsys, strip_exc):
    # The documented "never raises" contract: a failure inside the post-recalc
    # union-operator fix (BadZipFile / OSError / UnicodeDecodeError all escape a
    # RuntimeError-only catch) must degrade to the leave-formulas-un-evaluated
    # path — returning False AND rolling back the LibreOffice re-save, which may
    # still carry the '~' unions Excel would repair-strip on open.
    wb_path = _make_workbook(tmp_path / "c.xlsx", {"S": [["Total", "=1+1"]]})
    original = wb_path.read_bytes()
    monkeypatch.setattr(workbook_aggregator.shutil, "which", lambda name: "/usr/bin/soffice")

    def fake_convert(soffice, src, out_fmt, out_dir):
        wb = load_workbook(src)
        wb["S"]["C1"] = "RECALCED"
        wb.save(Path(out_dir) / f"{Path(src).stem}.xlsx")

    monkeypatch.setattr("excel_to_powerpoint._soffice_convert", fake_convert)

    def strip_boom(path):
        raise strip_exc

    monkeypatch.setattr(workbook_aggregator, "_strip_lo_union_operators", strip_boom)

    assert _recalc_with_libreoffice(wb_path) is False  # degraded, not raised
    assert wb_path.read_bytes() == original  # the pre-recalc merged file stands
    assert "LibreOffice recalc failed" in capsys.readouterr().err


# --- LibreOffice '~' union-operator fix ---------------------------------------


def test_excel_union_commas_rewrites_unions_outside_strings():
    from workbook_aggregator import _excel_union_commas

    assert (
        _excel_union_commas("_xlfn.PERCENTILE.INC((L10:L15~L20:L25~L30:L35),0.25)")
        == "_xlfn.PERCENTILE.INC((L10:L15,L20:L25,L30:L35),0.25)"
    )
    # '~' inside string literals is Excel's wildcard escape — it must survive.
    assert (
        _excel_union_commas('COUNTIF(A:A,"x~*")&(B1:B2~C1:C2)')
        == 'COUNTIF(A:A,"x~*")&(B1:B2,C1:C2)'
    )
    # Entity-escaped quotes (writer-dependent) delimit strings too.
    assert (
        _excel_union_commas("IF(A1=&quot;a~b&quot;,(A1~A2),0)")
        == "IF(A1=&quot;a~b&quot;,(A1,A2),0)"
    )


def test_strip_lo_union_operators_fixes_union_formulas(tmp_path: Path):
    from workbook_aggregator import _strip_lo_union_operators

    # Emulate what LibreOffice's recalc-on-load export does to the comps /
    # precedents quartile rows: the parenthesized multi-range union comes back
    # with LibreOffice's '~' operator, which Excel repairs away on open.
    wb_path = _make_workbook(
        tmp_path / "c.xlsx",
        {
            "comps": [["=_xlfn.PERCENTILE.INC((L10:L15~L20:L25~L30:L35),0.25)"]],
            "precedents": [['=IFERROR(_xlfn.PERCENTILE.INC((S8:S13~S17:S22),0.75),"n/a ~ ")']],
            "clean": [["=SUM(A1:A2)"]],
        },
    )
    assert _strip_lo_union_operators(wb_path) == 2  # two sheet parts rewritten

    wb = load_workbook(wb_path)
    assert wb["comps"]["A1"].value == "=_xlfn.PERCENTILE.INC((L10:L15,L20:L25,L30:L35),0.25)"
    # Union fixed, but the '~' inside the string literal survives.
    assert wb["precedents"]["A1"].value == '=IFERROR(_xlfn.PERCENTILE.INC((S8:S13,S17:S22),0.75),"n/a ~ ")'
    assert wb["clean"]["A1"].value == "=SUM(A1:A2)"


def test_strip_lo_union_operators_noop_without_tilde(tmp_path: Path):
    from workbook_aggregator import _strip_lo_union_operators

    wb_path = _make_workbook(tmp_path / "c.xlsx", {"S": [["=SUM(A1:A2)", 7]]})
    before = wb_path.read_bytes()
    assert _strip_lo_union_operators(wb_path) == 0
    assert wb_path.read_bytes() == before  # untouched — no rewrite churn


def test_recalc_strips_lo_union_operators(tmp_path: Path, monkeypatch):
    # End-to-end through _recalc_with_libreoffice: the emulated LibreOffice
    # export re-writes a union formula with '~'; the post-recalc fix restores ','.
    wb_path = _make_workbook(tmp_path / "c.xlsx", {"S": [["=AVERAGE(A1:A2)"]]})
    monkeypatch.setattr(workbook_aggregator.shutil, "which", lambda name: "/usr/bin/soffice")

    def fake_convert(soffice, src, out_fmt, out_dir):
        wb = load_workbook(src)
        wb["S"]["B1"] = "=_xlfn.PERCENTILE.INC((A1:A2~A4:A5),0.25)"
        wb.save(Path(out_dir) / f"{Path(src).stem}.xlsx")

    monkeypatch.setattr("excel_to_powerpoint._soffice_convert", fake_convert)
    assert _recalc_with_libreoffice(wb_path) is True
    assert (
        load_workbook(wb_path)["S"]["B1"].value
        == "=_xlfn.PERCENTILE.INC((A1:A2,A4:A5),0.25)"
    )
