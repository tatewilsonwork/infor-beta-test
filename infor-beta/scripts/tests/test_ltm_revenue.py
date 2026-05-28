"""Unit tests for the LTM revenue breakdown workbook helper."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from ltm_revenue import RevenueSegment, build_ltm_revenue_workbook


def _build(tmp_path: Path, **overrides) -> Path:
    kwargs = dict(
        company_name="SampleCo",
        period_label="LTM ended March 31, 2026",
        currency="US$MM",
        segmentation_basis="Service line",
        segments=[
            RevenueSegment("Cloud Services & Subscriptions", 1932.0),
            RevenueSegment("Customer Support", 1480.0),
            RevenueSegment("License", 360.0),
            RevenueSegment("Professional Service & Other", 290.0),
        ],
        output_dir=tmp_path,
    )
    kwargs.update(overrides)
    return build_ltm_revenue_workbook(**kwargs)


def test_build_ltm_revenue_workbook_writes_expected_file(tmp_path: Path):
    path = _build(tmp_path)
    assert path.exists()
    assert path.name == "SampleCo - LTM Revenue Breakdown.xlsx"


def test_workbook_uses_formulas_for_totals_and_percentages(tmp_path: Path):
    path = _build(tmp_path)
    wb = load_workbook(path)  # formulas, not computed values
    ws = wb.active

    # Header row 6, data rows 7-10, total row 11.
    assert ws["A6"].value == "Segment"
    assert ws["B7"].value == 1932.0
    # % of total references the total cell so Excel recomputes.
    assert ws["C7"].value == "=B7/B11"
    assert ws["B11"].value == "=SUM(B7:B10)"
    assert ws["C11"].value == "=SUM(C7:C10)"
    assert ws["A11"].value == "Total"


def test_segments_accept_plain_tuples(tmp_path: Path):
    path = _build(
        tmp_path,
        segmentation_basis="Geography",
        segments=[("Americas", 2500.0), ("EMEA", 1100.0), ("Asia-Pacific", 462.0)],
    )
    wb = load_workbook(path)
    ws = wb.active
    assert ws["A3"].value == "Segmentation: Geography"
    # 3 segments -> data rows 7-9, total row 10.
    assert ws["B10"].value == "=SUM(B7:B9)"


def test_empty_segments_raise(tmp_path: Path):
    with pytest.raises(ValueError):
        _build(tmp_path, segments=[])
