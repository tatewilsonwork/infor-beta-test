"""Unit tests for the cell-comment VIEW of a provenance record.

The cap-table skill hand-types web-sourced values into F7/F16, whose single
comment slot already holds the CapIQ refresh formula — the citation must be
APPENDED to that comment, never replace it (openpyxl allows one comment per
cell). Every assertion the v0.5.34 string-based helpers carried is still made
here, through the record-based entry point that replaced them: preserve the
existing comment and author, create one on a bare cell, and survive a save.

What is new is the direction of travel. `append_source_to_comment` takes a
`FigureSource`, not a sentence, so a citation cannot be written without a record
existing first — and `test_a_citation_string_is_rejected` pins that, because a
string was exactly what used to be accepted.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from comment_citations import append_source_to_comment, cite_cell
from provenance import FigureSource, ProvenanceLedger

_FX = FigureSource(url="https://example.com/fx", retrieved="2026-07-15")
_FILING = FigureSource(
    filing="FY2025 10-K", statement="Consolidated Statements of Operations"
)


def test_append_preserves_existing_formula_comment():
    # The cap-table case: F7 already carries the CapIQ refresh formula as its
    # comment. Appending the citation must keep the formula as the first line
    # and the original author.
    ws = Workbook().active
    ws["F7"] = 1.36
    ws["F7"].comment = Comment("=IQ_FX_RATE(...)", "INFOR")

    append_source_to_comment(ws["F7"], _FX)

    assert ws["F7"].comment.text == (
        "=IQ_FX_RATE(...)\nSource: https://example.com/fx — retrieved 2026-07-15"
    )
    assert ws["F7"].comment.author == "INFOR"


def test_append_creates_comment_when_cell_has_none():
    ws = Workbook().active
    ws["A1"] = 42.0
    append_source_to_comment(ws["A1"], FigureSource(url="https://example.com/quote",
                                                    retrieved="2026-07-15"))
    assert ws["A1"].comment is not None
    assert ws["A1"].comment.text == "Source: https://example.com/quote — retrieved 2026-07-15"
    assert ws["A1"].comment.author == "INFOR"


def test_filing_source_creates_comment_on_bare_cell():
    # The filing-sourced form used by financial-summary / ltm-metrics: no URL or
    # retrieval date, just the filing + statement the figure came from. The text
    # is byte-identical to what the v0.5.34 string helper produced.
    ws = Workbook().active
    ws["B6"] = 4520.0
    append_source_to_comment(ws["B6"], _FILING)
    assert ws["B6"].comment.text == (
        "Source: FY2025 10-K, Consolidated Statements of Operations"
    )
    assert ws["B6"].comment.author == "INFOR"


def test_filing_source_appends_to_existing_comment():
    ws = Workbook().active
    ws["B6"] = 4520.0
    ws["B6"].comment = Comment("Derived from operating income + D&A", "INFOR")
    append_source_to_comment(ws["B6"], FigureSource(filing="FY2025 10-K", statement="MD&A"))
    assert ws["B6"].comment.text == (
        "Derived from operating income + D&A\nSource: FY2025 10-K, MD&A"
    )
    assert ws["B6"].comment.author == "INFOR"


def test_page_number_is_part_of_the_citation():
    # The field the string convention could not enforce.
    ws = Workbook().active
    append_source_to_comment(
        ws["B6"], FigureSource(filing="FY2025 10-K", statement="Note 12", page=87)
    )
    assert ws["B6"].comment.text == "Source: FY2025 10-K, Note 12, p. 87"


def test_a_citation_string_is_rejected():
    ws = Workbook().active
    with pytest.raises(TypeError, match="FigureSource"):
        append_source_to_comment(ws["B6"], "FY2025 10-K, Consolidated Statements of Operations")


def test_the_pre_phase_g_three_argument_call_no_longer_works():
    # `append_source_to_comment(cell, url, retrieved)` was the v0.5.31 signature.
    # It has to fail loudly rather than quietly citing the wrong thing.
    ws = Workbook().active
    with pytest.raises(TypeError):
        append_source_to_comment(ws["F7"], "https://example.com/fx", "2026-07-15")


def test_cite_cell_writes_one_line_per_source():
    ws = Workbook().active
    ledger = ProvenanceLedger(stage="captable")
    entry = ledger.record("Share price", sources=[_FX, _FILING], value=42.18, location="captable!F16")
    cite_cell(ws["F16"], entry)
    assert ws["F16"].comment.text == (
        "Source: https://example.com/fx — retrieved 2026-07-15\n"
        "Source: FY2025 10-K, Consolidated Statements of Operations"
    )


def test_cite_cell_notes_the_derivation_of_a_computed_figure():
    # An LTM bridge total has no source of its own — its provenance is the
    # components' records — so the comment says how it was built.
    ws = Workbook().active
    ledger = ProvenanceLedger(stage="ltm-metrics")
    entry = ledger.record(
        "LTM Revenue",
        value=6062.0,
        derivation="FY2025 Revenue + Q3 2026 YTD Revenue − Q3 2025 YTD Revenue",
    )
    cite_cell(ws["B19"], entry)
    assert ws["B19"].comment.text == (
        "Derived: FY2025 Revenue + Q3 2026 YTD Revenue − Q3 2025 YTD Revenue"
    )


def test_appended_comment_round_trips_through_save(tmp_path: Path):
    # The skill saves the workbook after writing — the two-line comment must
    # survive an openpyxl save/load cycle intact.
    wb = Workbook()
    ws = wb.active
    ws["F16"] = 42.18
    ws["F16"].comment = Comment("=IQ_PRICE_CLOSE(...)", "INFOR")
    append_source_to_comment(
        ws["F16"], FigureSource(url="https://example.com/quote", retrieved="2026-07-15")
    )
    path = tmp_path / "cap.xlsx"
    wb.save(path)

    ws2 = load_workbook(path).active
    assert ws2["F16"].comment is not None
    assert ws2["F16"].comment.text == (
        "=IQ_PRICE_CLOSE(...)\nSource: https://example.com/quote — retrieved 2026-07-15"
    )
