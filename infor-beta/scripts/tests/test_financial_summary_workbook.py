"""Unit tests for the deal workbook's `financial-summary` tab."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from deal_workbook import TAB_FINANCIAL_SUMMARY, init_deal_workbook
from financial_summary_workbook import MetricSeries, build_financial_summary_workbook
from provenance import FigureSource, ProvenanceError, ProvenanceLedger

_FISCAL = ["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]


def _metrics() -> list[MetricSeries]:
    return [
        MetricSeries("Revenue", "US$MM", [3100.0, 3450.0, 3820.0, 4180.0, 4520.0],
                     result_label="LTM Revenue"),
        MetricSeries("Gross Profit", "US$MM", [1240.0, 1410.0, 1570.0, 1740.0, 1900.0],
                     result_label="LTM Gross Profit"),
        MetricSeries("Adjusted EBITDA", "US$MM", [820.0, 940.0, 1080.0, 1210.0, 1330.0],
                     result_label="LTM Adj. EBITDA"),
        # Non-flow metric: point-in-time latest value, no ltm-metrics bridge.
        MetricSeries("Combined Loan Balances", "US$MM", [9000.0, 10000.0, 11000.0, 12000.0, 12500.0],
                     ltm_value=12500.0),
    ]


def _build(tmp_path: Path, **overrides) -> Path:
    kwargs = dict(
        company_name="SampleCo",
        currency_note="Figures in US$MM unless noted",
        period_note="FY = fiscal year; LTM = trailing twelve months as of Q3 2026",
        fiscal_labels=_FISCAL,
        metrics=_metrics(),
        deal_workbook=init_deal_workbook(
            deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
        ),
    )
    kwargs.update(overrides)
    return build_financial_summary_workbook(**kwargs)


def _ws(path: Path):
    """The deal workbook's `financial-summary` tab."""
    return load_workbook(path)[TAB_FINANCIAL_SUMMARY]


def test_writes_the_tab_into_the_deal_workbook(tmp_path: Path):
    path = _build(tmp_path)
    assert path.exists()
    assert path.name == "pitch-Project Test.xlsx", "the deal owns one workbook"
    assert TAB_FINANCIAL_SUMMARY in load_workbook(path).sheetnames
    assert _ws(path).title == TAB_FINANCIAL_SUMMARY


def test_header_row_is_contiguous_period_axis(tmp_path: Path):
    ws = _ws(_build(tmp_path))
    assert ws["A5"].value == "Metric"
    # Five fiscal years oldest -> newest in B..F, LTM in G, Units in H.
    assert [ws.cell(row=5, column=c).value for c in range(2, 7)] == _FISCAL
    assert ws["G5"].value == "LTM"
    assert ws["H5"].value == "Units"


def test_metric_rows_are_numeric_with_units(tmp_path: Path):
    ws = _ws(_build(tmp_path))
    assert ws["A6"].value == "Revenue"
    assert [ws.cell(row=6, column=c).value for c in range(2, 7)] == [3100.0, 3450.0, 3820.0, 4180.0, 4520.0]
    assert all(isinstance(ws.cell(row=6, column=c).value, (int, float)) for c in range(2, 7))
    assert ws["H6"].value == "US$MM"


def test_flow_metric_ltm_links_label_keyed_to_ltm_metrics_tab(tmp_path: Path):
    ws = _ws(_build(tmp_path))
    # The LTM cell is a label-keyed lookup into the post-aggregation 'ltm-metrics' tab.
    assert ws["G6"].value == (
        "=INDEX('ltm-metrics'!$B:$B, MATCH(\"(=) LTM Revenue\", 'ltm-metrics'!$A:$A, 0))"
    )
    assert ws["G8"].value == (
        "=INDEX('ltm-metrics'!$B:$B, MATCH(\"(=) LTM Adj. EBITDA\", 'ltm-metrics'!$A:$A, 0))"
    )


def test_non_flow_metric_ltm_is_literal_value(tmp_path: Path):
    ws = _ws(_build(tmp_path))
    # Combined Loan Balances is the 4th metric -> row 9; its LTM is a point-in-time literal.
    assert ws["A9"].value == "Combined Loan Balances"
    assert ws["G9"].value == 12500.0


def test_suppression_drops_ltm_column_and_links_latest_fy(tmp_path: Path):
    ws = _ws(_build(tmp_path, show_ltm=False))
    # No LTM column: header runs Metric | FY1..FY5 | Units (Units lands in G).
    assert ws["G5"].value == "Units"
    row5 = [ws.cell(row=5, column=c).value for c in range(1, 8)]
    assert "LTM" not in row5
    # Flow metric: the most-recent FY cell (F) carries the ltm-metrics link.
    assert ws["F6"].value == (
        "=INDEX('ltm-metrics'!$B:$B, MATCH(\"(=) LTM Revenue\", 'ltm-metrics'!$A:$A, 0))"
    )
    # Non-flow metric keeps its literal most-recent FY value.
    assert ws["F9"].value == 12500.0


def test_value_cells_use_currency_number_format(tmp_path: Path):
    # v0.5.19: metric value cells carry the "$#,##0.0" currency format (FY values,
    # the flow-metric LTM link cell, and the non-flow LTM literal).
    ws = _ws(_build(tmp_path))
    expected = '$#,##0.0_);($#,##0.0);"--"'
    for c in range(2, 7):  # FY values B..F on the Revenue row
        assert ws.cell(row=6, column=c).number_format == expected
    assert ws["G6"].number_format == expected  # flow-metric LTM link cell
    assert ws["G9"].number_format == expected  # non-flow LTM literal


def test_no_merged_cells_in_data_block(tmp_path: Path):
    ws = _ws(_build(tmp_path))
    assert list(ws.merged_cells.ranges) == []


def test_accepts_plain_dicts(tmp_path: Path):
    path = _build(
        tmp_path,
        metrics=[
            {"label": "Revenue", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5], "result_label": "LTM Revenue"},
            {"label": "Gross Profit", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5], "result_label": "LTM Gross Profit"},
            {"label": "Adjusted EBITDA", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5], "result_label": "LTM Adj. EBITDA"},
            {"label": "Net Income", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5], "result_label": "LTM Net Income"},
        ],
    )
    ws = _ws(path)
    assert ws["A6"].value == "Revenue"


def test_wrong_metric_count_raises(tmp_path: Path):
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=_metrics()[:3])


def test_fiscal_value_count_mismatch_raises(tmp_path: Path):
    bad = _metrics()
    bad[0] = MetricSeries("Revenue", "US$MM", [1.0, 2.0, 3.0], result_label="LTM Revenue")
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=bad)


def test_non_flow_without_ltm_value_raises_when_ltm_shown(tmp_path: Path):
    bad = _metrics()
    bad[3] = MetricSeries("Combined Loan Balances", "US$MM", [1.0, 2.0, 3.0, 4.0, 5.0])
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=bad)


def test_combined_metric_fiscal_value_written_as_formula(tmp_path: Path):
    # A combined metric (loans + advances) passes each FY value as a formula of
    # its reported components, never pre-summed; openpyxl stores it as a formula.
    metrics = _metrics()
    metrics[3] = MetricSeries(
        "Combined Loan & Advance Bal.",
        "US$MM",
        ["=9000+800", "=10000+850", "=11000+900", "=12000+950", "=12500+1000"],
        ltm_value="=12500+1000",
    )
    ws = _ws(_build(tmp_path, metrics=metrics))
    assert ws["B9"].data_type == "f"
    assert ws["B9"].value == "=9000+800"
    # Non-flow combined LTM is likewise a formula.
    assert ws["G9"].data_type == "f"
    assert ws["G9"].value == "=12500+1000"


def test_bare_text_metric_value_raises(tmp_path: Path):
    bad = _metrics()
    bad[0] = MetricSeries("Revenue", "US$MM", [3100.0, 3450.0, 3820.0, 4180.0, "n/a"],
                          result_label="LTM Revenue")
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=bad)

# ─── Provenance records + the comments rendered from them (Phase G) ──────────
#
# v0.5.34 wrote the citation string straight onto the cell and that string WAS the
# record. Phase G inverts it: a `FigureSource` goes in, a record comes out into the
# run's ledger, and the comment is a view of the record. These tests assert both
# halves — the comment text is unchanged, and a record now exists for it.


def _fy_source(year: int) -> FigureSource:
    return FigureSource(filing=f"FY{year} 10-K", statement="Consolidated Statements of Operations")


def test_per_value_sources_written_as_cell_comments(tmp_path: Path):
    metrics = _metrics()
    metrics[0] = MetricSeries(
        "Revenue", "US$MM", [3100.0, 3450.0, 3820.0, 4180.0, 4520.0],
        result_label="LTM Revenue",
        sources=[
            _fy_source(2022),
            None,  # a value without a citation skips its comment
            _fy_source(2023),
            _fy_source(2024),
            _fy_source(2025),
        ],
    )
    ws = _ws(_build(tmp_path, metrics=metrics))
    assert ws["B6"].comment is not None
    assert ws["B6"].comment.text == "Source: FY2022 10-K, Consolidated Statements of Operations"
    assert ws["C6"].comment is None
    assert ws["F6"].comment.text == "Source: FY2025 10-K, Consolidated Statements of Operations"


def test_non_flow_ltm_source_written_on_ltm_cell(tmp_path: Path):
    metrics = _metrics()
    metrics[3] = MetricSeries(
        "Combined Loan Balances", "US$MM",
        [9000.0, 10000.0, 11000.0, 12000.0, 12500.0],
        ltm_value=12500.0,
        ltm_source=FigureSource(filing="Q3 2026 10-Q", statement="Consolidated Balance Sheets"),
    )
    ws = _ws(_build(tmp_path, metrics=metrics))
    assert ws["G9"].comment is not None
    assert ws["G9"].comment.text == "Source: Q3 2026 10-Q, Consolidated Balance Sheets"
    # Flow metrics' LTM link cells carry no comment (provenance lives on the
    # ltm-metrics bridge components).
    assert ws["G6"].comment is None


def test_sources_survive_dict_metrics(tmp_path: Path):
    metrics = [
        {"label": "Revenue", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5],
         "result_label": "LTM Revenue", "sources": [FigureSource(filing="FY 10-K")] * 5},
        {"label": "Gross Profit", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5],
         "result_label": "LTM Gross Profit"},
        {"label": "Adjusted EBITDA", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5],
         "result_label": "LTM Adj. EBITDA"},
        {"label": "Net Income", "units": "US$MM", "fiscal_values": [1, 2, 3, 4, 5],
         "result_label": "LTM Net Income"},
    ]
    ws = _ws(_build(tmp_path, metrics=metrics))
    assert ws["B6"].comment.text == "Source: FY 10-K"
    assert ws["B7"].comment is None


def test_source_count_mismatch_raises(tmp_path: Path):
    metrics = _metrics()
    metrics[0] = MetricSeries(
        "Revenue", "US$MM", [3100.0, 3450.0, 3820.0, 4180.0, 4520.0],
        result_label="LTM Revenue", sources=[_fy_source(2025)],  # 1 source for 5 values
    )
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=metrics)


def test_a_citation_string_is_rejected(tmp_path: Path):
    metrics = _metrics()
    metrics[0] = MetricSeries(
        "Revenue", "US$MM", [3100.0, 3450.0, 3820.0, 4180.0, 4520.0],
        result_label="LTM Revenue",
        sources=["FY2025 10-K, Consolidated Statements of Operations"] * 5,
    )
    with pytest.raises(ProvenanceError, match="FigureSource"):
        _build(tmp_path, metrics=metrics)


def test_page_number_reaches_the_comment(tmp_path: Path):
    metrics = _metrics()
    metrics[0] = MetricSeries(
        "Revenue", "US$MM", [3100.0, 3450.0, 3820.0, 4180.0, 4520.0],
        result_label="LTM Revenue",
        sources=[FigureSource(filing="FY2025 10-K", statement="Consolidated Statements "
                                                              "of Operations", page=61)] * 5,
    )
    ws = _ws(_build(tmp_path, metrics=metrics))
    assert ws["B6"].comment.text == (
        "Source: FY2025 10-K, Consolidated Statements of Operations, p. 61"
    )


def test_every_fiscal_value_lands_in_the_ledger(tmp_path: Path):
    ledger = ProvenanceLedger(stage="financial-summary")
    metrics = _metrics()
    metrics[0] = MetricSeries(
        "Revenue", "US$MM", [3100.0, 3450.0, 3820.0, 4180.0, 4520.0],
        result_label="LTM Revenue",
        sources=[_fy_source(y) for y in (2021, 2022, 2023, 2024, 2025)],
    )
    _build(tmp_path, metrics=metrics, provenance=ledger)
    by_figure = {f.figure: f for f in ledger.figures}
    assert by_figure["Revenue FY2021"].location == "financial-summary!B6"
    assert by_figure["Revenue FY2021"].value == 3100.0
    assert by_figure["Revenue FY2021"].units == "US$MM"
    assert by_figure["Revenue FY2025"].sources == (_fy_source(2025),)
    assert all(f.stage == "financial-summary" for f in ledger.figures)


def test_flow_metric_ltm_cell_is_recorded_as_a_link_not_a_source(tmp_path: Path):
    # The link's provenance is the bridge it points at, whose components carry the
    # filing records — so the record names the chain rather than inventing a source.
    ledger = ProvenanceLedger(stage="financial-summary")
    _build(tmp_path, provenance=ledger)
    ltm = next(f for f in ledger.figures if f.figure == "Revenue LTM")
    assert ltm.sources == ()
    assert "ltm-metrics" in ltm.derivation and "(=) LTM Revenue" in ltm.derivation
    assert ltm.location == "financial-summary!G6"


def test_suppressed_ltm_column_records_the_link_on_the_latest_fy_cell(tmp_path: Path):
    ledger = ProvenanceLedger(stage="financial-summary")
    _build(tmp_path, show_ltm=False, provenance=ledger)
    entry = next(f for f in ledger.figures if f.figure.startswith("Revenue FY2025 (LTM =="))
    assert entry.location == "financial-summary!F6"
    assert "suppressed" in entry.derivation


# ─── Configurable metric count (v0.5.26): 8 metrics = the two-slide deck ─────


def _metrics8() -> list[MetricSeries]:
    return _metrics() + [
        MetricSeries("Operating Income", "US$MM", [620.0, 710.0, 815.0, 920.0, 1010.0],
                     result_label="LTM Operating Income"),
        MetricSeries("Net Income", "US$MM", [410.0, 470.0, 540.0, 610.0, 680.0],
                     result_label="LTM Net Income"),
        MetricSeries("Free Cash Flow", "US$MM", [350.0, 400.0, 460.0, 520.0, 580.0],
                     result_label="LTM Free Cash Flow"),
        # Non-flow ratio metric.
        MetricSeries("Return on Equity", "%", [0.12, 0.13, 0.14, 0.15, 0.16],
                     ltm_value=0.16),
    ]


def test_eight_metrics_write_rows_6_to_13(tmp_path: Path):
    path = _build(tmp_path, metrics=_metrics8(), metric_count=8)
    ws = _ws(path)
    labels = [ws.cell(row=r, column=1).value for r in range(6, 14)]
    assert labels == [m.label for m in _metrics8()]
    assert ws.cell(row=14, column=1).value is None  # block ends after row 13


def test_metric_count_must_match_metrics(tmp_path: Path):
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=_metrics8(), metric_count=4)
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=_metrics(), metric_count=8)


def test_metric_count_must_be_positive_multiple_of_four(tmp_path: Path):
    with pytest.raises(ValueError):
        _build(tmp_path, metrics=_metrics8()[:6], metric_count=6)


# ─── Phase D: the LTM link is internal and its target row really exists ──────
# Before Phase D this link was `#N/A` until the aggregator merged the two
# standalone files, and v0.5.16 had to re-bind it after the COM merge turned it
# into an EXTERNAL reference that Excel's `.Formula` getter masked as internal.
# Both tabs now live in one file from the start, so the link is an ordinary
# internal reference and there is nothing to re-bind.


def test_ltm_link_targets_a_row_that_exists_on_the_ltm_metrics_tab(tmp_path: Path):
    from deal_workbook import TAB_LTM_METRICS
    from ltm_metrics import BridgeComponent, build_ltm_metrics_workbook

    deal = init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )
    build_ltm_metrics_workbook(
        company_name="SampleCo",
        period_label="LTM ended March 31, 2026",
        currency="US$MM",
        segmentation_basis="Service line",
        segments=[("Cloud", 1932.0), ("Support", 1480.0)],
        revenue_bridge=[
            BridgeComponent("FY2025 Revenue", 5400.0),
            BridgeComponent("Q3 2026 YTD Revenue", 3050.0),
            BridgeComponent("Q3 2025 YTD Revenue", 2388.0, subtract=True),
        ],
        deal_workbook=deal,
    )
    build_financial_summary_workbook(
        company_name="SampleCo",
        currency_note="Figures in US$MM unless noted",
        period_note="FY = fiscal year; LTM = trailing twelve months",
        fiscal_labels=_FISCAL,
        metrics=_metrics(),
        deal_workbook=deal,
    )

    wb = load_workbook(deal)
    assert TAB_LTM_METRICS in wb.sheetnames and TAB_FINANCIAL_SUMMARY in wb.sheetnames

    link = wb[TAB_FINANCIAL_SUMMARY]["G6"].value  # Revenue's LTM cell
    assert link.startswith("=INDEX('ltm-metrics'!")
    assert "[" not in link, f"the link must be internal, not an external ref: {link}"

    # The `(=) LTM Revenue` row the MATCH keys off is really on that tab.
    labels = [
        c.value
        for row in wb[TAB_LTM_METRICS].iter_rows(min_col=1, max_col=1)
        for c in row
        if isinstance(c.value, str)
    ]
    assert "(=) LTM Revenue" in labels, labels
