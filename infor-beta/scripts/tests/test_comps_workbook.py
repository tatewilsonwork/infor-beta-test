"""Tests for the comps workbook builder (deterministic; no Excel/CapIQ needed).

Capital IQ cannot be refreshed in this environment, so the builder only writes
the three input fields — vertical labels, CapIQ tickers, descriptions. These
tests pin that write and, critically, guard that the template's CapIQ array
formulas survive the openpyxl round-trip: they are what populate every metric
column once the analyst refreshes Capital IQ, so dropping them would gut the
workbook.
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from deal_workbook import TAB_COMPS, init_deal_workbook
from openpyxl.worksheet.formula import ArrayFormula

from comps_workbook import CompCompany, Vertical, build_comps_workbook

PLUGIN_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = PLUGIN_ROOT / "templates" / "INFOR Comps Template.xlsx"


def _deal(tmp_path: Path) -> Path:
    """A fresh deal workbook — the `comps` tab arrives with it, from the template."""
    return init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )


def _verticals():
    return [
        Vertical("Cloud ERP", [CompCompany(f"NYSE:T{i}", f"Desc {i}") for i in range(6)]),
        Vertical("Payments", [CompCompany(f"NasdaqGS:P{i}", f"Pay {i}") for i in range(6)]),
        Vertical("Capital Markets Tech", [CompCompany(f"TSX:C{i}", f"Cap {i}") for i in range(6)]),
    ]


def _build(tmp_path: Path, verticals=None) -> Path:
    return build_comps_workbook(
        deal_workbook=_deal(tmp_path),
        verticals=verticals if verticals is not None else _verticals(),
    )


def _formula_text(cell):
    value = cell.value
    return value.text if isinstance(value, ArrayFormula) else value


def test_writes_labels_tickers_and_descriptions(tmp_path: Path):
    ws = load_workbook(_build(tmp_path))[TAB_COMPS]  # formulas preserved (no data_only)

    # Vertical labels -> D9 / D19 / D29.
    assert ws["D9"].value == "Cloud ERP"
    assert ws["D19"].value == "Payments"
    assert ws["D29"].value == "Capital Markets Tech"

    # Six tickers per block in column B; descriptions in column AA.
    assert ws["B10"].value == "NYSE:T0"
    assert ws["B15"].value == "NYSE:T5"
    assert ws["AA10"].value == "Desc 0"
    assert ws["B20"].value == "NasdaqGS:P0"
    assert ws["B25"].value == "NasdaqGS:P5"
    assert ws["AA20"].value == "Pay 0"
    assert ws["B30"].value == "TSX:C0"
    assert ws["B35"].value == "TSX:C5"
    assert ws["AA35"].value == "Cap 5"


def test_capiq_array_formulas_preserved(tmp_path: Path):
    """The metric columns are CapIQ array formulas keyed off column B; the build
    must not drop them (openpyxl round-trip regression guard)."""
    ws = load_workbook(_build(tmp_path))[TAB_COMPS]

    # Company-name lookup at the top of each block keys off $B<row>.
    for cell in ("D10", "D20", "D30"):
        text = _formula_text(ws[cell])
        assert isinstance(text, str) and "SPG" in text and "$B" in text, cell

    # Group-average / global-statistic formulas survive too.
    assert "AVERAGE" in str(ws["L17"].value)
    assert "MEDIAN" in str(ws["L41"].value)


def test_only_intended_cells_written(tmp_path: Path):
    """Unused company rows, spacer rows, and unfilled verticals stay as-is."""
    ws = load_workbook(_build(tmp_path, [Vertical("Solo", [CompCompany("NYSE:AAA", "Only one")])]))[TAB_COMPS]

    assert ws["D9"].value == "Solo"
    assert ws["B10"].value == "NYSE:AAA"
    assert ws["B11"].value is None       # unused company row stays blank
    assert ws["B16"].value is None       # spacer row before the group average stays blank
    # Unfilled verticals keep the template's placeholder labels.
    assert ws["D19"].value == "[Group #2]"
    assert ws["D29"].value == "[Group #3]"


def test_empty_description_left_blank(tmp_path: Path):
    ws = load_workbook(_build(tmp_path, [Vertical("X", [CompCompany("NYSE:A")])]))[TAB_COMPS]
    assert ws["B10"].value == "NYSE:A"
    assert ws["AA10"].value is None


def test_too_many_companies_raises(tmp_path: Path):
    bad = [Vertical("X", [CompCompany(f"NYSE:T{i}", "d") for i in range(7)])]
    with pytest.raises(ValueError, match="per vertical"):
        _build(tmp_path, bad)


def test_too_many_verticals_raises(tmp_path: Path):
    bad = [Vertical(f"V{i}", [CompCompany("NYSE:A", "d")]) for i in range(4)]
    with pytest.raises(ValueError, match="vertical blocks"):
        _build(tmp_path, bad)


def test_empty_ticker_raises(tmp_path: Path):
    with pytest.raises(ValueError, match="empty ticker"):
        _build(tmp_path, [Vertical("X", [CompCompany("   ", "d")])])


def test_overlong_description_raises(tmp_path: Path):
    bad = [Vertical("X", [CompCompany("NYSE:A", "z" * 51)])]
    with pytest.raises(ValueError, match="max 50"):
        _build(tmp_path, bad)


def test_accepts_plain_dicts(tmp_path: Path):
    out = build_comps_workbook(
        deal_workbook=_deal(tmp_path),
        verticals=[{"name": "Dict Vertical", "companies": [{"ticker": "NYSE:ZZZ", "description": "From a dict"}]}],
    )
    ws = load_workbook(out)[TAB_COMPS]
    assert ws["D9"].value == "Dict Vertical"
    assert ws["B10"].value == "NYSE:ZZZ"
    assert ws["AA10"].value == "From a dict"


# ─── Provenance: what this tab asserts is the peer SET ───────────────────────
#
# The metric columns are CapIQ's own array formulas, so the claim this tab makes
# is which companies are comparable and what they do. A run's ledger held 70
# records and `comps` contributed zero, which is why the comps slide's figures
# were unauditable — the obligation existed only where a SKILL.md wrote it down.


def test_each_peer_is_recorded_with_the_source_for_its_inclusion(tmp_path: Path):
    from provenance import FigureSource, ProvenanceLedger

    ledger = ProvenanceLedger(stage="comps")
    build_comps_workbook(
        deal_workbook=_deal(tmp_path),
        verticals=[
            Vertical("Cloud ERP", [
                CompCompany("NYSE:AAAA", "Cloud ERP for mid-market",
                            source=FigureSource(url="https://example.com/aaaa",
                                                retrieved="2026-07-29")),
                CompCompany("NasdaqGS:BBBB", "Vertical SaaS",
                            source=FigureSource(filing="FY2025 10-K",
                                                statement="Item 1: Business", page=7)),
            ]),
        ],
        provenance=ledger,
    )

    assert [f.figure for f in ledger.figures] == [
        "Comps peer — NYSE:AAAA (Cloud ERP)",
        "Comps peer — NasdaqGS:BBBB (Cloud ERP)",
    ]
    assert [f.location for f in ledger.figures] == [f"{TAB_COMPS}!AA10", f"{TAB_COMPS}!AA11"]
    assert ledger.figures[0].citation_lines == (
        "https://example.com/aaaa — retrieved 2026-07-29",
    )
    assert ledger.figures[1].citation_lines == ("FY2025 10-K, Item 1: Business, p. 7",)


def test_a_peers_description_cell_comment_is_rendered_from_its_record(tmp_path: Path):
    from provenance import FigureSource, ProvenanceLedger

    out = build_comps_workbook(
        deal_workbook=_deal(tmp_path),
        verticals=[Vertical("Cloud ERP", [
            CompCompany("NYSE:AAAA", "Cloud ERP for mid-market",
                        source=FigureSource(url="https://example.com/aaaa", retrieved="2026-07-29")),
        ])],
        provenance=ProvenanceLedger(stage="comps"),
    )
    comment = load_workbook(out)[TAB_COMPS]["AA10"].comment
    assert comment is not None
    assert comment.text.endswith("Source: https://example.com/aaaa — retrieved 2026-07-29")


def test_a_citation_string_instead_of_a_source_record_raises(tmp_path: Path):
    from provenance import ProvenanceError

    with pytest.raises(ProvenanceError, match="no longer a source record"):
        build_comps_workbook(
            deal_workbook=_deal(tmp_path),
            verticals=[Vertical("X", [CompCompany("NYSE:A", "d", source="their website, 2026")])],
        )


def test_the_builder_still_works_with_no_ledger(tmp_path: Path):
    # Direct invocation, and every existing caller: provenance is opt-in at the
    # call site, required by the stage's SKILL.md.
    ws = load_workbook(_build(tmp_path))[TAB_COMPS]
    assert ws["B10"].value == "NYSE:T0"
