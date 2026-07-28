"""Both deck assemblers, fed the cap table the PIPELINE actually produces.

The coverage gap this closes
---------------------------
Three regressions of one class have now shipped green: a stale skip guard
(v0.5.36), the `summary_at` marker slip (v0.5.40), and the cap-table sheet name
(v0.5.45, fixed here). The common mechanism is not the individual mistake — it
is that the assembler tests built their companion-workbook fixtures from a
**source template** (`INFOR Cap Table Template.xlsx`, sheet `Cap with Links`, or
a hand-rolled `Workbook()` titled to match) while the conductor passes a
**pipeline artefact** (the deal workbook, tab `captable`). Any rename inside the
produced artefact therefore left the suite green and production broken.

So every test here builds its input with `init_deal_workbook` — the same call
the conductor makes at deal-init — and asserts the assembler consumes it. That
is the part that generalises past this one bug: Phases F/G/H can rename or move
a tab and these fail, where a source-template fixture would not notice.

Renderer-free on purpose
------------------------
The failure mode is the layout pre-flight and the sheet name handed to the
renderer, both of which happen *before* any conversion. So these spy on
`insert_excel_into_placeholder` and stop the run there rather than gating on
LibreOffice — no skip, and the assertion is on the exact argument that was
wrong. The full render against a deal workbook is covered by the insertion tests
in `test_earnings_update_assembler` / `test_slide_library_poc`, whose fixtures
this change moved onto `init_deal_workbook` too.
"""

from pathlib import Path

import pytest

import earnings_update_assembler
import pitch_deck_assembler
from deal_workbook import (
    TAB_CAPTABLE,
    TabSpec,
    init_deal_workbook,
    write_tab,
)
from template_layout import (
    CAP_TABLE_PICTURE_NAMES,
    CAP_TABLE_PICTURE_RANGE,
    NAME_CAP_OUTPUT_CCY,
    TemplateLayoutError,
    resolve_name_cell,
)

from tests.test_earnings_update_assembler import _assemble_sample_deck
from tests.test_slide_library_poc import _assemble, _sample_content


class _InsertionReached(Exception):
    """The spy reached the Excel insertion; the assembler need not finish.

    Everything under test — the workbook pre-flight, the picture-range
    resolution, and the sheet name passed to the renderer — has already happened
    by the time this is raised, and letting the deck finish would need a
    renderer for no extra coverage.
    """


def _spy_on_insertion(monkeypatch, module) -> list[dict]:
    """Replace `module.insert_excel_into_placeholder` with a recording stub."""
    calls: list[dict] = []

    def _spy(**kwargs):
        calls.append(kwargs)
        raise _InsertionReached

    monkeypatch.setattr(module, "insert_excel_into_placeholder", _spy)
    return calls


def _deal_workbook(tmp_path: Path, deliverable_type: str, *, currency: str | None = None) -> Path:
    """The deal's workbook, exactly as the conductor creates it at deal-init.

    No `stamp_defined_names` and no synthetic `Workbook()`: the point is that the
    `infor_` names travel with the real `captable` tab through
    `INFOR Deal Workbook Template.xlsx`, so the fixture has to be that file.
    """
    path = init_deal_workbook(
        deal_dir=tmp_path, deliverable_type=deliverable_type, deal_name="Project Test"
    )
    if currency is not None:
        def _fill(_wb, ws):
            ws[resolve_name_cell(ws, NAME_CAP_OUTPUT_CCY)] = currency

        write_tab(
            path,
            TAB_CAPTABLE,
            TabSpec(write=_fill, verify_names=(NAME_CAP_OUTPUT_CCY,)),
        )
    return path


# ─── The deal workbook is not the source template ────────────────────────────


def test_deal_workbook_renames_the_cap_table_sheet():
    """The premise of the bug, pinned so the fix cannot be read as cosmetic.

    `tools/build_deal_workbook_template.py` renames `Cap with Links` to
    `captable`. If these two ever became equal the tests below would still pass
    while proving nothing, so assert they differ.
    """
    from template_layout import CAP_TABLE_SOURCE_SHEET

    assert TAB_CAPTABLE != CAP_TABLE_SOURCE_SHEET
    assert CAP_TABLE_SOURCE_SHEET == "Cap with Links"
    assert TAB_CAPTABLE == "captable"


def test_shipped_deal_workbook_template_carries_the_captable_tab_and_its_names():
    """The repro from the v0.5.45 report, as a test.

    `verify_workbook_names` against the shipped deal-workbook template is the
    exact call both assemblers make; it raised for every build that supplied a
    cap table.
    """
    from deal_workbook import DEAL_WORKBOOK_TEMPLATE
    from template_layout import verify_workbook_names

    template = Path(__file__).resolve().parents[2] / "templates" / DEAL_WORKBOOK_TEMPLATE
    verify_workbook_names(template, sheet=TAB_CAPTABLE, names=CAP_TABLE_PICTURE_NAMES)


# ─── Earnings update ─────────────────────────────────────────────────────────


def test_earnings_assembler_takes_its_cap_table_from_the_deal_workbook(
    tmp_path: Path, monkeypatch
):
    """The earnings deck stage, fed what `earnings-update.yaml` really passes.

    Fails before v0.5.45 at the pre-flight: `verify_workbook_names(...,
    sheet='Cap with Links')` on a workbook whose tabs are
    `['captable', 'ltm-metrics', ...]`.
    """
    deal_workbook = _deal_workbook(tmp_path, "earnings-update")
    calls = _spy_on_insertion(monkeypatch, earnings_update_assembler)

    with pytest.raises(_InsertionReached):
        _assemble_sample_deck(tmp_path, captable_workbook_path=deal_workbook)

    assert len(calls) == 1
    assert calls[0]["sheet_name"] == TAB_CAPTABLE
    assert calls[0]["workbook_path"] == deal_workbook
    # Resolved through `infor_cap_picture_range` on the real tab, not the fallback.
    assert calls[0]["source_range"] == CAP_TABLE_PICTURE_RANGE


def test_earnings_cap_table_picture_range_resolves_on_the_deal_workbook(tmp_path: Path):
    """The range comes from the tab's own defined name.

    `resolve_workbook_range` returns its `fallback` for a sheet it cannot find,
    so this passed by accident while the sheet name was wrong. Asserting the tab
    carries the name is what makes the value meaningful rather than coincidental.
    """
    from template_layout import defined_name_ref
    from openpyxl import load_workbook

    deal_workbook = _deal_workbook(tmp_path, "earnings-update")
    wb = load_workbook(deal_workbook)
    try:
        assert defined_name_ref(wb[TAB_CAPTABLE], "infor_cap_picture_range") is not None
    finally:
        wb.close()
    assert (
        earnings_update_assembler.cap_table_picture_range(deal_workbook)
        == CAP_TABLE_PICTURE_RANGE
    )


# ─── Pitch ───────────────────────────────────────────────────────────────────


def test_pitch_assembler_takes_its_cap_table_from_the_deal_workbook(
    tmp_path: Path, monkeypatch
):
    """The pitch deck stage, fed what `pitch.yaml` really passes.

    Fails before v0.5.45 at the pre-flight, which additionally verifies
    `infor_cap_output_ccy` for the footnote letter.
    """
    deal_workbook = _deal_workbook(tmp_path, "pitch", currency="USD")
    calls = _spy_on_insertion(monkeypatch, pitch_deck_assembler)

    with pytest.raises(_InsertionReached):
        _assemble(
            tmp_path,
            _sample_content(),
            captable_workbook_path=deal_workbook,
            converge=False,
        )

    assert len(calls) == 1
    assert calls[0]["sheet_name"] == TAB_CAPTABLE
    assert calls[0]["workbook_path"] == deal_workbook
    assert calls[0]["source_range"] == CAP_TABLE_PICTURE_RANGE


def test_pitch_footnote_currency_reads_the_deal_workbooks_captable_tab(tmp_path: Path):
    """The client-facing `[x]$MM` footnote letter, off the real tab.

    This is the site that did not raise: it read
    `wb[CAP_TABLE_SHEET] if ... else wb.active`, and since Phase D the condition
    was always false — so the currency came from whichever tab happened to be
    active (`captable`, by luck of tab order) rather than from the cap table by
    name. A USD filer must read 'US'.
    """
    usd = _deal_workbook(tmp_path / "usd", "pitch", currency="USD")
    assert pitch_deck_assembler._output_currency_letter(usd) == "US"

    cad = _deal_workbook(tmp_path / "cad", "pitch", currency="CAD")
    assert pitch_deck_assembler._output_currency_letter(cad) == "C"

    gbp = _deal_workbook(tmp_path / "gbp", "pitch", currency="GBP")
    assert pitch_deck_assembler._output_currency_letter(gbp) == "GBP"


def test_pitch_footnote_currency_fails_loudly_without_a_captable_tab(tmp_path: Path):
    """No `wb.active` fallback: a workbook with no `captable` tab must raise.

    The deleted fallback made a missing tab indistinguishable from a present
    one, and mislabelling a US filer's figures `C$MM` on a client deck is worse
    than failing the deck stage.
    """
    from openpyxl import load_workbook

    path = _deal_workbook(tmp_path, "pitch")
    wb = load_workbook(path)
    try:
        del wb[TAB_CAPTABLE]
        wb.save(path)
    finally:
        wb.close()

    with pytest.raises(TemplateLayoutError, match=TAB_CAPTABLE):
        pitch_deck_assembler._output_currency_letter(path)


# ─── Drift lock ──────────────────────────────────────────────────────────────


def test_assemblers_never_address_a_source_template_sheet_name():
    """Neither assembler may reach for a `*_SOURCE_SHEET` constant.

    Both hold built artefacts only, so every sheet they name is a deal-workbook
    tab. This is the scan that would have caught v0.5.45 at review: the constants
    are now spelled apart on purpose, and this keeps them apart.
    """
    scripts = Path(__file__).resolve().parents[1]
    for module in ("earnings_update_assembler.py", "pitch_deck_assembler.py"):
        source = (scripts / module).read_text(encoding="utf-8")
        offenders = [
            line.strip()
            for line in source.splitlines()
            if "_SOURCE_SHEET" in line and not line.lstrip().startswith("#")
        ]
        assert not offenders, (
            f"{module} addresses a source template's sheet name: {offenders}. "
            f"It holds a deal workbook — use deal_workbook.TAB_*."
        )
