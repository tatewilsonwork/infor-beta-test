"""Tests for the deal's single workbook (Phase D step 1).

Three things are load-bearing and each is pinned here:

(a) **Every `infor_` defined name resolves on the deal workbook.** The Phase C
    writers all address cells through those names, and they are worksheet-scoped
    so they must travel with their sheet into the pre-assembled template. If they
    do not, every writer breaks — `resolve_name_range` raises rather than falling
    back, so the failure is loud, but it should be caught here and not three
    waves into a live run.
(b) **Writes are serialized.** A wave dispatches stages concurrently, and
    openpyxl rewrites the whole file on save, so two unsynchronised writers would
    lose a tab. The lock is exercised with real threads.
(c) **Template fidelity survived the assembly** — CapIQ `_xll.` formulas, the
    cap table's CapIQ-refresh cell comments, and the internal
    `'Bloomberg Output'!` reference that a naive per-sheet copy would have turned
    into an external link.

The shipped template is never modified: every test copies it under `tmp_path`.
"""

from __future__ import annotations

import threading
from pathlib import Path

import pytest
from openpyxl import load_workbook

import deal_workbook as dw
from deal_workbook import (
    DEAL_WORKBOOK_TEMPLATE,
    TAB_BLOOMBERG_OUTPUT,
    TAB_CAPTABLE,
    TAB_COMPS,
    TAB_FINANCIAL_SUMMARY,
    TAB_LTM_METRICS,
    TAB_ORDER,
    TAB_OWNERSHIP,
    TAB_PRECEDENTS,
    DealWorkbookError,
    TabSpec,
    deal_workbook_path,
    init_deal_workbook,
    workbook_filename,
    write_tab,
)
from template_layout import TEMPLATE_NAMED_RANGES, TemplateLayoutError, defined_name_ref, normalize_ref

PLUGIN_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = PLUGIN_ROOT / "templates" / DEAL_WORKBOOK_TEMPLATE

#: The source template each deal-workbook tab was copied from, so the expected
#: `infor_` names come from the same registry `add_template_named_ranges.py`
#: stamps rather than from a second hand-written list.
TAB_SOURCES: dict[str, tuple[str, str]] = {
    TAB_CAPTABLE: ("INFOR Cap Table Template.xlsx", "Cap with Links"),
    TAB_COMPS: ("INFOR Comps Template.xlsx", "Comps"),
    TAB_OWNERSHIP: ("INFOR Ownership Template.xlsx", "Ownership"),
    TAB_BLOOMBERG_OUTPUT: ("INFOR Ownership Template.xlsx", "Bloomberg Output"),
    TAB_PRECEDENTS: ("INFOR Precedents Template.xlsx", "Precedents"),
}


@pytest.fixture
def deal(tmp_path) -> Path:
    """A pitch deal workbook in a temp deal directory."""
    return init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )


# ─── (a) the defined names survive into the deal workbook ────────────────────


def test_shipped_deal_workbook_template_exists():
    assert TEMPLATE.is_file(), (
        f"{DEAL_WORKBOOK_TEMPLATE} is missing — build it with "
        f"tools/build_deal_workbook_template.py"
    )


def test_prep_tool_never_opens_a_shipped_template_in_excel():
    """The build must open STAGED COPIES, never the originals.

    The repo lives in a OneDrive-synced folder, and Excel 365 enables AutoSave
    by default for cloud-backed files — it writes the file back regardless of
    `Close(SaveChanges=False)`. An early revision of the tool opened the shipped
    templates directly and silently re-saved all four through Excel
    (`sharedStrings.xml` / `calcChain.xml` appeared, `styles.xml` and
    `printerSettings1.bin` changed), undoing Phase C's byte-level preservation.
    A source-level lock, because reproducing it needs Excel plus OneDrive.
    """
    tool = Path(__file__).resolve().parents[3] / "tools" / "build_deal_workbook_template.py"
    body = tool.read_text(encoding="utf-8")
    build_from = body.split("def _build_from(")[1].split("\ndef ")[0]
    assert "staged[template]" in build_from, "the build must open the staged copy"
    assert "TEMPLATES / template" not in build_from, (
        "_build_from opens a shipped template directly; Excel AutoSave will "
        "re-save it. Stage a copy outside OneDrive instead."
    )


def test_every_infor_name_resolves_on_the_deal_workbook(deal):
    """THE step-1 exit criterion: all 27 `infor_` names resolve, on the right cell.

    Not just "the name exists" — it must point at the same address it pointed at
    in its source template, because a name that survived the copy pointing
    somewhere else is worse than one that vanished.
    """
    wb = load_workbook(deal)
    checked = 0
    for tab, (source_template, source_sheet) in TAB_SOURCES.items():
        expected = TEMPLATE_NAMED_RANGES[source_template][source_sheet]
        assert expected, f"no registry entry for {source_template}/{source_sheet}"
        assert tab in wb.sheetnames, f"tab {tab!r} missing from the deal workbook"
        ws = wb[tab]
        for name, target in expected.items():
            ref = defined_name_ref(ws, name)
            assert ref is not None, f"{tab}: defined name {name!r} does not resolve"
            assert normalize_ref(ref) == normalize_ref(target), (
                f"{tab}: {name!r} resolves to {ref}, source template had {target}"
            )
            checked += 1
    assert checked == 27, f"expected 27 infor_ names across the tabs, checked {checked}"


def test_writers_resolve_names_through_the_deal_workbook(deal):
    """A writer's own resolution path works on a tab, not just a bare template."""
    from template_layout import NAME_FX_RATE, NAME_SHARE_PRICE, resolve_name_cell

    wb = load_workbook(deal)
    ws = wb[TAB_CAPTABLE]
    assert resolve_name_cell(ws, NAME_FX_RATE) == "F7"
    assert resolve_name_cell(ws, NAME_SHARE_PRICE) == "F16"


def test_write_tab_verifies_declared_names(deal):
    """A tab that lost a name fails at the first write, naming it."""
    wb = load_workbook(deal)
    del wb[TAB_CAPTABLE].defined_names["infor_fx_rate"]
    wb.save(deal)

    with pytest.raises(TemplateLayoutError, match="infor_fx_rate"):
        write_tab(deal, TAB_CAPTABLE, TabSpec(write=lambda _wb, _ws: None,
                                              verify_names=("infor_fx_rate",)))


# ─── (c) template fidelity ───────────────────────────────────────────────────


def test_capiq_formulas_survived_into_the_deal_workbook(deal):
    wb = load_workbook(deal)
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                text = value if isinstance(value, str) else getattr(value, "text", None)
                if isinstance(text, str) and "_xll." in text:
                    count += 1
    assert count > 500, f"only {count} CapIQ _xll. formulas survived; expected 560"


def test_captable_capiq_refresh_comments_survived(deal):
    """F7 / F16 carry the CapIQ refresh formula as a comment (the v0.5.3 design)."""
    ws = load_workbook(deal)[TAB_CAPTABLE]
    for cell in ("F7", "F16"):
        comment = ws[cell].comment
        assert comment is not None, f"{cell} lost its CapIQ refresh comment"
        assert "SPG(" in comment.text


def test_ownership_bloomberg_reference_stayed_internal(deal):
    """The pair was copied in one operation, so this is not an external link."""
    ws = load_workbook(deal)[TAB_OWNERSHIP]
    assert ws["B68"].value == "='Bloomberg Output'!C14"


def test_deal_workbook_has_no_external_references(deal):
    """An external link would prompt on open and leak the author's OneDrive path."""
    import re
    import zipfile

    with zipfile.ZipFile(deal) as zf:
        parts = [n for n in zf.namelist() if n.startswith("xl/externalLinks/")]
        assert parts == [], f"external-link parts present: {parts}"
        for name in zf.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                text = zf.read(name).decode("utf-8", "replace")
                assert not re.search(r"\[\d+\]", text), f"{name} has an external ref"


# ─── init ────────────────────────────────────────────────────────────────────


def test_workbook_filename_matches_the_old_combined_name():
    assert workbook_filename("pitch", "Project Atlas") == "pitch-Project Atlas.xlsx"
    assert (
        workbook_filename("earnings-update", "Project Atlas")
        == "earningsupdate-Project Atlas.xlsx"
    )


def test_pitch_init_carries_every_template_tab(deal):
    names = load_workbook(deal).sheetnames
    for tab in (TAB_CAPTABLE, TAB_COMPS, TAB_OWNERSHIP, TAB_BLOOMBERG_OUTPUT, TAB_PRECEDENTS):
        assert tab in names


def test_earnings_update_init_drops_the_pitch_only_tabs(tmp_path):
    path = init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="earnings-update", deal_name="Project EU"
    )
    names = load_workbook(path).sheetnames
    assert TAB_CAPTABLE in names
    for tab in (TAB_COMPS, TAB_PRECEDENTS, TAB_OWNERSHIP, TAB_BLOOMBERG_OUTPUT):
        assert tab not in names, f"{tab} should not be in an earnings-update workbook"


def test_init_is_idempotent_and_does_not_discard_written_tabs(tmp_path):
    path = init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )
    write_tab(path, TAB_LTM_METRICS, TabSpec(create=True,
                                             write=lambda _wb, ws: ws.cell(1, 1, "kept")))
    again = init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )
    assert again == path
    assert load_workbook(path)[TAB_LTM_METRICS]["A1"].value == "kept"


def test_init_overwrite_starts_clean(tmp_path):
    path = init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )
    write_tab(path, TAB_LTM_METRICS, TabSpec(create=True,
                                             write=lambda _wb, ws: ws.cell(1, 1, "gone")))
    init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test", overwrite=True
    )
    assert TAB_LTM_METRICS not in load_workbook(path).sheetnames


def test_deal_workbook_path_is_in_the_deal_directory(tmp_path):
    assert deal_workbook_path(tmp_path, "pitch", "Project X").parent == tmp_path


# ─── write_tab ───────────────────────────────────────────────────────────────


def test_write_tab_writes_into_a_template_tab(deal):
    write_tab(deal, TAB_CAPTABLE, TabSpec(write=lambda _wb, ws: ws.__setitem__("F3", "NYSE:X")))
    assert load_workbook(deal)[TAB_CAPTABLE]["F3"].value == "NYSE:X"


def test_write_tab_creates_an_authored_tab_in_canonical_order(deal):
    write_tab(deal, TAB_LTM_METRICS, TabSpec(create=True, write=lambda _wb, ws: None))
    write_tab(deal, TAB_FINANCIAL_SUMMARY, TabSpec(create=True, write=lambda _wb, ws: None))
    names = load_workbook(deal).sheetnames
    ordered = [n for n in names if n in TAB_ORDER]
    assert ordered == sorted(ordered, key=TAB_ORDER.index), names


def test_write_tab_replaces_a_created_tab_on_rerun(deal):
    write_tab(deal, TAB_LTM_METRICS, TabSpec(create=True,
                                             write=lambda _wb, ws: ws.cell(1, 1, "first")))
    write_tab(deal, TAB_LTM_METRICS, TabSpec(create=True,
                                             write=lambda _wb, ws: ws.cell(2, 1, "second")))
    ws = load_workbook(deal)[TAB_LTM_METRICS]
    assert ws["A1"].value is None and ws["A2"].value == "second"


def test_write_tab_refuses_an_unknown_tab_without_create(deal):
    with pytest.raises(DealWorkbookError, match="not in"):
        write_tab(deal, "nonesuch", TabSpec(write=lambda _wb, _ws: None))


def test_write_tab_refuses_a_missing_workbook(tmp_path):
    with pytest.raises(DealWorkbookError, match="does not exist"):
        write_tab(tmp_path / "nope.xlsx", TAB_CAPTABLE, TabSpec(write=lambda _wb, _ws: None))


def test_write_tab_sees_sibling_tabs(deal):
    """A producer can read another stage's tab — this is what killed the merge."""
    write_tab(deal, TAB_LTM_METRICS, TabSpec(create=True,
                                             write=lambda _wb, ws: ws.cell(1, 2, 123.4)))
    seen: list = []
    write_tab(
        deal,
        TAB_FINANCIAL_SUMMARY,
        TabSpec(create=True, write=lambda wb, _ws: seen.append(wb[TAB_LTM_METRICS]["B1"].value)),
    )
    assert seen == [123.4]


# ─── (b) serialization ───────────────────────────────────────────────────────


def test_concurrent_write_tab_calls_do_not_lose_a_tab(deal):
    """Six threads, six tabs, one file: every tab must survive.

    Without the lock each writer loads, mutates and saves the WHOLE workbook, so
    the last save wins and the other five tabs vanish. This is the real shape of
    a conductor wave — `comps`, `precedents` and `financial-summary` all reach
    for the deal workbook at once.
    """
    tabs = [f"thread-{i}" for i in range(6)]
    errors: list[BaseException] = []

    def writer(tab: str) -> None:
        try:
            write_tab(deal, tab, TabSpec(create=True,
                                         write=lambda _wb, ws: ws.cell(1, 1, tab)))
        except BaseException as exc:  # noqa: BLE001
            errors.append(exc)

    threads = [threading.Thread(target=writer, args=(tab,)) for tab in tabs]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=180)

    assert errors == [], f"writers raised: {errors}"
    wb = load_workbook(deal)
    for tab in tabs:
        assert tab in wb.sheetnames, f"{tab} was lost; have {wb.sheetnames}"
        assert wb[tab]["A1"].value == tab


def test_lock_is_released_when_the_writer_raises(deal):
    """A failing producer must not wedge the workbook for every later stage."""

    def boom(_wb, _ws):
        raise ValueError("producer failed")

    with pytest.raises(ValueError, match="producer failed"):
        write_tab(deal, TAB_CAPTABLE, TabSpec(write=boom))

    lock = deal.with_suffix(deal.suffix + ".lock")
    assert not lock.exists(), "the lock outlived the failed write"
    write_tab(deal, TAB_CAPTABLE, TabSpec(write=lambda _wb, ws: ws.__setitem__("F3", "ok")))
    assert load_workbook(deal)[TAB_CAPTABLE]["F3"].value == "ok"


def test_stale_lock_is_broken(deal, monkeypatch):
    """A lock from a killed writer is broken rather than waited out."""
    lock = deal.with_suffix(deal.suffix + ".lock")
    lock.write_text("99999")
    monkeypatch.setattr(dw, "_LOCK_STALE_S", -1.0)
    write_tab(deal, TAB_CAPTABLE, TabSpec(write=lambda _wb, ws: ws.__setitem__("F3", "after")))
    assert load_workbook(deal)[TAB_CAPTABLE]["F3"].value == "after"


def test_a_held_lock_times_out_with_a_useful_message(deal, monkeypatch):
    lock = deal.with_suffix(deal.suffix + ".lock")
    lock.write_text("12345")
    monkeypatch.setattr(dw, "_LOCK_TIMEOUT_S", 0.3)
    with pytest.raises(DealWorkbookError, match="timed out"):
        write_tab(deal, TAB_CAPTABLE, TabSpec(write=lambda _wb, _ws: None))
    lock.unlink()
