"""Tests for the precedents workbook builder (deterministic; no Excel/CapIQ needed).

The builder writes deal identity, the chosen family's source-FX $ metric inputs,
disclosed multiples (straight over the ratio formula), source hyperlinks, and the
3-letter HQ code. These tests pin that write and, critically, guard that the
template's live formulas survive the openpyxl round-trip: the column-C CapIQ FX
array formula, the ``=+I*C`` TEV conversion, the per-row ratio formulas in S–Z,
and the group / global statistic rows are what make the table compute once the
analyst refreshes Capital IQ, so dropping them would gut the workbook.
"""

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from deal_workbook import TAB_PRECEDENTS, init_deal_workbook
from openpyxl.worksheet.formula import ArrayFormula

from precedents_workbook import (
    PrecedentGroup,
    PrecedentTransaction,
    build_precedents_workbook,
)

PLUGIN_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = PLUGIN_ROOT / "templates" / "INFOR Precedents Template.xlsx"


def _deal(tmp_path: Path) -> Path:
    """A fresh deal workbook — the `precedents` tab arrives with it."""
    return init_deal_workbook(
        deal_dir=tmp_path, deliverable_type="pitch", deal_name="Project Test"
    )


def _operating_tx(**overrides) -> PrecedentTransaction:
    base = dict(
        input_currency="USD",
        announce_date=date(2025, 5, 6),
        target="Target Operating Co",
        acquiror="Acquiror LP",
        tev=1250.0,
        hq_country="USA",
        revenue_ltm=880.0,
        ebitda_ltm=195.0,
        tev_link="https://www.sec.gov/deal.htm",
        revenue_link="https://investors.example.com/pr",
        ebitda_link="https://investors.example.com/pr",
    )
    base.update(overrides)
    return PrecedentTransaction(**base)


def _financial_tx(**overrides) -> PrecedentTransaction:
    base = dict(
        input_currency="CAD",
        announce_date=date(2024, 9, 1),
        target="Target Bank",
        acquiror="BigBank Inc",
        tev=4000.0,
        hq_country="CAN",
        net_income_ltm=320.0,
        book_value=2100.0,
        tangible_book_value=1800.0,
        tev_link="https://sedarplus.ca/x",
        net_income_link="https://sedarplus.ca/x",
        book_value_link="https://sedarplus.ca/x",
        tangible_book_value_link="https://sedarplus.ca/x",
    )
    base.update(overrides)
    return PrecedentTransaction(**base)


def _build(tmp_path: Path, groups=None, **kwargs) -> Path:
    if groups is None:
        groups = [
            PrecedentGroup("Operating Peers", [_operating_tx()]),
            PrecedentGroup("Financial Peers", [_financial_tx()]),
        ]
    return build_precedents_workbook(
        deal_workbook=_deal(tmp_path),
        groups=groups,
        **kwargs,
    )


def _ftext(cell):
    value = cell.value
    return value.text if isinstance(value, ArrayFormula) else value


def _as_date(value):
    # openpyxl reads a date-formatted cell back as datetime; normalize to date.
    return value.date() if isinstance(value, datetime) else value


def test_writes_identity_metrics_and_groups(tmp_path: Path):
    ws = load_workbook(_build(tmp_path))[TAB_PRECEDENTS]

    # Output currency (C2) defaults to USD; group labels at E7 / E16.
    assert ws["C2"].value == "USD"
    assert ws["E7"].value == "Operating Peers"
    assert ws["E16"].value == "Financial Peers"

    # Group #1 row 8 — operating deal identity + chosen-family metric inputs.
    assert ws["B8"].value == "USD"
    assert _as_date(ws["E8"].value) == date(2025, 5, 6)
    assert ws["F8"].value == "Target Operating Co"
    assert ws["G8"].value == "Acquiror LP"
    assert ws["I8"].value == 1250.0
    assert ws["AI8"].value == "USA"
    assert ws["K8"].value == 880.0   # revenue LTM
    assert ws["O8"].value == 195.0   # adj. EBITDA LTM
    assert ws["H8"].value is None    # column H is left empty (HQ goes in AI)

    # Group #2 row 17 — financial deal: net income + book value + TBV.
    assert ws["B17"].value == "CAD"
    assert ws["AI17"].value == "CAN"
    assert ws["M17"].value == 320.0   # net income LTM
    assert ws["Q17"].value == 2100.0  # book value
    assert ws["R17"].value == 1800.0  # tangible book value


def test_disclosed_multiple_overwrites_ratio_formula(tmp_path: Path):
    """A disclosed multiple replaces the S–Z ratio formula with a literal; an
    un-disclosed ratio keeps its template formula so it computes on refresh."""
    groups = [PrecedentGroup("Ops", [_operating_tx(ev_ebitda_ltm=12.5, ebitda_ltm=None)])]
    ws = load_workbook(_build(tmp_path, groups))[TAB_PRECEDENTS]

    assert ws["U8"].value == 12.5                       # EV/EBITDA literal
    assert ws["O8"].value is None                        # no $ EBITDA written
    assert _ftext(ws["S8"]) == '=IF(K8="","n/a ",(K8*C8)/J8)'  # EV/Rev formula intact


def test_source_links_hyperlinked_unused_cleared(tmp_path: Path):
    ws = load_workbook(_build(tmp_path))[TAB_PRECEDENTS]

    # Provided links -> "Link" text + hyperlink target.
    assert ws["AB8"].value == "Link"
    assert ws["AB8"].hyperlink is not None
    assert ws["AB8"].hyperlink.target == "https://www.sec.gov/deal.htm"
    assert ws["AC8"].hyperlink.target == "https://investors.example.com/pr"

    # Unused link cells in a populated row are cleared (no dangling "Link").
    assert ws["AE8"].value is None   # net income link — not an operating metric
    assert ws["AF8"].value is None   # book value link
    assert ws["AG8"].value is None   # tangible book value link


def test_live_formulas_preserved(tmp_path: Path):
    """FX / TEV / ratio / statistic formulas must survive the openpyxl round-trip."""
    ws = load_workbook(_build(tmp_path))[TAB_PRECEDENTS]

    fx = _ftext(ws["C8"])
    assert isinstance(fx, str) and "SPG" in fx and "$C$2" in fx
    assert _ftext(ws["J8"]) == "=+I8*C8"
    assert _ftext(ws["U8"]).startswith("=IF(O8")    # EV/EBITDA ratio (not overwritten here)
    assert _ftext(ws["W8"]).startswith("=IF(M8")    # P/E ratio
    assert _ftext(ws["Z8"]).startswith("=IF(R8")    # P/TBV ratio
    assert "AVERAGE(S8:S13)" in _ftext(ws["S14"])    # group average
    assert "AVERAGE(S17:S22,S8:S13)" in _ftext(ws["S25"])  # global average
    assert "MEDIAN" in _ftext(ws["S26"])
    assert "PERCENTILE" in _ftext(ws["S29"])


def test_only_intended_cells_written(tmp_path: Path):
    """Unused rows stay blank and an unfilled group keeps its placeholder."""
    groups = [PrecedentGroup("Solo", [_operating_tx()])]
    ws = load_workbook(_build(tmp_path, groups))[TAB_PRECEDENTS]

    assert ws["E7"].value == "Solo"
    assert ws["F8"].value == "Target Operating Co"
    assert ws["F9"].value is None          # unused row in the filled group
    assert ws["E16"].value == "[Group #2]"  # unfilled group keeps placeholder


def test_iso_date_string_coerced_via_dict(tmp_path: Path):
    out = build_precedents_workbook(
        deal_workbook=_deal(tmp_path),
        groups=[{"name": "Dict Group", "transactions": [{
            "input_currency": "gbp",
            "announce_date": "2023-11-20",
            "target": "Dict Target",
            "acquiror": "Dict Acquiror",
            "tev": 500,
            "hq_country": "gbr",
            "ev_revenue_ltm": 3.1,
        }]}],
        output_currency="cad",
    )
    ws = load_workbook(out)[TAB_PRECEDENTS]
    assert ws["C2"].value == "CAD"          # lower-case input upper-cased
    assert ws["B8"].value == "GBP"
    assert ws["AI8"].value == "GBR"
    assert _as_date(ws["E8"].value) == date(2023, 11, 20)  # ISO string -> date
    assert ws["S8"].value == 3.1            # EV/Revenue multiple literal


def test_too_many_groups_raises(tmp_path: Path):
    bad = [PrecedentGroup(f"G{i}", [_operating_tx()]) for i in range(3)]
    with pytest.raises(ValueError, match="group blocks"):
        _build(tmp_path, bad)


def test_too_many_transactions_raises(tmp_path: Path):
    bad = [PrecedentGroup("G", [_operating_tx() for _ in range(7)])]
    with pytest.raises(ValueError, match="per group"):
        _build(tmp_path, bad)


def test_bad_hq_code_raises(tmp_path: Path):
    with pytest.raises(ValueError, match="hq_country must be a 3-letter"):
        _build(tmp_path, [PrecedentGroup("G", [_operating_tx(hq_country="US")])])


def test_bad_currency_code_raises(tmp_path: Path):
    with pytest.raises(ValueError, match="input_currency must be a 3-letter"):
        _build(tmp_path, [PrecedentGroup("G", [_operating_tx(input_currency="Dollar")])])


def test_non_positive_tev_raises(tmp_path: Path):
    with pytest.raises(ValueError, match="tev must be a positive number"):
        _build(tmp_path, [PrecedentGroup("G", [_operating_tx(tev=0)])])


def test_non_numeric_metric_raises(tmp_path: Path):
    with pytest.raises(ValueError, match="revenue_ltm must be a number"):
        _build(tmp_path, [PrecedentGroup("G", [_operating_tx(revenue_ltm="lots")])])


def test_bad_link_raises(tmp_path: Path):
    with pytest.raises(ValueError, match="tev_link must be an http"):
        _build(tmp_path, [PrecedentGroup("G", [_operating_tx(tev_link="sec.gov/x")])])


def test_tev_only_deal_raises(tmp_path: Path):
    """A deal with a TEV but no metric or multiple is rejected — it would add an
    empty row, and every kept deal must yield at least one multiple."""
    bare = PrecedentTransaction(
        input_currency="USD", announce_date=date(2025, 1, 1),
        target="No-Metric Co", acquiror="Buyer LP", tev=900.0, hq_country="USA",
    )
    with pytest.raises(ValueError, match="at least one disclosed multiple"):
        _build(tmp_path, [PrecedentGroup("G", [bare])])


def test_deal_with_only_a_disclosed_multiple_is_allowed(tmp_path: Path):
    """A disclosed multiple alone (no $ metric) satisfies the value requirement."""
    tx = PrecedentTransaction(
        input_currency="USD", announce_date=date(2025, 1, 1),
        target="Multiple-Only Co", acquiror="Buyer LP", tev=900.0, hq_country="USA",
        ev_ebitda_ltm=11.0,
    )
    ws = load_workbook(_build(tmp_path, [PrecedentGroup("G", [tx])]))[TAB_PRECEDENTS]
    assert ws["U8"].value == 11.0


def test_target_and_acquiror_use_palatino_9(tmp_path: Path):
    """Target / acquiror are written Palatino 9, not the template's stray Calibri 11."""
    ws = load_workbook(_build(tmp_path))[TAB_PRECEDENTS]
    for ref in ("F8", "G8"):
        assert ws[ref].font.name == "Palatino Linotype", f"{ref} must be Palatino"
        assert ws[ref].font.size == 9, f"{ref} must be 9 pt"


def test_unknown_dict_field_raises(tmp_path: Path):
    with pytest.raises(ValueError, match="unknown transaction field"):
        build_precedents_workbook(
            deal_workbook=_deal(tmp_path),
            groups=[{"name": "G", "transactions": [{
                "input_currency": "USD", "announce_date": "2025-01-01",
                "target": "T", "acquiror": "A", "tev": 100, "hq_country": "USA",
                "ev_revenue": 3.0,  # not a real field (should be ev_revenue_ltm/ntm)
            }]}],
            )
