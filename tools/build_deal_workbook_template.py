"""Assemble `templates/INFOR Deal Workbook Template.xlsx` from the four templates.

One-time prep tooling for Phase D, kept in the repo because it has to be
re-runnable: any time an analyst re-saves one of the four source templates from
Excel, run this again (after `add_template_named_ranges.py`) so the deal-workbook
template picks the change up. Like that tool it is **not** shipped plugin code,
so Phase D's COM deletion does not apply to it.

Usage (from the repo root):

    python tools/build_deal_workbook_template.py --check    # report, write nothing
    python tools/build_deal_workbook_template.py            # build the template
    python tools/build_deal_workbook_template.py --verify    # + full verification

Why a single pre-assembled template
-----------------------------------
Phase D gives the deal ONE workbook from stage one, so the template content has
to arrive as *tabs* rather than as five standalone files a later stage merges.
Nothing in the runtime can do that copy: openpyxl cannot move a sheet between
workbooks, and a cell-by-cell copy is exactly the lossy merge Phase D deletes.
So the copy happens once, here, and the runtime does a `shutil.copyfile`.

Why Excel COM, when `add_template_named_ranges.py` argued against it
--------------------------------------------------------------------
That tool's objection was specific and correct: it would have meant "opening a
Capital IQ workbook with the Cap IQ Office Tools add-in loaded and letting it
touch `CIQWBGuid` / `CIQWBInfo` on the way past."

That objection does not apply to an **automation-started** Excel, and Phase D's
step-0 probe measured why. A `DispatchEx("Excel.Application")` instance does not
load the `OPEN=`-registered SNL add-ins at all — the probe's first take was
discarded precisely because of it: `_xll.…SPG` read `#NAME?` in every subject,
including the pristine control, until the probe relaunched `excel.exe` normally.
Here that is the wanted property, not a defect:

  - the Cap IQ add-in never engages, so it cannot touch `CIQWBGuid` /
    `CIQWBInfo` / the `_xll.` formulas;
  - Excel still performs the sheet copy at full native fidelity — array
    formulas, styles, comments, merged cells, page setup, and the sheet-scoped
    `infor_` names, which travel with their sheet.

The alternatives were both worse. Zip-level surgery — this repo's preferred
conservative move — cannot be byte-for-byte across workbooks: sheet XML
addresses styles and strings by *index* into `xl/styles.xml` and
`xl/sharedStrings.xml`, so grafting a sheet means reindexing every cell in it,
which is neither conservative nor cheap. And a LibreOffice round-trip is
disqualified outright: it is what rewrote parenthesised range unions with `~`
and made Excel repair-strip 44 formulas on open (v0.5.23).

Why the brand theme has to be stamped
-------------------------------------
`Workbooks.Add()` hands back a workbook carrying **Office's default theme**, and
Excel resolves a copied sheet's theme-indexed colours against the DESTINATION
theme. The four source templates are full of them, so the assembled workbook
inherits INFOR's slot *numbers* and Office's slot *colours* — accent1 rendering
as Office 2024's `156082` instead of INFOR navy `0E213F`, in every tab, with
nothing in any cell wrong. The aggregator Phase D deleted stamped
`templates/INFORFG.thmx` on its way out; nothing replaced it, so v0.5.41 through
v0.5.51 shipped an Office-palette deal workbook and every deal copied it.

So the build applies `INFORFG.thmx` and then reads the saved theme part back to
confirm it took, falling back to grafting the thmx's theme part into the saved
zip. The two are equivalent by measurement: applying `INFORFG.thmx` by hand in
Excel produces an `xl/theme/theme1.xml` byte-identical to the thmx's
`theme/theme/theme1.xml` (md5 `2eb1697978066f80ed1a4ad3b6c64aec`).

`--verify` is the check that matters, and mirrors `add_template_named_ranges.py`'s
Excel oracle: every `infor_` name resolves to the same target it did in its
source template, every `_xll.` CapIQ formula's text is unchanged, the global
CapIQ names survive, the palette is INFOR's, there are no external-workbook
references back to the sources, and Excel opens the result with no repair record.
The palette check runs under plain `--check` too, with no Excel anywhere, so a
future re-run cannot quietly ship an Office-themed template again.

Deliberately preserved
----------------------
The comps template's 1,246 legacy artefacts and the cap table's Capital IQ
names are carried through untouched, exactly as Phase C carried them: they are
not ours to tidy. `CIQWBGuid` / `CIQWBInfo` are workbook-global and only one
survives a merge, so they are re-stamped explicitly from the cap table's values
(the status quo: every deal today copies a template carrying a fixed GUID).
"""

from __future__ import annotations

import argparse
import shutil
import sys
import tempfile
from pathlib import Path

from _excel_com import disable_autosave, excel_com_app

REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = REPO_ROOT / "infor-beta" / "scripts"
TEMPLATES = REPO_ROOT / "infor-beta" / "templates"
sys.path.insert(0, str(SCRIPTS))

OUTPUT_NAME = "INFOR Deal Workbook Template.xlsx"

from deal_workbook import (  # noqa: E402 — after the sys.path insert above
    TAB_BLOOMBERG_OUTPUT,
    TAB_CAPTABLE,
    TAB_COMPS,
    TAB_OWNERSHIP,
    TAB_PRECEDENTS,
)
from template_layout import (  # noqa: E402
    CAP_TABLE_SOURCE_SHEET,
    CAP_TABLE_TEMPLATE,
    COMPS_SOURCE_SHEET,
    COMPS_TEMPLATE,
    INFOR_THEME,
    OWNERSHIP_BBG_SOURCE_SHEET,
    OWNERSHIP_SOURCE_SHEET,
    OWNERSHIP_TEMPLATE,
    PRECEDENTS_SOURCE_SHEET,
    PRECEDENTS_TEMPLATE,
    THEME_COLOR_SLOTS,
    WORKBOOK_THEME_PART,
    read_theme,
    read_theme_part,
)

# (source template, source sheets to copy together, destination tab names).
#
# This table IS the rename — the single place a `*_SOURCE_SHEET` and a `TAB_*`
# legitimately meet — so both sides come from their constants rather than from
# literals. Spelling either side by hand here is how the two drift, and the
# v0.5.45 outage was exactly that drift going unnoticed: `Cap with Links` ->
# `captable` was recorded here and nowhere the assemblers could see.
#
# `Ownership` and `Bloomberg Output` are copied in ONE operation because the
# Ownership sheet's XLOOKUP rows reference `Bloomberg Output`; copying them
# separately would rebind that reference to the source workbook as an external
# link — the failure mode the aggregator's order-dependent merge suffered.
SHEET_PLAN: tuple[tuple[str, tuple[str, ...], tuple[str, ...]], ...] = (
    (CAP_TABLE_TEMPLATE, (CAP_TABLE_SOURCE_SHEET,), (TAB_CAPTABLE,)),
    # `__snloffice` is CapIQ's very-hidden metadata sheet. It rides along in the
    # copy and is deleted from the assembled workbook afterwards (DROP_SHEETS):
    # copying `Comps` alone made Excel record an external-workbook link back to
    # the comps template, listing both of the source's sheets.
    (COMPS_TEMPLATE, ("__snloffice", COMPS_SOURCE_SHEET), ("__snloffice", TAB_COMPS)),
    (
        OWNERSHIP_TEMPLATE,
        (OWNERSHIP_SOURCE_SHEET, OWNERSHIP_BBG_SOURCE_SHEET),
        (TAB_OWNERSHIP, TAB_BLOOMBERG_OUTPUT),
    ),
    (PRECEDENTS_TEMPLATE, (PRECEDENTS_SOURCE_SHEET,), (TAB_PRECEDENTS,)),
)

# CapIQ's very-hidden helper sheet — dropped, as the aggregator dropped it. No
# comps formula references it (verified over the shipped template).
DROP_SHEETS = ("__snloffice",)


def _expected_infor_names() -> dict[str, dict[str, str]]:
    """`{dest tab: {infor_name: target}}`, read off the SOURCE templates.

    The source of truth stays `template_layout.TEMPLATE_NAMED_RANGES` (the same
    registry `add_template_named_ranges.py` stamps and the writers resolve
    through), so this tool cannot expect a name the code does not, or miss one
    it does.
    """
    from template_layout import TEMPLATE_NAMED_RANGES

    expected: dict[str, dict[str, str]] = {}
    for template, sheets, dest_tabs in SHEET_PLAN:
        by_sheet = TEMPLATE_NAMED_RANGES.get(template, {})
        for sheet, dest in zip(sheets, dest_tabs):
            if by_sheet.get(sheet):
                expected[dest] = dict(by_sheet[sheet])
    return expected


def _xll_formulas(path: Path) -> dict[tuple[str, str], str]:
    """`{(sheet, cell): formula}` for every `_xll.` CapIQ formula in a workbook."""
    from openpyxl import load_workbook

    out: dict[tuple[str, str], str] = {}
    wb = load_workbook(path)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    text = value if isinstance(value, str) else getattr(value, "text", None)
                    if isinstance(text, str) and "_xll." in text:
                        out[(ws.title, cell.coordinate)] = text
    finally:
        wb.close()
    return out


def _copy_sheets(src, sheets: tuple[str, ...], dest) -> None:
    """Copy `sheets` from `src` onto the end of `dest`.

    Passing all of `sheets` in ONE call is load-bearing for the ownership pair:
    the `Ownership` sheet's XLOOKUP rows reference `Bloomberg Output`, and
    copying them separately would rebind that reference to the source workbook
    as an external link — the failure mode the aggregator's order-dependent
    merge suffered. Correct marshalling of the `After` keyword depends on the
    early binding `_excel_com.excel_com_app` guarantees.

    A **very-hidden** sheet is silently left out of a multi-sheet copy (measured:
    asking for `__snloffice` + `Comps` delivered only `Comps`), so any such sheet
    is made visible in the source first. `src` is always a staged COPY of the
    shipped template, so this cannot reach the original even if Excel saves it.
    """
    for name in sheets:
        sheet = src.Worksheets(name)
        if sheet.Visible != -1:  # xlSheetVisible
            sheet.Visible = -1
    selector = list(sheets) if len(sheets) > 1 else sheets[0]
    src.Worksheets(selector).Copy(After=dest.Sheets(dest.Sheets.Count))


def build(output: Path) -> None:
    """Assemble the deal-workbook template with an add-in-free automation Excel."""
    missing = [t for t, _s, _d in SHEET_PLAN if not (TEMPLATES / t).is_file()]
    if missing:
        raise SystemExit(f"source template(s) not found: {', '.join(missing)}")

    # NEVER let Excel open a shipped template in place. The repo lives in a
    # OneDrive-synced folder, where Excel 365 enables **AutoSave** by default —
    # and AutoSave writes the file back regardless of `Close(SaveChanges=False)`.
    # An early revision of this tool did exactly that and silently re-saved all
    # four source templates through Excel: `sharedStrings.xml` and
    # `calcChain.xml` appeared, `styles.xml` and `printerSettings1.bin` changed,
    # undoing the byte-level preservation Phase C's zip surgery went to real
    # trouble for. Working from copies in a temp directory puts the sources
    # outside OneDrive's reach entirely, which no in-Excel setting can undo.
    with tempfile.TemporaryDirectory() as staging:
        staged = {}
        for template, _sheets, _dest in SHEET_PLAN:
            staged[template] = Path(staging) / template
            shutil.copyfile(TEMPLATES / template, staged[template])
        _build_from(staged, output)

    stripped = _strip_orphan_external_links(output)
    if stripped:
        print(f"  stripped {stripped} vestigial external-link part(s)")
    _restamp_global_capiq_names(output)
    # Last, so nothing downstream can drop the theme part again: openpyxl carries
    # a loaded theme through a save, but "carries it through" is a property of
    # openpyxl, not something this tool should depend on for the palette.
    _ensure_infor_theme(output)


def _build_from(staged: dict[str, Path], output: Path) -> None:
    """Assemble `output` from the STAGED template copies (never the originals)."""
    with excel_com_app(purpose="deal-workbook template assembly") as excel:
        excel.ScreenUpdating = False
        dest = excel.Workbooks.Add()
        try:
            # Belt and braces: the automation instance does not load the SNL
            # add-ins (measured), but if a future Excel changes that, do not let
            # CapIQ recalculate its way through the copy. Settable only once a
            # workbook exists.
            excel.Calculation = -4135  # xlCalculationManual
            for template, sheets, dest_tabs in SHEET_PLAN:
                before = {dest.Sheets(i).Name for i in range(1, dest.Sheets.Count + 1)}
                src = excel.Workbooks.Open(str(staged[template]), UpdateLinks=0)
                try:
                    disable_autosave(src)
                    _copy_sheets(src, sheets, dest)
                finally:
                    src.Close(SaveChanges=False)

                # Rename by identifying which sheets are new, in workbook order —
                # index arithmetic off a pre-copy anchor is not reliable here.
                arrived = [
                    (i, dest.Sheets(i).Name)
                    for i in range(1, dest.Sheets.Count + 1)
                    if dest.Sheets(i).Name not in before
                ]
                if len(arrived) != len(dest_tabs):
                    raise SystemExit(
                        f"{template}: expected {len(dest_tabs)} new sheet(s), "
                        f"found {[n for _i, n in arrived]}"
                    )
                for (index, _old), dest_tab in zip(arrived, dest_tabs):
                    dest.Sheets(index).Name = dest_tab

            # Drop the blank sheet `Workbooks.Add` created, plus CapIQ's helper.
            keep = {tab for _t, _s, tabs in SHEET_PLAN for tab in tabs}
            for index in range(dest.Sheets.Count, 0, -1):
                sheet = dest.Sheets(index)
                if sheet.Name not in keep or sheet.Name in DROP_SHEETS:
                    sheet.Visible = -1  # a very-hidden sheet cannot be deleted
                    sheet.Delete()

            # INFOR's theme, over the Office default `Workbooks.Add()` gave us.
            # Deliberately AFTER the copies rather than before: the copy's
            # theme-colour behaviour is measured as it stands (the slot numbers
            # come across and resolve against the destination), and re-themeing
            # the destination mid-copy would put an unmeasured variable inside
            # the one operation this tool exists to get right. A raise here is
            # not fatal — `_ensure_infor_theme` grafts the part into the saved
            # zip instead — so a headless ApplyTheme failure costs a warning.
            try:
                dest.ApplyTheme(str(TEMPLATES / INFOR_THEME))
                print(f"  applied {INFOR_THEME}")
            except Exception as exc:  # noqa: BLE001
                print(f"  ApplyTheme({INFOR_THEME}) raised: {exc} — will graft the part")

            output.parent.mkdir(parents=True, exist_ok=True)
            if output.exists():
                output.unlink()
            dest.SaveAs(str(output), FileFormat=51)  # xlOpenXMLWorkbook
        finally:
            dest.Close(SaveChanges=False)


def _ensure_infor_theme(output: Path) -> None:
    """Confirm the saved workbook carries INFOR's palette; graft it if not.

    `Workbook.ApplyTheme` is the native path and the one that also updates the
    Normal style's font, so it is tried first (in `_build_from`). This is the
    check that it took, plus the fallback for a headless Excel where it did not:
    replace `xl/theme/theme1.xml` with the thmx's theme part, which is the same
    bytes Excel writes when the theme is applied by hand
    (md5 `2eb1697978066f80ed1a4ad3b6c64aec`).

    Zip surgery in `add_template_named_ranges.py`'s idiom — the theme part is
    rewritten and every other entry's payload copied through byte-for-byte. The
    part name does not change, so `[Content_Types].xml` and the workbook rels
    already describe it correctly and are left alone.
    """
    import zipfile

    theme = TEMPLATES / INFOR_THEME
    wanted = read_theme(theme)
    if read_theme(output).palette == wanted.palette:
        print(f"  theme palette is {wanted.color_scheme!r} (accent1 {wanted.palette['accent1']})")
        return

    payload = read_theme_part(theme)
    with zipfile.ZipFile(output) as zf:
        entries = [(item, zf.read(item.filename)) for item in zf.infolist()]
    if not any(item.filename == WORKBOOK_THEME_PART for item, _ in entries):
        raise SystemExit(
            f"{output.name} has no {WORKBOOK_THEME_PART} to replace — Excel writes "
            f"one into every workbook, so this file is not what this tool built."
        )
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as out:
        for item, entry in entries:
            out.writestr(item, payload if item.filename == WORKBOOK_THEME_PART else entry)

    got = read_theme(output)
    if got.palette != wanted.palette:  # pragma: no cover - the graft is a copy
        raise SystemExit(f"grafting {INFOR_THEME}'s theme part did not take: {got.palette}")
    print(f"  grafted {INFOR_THEME}'s theme part ({len(payload):,} bytes)")


def _strip_orphan_external_links(output: Path) -> int:
    """Remove vestigial external-workbook link parts. Returns how many went.

    Copying a sheet between workbooks makes Excel record an `externalBook` entry
    describing the source — here a link back to `INFOR Comps Template.xlsx`,
    listing both of its sheets, complete with a SharePoint absolute URL. Nothing
    references it: no sheet formula and no defined name contains a `[n]` index
    (asserted below, and the build fails rather than strips if that ever stops
    being true). Left in place it would make Excel offer to "update links" on
    every open of a client-facing workbook, and would leak the author's OneDrive
    path into the deal file.

    Zip surgery in `add_template_named_ranges.py`'s idiom: rewrite only the three
    parts that must change and copy every other entry's payload through
    byte-for-byte. Deliberately not done with openpyxl, whose external-link
    handling is exactly what is in question.
    """
    import re
    import zipfile

    with zipfile.ZipFile(output) as zf:
        entries = [(item, zf.read(item.filename)) for item in zf.infolist()]

    link_parts = [
        item.filename for item, _payload in entries if item.filename.startswith("xl/externalLinks/")
    ]
    if not link_parts:
        return 0

    # Refuse to strip a link something actually uses.
    referenced: list[str] = []
    for item, payload in entries:
        if item.filename.startswith("xl/worksheets/") or item.filename == "xl/workbook.xml":
            text = payload.decode("utf-8", "replace")
            if re.search(r"\[\d+\]", text):
                referenced.append(item.filename)
    if referenced:
        raise SystemExit(
            "external-workbook references are in USE by "
            f"{', '.join(referenced)} — refusing to strip them. The sheet copy "
            "rebound a formula to a source template; fix the copy, do not "
            "delete the link."
        )

    def rewrite(name: str, payload: bytes) -> bytes:
        text = payload.decode("utf-8")
        if name == "xl/workbook.xml":
            text = re.sub(r"<externalReferences>.*?</externalReferences>", "", text, flags=re.S)
        elif name == "xl/_rels/workbook.xml.rels":
            text = re.sub(
                r'<Relationship[^>]*Type="[^"]*/externalLink"[^>]*/>', "", text
            )
        elif name == "[Content_Types].xml":
            text = re.sub(r'<Override PartName="/xl/externalLinks/[^"]*"[^>]*/>', "", text)
        return text.encode("utf-8")

    touched = ("xl/workbook.xml", "xl/_rels/workbook.xml.rels", "[Content_Types].xml")
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as out:
        for item, payload in entries:
            if item.filename in link_parts:
                continue
            out.writestr(item, rewrite(item.filename, payload) if item.filename in touched else payload)
    return len(link_parts)


def _restamp_global_capiq_names(output: Path) -> None:
    """Re-add workbook-global CapIQ names the sheet copy did not carry.

    Excel copies the global names a copied sheet's formulas *reference*, so the
    `IQ_*` period constants come along. `CIQWBGuid` / `CIQWBInfo` are pure
    add-in metadata that no formula references, so they do not — and they are
    the two names the add-in uses to identify a workbook. Restamped from the
    cap table's values, keeping today's behaviour (every deal already copies a
    template carrying a fixed GUID).
    """
    from openpyxl import load_workbook
    from openpyxl.workbook.defined_name import DefinedName

    cap = load_workbook(TEMPLATES / "INFOR Cap Table Template.xlsx")
    comps = load_workbook(TEMPLATES / "INFOR Comps Template.xlsx")
    wanted: dict[str, str] = {}
    for source in (comps, cap):  # cap table wins on the two identity names
        for name, defined in source.defined_names.items():
            wanted[name] = defined.value

    wb = load_workbook(output)
    added = 0
    for name, value in wanted.items():
        if name not in wb.defined_names:
            wb.defined_names.add(DefinedName(name, attr_text=value))
            added += 1
    if added:
        wb.save(output)
    print(f"  re-stamped {added} workbook-global CapIQ name(s)")


def verify(output: Path, *, with_excel: bool) -> list[str]:
    """Verify the assembled template. Returns a list of problems (empty = good)."""
    from openpyxl import load_workbook

    from template_layout import defined_name_ref, normalize_ref

    problems: list[str] = []
    if not output.is_file():
        return [f"{output.name} does not exist — run without --check first"]

    wb = load_workbook(output)

    # 1. Every tab present, and nothing unexpected.
    expected_tabs = [
        tab for _t, _s, tabs in SHEET_PLAN for tab in tabs if tab not in DROP_SHEETS
    ]
    if wb.sheetnames != expected_tabs:
        problems.append(f"tabs are {wb.sheetnames}, expected {expected_tabs}")

    # 2. Every `infor_` name resolves to the same target it had in its source.
    expected_names = _expected_infor_names()
    total = 0
    for tab, names in expected_names.items():
        if tab not in wb.sheetnames:
            problems.append(f"tab {tab!r} missing, so its {len(names)} infor_ names are gone")
            continue
        ws = wb[tab]
        for name, target in names.items():
            total += 1
            ref = defined_name_ref(ws, name)
            if ref is None:
                problems.append(f"{tab}: defined name {name!r} does not resolve")
            elif normalize_ref(ref) != normalize_ref(target):
                problems.append(
                    f"{tab}: {name!r} resolves to {ref} but the source template had {target}"
                )
    print(f"  checked {total} infor_ defined names across {len(expected_names)} tabs")

    # 3. The two CapIQ identity names survived.
    for name in ("CIQWBGuid", "CIQWBInfo"):
        if name not in wb.defined_names:
            problems.append(f"workbook-global CapIQ name {name!r} is missing")

    # 4. Every `_xll.` CapIQ formula's text is unchanged from its source.
    dest_formulas = _xll_formulas(output)
    src_total = 0
    for template, sheets, dest_tabs in SHEET_PLAN:
        src_formulas = _xll_formulas(TEMPLATES / template)
        rename = {
            src_sheet: dest_tab
            for src_sheet, dest_tab in zip(sheets, dest_tabs)
            if dest_tab not in DROP_SHEETS
        }
        for (sheet, cell), formula in src_formulas.items():
            if sheet not in rename:
                continue  # a dropped helper sheet
            src_total += 1
            got = dest_formulas.get((rename[sheet], cell))
            if got is None:
                problems.append(f"{rename[sheet]}!{cell}: CapIQ formula lost")
            elif got != formula:
                problems.append(
                    f"{rename[sheet]}!{cell}: CapIQ formula changed\n"
                    f"      source: {formula}\n"
                    f"      built:  {got}"
                )
    print(f"  checked {src_total} _xll. CapIQ formulas")

    # 5. No external-workbook references back to the source templates.
    external = _external_refs(output)
    if external:
        problems.append(f"external-workbook references present: {external[:5]}")

    # 6. The palette every theme-indexed colour resolves against is INFOR's.
    problems.extend(_theme_problems(output))

    if with_excel:
        problems.extend(_verify_with_excel(output))
    return problems


def _theme_problems(output: Path) -> list[str]:
    """Fail the build unless the assembled workbook carries INFOR's theme.

    The palette, not the theme's name: all four source templates are named
    "Office Theme" while carrying the INFOR colour scheme, so a name assertion
    would be both wrong and easy to satisfy. `INFORFG.thmx` is the expectation
    rather than a constant copied out of it — there is then no second place for
    the brand palette to be declared, and no way for this check to disagree with
    what the build stamps. The fonts come with it: an Office-themed workbook
    renders the source templates' scheme-linked fonts as Aptos Narrow.
    """
    problems: list[str] = []
    theme = TEMPLATES / INFOR_THEME
    if not theme.is_file():
        return [f"{INFOR_THEME} is missing from templates/ — nothing to verify the palette against"]

    wanted = read_theme(theme)
    got = read_theme(output)
    drift = [
        f"{slot} {got.palette.get(slot, '(absent)')} != {wanted.palette[slot]}"
        for slot in THEME_COLOR_SLOTS
        if got.palette.get(slot) != wanted.palette.get(slot)
    ]
    if drift:
        problems.append(
            f"theme palette is not INFOR's — {got.name!r} / {got.color_scheme!r} "
            f"({', '.join(drift)}). Every theme-indexed colour in every tab resolves "
            f"against this, so the workbook renders in the wrong palette with nothing "
            f"in any cell wrong. Re-run this tool (it applies {INFOR_THEME})."
        )
    for label, mine, theirs in (
        ("major", got.major_font, wanted.major_font),
        ("minor", got.minor_font, wanted.minor_font),
    ):
        if mine != theirs:
            problems.append(f"theme {label} font is {mine!r}, expected {theirs!r} (from {INFOR_THEME})")
    if not problems:
        print(
            f"  theme palette is {got.color_scheme!r} from {INFOR_THEME} "
            f"(accent1 {got.palette['accent1']}, minor font {got.minor_font})"
        )
    return problems


def _external_refs(path: Path) -> list[str]:
    """External-workbook reference parts / `[n]`-indexed formulas in a workbook."""
    import re
    import zipfile

    found: list[str] = []
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            if name.startswith("xl/externalLinks/") and name.endswith(".xml"):
                found.append(name)
        for name in zf.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                text = zf.read(name).decode("utf-8", "replace")
                for match in re.finditer(r"\[\d+\]", text):
                    found.append(f"{name}: {match.group(0)}")
                    break
    return found


def _verify_with_excel(output: Path) -> list[str]:
    """Excel oracle: open the built template; fail if Excel will not have it.

    Same oracle `add_template_named_ranges.py --verify-excel` uses, and for the
    same reason: openpyxl reads a file Excel might still refuse. Opened as a copy
    in a temp dir, so AutoSave cannot write back to the repo.
    """
    problems: list[str] = []
    try:
        with excel_com_app(purpose="deal-workbook template verification") as excel:
            with tempfile.TemporaryDirectory() as tmp:
                probe = Path(tmp) / output.name
                shutil.copyfile(output, probe)
                wb = excel.Workbooks.Open(str(probe), UpdateLinks=0, CorruptLoad=0)
                try:
                    names = {n.Name for n in wb.Names}
                    for required in ("CIQWBGuid", "CIQWBInfo"):
                        if required not in names:
                            problems.append(f"Excel: {required!r} absent after open")
                    print(f"  Excel opened it: {wb.Sheets.Count} sheets, {len(names)} names")
                finally:
                    wb.Close(SaveChanges=False)
    except Exception as exc:  # noqa: BLE001
        problems.append(f"Excel could not open the built template: {exc}")
    return problems


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--check", action="store_true", help="verify only; write nothing")
    parser.add_argument(
        "--verify", action="store_true", help="after building, also run the Excel oracle"
    )
    args = parser.parse_args()

    output = TEMPLATES / OUTPUT_NAME

    if not args.check:
        if sys.platform != "win32":
            print("building needs Excel COM, so a Windows box; --check works anywhere.")
            return 2
        print(f"building {output.name} from {len(SHEET_PLAN)} source templates...")
        build(output)
        size = output.stat().st_size
        print(f"  wrote {output} ({size:,} bytes)")

    print("verifying...")
    problems = verify(output, with_excel=args.verify and sys.platform == "win32")
    if problems:
        print(f"\n{len(problems)} PROBLEM(S):")
        for problem in problems:
            print(f"  - {problem}")
        return 1
    print("\nall checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
