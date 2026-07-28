"""Stamp the shipped workbook templates with their `infor_` defined names.

One-time prep tooling for Phase C, kept in the repo because it has to be
re-runnable: any time an analyst re-saves one of the four templates from Excel,
run this again to restore the names the writers resolve through. It is **not**
shipped plugin code and is therefore exempt from Phase D's COM deletion — it
does not use COM at all.

Usage (from the repo root):

    python tools/add_template_named_ranges.py --check     # report, write nothing
    python tools/add_template_named_ranges.py             # stamp the templates
    python tools/add_template_named_ranges.py --verify-excel   # + Excel repair oracle

What the names are and where they come from
-------------------------------------------
`template_layout.TEMPLATE_NAMED_RANGES` — derived from the same `CellAnchor`
declarations the writers and the sentinel cross-check use, so this tool cannot
stamp a name the code does not expect, or miss one it does.

Why direct XML surgery, and not openpyxl or Excel COM
-----------------------------------------------------
The brief asked for Excel COM in preference to openpyxl, on the grounds that
openpyxl round-trip damage to these exact files is documented (the ownership
template needed cruft-stripping to stay Excel-openable; the precedents template
shed ~58.7k defined names and 174 external links during prep). This tool does
neither, because a third option is strictly more conservative than both:
rewrite the one XML part that has to change — `xl/workbook.xml` — and copy every
other zip entry's payload through byte-for-byte, with its original name, order,
compression type and timestamp.

(Two purely informational container fields are not preserved, because Python's
zip writer sets them itself: the deflate-level hint in the general-purpose flag
bits, and the Unix permission field, which a Windows-authored archive leaves at
zero. Neither is read by an OOXML consumer. Verified after stamping: in all
four templates exactly one entry's payload differs, `xl/workbook.xml`.)

The stamped files' **on-disk size changes** — the cap table 26,044 -> 21,663
bytes, comps 57,124 -> 48,160 — which looks alarming in a `git diff --stat` and
is not data loss. Every entry is re-deflated at Python's compression level rather
than Excel's, so the *compressed* representation differs while the decompressed
payload is identical, entry for entry. Re-running the tool is a no-op on content:
it strips its own previous `infor_` entries first, so the name count stays put.

Measured, on the shipped templates:

  - **openpyxl** preserves the defined names (34 -> 34 on the cap table,
    1,246 -> 1,246 on comps, `CIQWBGuid` included) but silently **drops
    `xl/printerSettings/printerSettings1.bin`** from both. Not fatal, and the
    runtime builders already round-trip these files, but there is no reason for
    a prep pass to lose the analyst's page setup.
  - **Excel COM** would rewrite the whole file through Excel, which means
    opening a Capital IQ workbook with the Cap IQ Office Tools add-in loaded
    and letting it touch `CIQWBGuid` / `CIQWBInfo` on the way past. The add-in
    is the delivery path this phase is explicitly told not to disturb.
  - **This tool** changes exactly one element of one part. Nothing else in the
    archive can differ, because nothing else is re-encoded.

Excel COM does still get used — as the **oracle**, under `--verify-excel`,
which is the check that actually matters: open each stamped template in Excel,
confirm no repair record was written, and confirm Excel itself resolves every
new name to the address the code expects.

Why the names are worksheet-scoped
----------------------------------
A workbook-scoped name travels with its sheet through the aggregator's Excel-COM
merge and lands in the destination's global namespace, which is how phantom
`[1]!name` external-workbook aliases appear — the failure that already blocks an
order-dependent sheet rename in `workbook_aggregator`. Worksheet-scoped names
travel with the sheet without ever entering that namespace. They are also
unambiguous in a dirty namespace on their own merits: the comps template carries
1,246 legacy names.
"""

from __future__ import annotations

import argparse
import pathlib
import re
import shutil
import sys
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "infor-beta" / "scripts"
TEMPLATES = REPO_ROOT / "infor-beta" / "templates"
sys.path.insert(0, str(SCRIPTS))

from openpyxl.utils import absolute_coordinate, quote_sheetname  # noqa: E402

from template_layout import TEMPLATE_NAMED_RANGES  # noqa: E402

WORKBOOK_PART = "xl/workbook.xml"
NAME_PREFIX = "infor_"

# `<definedNames>` sits between `<externalReferences>` and `<calcPr>` in the
# CT_Workbook sequence. Both shapes occur across the four templates: the two
# Excel-authored ones (cap table, comps) carry a populated element, the two
# openpyxl-authored ones (ownership, precedents) carry `<definedNames/>`.
_EMPTY_ELEMENT = re.compile(r"<definedNames\s*/>")
_CLOSE_TAG = "</definedNames>"


def _sheet_index(workbook_xml: str, sheet: str) -> int:
    """The sheet's position in `<sheets>`, which is what `localSheetId` indexes.

    Not the `sheetId` attribute — that is a stable internal id and is routinely
    different (the cap table's only sheet has `sheetId="7"` at position 0).
    """
    sheets = re.findall(r"<sheet\b[^>]*?\bname=\"([^\"]+)\"[^>]*>", workbook_xml)
    if sheet not in sheets:
        raise KeyError(f"sheet {sheet!r} not found in workbook.xml (have {sheets})")
    return sheets.index(sheet)


def _defined_name_xml(name: str, sheet: str, target: str, local_sheet_id: int) -> str:
    ref = f"{quote_sheetname(sheet)}!{absolute_coordinate(target)}"
    return (
        f'<definedName name="{escape(name)}" localSheetId="{local_sheet_id}">'
        f"{escape(ref)}</definedName>"
    )


def _strip_existing(workbook_xml: str) -> tuple[str, int]:
    """Drop any previously stamped `infor_` names so a re-run replaces them."""
    pattern = re.compile(
        r'<definedName\b[^>]*\bname="' + re.escape(NAME_PREFIX) + r'[^"]*"[^>]*>.*?</definedName>',
        re.S,
    )
    stripped, count = pattern.subn("", workbook_xml)
    return stripped, count


def stamp_workbook_xml(workbook_xml: str, sheet_names: dict[str, dict[str, str]]) -> str:
    """Return `workbook.xml` with the template's `infor_` names spliced in.

    Additive by construction: the only edit is inserting `<definedName>`
    children (and removing our own from a previous run). Every other byte of
    the part is left exactly as Excel or openpyxl wrote it.
    """
    updated, _ = _strip_existing(workbook_xml)
    entries = [
        _defined_name_xml(name, sheet, target, _sheet_index(updated, sheet))
        for sheet, targets in sheet_names.items()
        for name, target in sorted(targets.items())
    ]
    if not entries:
        return updated
    block = "".join(entries)
    if _EMPTY_ELEMENT.search(updated):
        return _EMPTY_ELEMENT.sub(f"<definedNames>{block}</definedNames>", updated, count=1)
    if _CLOSE_TAG in updated:
        return updated.replace(_CLOSE_TAG, f"{block}{_CLOSE_TAG}", 1)
    # No element at all: insert one in its schema position, after </sheets>.
    if "</sheets>" not in updated:
        raise ValueError("workbook.xml has no <sheets> element to anchor <definedNames> after")
    return updated.replace("</sheets>", f"</sheets><definedNames>{block}</definedNames>", 1)


def stamp_template(path: Path, sheet_names: dict[str, dict[str, str]]) -> None:
    """Rewrite `path` in place with the names added.

    Every zip entry other than `xl/workbook.xml` is copied with its original
    bytes, compression type and timestamp — the archive differs in exactly one
    part.
    """
    with zipfile.ZipFile(path) as src:
        infos = src.infolist()
        payloads = {info.filename: src.read(info.filename) for info in infos}
    if WORKBOOK_PART not in payloads:
        raise KeyError(f"{path.name} has no {WORKBOOK_PART}")

    payloads[WORKBOOK_PART] = stamp_workbook_xml(
        payloads[WORKBOOK_PART].decode("utf-8"), sheet_names
    ).encode("utf-8")

    tmp = path.with_suffix(path.suffix + ".stamping")
    with zipfile.ZipFile(tmp, "w") as out:
        for info in infos:
            # Reuse the original ZipInfo so date_time / compress_type /
            # external_attr survive; only the payload of one part changes.
            out.writestr(info, payloads[info.filename])
    shutil.move(str(tmp), str(path))


# ─── Verification ────────────────────────────────────────────────────────────


def check_with_openpyxl(path: Path, sheet_names: dict[str, dict[str, str]]) -> list[str]:
    """Resolve every name through openpyxl; return a list of problems."""
    from openpyxl import load_workbook

    from template_layout import defined_name_ref, normalize_ref

    problems: list[str] = []
    wb = load_workbook(path)
    try:
        for sheet, targets in sheet_names.items():
            if sheet not in wb.sheetnames:
                problems.append(f"sheet {sheet!r} missing")
                continue
            ws = wb[sheet]
            for name, target in sorted(targets.items()):
                found = defined_name_ref(ws, name)
                if found is None:
                    problems.append(f"{sheet}!{name}: not resolvable")
                elif found != normalize_ref(target):
                    problems.append(f"{sheet}!{name}: resolves to {found}, expected {target}")
    finally:
        wb.close()
    return problems


def _canonical_refers_to(refers_to: str) -> tuple[str, str]:
    """``(sheet, ref)`` from an Excel ``RefersTo`` string, quoting normalised.

    Excel canonicalises a name's destination on read: it keeps the single
    quotes around ``'Cap with Links'`` (a space forces them) and drops them
    around ``Ownership`` (they are optional). Comparing the raw string would
    report every unquoted sheet as a mismatch, so both sides are normalised to
    the unquoted sheet name plus the absolute reference.
    """
    body = refers_to.lstrip("=")
    sheet, _, ref = body.rpartition("!")
    return sheet.strip("'").replace("''", "'"), ref.upper()


def _repair_logs() -> set[Path]:
    """Excel's repair-record files in the user's temp dir.

    Excel writes `error<NNNNN>_<NN>.xml` when it repairs a workbook it *can*
    repair. Snapshotting before and after the open catches that case; the
    primary signal is the Open call itself (see `verify_with_excel`).
    """
    import tempfile

    return set(Path(tempfile.gettempdir()).glob("error*.xml"))


def verify_with_excel(paths_and_names) -> list[str]:
    """Open each template in Excel; return problems (repairs, unresolved names).

    The oracle for "opens with no repair prompt, pre-existing names intact".
    Windows + Excel only; used interactively during prep, never by the plugin.

    Three signals, primary first:

    1. **The Open call succeeds.** With `DisplayAlerts = False` Excel cannot put
       up the "We found a problem with some content" prompt, so a workbook that
       would have prompted fails `Workbooks.Open` instead. Self-tested during
       Phase C prep against two deliberately damaged copies of the precedents
       template — an unparseable `definedName` reference, and the v0.5.23
       LibreOffice `~` union operator inside a sheet formula — and both were
       refused, while all four stamped templates opened.
    2. **No new repair record** appears in the temp dir, for damage Excel
       repairs silently rather than refusing.
    3. **Excel's own view of the names** — the total count has not dropped
       below what was there before stamping (nothing was repair-stripped), and
       every new name resolves to the address the code expects.
    """
    # `excel_to_powerpoint.excel_com_app` is gone (Phase D deleted every COM path
    # from the shipped plugin). COM now lives only in `tools/`, with one owner.
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
    from _excel_com import excel_com_app

    problems: list[str] = []
    before = _repair_logs()
    with excel_com_app(purpose="template named-range verification", visible=False) as excel:
        for path, sheet_names, expected_existing in paths_and_names:
            try:
                wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=True)
            except Exception as exc:
                problems.append(
                    f"{path.name}: Excel refused to open the file ({exc}). With alerts "
                    f"suppressed this is what a repair prompt looks like — the workbook "
                    f"is damaged."
                )
                continue
            try:
                total = wb.Names.Count
                if total < expected_existing:
                    problems.append(
                        f"{path.name}: Excel reports {total} defined names, fewer than the "
                        f"{expected_existing} expected — names were repair-stripped"
                    )
                for sheet, targets in sheet_names.items():
                    for name, target in sorted(targets.items()):
                        try:
                            refers = str(wb.Worksheets(sheet).Names(name).RefersTo)
                        except Exception as exc:
                            problems.append(f"{path.name}: Excel cannot resolve {name!r} ({exc})")
                            continue
                        want = (sheet, absolute_coordinate(target).upper())
                        if _canonical_refers_to(refers) != want:
                            problems.append(
                                f"{path.name}: Excel resolves {name!r} to {refers!r}, "
                                f"expected {want[0]}!{want[1]}"
                            )
            finally:
                wb.Close(SaveChanges=False)
    new_logs = _repair_logs() - before
    if new_logs:
        problems.append(
            "Excel wrote repair records while opening the templates: "
            + ", ".join(sorted(p.name for p in new_logs))
        )
    return problems


def _existing_name_count(path: Path) -> int:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read(WORKBOOK_PART).decode("utf-8")
    return len(re.findall(r"<definedName\b", xml))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--check", action="store_true", help="report only; write nothing")
    parser.add_argument(
        "--verify-excel",
        action="store_true",
        help="after stamping, open each template in Excel and check for repair records",
    )
    args = parser.parse_args()

    excel_targets = []
    failures = 0
    for template, sheet_names in TEMPLATE_NAMED_RANGES.items():
        path = TEMPLATES / template
        if not path.exists():
            print(f"MISSING  {template}")
            failures += 1
            continue
        before = _existing_name_count(path)
        wanted = sum(len(t) for t in sheet_names.values())
        if not args.check:
            stamp_template(path, sheet_names)
        after = _existing_name_count(path)
        problems = check_with_openpyxl(path, sheet_names)
        status = "OK " if not problems else "FAIL"
        print(f"{status} {template}: {before} -> {after} defined names (+{wanted} infor_)")
        for problem in problems:
            print(f"       {problem}")
        failures += bool(problems)
        # Pre-existing count = whatever was there before minus any infor_ names
        # a previous run had already stamped.
        excel_targets.append((path, sheet_names, after - wanted))

    if args.verify_excel and not args.check:
        print("\nExcel repair oracle:")
        problems = verify_with_excel(excel_targets)
        if problems:
            failures += 1
            for problem in problems:
                print(f"  FAIL {problem}")
        else:
            print("  OK  every template opened with no repair record; Excel resolves every name")

    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
